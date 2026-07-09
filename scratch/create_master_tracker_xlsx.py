# -*- coding: utf-8 -*-
import os
import sys
import subprocess

# Auto-install openpyxl if not present
try:
    import openpyxl
except ImportError:
    print("openpyxl not found. Installing...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_master_tracker():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Master Tracker"

    # Show gridlines
    ws.views.sheetView[0].showGridLines = True

    # Styling definitions
    font_title = Font(name="Arial", size=15, bold=True, color="0D2240")
    font_header = Font(name="Arial", size=9.5, bold=True, color="FFFFFF")
    font_body = Font(name="Arial", size=9)
    font_body_bold = Font(name="Arial", size=9, bold=True)
    
    fill_header = PatternFill(start_color="0D2240", end_color="0D2240", fill_type="solid") # Deep Navy
    fill_gd1 = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid") # Light Green (Done GD 1)
    fill_gd2 = PatternFill(start_color="FFF4E5", end_color="FFF4E5", fill_type="solid") # Light Orange (GD 2 Plan)
    fill_gd3 = PatternFill(start_color="F1F3F4", end_color="F1F3F4", fill_type="solid") # Light Gray (GD 3 Plan)
    fill_zebra = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_title = Alignment(horizontal="center", vertical="center")
    
    border_thin = Side(border_style="thin", color="D1D5DB")
    border_thick = Side(border_style="medium", color="0D2240")
    border_double = Side(border_style="double", color="4B5563")
    
    box_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    header_border = Border(left=border_thin, right=border_thin, top=border_thick, bottom=border_thick)
    bottom_heavy_border = Border(bottom=border_double, left=border_thin, right=border_thin)

    # 1. Sheet Title block
    ws.merge_cells("A1:L2")
    title_cell = ws["A1"]
    title_cell.value = "BẢNG TIẾN ĐỘ THU THẬP & PHÂN TÍCH THƯ VIỆN HỆ CỬA NHÔM (MASTER TRACKER)"
    title_cell.font = font_title
    title_cell.alignment = align_title
    
    # 2. Header Row (Row 4)
    headers = [
        "STT", 
        'Nhóm Hệ', 
        "Tên hệ", 
        "Hãng tiêu biểu", 
        "Mã hệ", 
        "Loại sản phẩm", 
        "Catalog", 
        "CAD", 
        "Mặt cắt", 
        "Phụ kiện", 
        "Shopdrawing", 
        "Trạng thái"
    ]
    
    for col_idx, h_text in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = h_text
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = header_border

    # 3. Data definition (All systems grouped by phase and category)
    # Status levels:
    # "Đang thu thập" (GD1 - Yes/No fields filled accordingly)
    # "Kế hoạch GĐ 2" (GD2 - Yes/No fields marked empty/No)
    # "Kế hoạch GĐ 3" (GD3 - Yes/No fields marked empty/No)
    
    raw_data = []

    # --- PHASE 1 (Ưu tiên cao) ---
    # I. HỆ CỬA ĐI
    gd1_doors = [
        ("I. HỆ CỬA ĐI", "Cửa mở quay Casement - Hệ 45", "Xingfa, PMA, Viralwindow, Topal", "XF55, PMA55, Viralwindow VRA55", "Cửa đi mở quay thường", "Đang thu thập"),
        ("I. HỆ CỬA ĐI", "Cửa mở quay Casement - Hệ 50", "Schüco, Reynaers, Technal", "ADS50, CS59, Soleal", "Cửa đi mở quay cách nhiệt vừa", "Đang thu thập"),
        ("I. HỆ CỬA ĐI", "Cửa mở quay Casement - Hệ 55", "Xingfa, PMA, Viralwindow, Topal", "XF55, Platinum, Viralwindow VRA55", "Cửa đi mở quay tiêu chuẩn", "Đang thu thập"),
        ("I. HỆ CỬA ĐI", "Cửa mở quay Casement - Hệ 60", "Schüco, Reynaers, Cortizo", "ADS60, CS68, Millennium 60", "Cửa đi mở quay bản cánh vừa", "Đang thu thập"),
        ("I. HỆ CỬA ĐI", "Cửa mở quay Casement - Hệ 65", "Schüco, Reynaers, Technal", "ADS65, CS77, Soleal GY", "Cửa đi mở quay cao cấp", "Đang thu thập"),
        ("I. HỆ CỬA ĐI", "Cửa mở quay Casement - Hệ 70", "Schüco, Reynaers, Cortizo", "ADS70, ML8, Millennium 70", "Cửa đi mở quay bản cánh lớn", "Đang thu thập"),
        ("I. HỆ CỬA ĐI", "Cửa mở quay Casement - Hệ 75", "Schüco, Reynaers, Cortizo", "ADS75, ML8, Millennium 75", "Cửa đi mở quay Luxury", "Đang thu thập"),
        ("I. HỆ CỬA ĐI", "Cửa mở quay Casement - Hệ 80", "Schüco, Reynaers, Cortizo", "ADS80, ML8, Millennium 80", "Cửa đi mở quay bản siêu dày", "Đang thu thập"),
        ("I. HỆ CỬA ĐI", "Cửa mở quay Casement - Hệ 90", "Schüco, Reynaers, Wicona", "ADS90, ML10, Wicstyle 90", "Cửa đi mở quay siêu khủng", "Đang thu thập"),
        ("I. HỆ CỬA ĐI", "Cửa mở quay Casement - Cầu cách nhiệt", "Schüco, Reynaers, Cortizo, Viralwindow", "ADS 75.SI, ML8-HI, VRE65-TB", "Cửa đi mở quay cách âm cách nhiệt", "Đang thu thập"),
        ("I. HỆ CỬA ĐI", "Cửa mở quay Casement - Siêu kín khít", "Schüco, Reynaers, Technal", "ADS 90.SI, ML10-Passive", "Cửa đi kín khí Passive House", "Đang thu thập"),
        ("I. HỆ CỬA ĐI", "Cửa mở quay Casement - Chống cháy", "Aluprof, Wicona, Cortizo", "MB-78EI, Wicstyle 77FP", "Cửa đi chống cháy chuyên dụng", "Đang thu thập"),
        ("I. HỆ CỬA ĐI", "Cửa mở quay Casement - Chống đạn", "Schüco, Cortizo, Reynaers", "ADS 90.BR, Bullet-Proof", "Cửa đi chống đạn an ninh cao", "Đang thu thập"),
        ("I. HỆ CỬA ĐI", "Cửa mở quay Casement - Chống bão", "Schüco, Cortizo, Technal", "ADS 90.HD, High-Performance", "Cửa đi chịu áp lực gió bão biển", "Đang thu thập"),
        ("I. HỆ CỬA ĐI", "Cửa mở 2 chiều - Swing Door", "Schüco, Reynaers, Technal", "ADS 65 Swing, CS Swing", "Cửa đi mở 2 chiều bản lề sàn", "Đang thu thập"),
        ("I. HỆ CỬA ĐI", "Cửa mở 2 chiều - Double Swing", "Schüco, Reynaers, Technal", "ADS 65 HD Double, CS Swing", "Cửa đi 2 cánh mở 2 chiều", "Đang thu thập"),
        ("I. HỆ CỬA ĐI", "Cửa mở 2 chiều - Spring Door", "Xingfa, PMA, Viralwindow", "XF55 Spring, PMA Luxury", "Cửa bản lề lò xo tự đóng", "Đang thu thập"),
        ("I. HỆ CỬA ĐI", "Cửa Pivot - Pivot thường", "Schüco, Reynaers, Technal", "ADS 75 Pivot, CS Pivot", "Cửa xoay trục lệch tâm thường", "Đang thu thập"),
        ("I. HỆ CỬA ĐI", "Cửa Pivot - Pivot âm sàn", "Cortizo, Reynaers, Viralwindow", "Pivot XXL, ML Pivot, VR Pivot", "Cửa xoay trục lực tải âm sàn", "Đang thu thập"),
        ("I. HỆ CỬA ĐI", "Cửa Pivot - Pivot siêu lớn", "Cortizo, Reynaers, Schüco", "Pivot XXL, Hi-Finity Pivot", "Cửa xoay đại sảnh biệt thự", "Đang thu thập"),
        ("I. HỆ CỬA ĐI", "Cửa Pivot - Pivot biệt thự", "Cortizo, Reynaers, Schüco", "Pivot XXL, MasterLine Pivot", "Cửa chính biệt thự sang trọng", "Đang thu thập"),
        ("I. HỆ CỬA ĐI", "Cửa trượt Sliding - 2 cánh lùa", "Xingfa, PMA, Viralwindow, Topal", "XF93, PMA93, Viralwindow VRA94 / VRE120, Topal XF", "Cửa lùa 2 cánh trượt 2 ray", "Đang thu thập"),
        ("I. HỆ CỬA ĐI", "Cửa trượt Sliding - 3 cánh lùa", "Xingfa, PMA, Viralwindow, Topal", "XF93 3-Ray, PMA93 3-Ray", "Cửa lùa 3 cánh trượt 3 ray", "Đang thu thập"),
        ("I. HỆ CỬA ĐI", "Cửa trượt Sliding - 4 cánh lùa", "Xingfa, PMA, Viralwindow, Topal", "XF93 4-Cánh, PMA93 4-Cánh", "Cửa lùa 4 cánh trượt 2 ray", "Đang thu thập"),
        ("I. HỆ CỬA ĐI", "Cửa trượt Sliding - 6 cánh lùa", "Xingfa, PMA, Viralwindow", "XF93 6-Cánh, PMA93 6-Cánh", "Cửa lùa 6 cánh trượt 3 ray", "Đang thu thập"),
        ("I. HỆ CỬA ĐI", "Cửa trượt Sliding - Pocket Door", "Reynaers, Schüco, Cortizo", "CP130 Pocket, ASS50 Pocket", "Cửa trượt lùa âm tường hốc xây", "Đang thu thập"),
        ("I. HỆ CỬA ĐI", "Cửa trượt Sliding - Corner Sliding", "Reynaers, Cortizo, Schüco", "CP130-LS Corner, Cor Vision Corner", "Cửa trượt góc 90 độ không cột", "Đang thu thập"),
        ("I. HỆ CỬA ĐI", "Cửa trượt Sliding - Slim Sliding", "Alugood, QueenViet, Viralwindow, PMA", "Slim Alugood, PMA Slim", "Cửa lùa nhôm mảnh tối giản", "Đang thu thập"),
        ("I. HỆ CỬA ĐI", "Cửa trượt Sliding - Heavy Duty Sliding", "Schüco, Reynaers, Cortizo", "ASE80 HD, CP155 Heavy Duty", "Cửa trượt chịu tải trọng nặng", "Đang thu thập"),
        ("I. HỆ CỬA ĐI", "Cửa trượt Sliding - Panorama Sliding", "Schüco, Reynaers, Cortizo", "ASS77 PD, Hi-Finity, Cor Vision Plus", "Cửa trượt lùa vô cực đố 20mm", "Đang thu thập"),
        ("I. HỆ CỬA ĐI", "Lift & Slide - 2 ray trượt nâng", "Schüco, Reynaers, Cortizo", "ASE80 Lift, CP155 Lift", "Cửa trượt nâng hạ 2 ray", "Đang thu thập"),
        ("I. HỆ CỬA ĐI", "Lift & Slide - 3 ray trượt nâng", "Schüco, Reynaers, Cortizo", "ASE80 3-Ray, CP155 3-Ray", "Cửa trượt nâng hạ 3 ray", "Đang thu thập"),
        ("I. HỆ CỬA ĐI", "Lift & Slide - 4 ray trượt nâng", "Reynaers, Cortizo, Wicona", "CP155 4-Ray, 4600 Slider", "Cửa trượt nâng hạ 4 ray", "Đang thu thập"),
        ("I. HỆ CỬA ĐI", "Lift & Slide - Góc trượt nâng", "Schüco, Reynaers, Cortizo", "ASE80 Corner, CP155 Corner", "Cửa trượt nâng hạ vuông góc", "Đang thu thập"),
        ("I. HỆ CỬA ĐI", "Cửa xếp trượt - Folding Door", "Schüco, Reynaers, Cortizo, Xingfa", "ASS70 FD, CF77, XF63", "Cửa xếp gấp trượt xếp lùa", "Đang thu thập"),
        ("I. HỆ CỬA ĐI", "Cửa xếp trượt - Bi-fold", "Cortizo, Wicona, Aluprof", "Bi-Fold, Wicslide Folding", "Cửa xếp trượt gấp kiểu Bi-fold", "Đang thu thập"),
        ("I. HỆ CỬA ĐI", "Cửa xếp trượt - Multi-fold", "Schüco, Reynaers, Cortizo", "ASS70 FD, CF77, Bi-Fold", "Cửa xếp trượt nhiều cánh xếp ghép", "Đang thu thập"),
        ("I. HỆ CỬA ĐI", "Cửa xếp trượt - Accordion", "Xingfa, PMA, Viralwindow", "XF63 Accordion, PMA Folding", "Cửa xếp gấp trượt kiểu xếp lớp", "Đang thu thập"),
        ("I. HỆ CỬA ĐI", "Cửa thủy lực - Khung nhôm", "Owin, Xingfa, PMA", "Owin thủy lực, XF Thủy lực", "Cửa đi bản cánh lớn 180mm", "Đang thu thập"),
        ("I. HỆ CỬA ĐI", "Cửa thủy lực - Khung Slim", "PMA, Viralwindow, QueenViet", "PMA Slim thủy lực, Viralwindow VR Slim", "Cửa đi bản cánh slim thủy lực", "Đang thu thập"),
        ("I. HỆ CỬA ĐI", "Cửa thủy lực - Khung lớn", "Owin, Xingfa, PMA", "Owin bản lớn mạ titan gold", "Cửa thủy lực đại sảnh biệt thự", "Đang thu thập"),
        ("I. HỆ CỬA ĐI", "Cửa tự động - Trượt tự động", "Schüco, Nabco, Hafele", "ADS75 Auto, Nabco Sliding", "Cửa lùa trượt mở bằng motor", "Đang thu thập"),
        ("I. HỆ CỬA ĐI", "Cửa tự động - Mở quay tự động", "Schüco, Hafele, Nabco", "ADS75 Auto Swing", "Cửa mở quay tích hợp motor khuỷu", "Đang thu thập"),
        ("I. HỆ CỬA ĐI", "Cửa tự động - Cửa bệnh viện", "Nabco, Hafele, Kaba", "Hospital Door, Hermetic Door", "Cửa tự động kín khít phòng mổ", "Đang thu thập"),
        ("I. HỆ CỬA ĐI", "Cửa tự động - Cửa siêu thị", "Nabco, Hafele, Dorma", "Supermarket Auto Door", "Cửa tự động trượt lùa cảm biến", "Đang thu thập"),
    ]
    for cat, item, brand, code, p_type, status in gd1_doors:
        raw_data.append((cat, item, brand, code, p_type, "Đã có", "Đã có", "Đã có", "Đã có", "Đã có", "Kế hoạch", status))

    # II. HỆ CỬA SỔ
    gd1_windows = [
        ("II. HỆ CỬA SỔ", "Cửa sổ mở quay - 1 cánh", "Xingfa, PMA, Viralwindow, Topal", "XF55, PMA55, Viralwindow VRA55", "Cửa sổ mở quay 1 cánh định vị A", "Đang thu thập"),
        ("II. HỆ CỬA SỔ", "Cửa sổ mở quay - 2 cánh", "Xingfa, PMA, Viralwindow, Topal", "XF55, PMA55, Viralwindow VRA55", "Cửa sổ mở quay 2 cánh có đố động", "Đang thu thập"),
        ("II. HỆ CỬA SỔ", "Cửa sổ mở quay - 3 cánh", "Xingfa, PMA, Viralwindow", "XF55, PMA55, Viralwindow VRA55", "Cửa sổ mở quay 3 cánh (2 quay + 1 hất)", "Đang thu thập"),
        ("II. HỆ CỬA SỔ", "Cửa sổ mở quay - 4 cánh", "Xingfa, PMA, Viralwindow", "XF55, PMA55, Viralwindow VRA55", "Cửa sổ mở quay 4 cánh mặt tiền", "Đang thu thập"),
        ("II. HỆ CỬA SỔ", "Cửa sổ mở hất - Top Hung", "Xingfa, PMA, Viralwindow, Topal", "XF55, PMA55, Viralwindow VRA55", "Cửa sổ hất trên định vị chống sập", "Đang thu thập"),
        ("II. HỆ CỬA SỔ", "Cửa sổ mở hất - Bottom Hung", "Schüco, Reynaers, Cortizo", "AWS 65, CS68", "Cửa sổ hất dưới mở trong thông gió", "Đang thu thập"),
        ("II. HỆ CỬA SỔ", "Cửa sổ mở lật - Tilt", "Schüco, Reynaers, Cortizo", "AWS 75.SI, ML8, Cor 70 T&T", "Cửa sổ chỉ mở lật trên 15 độ", "Đang thu thập"),
        ("II. HỆ CỬA SỔ", "Cửa sổ mở lật - Tilt & Turn", "Schüco, Reynaers, Cortizo", "AWS 75.SI, MasterLine 8 T&T", "Cửa sổ quay lật đa năng châu Âu", "Đang thu thập"),
        ("II. HỆ CỬA SỔ", "Cửa sổ trượt - 2 ray lùa", "Xingfa, PMA, Viralwindow, Topal", "XF93, PMA93, Viralwindow VRA94 / VRE120", "Cửa sổ lùa 2 cánh ray inox", "Đang thu thập"),
        ("II. HỆ CỬA SỔ", "Cửa sổ trượt - 3 ray lùa", "Xingfa, PMA, Viralwindow", "XF93 3-Ray, PMA93 3-Ray", "Cửa sổ lùa 3 cánh lồng ghép", "Đang thu thập"),
        ("II. HỆ CỬA SỔ", "Cửa sổ trượt - Slim", "Reynaers, Cortizo, VR", "Hi-Finity Window, Viralwindow VR Slim", "Cửa sổ lùa đố dọc siêu mỏng", "Đang thu thập"),
        ("II. HỆ CỬA SỔ", "Cửa sổ trượt - Panorama", "Schüco, Reynaers, Cortizo", "ASS77 PD, Hi-Finity, Cor Vision", "Cửa sổ trùa lùa góc view vô cực", "Đang thu thập"),
        ("II. HỆ CỬA SỔ", "Cửa sổ cố định - Fixed Window", "Xingfa, PMA, Viralwindow, Schüco", "XF55, PMA55, AWS50 Fixed", "Vách kính lấy sáng không mở", "Đang thu thập"),
        ("II. HỆ CỬA SỔ", "Cửa sổ góc - Glass to Glass", "Cortizo, Reynaers, Schüco", "Fixed Facade, Corner Window", "Vách kính bo góc không khung nhôm", "Đang thu thập"),
        ("II. HỆ CỬA SỔ", "Cửa sổ thông gió - Louvre", "AluK, SAPA, Xingfa", "AluK Louvre, Xingfa chớp nhôm", "Cửa chớp nhôm lấy gió cố định", "Đang thu thập"),
        ("II. HỆ CỬA SỔ", "Cửa sổ thông gió - Jalousie", "AluK, SAPA, Wicona", "Wicona Ventilation, Jalousie", "Cửa sổ chớp kính xoay lật góc", "Đang thu thập"),
    ]
    for cat, item, brand, code, p_type, status in gd1_windows:
        raw_data.append((cat, item, brand, code, p_type, "Đã có", "Đã có", "Đã có", "Đã có", "Đã có", "Kế hoạch", status))

    # III. HỆ SLIM
    gd1_slim = [
        ("III. HỆ SLIM", "Slim Interior - Cửa đi lùa treo", "Alugood, QueenViet, Viralwindow, PMA", "Alugood Slim, PMA Slim", "Cửa đi lùa treo trong nhà giảm chấn", "Đang thu thập"),
        ("III. HỆ SLIM", "Slim Interior - Cửa phòng ngủ", "Alugood, QueenViet, Viralwindow, PMA", "Slim Lùa Treo, PMA Slim", "Cửa đi thông phòng ngủ cách âm nhẹ", "Đang thu thập"),
        ("III. HỆ SLIM", "Slim Interior - Vách ngăn slim", "Alugood, QueenViet, Viralwindow, PMA", "Slim Fixed Partition", "Vách ngăn kính khung nhôm siêu mảnh", "Đang thu thập"),
        ("III. HỆ SLIM", "Slim Interior - Cửa trượt âm tường", "Alugood, QueenViet, VR", "Slim Pocket Door", "Cửa trượt lùa treo giấu tường", "Đang thu thập"),
        ("III. HỆ SLIM", "Slim Exterior - Cửa đi lùa bão", "Schüco, Reynaers, Cortizo, Viralwindow", "ASS77 PD, Hi-Finity, Cor Vision", "Cửa đi lùa slim chịu lực ngoài trời", "Đang thu thập"),
        ("III. HỆ SLIM", "Slim Exterior - Cửa sổ lùa bão", "Reynaers, Cortizo, VR", "Hi-Finity Window, Viralwindow VR Slim", "Cửa sổ lùa slim chịu lực ngoài trời", "Đang thu thập"),
        ("III. HỆ SLIM", "Slim Exterior - Lift & Slide slim", "Cortizo, Reynaers, Wicona", "Cor Vision Plus L&S, CP155 Slim", "Cửa trượt nâng slim ngoại thất", "Đang thu thập"),
        ("III. HỆ SLIM", "Slim Panorama (Tầm nhìn vô cực)", "Schüco, Reynaers, Cortizo", "ASS77 PD, Hi-Finity, Cor Vision", "Cửa trượt lùa góc mở rộng view 98%", "Đang thu thập"),
        ("III. HỆ SLIM", "Slim Minimal (Tối giản cực hạn)", "Cortizo, Reynaers, SAPA", "Cor Vision, SAPA Artline", "Khung bao cánh ngậm sàn trần", "Đang thu thập"),
        ("III. HỆ SLIM", "Slim Hidden Frame (Ẩn khung bao)", "Technal, Wicona, Schüco", "Lumeal Hidden, ASS77 PD", "Cửa đi trượt lùa giấu khung bao cánh", "Đang thu thập"),
        ("III. HỆ SLIM", "Slim Steel Look (Mô phỏng sắt)", "Reynaers, Technal, Cortizo", "MasterLine Steel Look", "Nhôm phay gờ kiểu cổ điển sắt rèn", "Đang thu thập"),
    ]
    for cat, item, brand, code, p_type, status in gd1_slim:
        raw_data.append((cat, item, brand, code, p_type, "Đã có", "Đã có", "Đã có", "Đã có", "Đã có", "Kế hoạch", status))

    # IV. HỆ CẦU CÁCH NHIỆT
    gd1_thermal = [
        ("IV. HỆ CẦU CÁCH NHIỆT", "Thermal Break Door", "Schüco, Reynaers, Cortizo, Viralwindow", "ADS 75.SI, ML8-HI, VRE65-TB", "Cửa đi mở quay/lùa có cầu cách nhiệt", "Đang thu thập"),
        ("IV. HỆ CẦU CÁCH NHIỆT", "Thermal Break Window", "Schüco, Reynaers, Cortizo, Viralwindow", "AWS 75.SI, ML8-HI, VRX75-TB", "Cửa sổ mở quay/lật có cầu cách nhiệt", "Đang thu thập"),
        ("IV. HỆ CẦU CÁCH NHIỆT", "Thermal Break Sliding", "Schüco, Reynaers, Cortizo, Wicona", "ASE80 Lift, CP155, Cor 4600", "Cửa đi trượt lùa cách âm cách nhiệt", "Đang thu thập"),
        ("IV. HỆ CẦU CÁCH NHIỆT", "Thermal Break Curtain Wall", "Schüco, Reynaers, Wicona", "FW 50+.SI, CW 50-HI", "Mặt dựng kính khung nhôm cách nhiệt", "Đang thu thập"),
        ("IV. HỆ CẦU CÁCH NHIỆT", "Passive House System", "Schüco, Reynaers, Wicona", "AWS 90.SI+, ML10 Passive", "Hệ thống nhôm cản nhiệt Passive House", "Đang thu thập"),
    ]
    for cat, item, brand, code, p_type, status in gd1_thermal:
        raw_data.append((cat, item, brand, code, p_type, "Đã có", "Đã có", "Đã có", "Đã có", "Đã có", "Kế hoạch", status))

    # XII. HỆ BIỆT THỰ
    gd1_villa = [
        ("XII. HỆ BIỆT THỰ", "Cửa chính biệt thự", "Schüco, Reynaers, Cortizo", "ADS 90 Door, Millennium Premium", "Cửa đi pano nhôm đúc hoặc hoa văn kính", "Đang thu thập"),
        ("XII. HỆ BIỆT THỰ", "Cửa đại sảnh lâu đài", "Schüco, Reynaers, Cortizo", "ADS 90 Door, Pivot XXL", "Cửa cánh mở xoay rộng trục lớn", "Đang thu thập"),
        ("XII. HỆ BIỆT THỰ", "Cửa siêu cao biệt thự", "Cortizo, Reynaers, Schüco", "Pivot XXL, ML Pivot", "Cửa cánh cao trên 3.5m", "Đang thu thập"),
        ("XII. HỆ BIỆT THỰ", "Cửa siêu rộng biệt thự", "Cortizo, Reynaers, Schüco", "Pivot XXL, Hi-Finity", "Cửa cánh rộng trên 2m", "Đang thu thập"),
        ("XII. HỆ BIỆT THỰ", "Pivot Luxury biệt thự", "Reynaers, Cortizo, Schüco", "Hi-Finity Pivot, ML Pivot", "Cửa xoay trục lệch tâm khóa vân tay", "Đang thu thập"),
        ("XII. HỆ BIỆT THỰ", "Lift & Slide Luxury", "Schüco, Reynaers, Cortizo", "ASE80 Lift, CP155, Cor 4600", "Cửa trượt nâng rãnh âm phẳng sàn", "Đang thu thập"),
    ]
    for cat, item, brand, code, p_type, status in gd1_villa:
        raw_data.append((cat, item, brand, code, p_type, "Đã có", "Đã có", "Đã có", "Đã có", "Đã có", "Kế hoạch", status))

    # XIII. HỆ SIÊU TRƯỜNG SIÊU TRỌNG
    gd1_heavy = [
        ("XIII. HỆ SIÊU TRƯỜNG SIÊU TRỌNG", "Heavy Sliding (Trượt siêu nặng)", "Schüco, Reynaers, Cortizo", "ASE80 HD, CP155 Heavy Duty", "Cửa trượt cánh nặng tới 400kg", "Đang thu thập"),
        ("XIII. HỆ SIÊU TRƯỜNG SIÊU TRỌNG", "Heavy Casement (Quay siêu nặng)", "Schüco, Reynaers, Cortizo", "ADS90 HD, Millennium Plus", "Cửa đi mở quay chịu tải bản lề lớn", "Đang thu thập"),
        ("XIII. HỆ SIÊU TRƯỜNG SIÊU TRỌNG", "Jumbo Door (Kính khổ lớn)", "Reynaers, Cortizo, Schüco", "Hi-Finity Jumbo, Cor Vision XXL", "Cửa trượt lùa kính nguyên tấm khổ cực đại", "Đang thu thập"),
        ("XIII. HỆ SIÊU TRƯỜNG SIÊU TRỌNG", "Mega Door (Cửa đại cực lớn)", "Cortizo, Reynaers, Wicona", "Pivot XXL, Mega Slider", "Cửa đi xoay đại sảnh khẩu độ rộng", "Đang thu thập"),
        ("XIII. HỆ SIÊU TRƯỜNG SIÊU TRỌNG", "XXL Window (Sổ kính cực lớn)", "Schüco, Reynaers, Cortizo", "AWS 75 fixed XXL, Fixed Facade", "Vách kính lấy sáng thông tầng biệt thự", "Đang thu thập"),
    ]
    for cat, item, brand, code, p_type, status in gd1_heavy:
        raw_data.append((cat, item, brand, code, p_type, "Đã có", "Đã có", "Đã có", "Đã có", "Đã có", "Kế hoạch", status))


    # --- PHASE 2 (Kế hoạch GĐ 2) ---
    # V. HỆ MẶT DỰNG
    gd2_facade = [
        ("V. HỆ MẶT DỰNG", "Stick Curtain Wall", "Xingfa, Schüco, Reynaers", "XF65 Stick, FW50+ Stick", "Mặt dựng kính ghép thanh tại công trường", "Kế hoạch GĐ 2"),
        ("V. HỆ MẶT DỰNG", "Semi Unitized", "Schüco, Reynaers, Eurowindow", "FW50 Semi-Unitized, CW60 Semi", "Mặt dựng kính bán lắp ghép định hình", "Kế hoạch GĐ 2"),
        ("V. HỆ MẶT DỰNG", "Unitized", "Schüco, Reynaers, BM Windows", "FW50 Unitized, CW60 Unitized", "Mặt dựng kính modul hóa sản xuất tại xưởng", "Kế hoạch GĐ 2"),
        ("V. HỆ MẶT DỰNG", "Structural Glazing", "Schüco, Reynaers, Dow Corning", "Structural Glazing Facade", "Mặt dựng kính keo chịu lực giấu đố hoàn toàn", "Kế hoạch GĐ 2"),
        ("V. HỆ MẶT DỰNG", "Spider Glass", "Kinlong, Hafele, Adler", "Spider Fittings System", "Mặt dựng kính chân nhện liên kết điểm", "Kế hoạch GĐ 2"),
        ("V. HỆ MẶT DỰNG", "Point Fixed", "Hafele, Kinlong, Dorma", "Point Fixed Glazing", "Vách mặt tiền liên kết bu-lông xuyên kính", "Kế hoạch GĐ 2"),
        ("V. HỆ MẶT DỰNG", "Cable Facade", "Schüco, Reynaers, Wicona", "Cable Supported Facade", "Vách kính cường lực liên kết hệ cáp chịu lực", "Kế hoạch GĐ 2"),
        ("V. HỆ MẶT DỰNG", "Double Skin Facade", "Schüco, Reynaers, Wicona", "Double Skin Facade System", "Mặt dựng kính 2 lớp cách âm cản nhiệt cao", "Kế hoạch GĐ 2"),
    ]
    for cat, item, brand, code, p_type, status in gd2_facade:
        raw_data.append((cat, item, brand, code, p_type, "Chưa", "Chưa", "Chưa", "Chưa", "Chưa", "Chưa", status))

    # VI. HỆ MÁI
    gd2_roof = [
        ("VI. HỆ MÁI", "Glass Roof (Mái kính)", "Sao Vàng CNC, Schüco, Reynaers", "Glass Roof SV-R1", "Mái kính chịu lực che mưa đón sáng", "Kế hoạch GĐ 2"),
        ("VI. HỆ MÁI", "Skylight (Giếng trời)", "Schüco, Reynaers, Velux", "FW50+ Skylight, Velux Skylight", "Mái kính giếng trời lấy sáng tự động", "Kế hoạch GĐ 2"),
        ("VI. HỆ MÁI", "Sunroom (Phòng kính)", "Schüco, Reynaers, Wicona", "Sunroom / Winter Garden", "Phòng kính sảnh vườn ngắm cảnh cản nhiệt", "Kế hoạch GĐ 2"),
        ("VI. HỆ MÁI", "Pergola nhôm kính", "Sao Vàng, Euro-Pergola", "Bioclimatic Pergola System", "Mái hiên nhôm kính điều khiển lá chớp lật", "Kế hoạch GĐ 2"),
        ("VI. HỆ MÁI", "Canopy (Mái đón hiên)", "Sao Vàng CNC, Adler, VVP", "Glass Canopy SV-C1", "Mái đón sảnh chính liên kết thanh treo inox", "Kế hoạch GĐ 2"),
        ("VI. HỆ MÁI", "Roof Window (Cửa sổ mái)", "Velux, Schüco, Roto Window", "Velux GGL, Schüco AWS 57 RO", "Cửa sổ mở lật hất thoát hiểm trên mái", "Kế hoạch GĐ 2"),
    ]
    for cat, item, brand, code, p_type, status in gd2_roof:
        raw_data.append((cat, item, brand, code, p_type, "Chưa", "Chưa", "Chưa", "Chưa", "Chưa", "Chưa", status))

    # VII. HỆ LAM
    gd2_lam = [
        ("VII. HỆ LAM", "Lam đứng", "Xingfa, PMA, Sao Vàng", "Lam chớp đứng trang trí", "Hệ lam nhôm dọc mặt tiền cản nắng", "Kế hoạch GĐ 2"),
        ("VII. HỆ LAM", "Lam ngang", "Xingfa, PMA, Sao Vàng", "Lam chớp ngang che cục nóng", "Hệ lam nhôm ngang cản nắng che bụi", "Kế hoạch GĐ 2"),
        ("VII. HỆ LAM", "Lam chắn nắng Louvres", "Hunter Douglas, SAPA, Xingfa", "Aerofoil Louvres, Sunshade", "Hệ lam chắn nắng nhôm lá liễu cố định", "Kế hoạch GĐ 2"),
        ("VII. HỆ LAM", "Aerofoil (Lam cánh máy bay)", "Hunter Douglas, SAPA, Topal", "Aerofoil Louvres 150/200/300", "Lam chắn nắng hình cánh máy bay chịu gió lớn", "Kế hoạch GĐ 2"),
        ("VII. HỆ LAM", "Facade Louvre (Lam trang trí)", "Hunter Douglas, Wicona, SAPA", "Facade Decorative Louvre", "Hệ lam nhôm ốp tạo hình nghệ thuật mặt tiền", "Kế hoạch GĐ 2"),
    ]
    for cat, item, brand, code, p_type, status in gd2_lam:
        raw_data.append((cat, item, brand, code, p_type, "Chưa", "Chưa", "Chưa", "Chưa", "Chưa", "Chưa", status))

    # VIII. HỆ LAN CAN
    gd2_railing = [
        ("VIII. HỆ LAN CAN", "Lan can kính", "Adler, VVP, Hafele, Sao Vàng", "Glass Railing SV-R1", "Lan can kính cường lực tay vịn inox/nhôm", "Kế hoạch GĐ 2"),
        ("VIII. HỆ LAN CAN", "Lan can nhôm đúc", "Sao Vàng CNC, Nhôm đúc mỹ nghệ", "Alu-Railing SV-A1", "Lan can nhôm đúc hoa văn biệt thự tân cổ", "Kế hoạch GĐ 2"),
        ("VIII. HỆ LAN CAN", "Lan can Slim", "PMA, Viralwindow, Bogo, Sigico", "Slim Glass Railing", "Lan can kính tay vịn siêu mảnh thẩm mỹ", "Kế hoạch GĐ 2"),
        ("VIII. HỆ LAN CAN", "Lan can không trụ", "Hafele, Adler, Sao Vàng", "U-Channel Glass Railing", "Lan can u-nhôm âm sàn kẹp kính không trụ", "Kế hoạch GĐ 2"),
        ("VIII. HỆ LAN CAN", "Lan can âm sàn", "Hafele, Adler, BM Windows", "Embedded Glass Railing", "Lan can kẹp kính âm sàn giấu khung u nhôm", "Kế hoạch GĐ 2"),
    ]
    for cat, item, brand, code, p_type, status in gd2_railing:
        raw_data.append((cat, item, brand, code, p_type, "Chưa", "Chưa", "Chưa", "Chưa", "Chưa", "Chưa", status))

    # IX. HỆ VÁCH
    gd2_partition = [
        ("IX. HỆ VÁCH", "Vách kính cố định", "Xingfa, PMA, Viralwindow, Topal", "XF55 Vách, PMA55 Vách", "Vách kính ngăn che ngoài trời tĩnh", "Kế hoạch GĐ 2"),
        ("IX. HỆ VÁCH", "Vách ngăn văn phòng", "PMA, VR, Xingfa", "Office Partition SV-P1", "Vách ngăn thạch cao kết hợp khung nhôm kính", "Kế hoạch GĐ 2"),
        ("IX. HỆ VÁCH", "Vách Slim nội thất", "Alugood, QueenViet, Viralwindow, PMA", "Slim Interior Partition", "Vách nhôm slim siêu mảnh ngăn phòng họp", "Kế hoạch GĐ 2"),
        ("IX. HỆ VÁCH", "Vách kính mặt tiền", "Xingfa, PMA, Viralwindow, Schüco", "XF65 Facade, FW50+ Vách", "Vách kính mặt tiền thông tầng shop/showroom", "Kế hoạch GĐ 2"),
    ]
    for cat, item, brand, code, p_type, status in gd2_partition:
        raw_data.append((cat, item, brand, code, p_type, "Chưa", "Chưa", "Chưa", "Chưa", "Chưa", "Chưa", status))

    # X. HỆ PHÒNG TẮM
    gd2_shower = [
        ("X. HỆ PHÒNG TẮM", "Shower Screen (Vách kính tắm)", "Hafele, VVP, Adler, Sao Vàng", "Shower Screen SV-S1", "Vách kính phòng tắm cường lực 8-10mm", "Kế hoạch GĐ 2"),
        ("X. HỆ PHÒNG TẮM", "Shower Door (Cửa phòng tắm)", "Hafele, VVP, Adler, Sao Vàng", "Shower Door SV-S2", "Cửa đi kính cường lực vào phòng tắm", "Kế hoạch GĐ 2"),
        ("X. HỆ PHÒNG TẮM", "Cabin trượt phòng tắm", "Hafele, VVP, Adler, Bogo", "Sliding Shower Cabin", "Cabin phòng tắm cửa lùa trượt inox/nhôm", "Kế hoạch GĐ 2"),
        ("X. HỆ PHÒNG TẮM", "Cabin mở quay phòng tắm", "Hafele, VVP, Adler, Bogo", "Pivot Shower Cabin", "Cabin phòng tắm cửa mở quay góc 90/135/180", "Kế hoạch GĐ 2"),
    ]
    for cat, item, brand, code, p_type, status in gd2_shower:
        raw_data.append((cat, item, brand, code, p_type, "Chưa", "Chưa", "Chưa", "Chưa", "Chưa", "Chưa", status))

    # XI. HỆ PHÒNG (Room Doors)
    gd2_room = [
        ("XI. HỆ PHÒNG", "Cửa phòng ngủ", "Xingfa, PMA, Viralwindow, Topal", "XF55 Cửa phòng, PMA55", "Cửa đi mở quay cách âm 1 cánh phòng ngủ", "Kế hoạch GĐ 2"),
        ("XI. HỆ PHÒNG", "Cửa toilet chống nước", "PMA, VR, Xingfa", "PMA55 Toilet, XF55 Toilet", "Cửa toilet nhôm kính tích hợp lá chớp", "Kế hoạch GĐ 2"),
        ("XI. HỆ PHÒNG", "Cửa phòng kho", "Xingfa, PMA, Viralwindow", "XF55 Kho, PMA55 Kho", "Cửa đi pano nhôm kín che phòng kỹ thuật/kho", "Kế hoạch GĐ 2"),
        ("XI. HỆ PHÒNG", "Cửa ra ban công", "Schüco, Reynaers, Cortizo, Xingfa", "ADS65, CS77, XF55", "Cửa đi mở quay/lùa ra ban công kín nước bão", "Kế hoạch GĐ 2"),
    ]
    for cat, item, brand, code, p_type, status in gd2_room:
        raw_data.append((cat, item, brand, code, p_type, "Chưa", "Chưa", "Chưa", "Chưa", "Chưa", "Chưa", status))


    # --- PHASE 3 (Kế hoạch GĐ 3) ---
    # XIV. HỆ CHUYÊN DỤNG
    gd3_special = [
        ("XIV. HỆ CHUYÊN DỤNG", "Fire Rated (Chống cháy)", "Aluprof, Wicona, Cortizo", "MB-78EI, Wicstyle 77FP", "Cửa đi/sổ nhôm kính ngăn lửa EI30/60/90", "Kế hoạch GĐ 3"),
        ("XIV. HỆ CHUYÊN DỤNG", "Bullet Resistant (Chống đạn)", "Schüco, Cortizo, Wicona", "ADS 90.BR, Bullet-Proof", "Cửa đi nhôm lõi thép chống đạn cấp FB4/FB6", "Kế hoạch GĐ 3"),
        ("XIV. HỆ CHUYÊN DỤNG", "Blast Resistant (Chống nổ)", "Schüco, Wicona, SAPA", "Blast Resistant System", "Cửa đi gia cường chống lực chấn động nổ", "Kế hoạch GĐ 3"),
        ("XIV. HỆ CHUYÊN DỤNG", "Hurricane Resistant (Chống bão)", "Schüco, Cortizo, Reynaers", "ADS 90.HD, High-Performance", "Cửa đi chịu áp lực gió bão biển cấp 15", "Kế hoạch GĐ 3"),
        ("XIV. HỆ CHUYÊN DỤNG", "Sound Proof (Cách âm phòng thu)", "Schüco, Reynaers, Wicona", "AWS 90.SI Acoustic", "Cửa đi nhôm kính cách âm đặc biệt Rw > 45dB", "Kế hoạch GĐ 3"),
        ("XIV. HỆ CHUYÊN DỤNG", "Smoke Proof (Ngăn khói)", "Aluprof, Wicona, Schüco", "MB-45D Smoke Control", "Cửa đi ngăn khói hành lang chung cư thoát hiểm", "Kế hoạch GĐ 3"),
        ("XIV. HỆ CHUYÊN DỤNG", "Hospital Door (Phòng mổ)", "Nabco, Hafele, Kaba", "Hospital Door, Hermetic", "Cửa lùa tự động gioăng đệm nén khí phòng mổ", "Kế hoạch GĐ 3"),
        ("XIV. HỆ CHUYÊN DỤNG", "Clean Room (Phòng sạch)", "Nabco, Hafele, Kaba", "Clean Room Door", "Cửa đi phẳng mặt inox/nhôm chống bám bụi bẩn", "Kế hoạch GĐ 3"),
    ]
    for cat, item, brand, code, p_type, status in gd3_special:
        raw_data.append((cat, item, brand, code, p_type, "Chưa", "Chưa", "Chưa", "Chưa", "Chưa", "Chưa", status))

    # XV. HỆ THÔNG MINH
    gd3_smart = [
        ("XV. HỆ THÔNG MINH", "Smart Door (Khóa vân tay/Face ID)", "CMECH, Hopo, Philips, Kaadas", "Smart Lock Integrated Door", "Cửa đi tích hợp hệ khóa thông minh Face ID", "Kế hoạch GĐ 3"),
        ("XV. HỆ THÔNG MINH", "Smart Window (Cảm biến mưa tự đóng)", "Hopo, CMECH, Siegenia", "Smart Motorized Window", "Cửa sổ động cơ điện tự đóng khi cảm biến mưa", "Kế hoạch GĐ 3"),
        ("XV. HỆ THÔNG MINH", "Fingerprint (Mở khóa vân tay)", "Philips, Kaadas, Samsung", "Fingerprint Door Entry", "Mở khóa bằng tay nắm tích hợp nhận diện vân tay", "Kế hoạch GĐ 3"),
        ("XV. HỆ THÔNG MINH", "Face ID camera ẩn", "Philips, Kaadas, Samsung", "Face ID Door System", "Mở khóa bằng camera quét khuôn mặt hồng ngoại", "Kế hoạch GĐ 3"),
        ("XV. HỆ THÔNG MINH", "Motorized Sliding (Cửa lùa tự động)", "Schüco, Reynaers, Cortizo", "e-Drive, Motorized Sliding", "Cửa lùa slim lớn đóng mở bằng động cơ điện", "Kế hoạch GĐ 3"),
        ("XV. HỆ THÔNG MINH", "Home Automation (Tích hợp SmartHome)", "Schüco, Reynaers, Lumi, BKAV", "Smart Home Integrated Windows", "Hệ thống cửa kết nối điều khiển trung tâm KNX", "Kế hoạch GĐ 3"),
    ]
    for cat, item, brand, code, p_type, status in gd3_smart:
        raw_data.append((cat, item, brand, code, p_type, "Chưa", "Chưa", "Chưa", "Chưa", "Chưa", "Chưa", status))

    # XVI. HỆ PHỤ TRỢ
    gd3_auxiliary = [
        ("XVI. HỆ PHỤ TRỢ", "Cửa lưới chống muỗi (côn trùng)", "Metaco (Nhật Bản), Quang Minh, Seiki", "Insect Screen System", "Lưới sợi thủy tinh tự cuốn hoặc lưới inox lá lùa", "Kế hoạch GĐ 3"),
        ("XVI. HỆ PHỤ TRỢ", "Cửa lưới bảo vệ an toàn", "Xingfa, PMA, Quang Minh", "Safety Screen Integrated", "Lưới thép inox 304 dày ngăn trộm cắt phá", "Kế hoạch GĐ 3"),
        ("XVI. HỆ PHỤ TRỢ", "Rèm tích hợp trong hộp kính", "Magnetic Blinds, Sunshade Glass", "Integrated Blind Glass", "Rèm nhôm nằm trong hộp kính hút chân không", "Kế hoạch GĐ 3"),
        ("XVI. HỆ PHỤ TRỢ", "Lam thông gió điều hòa", "Xingfa, PMA, Sao Vàng", "AC Louvres System", "Chớp nhôm che cục nóng và lấy gió tươi kỹ thuật", "Kế hoạch GĐ 3"),
        ("XVI. HỆ PHỤ TRỢ", "Ô lấy sáng mái kính", "Sao Vàng, Adler, Velux", "Skylight Roof Light", "Ô lấy sáng cố định kết cấu kính hộp cách nhiệt", "Kế hoạch GĐ 3"),
        ("XVI. HỆ PHỤ TRỢ", "Kính điện thông minh", "Polytron, Switchable Smart Glass", "Smart Glass Wall", "Kính chuyển đổi đục/trong bằng nút nhấn điện", "Kế hoạch GĐ 3"),
    ]
    for cat, item, brand, code, p_type, status in gd3_auxiliary:
        raw_data.append((cat, item, brand, code, p_type, "Chưa", "Chưa", "Chưa", "Chưa", "Chưa", "Chưa", status))

    # XVII. HỆ ĐẶC BIỆT
    gd3_highlights = [
        ("XVII. HỆ ĐẶC BIỆT", "Glass Corner (Bo góc kính-kính)", "Sao Vàng CNC, Dow Corning, Adler", "Glass to Glass Corner", "Góc vách kính vuông góc liên kết keo cấu trúc", "Kế hoạch GĐ 3"),
        ("XVII. HỆ ĐẶC BIỆT", "Invisible Frame (Ẩn khung bao)", "Schüco, Reynaers, Cortizo", "ASS77 PD, Hi-Finity, Cor Vision", "Khung bao cánh lùa chôn giấu sâu vào tường xây", "Kế hoạch GĐ 3"),
        ("XVII. HỆ ĐẶC BIỆT", "Hidden Sash (Ẩn cánh cửa sổ)", "Schüco, Reynaers, Cortizo", "AWS 75 BS.HI (Block System)", "Cánh cửa sổ hất ẩn giấu chìm sau khung bao bao", "Kế hoạch GĐ 3"),
        ("XVII. HỆ ĐẶC BIỆT", "Frameless (Không hệ khung nhôm)", "Sao Vàng, Adler, VVP, Hafele", "Frameless Glass System", "Hệ thống vách cửa kính cường lực chịu lực thuần", "Kế hoạch GĐ 3"),
        ("XVII. HỆ ĐẶC BIỆT", "Minimal Frame (Khung tối giản cực mảnh)", "Cortizo, Reynaers, Schüco", "Cor Vision, Hi-Finity, ASS77", "Đố nhôm cực mỏng 20mm mang phong cách tối giản", "Kế hoạch GĐ 3"),
        ("XVII. HỆ ĐẶC BIỆT", "Zero Threshold (Ray lùa phẳng sàn)", "Schüco, Reynaers, Cortizo", "Zero Threshold System", "Hệ ray dưới cửa lùa âm phẳng hoàn toàn mặt sàn", "Kế hoạch GĐ 3"),
        ("XVII. HỆ ĐẶC BIỆT", "Pocket Wall (Trượt giấu tường)", "Schüco, Reynaers, Cortizo", "ASS50 Pocket, CP130 Pocket", "Cánh cửa trượt lùa ẩn giấu hoàn toàn vào tường", "Kế hoạch GĐ 3"),
        ("XVII. HỆ ĐẶC BIỆT", "Pocket Window (Trượt cửa sổ giấu tường)", "Reynaers, Cortizo, Schüco", "CP130 Pocket Window", "Cửa sổ lùa giấu cánh vào trong vách xây thạch cao", "Kế hoạch GĐ 3"),
    ]
    for cat, item, brand, code, p_type, status in gd3_highlights:
        raw_data.append((cat, item, brand, code, p_type, "Chưa", "Chưa", "Chưa", "Chưa", "Chưa", "Chưa", status))

    # XVIII. HỆ CÔNG NGHIỆP
    gd3_industrial = [
        ("XVIII. HỆ CÔNG NGHIỆP", "Nhà xưởng (Cửa cánh thép/nhôm lớn)", "Sao Vàng, Doorhan, Hörmann", "Industrial Factory Door", "Cửa đi chính cánh xếp/cuốn lớn cho nhà máy", "Kế hoạch GĐ 3"),
        ("XVIII. HỆ CÔNG NGHIỆP", "Nhà kho (Cửa cách nhiệt kho lạnh)", "Hörmann, Doorhan, Sao Vàng", "Cold Room Insulation Door", "Cửa đi bản dày cách nhiệt polyurethane chống lạnh", "Kế hoạch GĐ 3"),
        ("XVIII. HỆ CÔNG NGHIỆP", "Hangar Door (Cửa hangar lớn)", "Sao Vàng, Megadoor, Shipyard", "Hangar Sliding Door", "Cửa đi xếp trượt khổng lồ cho gara máy bay/tàu", "Kế hoạch GĐ 3"),
        ("XVIII. HỆ CÔNG NGHIỆP", "Logistic Door (Cửa cuốn tốc độ cao)", "Hörmann, Nabco, Hafele", "High Speed Roll-up Door", "Cửa cuốn nhựa PVC đóng mở nhanh cản bụi kho", "Kế hoạch GĐ 3"),
        ("XVIII. HỆ CÔNG NGHIỆP", "Factory Window (Cửa sổ nhà máy)", "Xingfa, PMA, Sao Vàng", "Factory Steel/Alu Window", "Cửa sổ sắt/nhôm kính mở quay hất lấy gió nhà máy", "Kế hoạch GĐ 3"),
    ]
    for cat, item, brand, code, p_type, status in gd3_industrial:
        raw_data.append((cat, item, brand, code, p_type, "Chưa", "Chưa", "Chưa", "Chưa", "Chưa", "Chưa", status))

    # XIX. HỆ PHỤ KIỆN (TÁCH RIÊNG)
    gd3_accessories = [
        ("XIX. HỆ PHỤ KIỆN", "Bản lề", "CMECH, Roto, Sobinco, Kinlong", "Bản lề 3D, Bản lề cối, Bản lề ẩn", "Bản lề chịu tải lực cửa mở quay/xếp trượt", "Kế hoạch GĐ 3"),
        ("XIX. HỆ PHỤ KIỆN", "Khóa", "CMECH, Philips, Roto, Kinlong", "Khóa đa điểm, Khóa từ, Khóa gạt", "Hệ khóa chốt an toàn chống đột nhập cửa nhôm", "Kế hoạch GĐ 3"),
        ("XIX. HỆ PHỤ KIỆN", "Tay nắm", "CMECH, Roto, Bogo, Sigico", "Tay nắm gạt, Tay kéo, Tay gạt chốt", "Tay nắm vận hành khóa đa điểm và chốt cửa", "Kế hoạch GĐ 3"),
        ("XIX. HỆ PHỤ KIỆN", "Bánh xe", "CMECH, Roto, Sobinco, Kinlong", "Bánh xe trượt nâng, Bánh xe xếp lùa", "Bánh xe chịu lực tải cánh trượt nâng lùa", "Kế hoạch GĐ 3"),
        ("XIX. HỆ PHỤ KIỆN", "Ray", "CMECH, Schüco, Reynaers, Cortizo", "Ray inox âm sàn, Ray lùa phẳng", "Ray inox dẫn hướng chuyển động bánh xe lùa", "Kế hoạch GĐ 3"),
        ("XIX. HỆ PHỤ KIỆN", "Chốt", "CMECH, Roto, Kinlong, Draho", "Chốt chìm, Chốt biên, Chốt âm", "Cơ cấu chốt giữ cánh phụ cửa đi/sổ đa cánh", "Kế hoạch GĐ 3"),
        ("XIX. HỆ PHỤ KIỆN", "Gioăng", "EPDM, Cao su tổng hợp, Gioăng lông", "Gioăng chèn kính, Gioăng khung bao", "Gioăng làm kín co giãn đàn hồi chống rò rỉ nước khí", "Kế hoạch GĐ 3"),
        ("XIX. HỆ PHỤ KIỆN", "Ke ép góc", "Nhôm hợp kim đúc, Ke nhảy inox", "Ke góc đúc ép thủy lực, Ke nhảy", "Ke gia cường liên kết góc thanh nhôm đùn định hình", "Kế hoạch GĐ 3"),
        ("XIX. HỆ PHỤ KIỆN", "Góc nhảy", "Nhôm đúc, Ke nhảy thông minh", "Ke nhảy inox 304 định vị góc", "Ke định góc ghép khít bằng ốc vít căn siết", "Kế hoạch GĐ 3"),
        ("XIX. HỆ PHỤ KIỆN", "Nẹp kính", "Xingfa, PMA, Viralwindow, Topal", "Nẹp vuông, Nẹp bo tròn", "Nẹp che sập giữ cố định mép kính vào khung bao cánh", "Kế hoạch GĐ 3"),
        ("XIX. HỆ PHỤ KIỆN", "Phụ kiện đồng bộ theo hãng", "Schüco, Reynaers, CMECH, Hopo", "Hệ phụ kiện đồng bộ rãnh C", "Tập hợp đồng bộ chốt bản lề bánh xe tay nắm theo hãng", "Kế hoạch GĐ 3"),
    ]
    for cat, item, brand, code, p_type, status in gd3_accessories:
        raw_data.append((cat, item, brand, code, p_type, "Chưa", "Chưa", "Chưa", "Chưa", "Chưa", "Chưa", status))

    # XX. HỆ PROFILE (TÁCH RIÊNG)
    gd3_profile = [
        ("XX. HỆ PROFILE", "Khung bao", "Xingfa, PMA, Viralwindow, Topal, Schüco", "Profiles Khung bao đi/sổ", "Thanh nhôm biên viền bao quanh lắp tường xây", "Kế hoạch GĐ 3"),
        ("XX. HỆ PROFILE", "Khung cánh", "Xingfa, PMA, Viralwindow, Topal, Schüco", "Profiles Khung cánh mở/lùa", "Thanh nhôm tạo khung chuyển động đóng mở giữ kính", "Kế hoạch GĐ 3"),
        ("XX. HỆ PROFILE", "Đố đứng", "Xingfa, PMA, Viralwindow, Topal", "Profiles Đố đứng dọc phân ô", "Thanh nhôm chia nhỏ khung ô vách vòm thẳng đứng", "Kế hoạch GĐ 3"),
        ("XX. HỆ PROFILE", "Đố ngang", "Xingfa, PMA, Viralwindow, Topal", "Profiles Đố ngang chia ô", "Thanh nhôm nằm ngang chia ô vách cửa sổ đi", "Kế hoạch GĐ 3"),
        ("XX. HỆ PROFILE", "Ray dưới", "Xingfa, PMA, Viralwindow, Topal, Schüco", "Profiles Ray dưới cửa lùa", "Thanh nhôm định hình rãnh trượt dẫn hướng trượt lùa", "Kế hoạch GĐ 3"),
        ("XX. HỆ PROFILE", "Ray trên", "Xingfa, PMA, Viralwindow, Topal, Schüco", "Profiles Ray trên cửa lùa treo", "Thanh nhôm treo lực tải cửa lùa treo trong nhà", "Kế hoạch GĐ 3"),
        ("XX. HỆ PROFILE", "Nẹp kính", "Xingfa, PMA, Viralwindow, Topal", "Profiles Nẹp gài sập kính", "Thanh nhôm biên mỏng gạt chốt sập giữ kính", "Kế hoạch GĐ 3"),
        ("XX. HỆ PROFILE", "Thanh tăng cứng", "Xingfa, PMA, Viralwindow, Topal, Schüco", "Reinforcement Profiles", "Thanh nhôm bản dày chèn gia cường lòng khoang trống nhôm", "Kế hoạch GĐ 3"),
        ("XX. HỆ PROFILE", "Thanh liên kết", "Xingfa, PMA, Viralwindow, Topal", "Connection Profiles", "Ke liên kết đố ghép nối bu-lông đai ốc âm nhôm", "Kế hoạch GĐ 3"),
        ("XX. HỆ PROFILE", "Thanh trang trí", "Xingfa, PMA, Viralwindow, Topal, Schüco", "Decorative Profiles", "Thanh nhôm chớp ốp ngoài bọc trang trí thẩm mỹ đố", "Kế hoạch GĐ 3"),
        ("XX. HỆ PROFILE", "Thanh ghép chuyển góc", "Xingfa, PMA, Viralwindow, Topal", "Corner Joint Profiles", "Thanh nhôm ghép tròn/vuông xoay góc vách kính liên tục", "Kế hoạch GĐ 3"),
        ("XX. HỆ PROFILE", "Thanh góc vuông", "Xingfa, PMA, Viralwindow, Topal", "90 Degree Profiles", "Thanh nhôm ghép nối góc tường bao vuông 90 độ", "Kế hoạch GĐ 3"),
        ("XX. HỆ PROFILE", "Thanh che nước bão", "Xingfa, PMA, Viralwindow, Topal", "Drip Profiles", "Thanh nhôm che máng hướng chảy nước mưa ngoài cánh", "Kế hoạch GĐ 3"),
        ("XX. HỆ PROFILE", "Thanh chống xệ cánh", "Xingfa, PMA, Viralwindow, Topal", "Anti-sag Profiles", "Thanh nhôm đệm tăng cứng chống xệ chân góc bản lề", "Kế hoạch GĐ 3"),
        ("XX. HỆ PROFILE", "Thanh thoát nước", "Xingfa, PMA, Viralwindow, Topal", "Drainage Profiles", "Thanh nhôm thiết kế máng thoát nước ngầm khung bao dốc", "Kế hoạch GĐ 3"),
    ]
    for cat, item, brand, code, p_type, status in gd3_profile:
        raw_data.append((cat, item, brand, code, p_type, "Chưa", "Chưa", "Chưa", "Chưa", "Chưa", "Chưa", status))


    # Write data rows (Starting from Row 5)
    for idx, row_data in enumerate(raw_data, start=5):
        stt = idx - 4
        fill_to_apply = fill_zebra if idx % 2 == 1 else None
        
        status_text = row_data[11]
        
        # Color based on status
        if status_text == "Đang thu thập":
            status_fill = fill_gd1
        elif status_text == "Kế hoạch GĐ 2":
            status_fill = fill_gd2
        else:
            status_fill = fill_gd3

        # Write STT (Col 1)
        cell_stt = ws.cell(row=idx, column=1)
        cell_stt.value = stt
        cell_stt.font = font_body_bold
        cell_stt.alignment = align_center
        cell_stt.border = box_border
        if fill_to_apply:
            cell_stt.fill = fill_to_apply

        # Write rest of columns (Col 2 to 12)
        for col_idx, cell_value in enumerate(row_data, start=2):
            cell = ws.cell(row=idx, column=col_idx)
            cell.value = cell_value
            cell.font = font_body
            cell.border = box_border
            
            # Apply fill
            if col_idx == 12: # Status column gets specific fill
                cell.fill = status_fill
                cell.font = font_body_bold
            elif fill_to_apply:
                cell.fill = fill_to_apply
                
            # Alignment rules
            if col_idx in [2, 4, 5, 7, 8, 9, 10, 11, 12]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left

    # Add a heavy bottom border to the last data row
    last_row_idx = len(raw_data) + 4
    for col_idx in range(1, 13):
        ws.cell(row=last_row_idx, column=col_idx).border = bottom_heavy_border
        
    # 4. Auto-fit column widths with padding
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row in [1, 2]:
                continue
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        calculated_width = max(max_len * 1.05 + 4, 10)
        ws.column_dimensions[col_letter].width = min(calculated_width, 40)
        
    # Save the Excel Workbook
    output_path = r"d:\Sao Vàng\Website-SaoVang\BaoCao_ThuVien_CuaNhom\Master_Tracker_HeCuaNhom.xlsx"
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"Master Tracker Excel successfully saved to {output_path}")

if __name__ == "__main__":
    create_master_tracker()