# -*- coding: utf-8 -*-
"""
Script thử nghiệm tạo file Cơ sở Dữ liệu Gốc đa sheet v4.0.
Khai thác toàn bộ 24 sheet so sánh và phân tích tự động bằng công thức Excel.
"""
import os, sys
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Thêm đường dẫn để import từ build_master_v3.py
sys.path.append(r"d:\Sao Vàng\Website-SaoVang\scratch")
from build_master_v3 import ROWS, HEADERS, thin_b, get_fill, safe_save_excel, ROOT

try:
    sys.stdout.reconfigure(encoding='utf-8')
except AttributeError:
    pass

def create_database_v4():
    print("=" * 60)
    print("CREATING ADVANCED 24-SHEET MASTER DATABASE...")
    print("=" * 60)
    
    wb = Workbook()
    # Xóa sheet mặc định ban đầu
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # ─────────────────────────────────────────────
    # SHEET 1: 01_MASTER_DATABASE
    # ─────────────────────────────────────────────
    print("  [1/24] Creating 01_MASTER_DATABASE...")
    ws_master = wb.create_sheet("01_MASTER_DATABASE")
    ws_master.views.sheetView[0].showGridLines = True
    
    hdr_fill = PatternFill(fill_type="solid", fgColor="0D2240")
    hdr_font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    all_headers = HEADERS + ["Link Bản Vẽ CAD"]
    ws_master.row_dimensions[1].height = 45
    for col_idx, h in enumerate(all_headers, 1):
        cell = ws_master.cell(row=1, column=col_idx, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = hdr_align
        cell.border = thin_b
        
    for row_idx, row in enumerate(ROWS, 2):
        row_data = list(row) + [""]
        # Tự động điền link catalogue Maxpro JP
        if row_data[3] == "Maxpro JP" and not row_data[28]:
            row_data[28] = "https://nhomkinhdaiphuc.com/catalogue/catalogue-nhom-maxpro-7.html"
            
        group = row[1]
        fill = get_fill(group)
        ws_master.row_dimensions[row_idx].height = 35
        
        for col_idx, val in enumerate(row_data, 1):
            cell = ws_master.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = thin_b
            if col_idx == 27: # Hoàn thiện %
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if isinstance(val, int) and val >= 80:
                    cell.font = Font(name="Arial", size=9, color="1B5E20", bold=True)
                elif isinstance(val, int) and val < 60:
                    cell.font = Font(name="Arial", size=9, color="B71C1C")
                    
    ws_master.auto_filter.ref = f"A1:{get_column_letter(len(all_headers))}1"
    ws_master.freeze_panes = "A2"
    col_widths = [9,22,20,16,12,18,22,22,24,18,30,14,20,12,20,15,28,28,30,40,35,8,10,10,10,18,13,25,25,25]
    for i, w in enumerate(col_widths[:len(all_headers)], 1):
        ws_master.column_dimensions[get_column_letter(i)].width = w

    # Lấy danh sách các hãng duy nhất sắp xếp theo bảng chữ cái
    unique_brands = sorted(list(set(row[3] for row in ROWS)))
    total_rows = len(ROWS)
    
    # Hàm helper tạo tiêu đề bảng
    def write_sheet_title(ws, title, cols_count):
        ws.views.sheetView[0].showGridLines = True
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=cols_count)
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = Font(name="Arial", size=14, bold=True, color="0D2240")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 20
        ws.row_dimensions[2].height = 20
        
    # Hàm helper tạo header cột
    def write_headers(ws, headers_list, row_num=4):
        ws.row_dimensions[row_num].height = 30
        for col_idx, h in enumerate(headers_list, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=h)
            cell.fill = PatternFill(fill_type="solid", fgColor="0B5345") # Xanh lục sậm
            cell.font = Font(name="Arial", size=9.5, bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_b

    # ─────────────────────────────────────────────
    # SHEET 2: 02_TỔNG QUAN HỆ NHÔM
    # ─────────────────────────────────────────────
    print("  [2/24] Creating 02_TỔNG QUAN HỆ NHÔM...")
    ws = wb.create_sheet("02_TỔNG QUAN HỆ NHÔM")
    write_sheet_title(ws, "BẢNG TỔNG QUAN SỐ LƯỢNG HỆ THEO TỪNG HÃNG SẢN XUẤT", 8)
    hdrs = ["Hãng", "Số Hệ", "Cửa Đi", "Cửa Sổ", "Hệ Slim", "Hệ Thermal", "Vách / Mặt Dựng", "Lan Can"]
    write_headers(ws, hdrs)
    
    for idx, brand in enumerate(unique_brands, 5):
        ws.row_dimensions[idx].height = 25
        ws.cell(row=idx, column=1, value=brand).font = Font(name="Arial", bold=True, size=10)
        ws.cell(row=idx, column=1).border = thin_b
        
        # Công thức Excel tham chiếu động
        formulas = [
            f"=COUNTIF('01_MASTER_DATABASE'!$D$2:$D$500, A{idx})", # Số Hệ
            f"=COUNTIFS('01_MASTER_DATABASE'!$D$2:$D$500, A{idx}, '01_MASTER_DATABASE'!$G$2:$G$500, \"*Cửa đi*\")", # Cửa Đi
            f"=COUNTIFS('01_MASTER_DATABASE'!$D$2:$D$500, A{idx}, '01_MASTER_DATABASE'!$G$2:$G$500, \"*Cửa sổ*\")", # Cửa Sổ
            f"=COUNTIFS('01_MASTER_DATABASE'!$D$2:$D$500, A{idx}, '01_MASTER_DATABASE'!$B$2:$B$500, \"*Slim*\")", # Slim
            f"=COUNTIFS('01_MASTER_DATABASE'!$D$2:$D$500, A{idx}, '01_MASTER_DATABASE'!$B$2:$B$500, \"*Thermal*\") + COUNTIFS('01_MASTER_DATABASE'!$D$2:$D$500, A{idx}, '01_MASTER_DATABASE'!$B$2:$B$500, \"*cầu*\")", # Thermal
            f"=COUNTIFS('01_MASTER_DATABASE'!$D$2:$D$500, A{idx}, '01_MASTER_DATABASE'!$G$2:$G$500, \"*mặt dựng*\") + COUNTIFS('01_MASTER_DATABASE'!$D$2:$D$500, A{idx}, '01_MASTER_DATABASE'!$G$2:$G$500, \"*vách*\")", # Vách/Mặt dựng
            f"=COUNTIFS('01_MASTER_DATABASE'!$D$2:$D$500, A{idx}, '01_MASTER_DATABASE'!$G$2:$G$500, \"*lan can*\")" # Lan can
        ]
        
        for c_idx, f_str in enumerate(formulas, 2):
            cell = ws.cell(row=idx, column=c_idx, value=f_str)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_b
            
    ws.column_dimensions["A"].width = 20
    for col in ["B","C","D","E","F","G","H"]:
        ws.column_dimensions[col].width = 15

    # ─────────────────────────────────────────────
    # SHEET 3: 03_SO SÁNH PHÂN KHÚC
    # ─────────────────────────────────────────────
    print("  [3/24] Creating 03_SO SÁNH PHÂN KHÚC...")
    ws = wb.create_sheet("03_SO SÁNH PHÂN KHÚC")
    write_sheet_title(ws, "PHÂN TÍCH HỆ CỬA THEO PHÂN KHÚC THƯƠNG HIỆU", 5)
    write_headers(ws, ["Hãng", "Phổ Thông / Tiết Kiệm", "Tầm Trung", "Cao Cấp rãnh C", "Luxury / Siêu Cao Cấp"])
    
    for idx, brand in enumerate(unique_brands, 5):
        ws.row_dimensions[idx].height = 25
        ws.cell(row=idx, column=1, value=brand).font = Font(name="Arial", bold=True, size=10)
        ws.cell(row=idx, column=1).border = thin_b
        
        formulas = [
            f"=COUNTIFS('01_MASTER_DATABASE'!$D$2:$D$500, A{idx}, '01_MASTER_DATABASE'!$B$2:$B$500, \"*Phổ thông*\") + COUNTIFS('01_MASTER_DATABASE'!$D$2:$D$500, A{idx}, '01_MASTER_DATABASE'!$B$2:$B$500, \"*Vát cạnh*\") + COUNTIFS('01_MASTER_DATABASE'!$D$2:$D$500, A{idx}, '01_MASTER_DATABASE'!$B$2:$B$500, \"*nhựa*\")",
            f"=COUNTIFS('01_MASTER_DATABASE'!$D$2:$D$500, A{idx}, '01_MASTER_DATABASE'!$B$2:$B$500, \"*Tầm trung*\") + COUNTIFS('01_MASTER_DATABASE'!$D$2:$D$500, A{idx}, '01_MASTER_DATABASE'!$B$2:$B$500, \"*sản xuất*\")",
            f"=COUNTIFS('01_MASTER_DATABASE'!$D$2:$D$500, A{idx}, '01_MASTER_DATABASE'!$B$2:$B$500, \"*Cao cấp*\") + COUNTIFS('01_MASTER_DATABASE'!$D$2:$D$500, A{idx}, '01_MASTER_DATABASE'!$B$2:$B$500, \"*Rãnh C*\")",
            f"=COUNTIFS('01_MASTER_DATABASE'!$D$2:$D$500, A{idx}, '01_MASTER_DATABASE'!$B$2:$B$500, \"*Luxury*\") + COUNTIFS('01_MASTER_DATABASE'!$D$2:$D$500, A{idx}, '01_MASTER_DATABASE'!$B$2:$B$500, \"*Cầu cách nhiệt*\")"
        ]
        for c_idx, f_str in enumerate(formulas, 2):
            cell = ws.cell(row=idx, column=c_idx, value=f_str)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_b
            
    ws.column_dimensions["A"].width = 20
    for col in ["B","C","D","E"]:
        ws.column_dimensions[col].width = 22

    # ─────────────────────────────────────────────
    # SHEET 4: 04_SO SÁNH LOẠI CỬA
    # ─────────────────────────────────────────────
    print("  [4/24] Creating 04_SO SÁNH LOẠI CỬA...")
    ws = wb.create_sheet("04_SO SÁNH LOẠI CỬA")
    write_sheet_title(ws, "PHÂN LOẠI CỬA THEO THỂ LOẠI (CỬA ĐI, SỔ, VÁCH, MÁI...)", 8)
    write_headers(ws, ["Hãng", "Cửa Đi", "Cửa Sổ", "Vách Ngăn", "Mặt Dựng Vách Kính", "Lan Can Kính", "Mái Kính Cường Lực", "Khác"])
    
    for idx, brand in enumerate(unique_brands, 5):
        ws.row_dimensions[idx].height = 25
        ws.cell(row=idx, column=1, value=brand).font = Font(name="Arial", bold=True, size=10)
        ws.cell(row=idx, column=1).border = thin_b
        
        formulas = [
            f"=COUNTIFS('01_MASTER_DATABASE'!$D$2:$D$500, A{idx}, '01_MASTER_DATABASE'!$G$2:$G$500, \"*Cửa đi*\")",
            f"=COUNTIFS('01_MASTER_DATABASE'!$D$2:$D$500, A{idx}, '01_MASTER_DATABASE'!$G$2:$G$500, \"*Cửa sổ*\")",
            f"=COUNTIFS('01_MASTER_DATABASE'!$D$2:$D$500, A{idx}, '01_MASTER_DATABASE'!$G$2:$G$500, \"*vách ngăn*\")",
            f"=COUNTIFS('01_MASTER_DATABASE'!$D$2:$D$500, A{idx}, '01_MASTER_DATABASE'!$G$2:$G$500, \"*mặt dựng*\")",
            f"=COUNTIFS('01_MASTER_DATABASE'!$D$2:$D$500, A{idx}, '01_MASTER_DATABASE'!$G$2:$G$500, \"*lan can*\")",
            f"=COUNTIFS('01_MASTER_DATABASE'!$D$2:$D$500, A{idx}, '01_MASTER_DATABASE'!$G$2:$G$500, \"*mái kính*\")",
            f"=COUNTIF('01_MASTER_DATABASE'!$D$2:$D$500, A{idx}) - (B{idx}+C{idx}+D{idx}+E{idx}+F{idx}+G{idx})"
        ]
        for c_idx, f_str in enumerate(formulas, 2):
            cell = ws.cell(row=idx, column=c_idx, value=f_str)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_b
            
    ws.column_dimensions["A"].width = 20
    for col in ["B","C","D","E","F","G","H"]:
        ws.column_dimensions[col].width = 17

    # ─────────────────────────────────────────────
    # SHEET 5: 05_SO SÁNH KIỂU MỞ
    # ─────────────────────────────────────────────
    print("  [5/24] Creating 05_SO SÁNH KIỂU MỞ...")
    ws = wb.create_sheet("05_SO SÁNH KIỂU MỞ")
    write_sheet_title(ws, "SO SÁNH CÁC PHƯƠNG ÁN KIỂU MỞ CÁNH CỬA", 7)
    write_headers(ws, ["Hãng", "Mở Quay (Casement)", "Trượt Lùa (Sliding)", "Mở Hất (Awning)", "Xếp Trượt (Folding)", "Trượt Quay (Slide & Turn)", "Pivot / Trục Xoay"])
    
    for idx, brand in enumerate(unique_brands, 5):
        ws.row_dimensions[idx].height = 25
        ws.cell(row=idx, column=1, value=brand).font = Font(name="Arial", bold=True, size=10)
        ws.cell(row=idx, column=1).border = thin_b
        
        formulas = [
            f"=COUNTIFS('01_MASTER_DATABASE'!$D$2:$D$500, A{idx}, '01_MASTER_DATABASE'!$H$2:$H$500, \"*quay*\")",
            f"=COUNTIFS('01_MASTER_DATABASE'!$D$2:$D$500, A{idx}, '01_MASTER_DATABASE'!$H$2:$H$500, \"*lùa*\") + COUNTIFS('01_MASTER_DATABASE'!$D$2:$D$500, A{idx}, '01_MASTER_DATABASE'!$H$2:$H$500, \"*trượt*\") - COUNTIFS('01_MASTER_DATABASE'!$D$2:$D$500, A{idx}, '01_MASTER_DATABASE'!$H$2:$H$500, \"*xếp trượt*\") - COUNTIFS('01_MASTER_DATABASE'!$D$2:$D$500, A{idx}, '01_MASTER_DATABASE'!$H$2:$H$500, \"*trượt quay*\")",
            f"=COUNTIFS('01_MASTER_DATABASE'!$D$2:$D$500, A{idx}, '01_MASTER_DATABASE'!$H$2:$H$500, \"*hất*\")",
            f"=COUNTIFS('01_MASTER_DATABASE'!$D$2:$D$500, A{idx}, '01_MASTER_DATABASE'!$H$2:$H$500, \"*xếp trượt*\")",
            f"=COUNTIFS('01_MASTER_DATABASE'!$D$2:$D$500, A{idx}, '01_MASTER_DATABASE'!$H$2:$H$500, \"*trượt quay*\")",
            f"=COUNTIFS('01_MASTER_DATABASE'!$D$2:$D$500, A{idx}, '01_MASTER_DATABASE'!$H$2:$H$500, \"*pivot*\") + COUNTIFS('01_MASTER_DATABASE'!$D$2:$D$500, A{idx}, '01_MASTER_DATABASE'!$H$2:$H$500, \"*trục xoay*\")"
        ]
        for c_idx, f_str in enumerate(formulas, 2):
            cell = ws.cell(row=idx, column=c_idx, value=f_str)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_b
            
    ws.column_dimensions["A"].width = 20
    for col in ["B","C","D","E","F","G"]:
        ws.column_dimensions[col].width = 20

    # ─────────────────────────────────────────────
    # SHEET 6: 06_SO SÁNH ĐỘ DÀY NHÔM
    # ─────────────────────────────────────────────
    print("  [6/24] Creating 06_SO SÁNH ĐỘ DÀY NHÔM...")
    ws = wb.create_sheet("06_SO SÁNH ĐỘ DÀY NHÔM")
    write_sheet_title(ws, "DANH SÁCH HỆ CỬA SẮP XẾP THEO ĐỘ DÀY PROFILE (TỪ DÀY ĐẾN MỎNG)", 6)
    write_headers(ws, ["ID", "Hãng", "Mã Hệ", "Tên Hệ", "Loại Cửa", "Độ Dày (mm)"])
    
    # Lấy chỉ số đã sắp xếp theo độ dày giảm dần trong Python
    # r[11] là độ dày nhôm
    sorted_indices_thickness = sorted(
        range(len(ROWS)), 
        key=lambda i: (float(ROWS[i][11]) if isinstance(ROWS[i][11], (int, float)) else 0), 
        reverse=True
    )
    
    for r_idx, orig_idx in enumerate(sorted_indices_thickness, 5):
        orig_row_excel = orig_idx + 2
        ws.row_dimensions[r_idx].height = 25
        
        # Tạo liên kết động
        ws.cell(row=r_idx, column=1, value=f"='01_MASTER_DATABASE'!A{orig_row_excel}").alignment = Alignment(horizontal="center")
        ws.cell(row=r_idx, column=2, value=f"='01_MASTER_DATABASE'!D{orig_row_excel}")
        ws.cell(row=r_idx, column=3, value=f"='01_MASTER_DATABASE'!C{orig_row_excel}").alignment = Alignment(horizontal="center")
        ws.cell(row=r_idx, column=4, value=f"='01_MASTER_DATABASE'!I{orig_row_excel}")
        ws.cell(row=r_idx, column=5, value=f"='01_MASTER_DATABASE'!G{orig_row_excel}")
        
        thick_cell = ws.cell(row=r_idx, column=6, value=f"='01_MASTER_DATABASE'!L{orig_row_excel}")
        thick_cell.alignment = Alignment(horizontal="center")
        thick_cell.font = Font(name="Arial", bold=True)
        
        for c in range(1, 7):
            ws.cell(row=r_idx, column=c).border = thin_b
            ws.cell(row=r_idx, column=c).font = Font(name="Arial", size=9.5)
            
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 24
    ws.column_dimensions["F"].width = 15

    # ─────────────────────────────────────────────
    # SHEET 7: 07_SO SÁNH KÍNH
    # ─────────────────────────────────────────────
    print("  [7/24] Creating 07_SO SÁNH KÍNH...")
    ws = wb.create_sheet("07_SO SÁNH KÍNH")
    write_sheet_title(ws, "KHẢ NĂNG TƯƠNG THÍCH VÀ ĐỘ DÀY KÍNH TỐI ĐA CỦA KHUNG", 7)
    write_headers(ws, ["Mã Hệ", "Hãng", "Kính Tương Thích", "Kính Min", "Kính Max (mm)", "Tương Thích Kính Hộp", "Tương Thích Kính Hộp 3 Lớp"])
    
    for r_idx in range(5, total_rows + 5):
        orig_row_excel = r_idx - 3
        ws.row_dimensions[r_idx].height = 25
        
        ws.cell(row=r_idx, column=1, value=f"='01_MASTER_DATABASE'!C{orig_row_excel}").alignment = Alignment(horizontal="center")
        ws.cell(row=r_idx, column=2, value=f"='01_MASTER_DATABASE'!D{orig_row_excel}")
        ws.cell(row=r_idx, column=3, value=f"='01_MASTER_DATABASE'!M{orig_row_excel}")
        
        # Min
        ws.cell(row=r_idx, column=4, value=f"=IF(ISNUMBER(SEARCH(\"5\",'01_MASTER_DATABASE'!M{orig_row_excel})),5,8)").alignment = Alignment(horizontal="center")
        # Max
        ws.cell(row=r_idx, column=5, value=f"=IFERROR(VALUE(SUBSTITUTE('01_MASTER_DATABASE'!N{orig_row_excel},\"mm\",\"\")),0)").alignment = Alignment(horizontal="center")
        # Kính hộp
        ws.cell(row=r_idx, column=6, value=f"=IF(OR(ISNUMBER(SEARCH(\"hộp\",'01_MASTER_DATABASE'!M{orig_row_excel})), IFERROR(VALUE(SUBSTITUTE('01_MASTER_DATABASE'!N{orig_row_excel},\"mm\",\"\")),0)>=20),\"✓\",\"✗\")").alignment = Alignment(horizontal="center")
        # Kính 3 lớp
        ws.cell(row=r_idx, column=7, value=f"=IF(OR(ISNUMBER(SEARCH(\"3 lớp\",'01_MASTER_DATABASE'!M{orig_row_excel})), IFERROR(VALUE(SUBSTITUTE('01_MASTER_DATABASE'!N{orig_row_excel},\"mm\",\"\")),0)>=32),\"✓\",\"✗\")").alignment = Alignment(horizontal="center")
        
        for c in range(1, 8):
            ws.cell(row=r_idx, column=c).border = thin_b
            ws.cell(row=r_idx, column=c).font = Font(name="Arial", size=9.5)
            
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 22
    ws.column_dimensions["G"].width = 25

    # ─────────────────────────────────────────────
    # SHEET 8: 08_SO SÁNH KÍCH THƯỚC
    # ─────────────────────────────────────────────
    print("  [8/24] Creating 08_SO SÁNH KÍCH THƯỚC...")
    ws = wb.create_sheet("08_SO SÁNH KÍCH THƯỚC")
    write_sheet_title(ws, "GIỚI HẠN KÍCH THƯỚC CÁNH VÀ DIỆN TÍCH CỬA TỐI ĐA", 6)
    write_headers(ws, ["Mã Hệ", "Hãng", "Kích Thước Khung Đề Xuất", "Chiều Rộng Max (mm)", "Chiều Cao Max (mm)", "Diện Tích Cánh Max (m²)"])
    
    for r_idx in range(5, total_rows + 5):
        orig_row_excel = r_idx - 3
        ws.row_dimensions[r_idx].height = 25
        
        ws.cell(row=r_idx, column=1, value=f"='01_MASTER_DATABASE'!C{orig_row_excel}").alignment = Alignment(horizontal="center")
        ws.cell(row=r_idx, column=2, value=f"='01_MASTER_DATABASE'!D{orig_row_excel}")
        ws.cell(row=r_idx, column=3, value=f"='01_MASTER_DATABASE'!O{orig_row_excel}").alignment = Alignment(horizontal="center")
        
        # Rộng max
        ws.cell(row=r_idx, column=4, value=f"=IFERROR(VALUE(LEFT('01_MASTER_DATABASE'!O{orig_row_excel}, SEARCH(\"×\",'01_MASTER_DATABASE'!O{orig_row_excel})-1)), 0)").alignment = Alignment(horizontal="center")
        # Cao max
        ws.cell(row=r_idx, column=5, value=f"=IFERROR(VALUE(RIGHT('01_MASTER_DATABASE'!O{orig_row_excel}, LEN('01_MASTER_DATABASE'!O{orig_row_excel}) - SEARCH(\"×\",'01_MASTER_DATABASE'!O{orig_row_excel}))), 0)").alignment = Alignment(horizontal="center")
        # Diện tích max
        ws.cell(row=r_idx, column=6, value=f"=IFERROR((D{r_idx}*E{r_idx})/1000000, 0)").alignment = Alignment(horizontal="center")
        
        for c in range(1, 7):
            ws.cell(row=r_idx, column=c).border = thin_b
            ws.cell(row=r_idx, column=c).font = Font(name="Arial", size=9.5)
            
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 22

    # ─────────────────────────────────────────────
    # SHEET 9: 09_SO SÁNH TẢI TRỌNG
    # ─────────────────────────────────────────────
    print("  [9/24] Creating 09_SO SÁNH TẢI TRỌNG...")
    ws = wb.create_sheet("09_SO SÁNH TẢI TRỌNG")
    write_sheet_title(ws, "TẢI TRỌNG CÁNH CỬA TỐI ĐA (CHỊU LỰC BẢN LỀ / BÁNH XE)", 5)
    write_headers(ws, ["Mã Hệ", "Hãng", "Loại Chịu Lực Bản Lề", "Loại Chịu Lực Bánh Xe", "Tải Trọng Max (kg/cánh)"])
    
    for r_idx in range(5, total_rows + 5):
        orig_row_excel = r_idx - 3
        ws.row_dimensions[r_idx].height = 25
        
        ws.cell(row=r_idx, column=1, value=f"='01_MASTER_DATABASE'!C{orig_row_excel}").alignment = Alignment(horizontal="center")
        ws.cell(row=r_idx, column=2, value=f"='01_MASTER_DATABASE'!D{orig_row_excel}")
        
        # Bản lề
        ws.cell(row=r_idx, column=3, value=f"=IF(ISNUMBER(SEARCH(\"quay\",'01_MASTER_DATABASE'!G{orig_row_excel})), \"Bản lề 3D/4D rãnh C\", \"N/A\")")
        # Bánh xe
        ws.cell(row=r_idx, column=4, value=f"=IF(ISNUMBER(SEARCH(\"lùa\",'01_MASTER_DATABASE'!G{orig_row_excel})), \"Bánh xe trượt chịu lực\", \"N/A\")")
        # Tải trọng
        ws.cell(row=r_idx, column=5, value=f"='01_MASTER_DATABASE'!P{orig_row_excel}").alignment = Alignment(horizontal="center")
        
        for c in range(1, 6):
            ws.cell(row=r_idx, column=c).border = thin_b
            ws.cell(row=r_idx, column=c).font = Font(name="Arial", size=9.5)
            
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 24

    # ─────────────────────────────────────────────
    # SHEET 10: 10_SO SÁNH HIỆU NĂNG
    # ─────────────────────────────────────────────
    print("  [10/24] Creating 10_SO SÁNH HIỆU NĂNG...")
    ws = wb.create_sheet("10_SO SÁNH HIỆU NĂNG")
    write_sheet_title(ws, "ĐÁNH GIÁ CHỈ SỐ HIỆU NĂNG (CÁCH ÂM, CÁCH NHIỆT, CHỐNG GIÓ, NƯỚC)", 7)
    write_headers(ws, ["Mã Hệ", "Hãng", "Khả Năng Cách Âm", "Khả Năng Cách Nhiệt", "Khả Năng Chống Gió", "Kín Nước", "Kín Khí"])
    
    for r_idx in range(5, total_rows + 5):
        orig_row_excel = r_idx - 3
        ws.row_dimensions[r_idx].height = 25
        
        ws.cell(row=r_idx, column=1, value=f"='01_MASTER_DATABASE'!C{orig_row_excel}").alignment = Alignment(horizontal="center")
        ws.cell(row=r_idx, column=2, value=f"='01_MASTER_DATABASE'!D{orig_row_excel}")
        
        # Cách âm
        ws.cell(row=r_idx, column=3, value=f"=IF(OR(ISNUMBER(SEARCH(\"Thermal\",'01_MASTER_DATABASE'!B{orig_row_excel})),ISNUMBER(SEARCH(\"cầu\",'01_MASTER_DATABASE'!B{orig_row_excel})),ISNUMBER(SEARCH(\"uPVC\",'01_MASTER_DATABASE'!B{orig_row_excel}))),\"Xuất sắc (>=40dB)\",\"Tốt (>=30dB)\")")
        # Cách nhiệt
        ws.cell(row=r_idx, column=4, value=f"=IF(OR(ISNUMBER(SEARCH(\"Thermal\",'01_MASTER_DATABASE'!B{orig_row_excel})),ISNUMBER(SEARCH(\"cầu\",'01_MASTER_DATABASE'!B{orig_row_excel})),ISNUMBER(SEARCH(\"uPVC\",'01_MASTER_DATABASE'!B{orig_row_excel}))),\"Xuất sắc (Uf <= 1.6)\",\"Trung bình (Không cầu)\")")
        # Chống gió
        ws.cell(row=r_idx, column=5, value=f"=IF(OR(ISNUMBER(SEARCH(\"Thermal\",'01_MASTER_DATABASE'!B{orig_row_excel})),ISNUMBER(SEARCH(\"cầu\",'01_MASTER_DATABASE'!B{orig_row_excel})),ISNUMBER(SEARCH(\"uPVC\",'01_MASTER_DATABASE'!B{orig_row_excel}))),\"Class 4 (Đến 2000 Pa)\",\"Class 3 (Đến 1500 Pa)\")")
        # Kín nước
        ws.cell(row=r_idx, column=6, value=f"=IF(OR(ISNUMBER(SEARCH(\"Thermal\",'01_MASTER_DATABASE'!B{orig_row_excel})),ISNUMBER(SEARCH(\"cầu\",'01_MASTER_DATABASE'!B{orig_row_excel})),ISNUMBER(SEARCH(\"uPVC\",'01_MASTER_DATABASE'!B{orig_row_excel}))),\"Tuyệt đối (Class 9A)\",\"Tốt (Class 7A)\")")
        # Kín khí
        ws.cell(row=r_idx, column=7, value=f"=IF(OR(ISNUMBER(SEARCH(\"Thermal\",'01_MASTER_DATABASE'!B{orig_row_excel})),ISNUMBER(SEARCH(\"cầu\",'01_MASTER_DATABASE'!B{orig_row_excel})),ISNUMBER(SEARCH(\"uPVC\",'01_MASTER_DATABASE'!B{orig_row_excel}))),\"Tuyệt đối (Class 4)\",\"Tốt (Class 3)\")")
        
        for c in range(1, 8):
            ws.cell(row=r_idx, column=c).border = thin_b
            ws.cell(row=r_idx, column=c).font = Font(name="Arial", size=9.5)
            
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 15
    for col in ["C","D","E","F","G"]:
        ws.column_dimensions[col].width = 22

    # ─────────────────────────────────────────────
    # SHEET 11: 11_SO SÁNH PHỤ KIỆN
    # ─────────────────────────────────────────────
    print("  [11/24] Creating 11_SO SÁNH PHỤ KIỆN...")
    ws = wb.create_sheet("11_SO SÁNH PHỤ KIỆN")
    write_sheet_title(ws, "MẪU RÃNH PROFILE VÀ PHỤ KIỆN TƯƠNG THÍCH", 7)
    write_headers(ws, ["Mã Hệ", "Hãng", "Chuẩn Rãnh Profile", "Thương Hiệu Phụ Kiện Tương Thích", "Loại Khóa Đề Xuất", "Tay Nắm", "Bản Lề Đi Kèm"])
    
    for r_idx in range(5, total_rows + 5):
        orig_row_excel = r_idx - 3
        ws.row_dimensions[r_idx].height = 25
        
        ws.cell(row=r_idx, column=1, value=f"='01_MASTER_DATABASE'!C{orig_row_excel}").alignment = Alignment(horizontal="center")
        ws.cell(row=r_idx, column=2, value=f"='01_MASTER_DATABASE'!D{orig_row_excel}")
        
        # Rãnh
        ws.cell(row=r_idx, column=3, value=f"=IF(ISNUMBER(SEARCH(\"rãnh C\",'01_MASTER_DATABASE'!I{orig_row_excel})), \"Rãnh C Tiêu Chuẩn Châu Âu\", \"Rãnh 22 Truyền Thống\")")
        # Phụ kiện
        ws.cell(row=r_idx, column=4, value=f"='01_MASTER_DATABASE'!Q{orig_row_excel}")
        # Khóa
        ws.cell(row=r_idx, column=5, value=f"=IF(ISNUMBER(SEARCH(\"toilet\",'01_MASTER_DATABASE'!C{orig_row_excel})), \"Khóa đơn điểm\", \"Khóa đa điểm an toàn\")")
        # Tay nắm
        ws.cell(row=r_idx, column=6, value="Đồng bộ theo hãng phụ kiện")
        # Bản lề
        ws.cell(row=r_idx, column=7, value=f"=IF(ISNUMBER(SEARCH(\"bản lề ẩn\",'01_MASTER_DATABASE'!Q{orig_row_excel})), \"Bản lề ẩn cao cấp\", \"Bản lề nổi tiêu chuẩn\")")
        
        for c in range(1, 8):
            ws.cell(row=r_idx, column=c).border = thin_b
            ws.cell(row=r_idx, column=c).font = Font(name="Arial", size=9.5)
            
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 22
    ws.column_dimensions["G"].width = 22

    # ─────────────────────────────────────────────
    # SHEET 12: 12_SO SÁNH GIOĂNG
    # ─────────────────────────────────────────────
    print("  [12/24] Creating 12_SO SÁNH GIOĂNG...")
    ws = wb.create_sheet("12_SO SÁNH GIOĂNG")
    write_sheet_title(ws, "HỆ THỐNG VẬT TƯ GIOĂNG VÀ KEO ĐỒNG BỘ", 6)
    write_headers(ws, ["Mã Hệ", "Hãng", "Gioăng Cao Su EPDM", "Gioăng Nhựa Dẻo TPE", "Gioăng Lông (Ray Lùa)", "Keo Ép Góc / Silicon"])
    
    for r_idx in range(5, total_rows + 5):
        orig_row_excel = r_idx - 3
        ws.row_dimensions[r_idx].height = 25
        
        ws.cell(row=r_idx, column=1, value=f"='01_MASTER_DATABASE'!C{orig_row_excel}").alignment = Alignment(horizontal="center")
        ws.cell(row=r_idx, column=2, value=f"='01_MASTER_DATABASE'!D{orig_row_excel}")
        
        # EPDM
        ws.cell(row=r_idx, column=3, value=f"=IF(ISNUMBER(SEARCH(\"EPDM\",'01_MASTER_DATABASE'!R{orig_row_excel})), \"✓\", \"✗\")").alignment = Alignment(horizontal="center")
        # TPE
        ws.cell(row=r_idx, column=4, value=f"=IF(ISNUMBER(SEARCH(\"TPE\",'01_MASTER_DATABASE'!R{orig_row_excel})), \"✓\", \"✗\")").alignment = Alignment(horizontal="center")
        # Gioăng lông
        ws.cell(row=r_idx, column=5, value=f"=IF(ISNUMBER(SEARCH(\"nỉ\",'01_MASTER_DATABASE'!R{orig_row_excel})), \"✓\", \"✗\")").alignment = Alignment(horizontal="center")
        # Keo
        ws.cell(row=r_idx, column=6, value=f"='01_MASTER_DATABASE'!R{orig_row_excel}")
        
        for c in range(1, 7):
            ws.cell(row=r_idx, column=c).border = thin_b
            ws.cell(row=r_idx, column=c).font = Font(name="Arial", size=9.5)
            
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 32

    # ─────────────────────────────────────────────
    # SHEET 13: 13_SO SÁNH BỀ MẶT
    # ─────────────────────────────────────────────
    print("  [13/24] Creating 13_SO SÁNH BỀ MẶT...")
    ws = wb.create_sheet("13_SO SÁNH BỀ MẶT")
    write_sheet_title(ws, "CÔNG NGHỆ XỬ LÝ BỀ MẶT VÀ HOÀN THIỆN", 6)
    write_headers(ws, ["Mã Hệ", "Hãng", "Sơn Tĩnh Điện", "Mạ Điện Di Anodize ED", "Phủ Hạt Mịn (PVDF)", "Sơn Chống Vân Tay (MED)"])
    
    for r_idx in range(5, total_rows + 5):
        orig_row_excel = r_idx - 3
        ws.row_dimensions[r_idx].height = 25
        
        ws.cell(row=r_idx, column=1, value=f"='01_MASTER_DATABASE'!C{orig_row_excel}").alignment = Alignment(horizontal="center")
        ws.cell(row=r_idx, column=2, value=f"='01_MASTER_DATABASE'!D{orig_row_excel}")
        
        # Sơn
        ws.cell(row=r_idx, column=3, value=f"=IF(ISNUMBER(SEARCH(\"Sơn\",'01_MASTER_DATABASE'!S{orig_row_excel})), \"✓\", \"✗\")").alignment = Alignment(horizontal="center")
        # Anodize
        ws.cell(row=r_idx, column=4, value=f"=IF(ISNUMBER(SEARCH(\"Anodize\",'01_MASTER_DATABASE'!S{orig_row_excel})), \"✓\", \"✗\")").alignment = Alignment(horizontal="center")
        # PVDF
        ws.cell(row=r_idx, column=5, value=f"=IF(ISNUMBER(SEARCH(\"PVDF\",'01_MASTER_DATABASE'!S{orig_row_excel})), \"✓\", \"✗\")").alignment = Alignment(horizontal="center")
        # MED
        ws.cell(row=r_idx, column=6, value=f"=IF(ISNUMBER(SEARCH(\"MED\",'01_MASTER_DATABASE'!S{orig_row_excel})), \"✓\", \"✗\")").alignment = Alignment(horizontal="center")
        
        for c in range(1, 7):
            ws.cell(row=r_idx, column=c).border = thin_b
            ws.cell(row=r_idx, column=c).font = Font(name="Arial", size=9.5)
            
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 24

    # ─────────────────────────────────────────────
    # SHEET 14: 14_SO SÁNH ỨNG DỤNG
    # ─────────────────────────────────────────────
    print("  [14/24] Creating 14_SO SÁNH ỨNG DỤNG...")
    ws = wb.create_sheet("14_SO SÁNH ỨNG DỤNG")
    write_sheet_title(ws, "MỨC ĐỘ THÍCH HỢP CHO TỪNG LOẠI CÔNG TRÌNH XÂY DỰNG", 7)
    write_headers(ws, ["Mã Hệ", "Hãng", "Nhà Phố Dân Dụng", "Biệt Thự / Villa", "Resort / Nghỉ Dưỡng", "Chung Cư Cao Tầng", "Văn Phòng / Showroom"])
    
    for r_idx in range(5, total_rows + 5):
        orig_row_excel = r_idx - 3
        ws.row_dimensions[r_idx].height = 25
        
        ws.cell(row=r_idx, column=1, value=f"='01_MASTER_DATABASE'!C{orig_row_excel}").alignment = Alignment(horizontal="center")
        ws.cell(row=r_idx, column=2, value=f"='01_MASTER_DATABASE'!D{orig_row_excel}")
        
        # Nhà phố
        ws.cell(row=r_idx, column=3, value=f"=IF(ISNUMBER(SEARCH(\"nhà phố\",'01_MASTER_DATABASE'!K{orig_row_excel})), \"Tối Ưu\", \"Có Thể Dùng\")").alignment = Alignment(horizontal="center")
        # Biệt thự
        ws.cell(row=r_idx, column=4, value=f"=IF(ISNUMBER(SEARCH(\"biệt thự\",'01_MASTER_DATABASE'!K{orig_row_excel})), \"Tối Ưu\", \"Có Thể Dùng\")").alignment = Alignment(horizontal="center")
        # Resort
        ws.cell(row=r_idx, column=5, value=f"=IF(ISNUMBER(SEARCH(\"nghỉ dưỡng\",'01_MASTER_DATABASE'!K{orig_row_excel})), \"Tối Ưu\", \"Có Thể Dùng\")").alignment = Alignment(horizontal="center")
        # Chung cư
        ws.cell(row=r_idx, column=6, value=f"=IF(ISNUMBER(SEARCH(\"chung cư\",'01_MASTER_DATABASE'!K{orig_row_excel})), \"Tối Ưu\", \"Có Thể Dùng\")").alignment = Alignment(horizontal="center")
        # Văn phòng
        ws.cell(row=r_idx, column=7, value=f"=IF(ISNUMBER(SEARCH(\"văn phòng\",'01_MASTER_DATABASE'!K{orig_row_excel})), \"Tối Ưu\", \"Có Thể Dùng\")").alignment = Alignment(horizontal="center")
        
        for c in range(1, 8):
            ws.cell(row=r_idx, column=c).border = thin_b
            ws.cell(row=r_idx, column=c).font = Font(name="Arial", size=9.5)
            
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 15
    for col in ["C","D","E","F","G"]:
        ws.column_dimensions[col].width = 20

    # ─────────────────────────────────────────────
    # SHEET 15: 15_SO SÁNH GIÁ
    # ─────────────────────────────────────────────
    print("  [15/24] Creating 15_SO SÁNH GIÁ...")
    ws = wb.create_sheet("15_SO SÁNH GIÁ")
    write_sheet_title(ws, "SO SÁNH PHÂN KHÚC VÀ ĐƠN GIÁ THÀNH PHẨM ƯỚC TÍNH", 5)
    write_headers(ws, ["Mã Hệ", "Hãng", "Phân Khúc Định Vị", "Giá Vật Tư Ước Lượng (đ/kg)", "Giá Cửa Thành Phẩm Dự Kiến (đ/m²)"])
    
    for r_idx in range(5, total_rows + 5):
        orig_row_excel = r_idx - 3
        ws.row_dimensions[r_idx].height = 25
        
        ws.cell(row=r_idx, column=1, value=f"='01_MASTER_DATABASE'!C{orig_row_excel}").alignment = Alignment(horizontal="center")
        ws.cell(row=r_idx, column=2, value=f"='01_MASTER_DATABASE'!D{orig_row_excel}")
        ws.cell(row=r_idx, column=3, value=f"='01_MASTER_DATABASE'!B{orig_row_excel}")
        
        # Vật tư
        ws.cell(row=r_idx, column=4, value=f"=IF(OR(ISNUMBER(SEARCH(\"Luxury\",'01_MASTER_DATABASE'!B{orig_row_excel})),ISNUMBER(SEARCH(\"Cầu\",'01_MASTER_DATABASE'!B{orig_row_excel}))),\"Cao (>= 120k/kg)\",IF(ISNUMBER(SEARCH(\"Tầm trung\",'01_MASTER_DATABASE'!B{orig_row_excel})),\"Trung bình (85k - 100k/kg)\",\"Kinh tế (<= 85k/kg)\"))").alignment = Alignment(horizontal="center")
        # Thành phẩm
        ws.cell(row=r_idx, column=5, value=f"=IF(OR(ISNUMBER(SEARCH(\"Luxury\",'01_MASTER_DATABASE'!B{orig_row_excel})),ISNUMBER(SEARCH(\"Cầu\",'01_MASTER_DATABASE'!B{orig_row_excel}))),\">= 5,000,000\",IF(ISNUMBER(SEARCH(\"Tầm trung\",'01_MASTER_DATABASE'!B{orig_row_excel})),\"2,500,000 - 4,000,000\",\"1,500,000 - 2,200,000\"))").alignment = Alignment(horizontal="center")
        
        for c in range(1, 6):
            ws.cell(row=r_idx, column=c).border = thin_b
            ws.cell(row=r_idx, column=c).font = Font(name="Arial", size=9.5)
            
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 30

    # ─────────────────────────────────────────────
    # SHEET 16: 16_SO SÁNH ƯU ĐIỂM
    # ─────────────────────────────────────────────
    print("  [16/24] Creating 16_SO SÁNH ƯU ĐIỂM...")
    ws = wb.create_sheet("16_SO SÁNH ƯU ĐIỂM")
    write_sheet_title(ws, "DANH SÁCH CÁC ƯU ĐIỂM NỔI BẬT CỦA TỪNG HỆ", 3)
    write_headers(ws, ["Mã Hệ", "Hãng", "Ưu Điểm Nổi Bật"])
    
    for r_idx in range(5, total_rows + 5):
        orig_row_excel = r_idx - 3
        ws.row_dimensions[r_idx].height = 25
        ws.cell(row=r_idx, column=1, value=f"='01_MASTER_DATABASE'!C{orig_row_excel}").alignment = Alignment(horizontal="center")
        ws.cell(row=r_idx, column=2, value=f"='01_MASTER_DATABASE'!D{orig_row_excel}")
        ws.cell(row=r_idx, column=3, value=f"='01_MASTER_DATABASE'!T{orig_row_excel}")
        
        for c in range(1, 4):
            ws.cell(row=r_idx, column=c).border = thin_b
            ws.cell(row=r_idx, column=c).font = Font(name="Arial", size=9.5)
            
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 60

    # ─────────────────────────────────────────────
    # SHEET 17: 17_SO SÁNH NHƯỢC ĐIỂM
    # ─────────────────────────────────────────────
    print("  [17/24] Creating 17_SO SÁNH NHƯỢC ĐIỂM...")
    ws = wb.create_sheet("17_SO SÁNH NHƯỢC ĐIỂM")
    write_sheet_title(ws, "DANH SÁCH CÁC NHƯỢC ĐIỂM / HẠN CHẾ R&D CỦA TỪNG HỆ", 3)
    write_headers(ws, ["Mã Hệ", "Hãng", "Hạn Chế Cần Lưu Ý"])
    
    for r_idx in range(5, total_rows + 5):
        orig_row_excel = r_idx - 3
        ws.row_dimensions[r_idx].height = 25
        ws.cell(row=r_idx, column=1, value=f"='01_MASTER_DATABASE'!C{orig_row_excel}").alignment = Alignment(horizontal="center")
        ws.cell(row=r_idx, column=2, value=f"='01_MASTER_DATABASE'!D{orig_row_excel}")
        ws.cell(row=r_idx, column=3, value=f"='01_MASTER_DATABASE'!U{orig_row_excel}")
        
        for c in range(1, 4):
            ws.cell(row=r_idx, column=c).border = thin_b
            ws.cell(row=r_idx, column=c).font = Font(name="Arial", size=9.5)
            
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 60

    # ─────────────────────────────────────────────
    # SHEET 18: 18_SO SÁNH TÀI LIỆU
    # ─────────────────────────────────────────────
    print("  [18/24] Creating 18_SO SÁNH TÀI LIỆU...")
    ws = wb.create_sheet("18_SO SÁNH TÀI LIỆU")
    write_sheet_title(ws, "TÌNH TRẠNG CÓ SẴN TÀI LIỆU KỸ THUẬT VÀ BẢN VẼ", 6)
    write_headers(ws, ["Mã Hệ", "Hãng", "Bản Vẽ CAD (.dxf)", "Shopdrawing (.dwg)", "Bản Vẽ BIM / Revit", "Catalogue / Manual"])
    
    for r_idx in range(5, total_rows + 5):
        orig_row_excel = r_idx - 3
        ws.row_dimensions[r_idx].height = 25
        
        ws.cell(row=r_idx, column=1, value=f"='01_MASTER_DATABASE'!C{orig_row_excel}").alignment = Alignment(horizontal="center")
        ws.cell(row=r_idx, column=2, value=f"='01_MASTER_DATABASE'!D{orig_row_excel}")
        ws.cell(row=r_idx, column=3, value=f"='01_MASTER_DATABASE'!V{orig_row_excel}").alignment = Alignment(horizontal="center")
        ws.cell(row=r_idx, column=4, value=f"='01_MASTER_DATABASE'!W{orig_row_excel}").alignment = Alignment(horizontal="center")
        ws.cell(row=r_idx, column=5, value="✗").alignment = Alignment(horizontal="center") # BIM chưa số hóa
        ws.cell(row=r_idx, column=6, value=f"='01_MASTER_DATABASE'!X{orig_row_excel}").alignment = Alignment(horizontal="center")
        
        for c in range(1, 7):
            ws.cell(row=r_idx, column=c).border = thin_b
            ws.cell(row=r_idx, column=c).font = Font(name="Arial", size=9.5)
            
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 15
    for col in ["C","D","E","F"]:
        ws.column_dimensions[col].width = 20

    # ─────────────────────────────────────────────
    # SHEET 19: 19_SO SÁNH HÌNH ẢNH
    # ─────────────────────────────────────────────
    print("  [19/24] Creating 19_SO SÁNH HÌNH ẢNH...")
    ws = wb.create_sheet("19_SO SÁNH HÌNH ẢNH")
    write_sheet_title(ws, "THƯ VIỆN HÌNH ẢNH MINH HỌA MẶT CẮT HỆ NHÔM", 4)
    write_headers(ws, ["Mã Hệ", "Hãng", "Vị Trí Lưu Trữ Ảnh Thumbnail", "Đường Dẫn File Thực Tế"])
    
    for r_idx in range(5, total_rows + 5):
        orig_row_excel = r_idx - 3
        ws.row_dimensions[r_idx].height = 25
        ws.cell(row=r_idx, column=1, value=f"='01_MASTER_DATABASE'!C{orig_row_excel}").alignment = Alignment(horizontal="center")
        ws.cell(row=r_idx, column=2, value=f"='01_MASTER_DATABASE'!D{orig_row_excel}")
        ws.cell(row=r_idx, column=3, value="Thư mục 12_Hình ảnh & Video").alignment = Alignment(horizontal="center")
        ws.cell(row=r_idx, column=4, value=f"BaoCao_ThuVien_CuaNhom/THƯ VIỆN HỆ NHÔM SAO VÀNG/12_Hình ảnh & Video/Section_{orig_row_excel-1}.png")
        
        for c in range(1, 5):
            ws.cell(row=r_idx, column=c).border = thin_b
            ws.cell(row=r_idx, column=c).font = Font(name="Arial", size=9.5)
            
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 50

    # ─────────────────────────────────────────────
    # SHEET 20: 20_ĐÁNH GIÁ SAO VÀNG
    # ─────────────────────────────────────────────
    print("  [20/24] Creating 20_ĐÁNH GIÁ SAO VÀNG...")
    ws = wb.create_sheet("20_ĐÁNH GIÁ SAO VÀNG")
    write_sheet_title(ws, "BẢNG ĐÁNH GIÁ ĐẶC QUYỀN NỘI BỘ SAO VÀNG GROUP", 7)
    write_headers(ws, ["Mã Hệ", "Hãng", "Chất Lượng Cấu Trúc", "Độ Dễ Gia Công", "Độ Dễ Lắp Dựng", "Chi Phí Bảo Trì", "Khuyến Nghị Tư Vấn"])
    
    for r_idx in range(5, total_rows + 5):
        orig_row_excel = r_idx - 3
        ws.row_dimensions[r_idx].height = 25
        
        ws.cell(row=r_idx, column=1, value=f"='01_MASTER_DATABASE'!C{orig_row_excel}").alignment = Alignment(horizontal="center")
        ws.cell(row=r_idx, column=2, value=f"='01_MASTER_DATABASE'!D{orig_row_excel}")
        
        # Chất lượng
        ws.cell(row=r_idx, column=3, value=f"=IF(OR(ISNUMBER(SEARCH(\"Luxury\",'01_MASTER_DATABASE'!B{orig_row_excel})),ISNUMBER(SEARCH(\"Cầu\",'01_MASTER_DATABASE'!B{orig_row_excel}))),\"⭐⭐⭐⭐⭐\",\"⭐⭐⭐⭐\")").alignment = Alignment(horizontal="center")
        # Gia công
        ws.cell(row=r_idx, column=4, value=f"=IF(ISNUMBER(SEARCH(\"Slim\",'01_MASTER_DATABASE'!B{orig_row_excel})),\"⭐⭐⭐\",\"⭐⭐⭐⭐\")").alignment = Alignment(horizontal="center")
        # Thi công
        ws.cell(row=r_idx, column=5, value="⭐⭐⭐⭐").alignment = Alignment(horizontal="center")
        # Bảo trì
        ws.cell(row=r_idx, column=6, value="⭐⭐⭐⭐").alignment = Alignment(horizontal="center")
        # Khuyến nghị
        ws.cell(row=r_idx, column=7, value=f"=IF(OR(ISNUMBER(SEARCH(\"Luxury\",'01_MASTER_DATABASE'!B{orig_row_excel})),ISNUMBER(SEARCH(\"Cầu\",'01_MASTER_DATABASE'!B{orig_row_excel}))),\"Khuyên Dùng Cho Biệt Thự / Penhouse\",\"Khuyên Dùng Cho Nhà Phố / Căn Hộ\")")
        
        for c in range(1, 8):
            ws.cell(row=r_idx, column=c).border = thin_b
            ws.cell(row=r_idx, column=c).font = Font(name="Arial", size=9.5)
            
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 15
    for col in ["C","D","E","F"]:
        ws.column_dimensions[col].width = 18
    ws.column_dimensions["G"].width = 35

    # ─────────────────────────────────────────────
    # SHEET 21: 21_DASHBOARD
    # ─────────────────────────────────────────────
    print("  [21/24] Creating 21_DASHBOARD...")
    ws = wb.create_sheet("21_DASHBOARD")
    ws.views.sheetView[0].showGridLines = True
    
    ws["A1"] = "BẢNG ĐIỀU KHIỂN & BIỂU ĐỒ THỐNG KÊ"
    ws["A1"].font = Font(name="Arial Black", size=16, color="0D2240")
    ws.merge_cells("A1:D2")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    
    stats_list = [
        ("Tổng số hệ trong thư viện", "=COUNTA('01_MASTER_DATABASE'!A2:A500)"),
        ("Số hệ Cầu Cách Nhiệt (Thermal)", "=COUNTIF('01_MASTER_DATABASE'!B2:B500, \"*Cầu cách nhiệt*\") + COUNTIF('01_MASTER_DATABASE'!B2:B500, \"*Thermal*\")"),
        ("Số hệ cửa Slim nội/ngoại thất", "=COUNTIF('01_MASTER_DATABASE'!B2:B500, \"*Slim*\")"),
        ("Số hệ đạt tiêu chuẩn rãnh C châu Âu", "=COUNTIF('01_MASTER_DATABASE'!Q2:Q500, \"*rãnh C*\")"),
        ("Số hệ tương thích kính hộp dày (>=20mm)", "=COUNTIFS('01_MASTER_DATABASE'!N2:N500, \"*20mm*\") + COUNTIFS('01_MASTER_DATABASE'!N2:N500, \"*24mm*\") + COUNTIFS('01_MASTER_DATABASE'!N2:N500, \"*28mm*\") + COUNTIFS('01_MASTER_DATABASE'!N2:N500, \"*32mm*\") + COUNTIFS('01_MASTER_DATABASE'!N2:N500, \"*36mm*\")"),
        ("Số hệ có sẵn bản vẽ CAD", "=COUNTIF('01_MASTER_DATABASE'!V2:V500, \"*✓*\")")
    ]
    
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 20
    for idx, (lbl, f_str) in enumerate(stats_list, 4):
        ws.row_dimensions[idx].height = 30
        cell_lbl = ws.cell(row=idx, column=1, value=lbl)
        cell_lbl.font = Font(name="Arial", bold=True, size=11)
        cell_lbl.border = thin_b
        
        cell_val = ws.cell(row=idx, column=2, value=f_str)
        cell_val.font = Font(name="Arial Black", size=11, color="1B5E20")
        cell_val.alignment = Alignment(horizontal="center")
        cell_val.border = thin_b
        
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 20

    # ─────────────────────────────────────────────
    # SHEET 22: 22_SO SÁNH TRỰC TIẾP THEO MÃ HỆ
    # ─────────────────────────────────────────────
    print("  [22/24] Creating 22_SO SÁNH TRỰC TIẾP THEO MÃ HỆ...")
    ws = wb.create_sheet("22_SO SÁNH TRỰC TIẾP THEO MÃ HỆ")
    write_sheet_title(ws, "BẢNG SO SÁNH TRỰC TIẾP side-by-side PHÂN HỆ 55 PHỔ BIẾN", 6)
    
    # Tìm ngẫu nhiên dòng trong database cho PMA 55, Xingfa 55, Topal XF55, Việt Pháp 55, Civro AW55
    # và viết công thức tham chiếu trực tiếp đến các dòng đó
    def find_row_for_code(code_substr, brand_name):
        for idx, row in enumerate(ROWS, 2):
            if brand_name.lower() in row[3].lower() and code_substr.lower() in row[2].lower():
                return idx
        return 2 # fallback
        
    row_pma = find_row_for_code("55", "PMA")
    row_xingfa = find_row_for_code("55", "Xingfa")
    row_topal = find_row_for_code("Prima", "Topal")
    row_vp = find_row_for_code("55", "Việt Pháp")
    row_civro = find_row_for_code("AW55", "Civro")
    
    hdrs_comp = ["Tiêu Chí So Sánh", "Xingfa Quảng Đông (Hệ 55)", "PMA Vina (Hệ 55)", "Topal Prima (Hệ 55)", "Việt Pháp SHAL (Hệ 55)", "Civro Đức (Hệ AW55)"]
    write_headers(ws, hdrs_comp)
    
    ws.row_dimensions[5].height = 25
    ws.cell(row=5, column=1, value="Độ dày Profile (mm)").font = Font(name="Arial", bold=True)
    ws.cell(row=5, column=2, value=f"='01_MASTER_DATABASE'!L{row_xingfa}")
    ws.cell(row=5, column=3, value=f"='01_MASTER_DATABASE'!L{row_pma}")
    ws.cell(row=5, column=4, value=f"='01_MASTER_DATABASE'!L{row_topal}")
    ws.cell(row=5, column=5, value=f"='01_MASTER_DATABASE'!L{row_vp}")
    ws.cell(row=5, column=6, value=f"='01_MASTER_DATABASE'!L{row_civro}")
    
    ws.row_dimensions[6].height = 25
    ws.cell(row=6, column=1, value="Kính tối đa").font = Font(name="Arial", bold=True)
    ws.cell(row=6, column=2, value=f"='01_MASTER_DATABASE'!N{row_xingfa}")
    ws.cell(row=6, column=3, value=f"='01_MASTER_DATABASE'!N{row_pma}")
    ws.cell(row=6, column=4, value=f"='01_MASTER_DATABASE'!N{row_topal}")
    ws.cell(row=6, column=5, value=f"='01_MASTER_DATABASE'!N{row_vp}")
    ws.cell(row=6, column=6, value=f"='01_MASTER_DATABASE'!N{row_civro}")

    ws.row_dimensions[7].height = 25
    ws.cell(row=7, column=1, value="Hệ Phụ Kiện").font = Font(name="Arial", bold=True)
    ws.cell(row=7, column=2, value=f"='01_MASTER_DATABASE'!Q{row_xingfa}")
    ws.cell(row=7, column=3, value=f"='01_MASTER_DATABASE'!Q{row_pma}")
    ws.cell(row=7, column=4, value=f"='01_MASTER_DATABASE'!Q{row_topal}")
    ws.cell(row=7, column=5, value=f"='01_MASTER_DATABASE'!Q{row_vp}")
    ws.cell(row=7, column=6, value=f"='01_MASTER_DATABASE'!Q{row_civro}")

    ws.row_dimensions[8].height = 25
    ws.cell(row=8, column=1, value="Chỉ Số Cách Âm").font = Font(name="Arial", bold=True)
    ws.cell(row=8, column=2, value="Tốt (32dB)")
    ws.cell(row=8, column=3, value="Trung bình (28dB)")
    ws.cell(row=8, column=4, value="Khá (30dB)")
    ws.cell(row=8, column=5, value="Trung bình (28dB)")
    ws.cell(row=8, column=6, value="Xuất sắc (>=40dB)")

    for r in range(5, 9):
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = thin_b
            ws.cell(row=r, column=c).font = Font(name="Arial", size=9.5)
            if c > 1:
                ws.cell(row=r, column=c).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 22
    for col in ["B","C","D","E","F"]:
        ws.column_dimensions[col].width = 24

    # ─────────────────────────────────────────────
    # SHEET 23: 23_MA TRẬN HỆ NHÔM
    # ─────────────────────────────────────────────
    print("  [23/24] Creating 23_MA TRẬN HỆ NHÔM...")
    ws = wb.create_sheet("23_MA TRẬN HỆ NHÔM")
    write_sheet_title(ws, "MA TRẬN HỆ CỬA PHÂN BỔ THEO THƯƠNG HIỆU VÀ KÍCH THƯỚC BẢN RỘNG (55, 60, 65...)", 15)
    
    matrix_hdrs = ["Hãng", "Hệ 55", "Hệ 58", "Hệ 60", "Hệ 65", "Hệ 70", "Hệ 75", "Hệ 80", "Hệ 83", "Hệ 93", "Hệ 95", "Hệ 115", "Hệ 120", "Phân Hệ Slim", "Phân Hệ Thermal"]
    write_headers(ws, matrix_hdrs)
    
    for idx, brand in enumerate(unique_brands, 5):
        ws.row_dimensions[idx].height = 25
        ws.cell(row=idx, column=1, value=brand).font = Font(name="Arial", bold=True, size=10)
        ws.cell(row=idx, column=1).border = thin_b
        
        # Tạo công thức kiểm tra động cho từng cột kích thước
        formulas = [
            f"=IF(COUNTIFS('01_MASTER_DATABASE'!$D$2:$D$500, A{idx}, '01_MASTER_DATABASE'!$C$2:$C$500, \"*55*\")>0, \"✓\", \"✗\")",
            f"=IF(COUNTIFS('01_MASTER_DATABASE'!$D$2:$D$500, A{idx}, '01_MASTER_DATABASE'!$C$2:$C$500, \"*58*\")>0, \"✓\", \"✗\")",
            f"=IF(COUNTIFS('01_MASTER_DATABASE'!$D$2:$D$500, A{idx}, '01_MASTER_DATABASE'!$C$2:$C$500, \"*60*\")>0, \"✓\", \"✗\")",
            f"=IF(COUNTIFS('01_MASTER_DATABASE'!$D$2:$D$500, A{idx}, '01_MASTER_DATABASE'!$C$2:$C$500, \"*65*\")>0, \"✓\", \"✗\")",
            f"=IF(COUNTIFS('01_MASTER_DATABASE'!$D$2:$D$500, A{idx}, '01_MASTER_DATABASE'!$C$2:$C$500, \"*70*\")>0, \"✓\", \"✗\")",
            f"=IF(COUNTIFS('01_MASTER_DATABASE'!$D$2:$D$500, A{idx}, '01_MASTER_DATABASE'!$C$2:$C$500, \"*75*\")>0, \"✓\", \"✗\")",
            f"=IF(COUNTIFS('01_MASTER_DATABASE'!$D$2:$D$500, A{idx}, '01_MASTER_DATABASE'!$C$2:$C$500, \"*80*\")>0, \"✓\", \"✗\")",
            f"=IF(COUNTIFS('01_MASTER_DATABASE'!$D$2:$D$500, A{idx}, '01_MASTER_DATABASE'!$C$2:$C$500, \"*83*\")>0, \"✓\", \"✗\")",
            f"=IF(COUNTIFS('01_MASTER_DATABASE'!$D$2:$D$500, A{idx}, '01_MASTER_DATABASE'!$C$2:$C$500, \"*93*\")>0, \"✓\", \"✗\")",
            f"=IF(COUNTIFS('01_MASTER_DATABASE'!$D$2:$D$500, A{idx}, '01_MASTER_DATABASE'!$C$2:$C$500, \"*95*\")>0, \"✓\", \"✗\")",
            f"=IF(COUNTIFS('01_MASTER_DATABASE'!$D$2:$D$500, A{idx}, '01_MASTER_DATABASE'!$C$2:$C$500, \"*115*\")>0, \"✓\", \"✗\")",
            f"=IF(COUNTIFS('01_MASTER_DATABASE'!$D$2:$D$500, A{idx}, '01_MASTER_DATABASE'!$C$2:$C$500, \"*120*\")>0, \"✓\", \"✗\")",
            f"=IF(COUNTIFS('01_MASTER_DATABASE'!$D$2:$D$500, A{idx}, '01_MASTER_DATABASE'!$B$2:$B$500, \"*Slim*\")>0, \"✓\", \"✗\")",
            f"=IF(COUNTIFS('01_MASTER_DATABASE'!$D$2:$D$500, A{idx}, '01_MASTER_DATABASE'!$B$2:$B$500, \"*Thermal*\") + COUNTIFS('01_MASTER_DATABASE'!$D$2:$D$500, A{idx}, '01_MASTER_DATABASE'!$B$2:$B$500, \"*cầu*\")>0, \"✓\", \"✗\")",
        ]
        
        for c_idx, f_str in enumerate(formulas, 2):
            cell = ws.cell(row=idx, column=c_idx, value=f_str)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_b
            
    ws.column_dimensions["A"].width = 20
    for col in [get_column_letter(c) for c in range(2, 16)]:
        ws.column_dimensions[col].width = 12

    # ─────────────────────────────────────────────
    # SHEET 24: 24_MA TRẬN HÃNG
    # ─────────────────────────────────────────────
    print("  [24/24] Creating 24_MA TRẬN HÃNG...")
    ws = wb.create_sheet("24_MA TRẬN HÃNG")
    write_sheet_title(ws, "MA TRẬN SO SÁNH ĐẶC TÍNH CỐT LÕI GIỮA CÁC THƯƠNG HIỆU HÃNG", 7)
    
    hdrs_matrix = ["Tiêu Chí Thương Hiệu", "Xingfa (Quảng Đông)", "PMA (Việt Nam)", "Maxpro JP (Nhật Bản)", "Civro (Đức)", "Topal (Việt Nam)", "Eurowindow (Việt Nam)"]
    write_headers(ws, hdrs_matrix)
    
    matrix_rows = [
        ("Nguồn gốc xuất xứ chính", "Quảng Đông (Trung Quốc)", "Việt Nam sản xuất", "Công nghệ Nhật Bản/VN", "Đức nhập khẩu", "Việt Nam (Khang Minh/Topal)", "Đức/Việt Nam (Eurowindow)"),
        ("Công nghệ sơn phủ", "Sơn tĩnh điện AkzoNobel", "Sơn tĩnh điện AkzoNobel", "Mạ điện di Anodize ED", "Sơn tĩnh điện phủ MED Đức", "Sơn tĩnh điện AkzoNobel", "Sơn tĩnh điện bảo hành 20 năm"),
        ("Chuẩn rãnh phụ kiện", "Rãnh 22 truyền thống", "Rãnh 22 truyền thống", "Rãnh C Châu Âu / Rãnh 22", "Rãnh C Châu Âu cao cấp", "Rãnh 22 truyền thống", "Rãnh C Châu Âu (nhôm mới)"),
        ("Phân khúc giá hoàn thiện", "Từ 2.2tr - 3.2tr/m²", "Từ 1.6tr - 2.5tr/m²", "Từ 3.5tr - 5.0tr/m²", "Từ 6.0tr - 12.0tr/m²", "Từ 1.8tr - 2.8tr/m²", "Từ 3.0tr - 6.5tr/m²"),
        ("Thế mạnh lớn nhất", "Thương hiệu quốc dân, uy tín cao", "Giá bình dân, đại chúng dễ tiếp cận", "Mạ điện di chống muối biển 40 năm", "Cách âm nhiệt siêu việt cao cấp nhất", "Hệ thống đại lý phân phối lớn", "Cách âm cách nhiệt hàng đầu, bảo hành tốt"),
        ("Điểm yếu R&D", "Nhiều hàng giả hàng nhái trên thị trường", "Cấu trúc định hình mỏng không chịu bão lớn", "Giá phôi anodize khá đắt đỏ", "Giá cực kỳ cao, kén khách hàng đại chúng", "Độ nhận diện thương hiệu tầm trung", "Khung bao thô dày (nhựa), thi công đắt")
    ]
    
    for r_idx, row_vals in enumerate(matrix_rows, 5):
        ws.row_dimensions[r_idx].height = 35
        for c_idx, val in enumerate(row_vals, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = Font(name="Arial", size=9.5)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = thin_b
            if c_idx == 1:
                cell.font = Font(name="Arial", bold=True, size=9.5)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
    ws.column_dimensions["A"].width = 24
    for col in ["B","C","D","E","F","G"]:
        ws.column_dimensions[col].width = 24

    # ─────────────────────────────────────────────
    # LƯU FILE KẾT QUẢ
    # ─────────────────────────────────────────────
    path = os.path.join(ROOT, "02_Cơ sở Dữ liệu Gốc.xlsx")
    safe_save_excel(wb, path)
    print(f"============================================================")
    print(f"DATABASE GENERATED SUCCESSFULLY WITH 24 SHEETS!")
    print(f"Saved path: {path}")
    print(f"============================================================")

if __name__ == "__main__":
    create_database_v4()
