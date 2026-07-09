# -*- coding: utf-8 -*-
import os
import sys
import subprocess

# Auto-install openpyxl if not present
try:
    import openpyxl
except ImportError:
    print("openpyxl not found. Installing...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_catalogue_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Catalogue Hệ Cửa Nhôm"

    # Show gridlines
    ws.views.sheetView[0].showGridLines = True

    # Styling definitions
    font_title = Font(name="Arial", size=16, bold=True, color="0D2240")
    font_header = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    font_body = Font(name="Arial", size=9)
    font_body_bold = Font(name="Arial", size=9, bold=True)
    
    fill_header = PatternFill(start_color="0D2240", end_color="0D2240", fill_type="solid") # Deep Navy
    fill_zebra = PatternFill(start_color="F7F9FC", end_color="F7F9FC", fill_type="solid") # Light Blue/Gray
    fill_accent = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid") # Light Gold Accent
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_title = Alignment(horizontal="center", vertical="center")
    
    border_thin = Side(border_style="thin", color="D1D5DB")
    border_double = Side(border_style="double", color="4B5563")
    border_thick = Side(border_style="medium", color="0D2240")
    
    box_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    bottom_heavy_border = Border(bottom=border_double, left=border_thin, right=border_thin)
    header_border = Border(left=border_thin, right=border_thin, top=border_thick, bottom=border_thick)

    # 1. Sheet Title block
    ws.merge_cells("A1:J2")
    title_cell = ws["A1"]
    title_cell.value = "DANH MỤC HỆ THỐNG & CƠ SỞ DỮ LIỆU CỬA NHÔM KÍNH SAO VÀNG"
    title_cell.font = font_title
    title_cell.alignment = align_title
    
    # 2. Header Row (Row 4)
    headers = [
        "Hãng nhôm", 
        "Hệ nhôm (Series)", 
        "Hệ sản phẩm", 
        "Sản phẩm chi tiết", 
        "Công năng chính", 
        "Mã mặt cắt tiêu biểu", 
        "Hình ảnh minh họa", 
        'Phụ Kiện', 
        "Phân khúc", 
        'Đặc Điểm Kỹ Thuật'
    ]
    
    for col_idx, h_text in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = h_text
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = header_border
        
    # 3. Data definition
    data = [
        # SCHUCO
        ("Schüco (Đức)", "AWS 75.SI", "Cửa sổ mở quay/quay lật", "Cửa sổ 1 cánh mở quay lật rãnh C", 
         "Cách âm và cách nhiệt siêu cao, mở lật góc 15 độ thông gió an toàn, khóa chốt đa điểm ôm khít chống trộm.", 
         "MC-AWS75-01", "https://nhomkinhdaiphuc.com/upload/henhom/schuco-aws75.webp", 
         "Phụ kiện Schüco đồng bộ, Roto rãnh C", "Luxury (Siêu cao cấp)", "Dải cầu cách nhiệt đa khoang polyamide chèn xốp cách nhiệt."),
        
        ("Schüco (Đức)", "ADS 75.SI", "Cửa đi mở quay/Pivot", "Cửa đi 1 cánh mở quay ra ngoài bản lớn", 
         "Đại sảnh hoặc cửa chính biệt thự, chịu lực cực cao, thiết kế phẳng mặt trong ngoài sang trọng.", 
         "MC-ADS75-02", "https://nhomkinhdaiphuc.com/upload/henhom/schuco-ads75.webp", 
         "Phụ kiện khóa tự động Schüco, CMECH", "Luxury (Siêu cao cấp)", "Nhôm bản dày 2.5mm - 3.0mm chịu lực gió tốt."),
        
        ("Schüco (Đức)", "ASE 80.HI", "Cửa trượt nâng (Lift & Slide)", "Cửa đi trượt nâng 2 ray khẩu độ lớn", 
         "Cơ cấu nâng cánh khi trượt giúp vận hành siêu nhẹ cho cánh kính nặng tới 400kg, hạ cánh ép gioăng cách âm cách nhiệt tốt.", 
         "MC-ASE80-03", "https://nhomkinhdaiphuc.com/upload/henhom/schuco-ase80.webp", 
         "Phụ kiện nâng hạ Roto, CMECH, Schüco", "Luxury (Siêu cao cấp)", "Ray inox chịu tải trọng cao trượt êm ái."),
        
        ("Schüco (Đức)", "ASS 77 PD", "Cửa lùa slim (Panorama)", "Cửa lùa góc 90 độ âm tường không cột", 
         "Đố giữa siêu mảnh 30mm, khung bao giấu hoàn toàn vào tường trần, view rộng panorama 98%, phẳng sàn không vấp.", 
         "MC-ASS77-04", "https://nhomkinhdaiphuc.com/upload/henhom/schuco-ass77.webp", 
         "Phụ kiện khóa âm tự động Schüco", "Luxury (Siêu cao cấp)", "Tích hợp động cơ đóng mở tự động thông minh."),

        # REYNAERS
        ("Reynaers (Bỉ)", "MasterLine 8", "Cửa đi/sổ mở quay", "Cửa đi ban công mở trong cách nhiệt", 
         "Tiêu chuẩn sinh thái xanh châu Âu, độ kín nước bão gió Class E900 cực cao, rãnh C phụ kiện chìm đẹp mắt.", 
         "MC-ML8-01", "https://nhomkinhdaiphuc.com/upload/henhom/reynaers-ml8.webp", 
         "Phụ kiện Reynaers đồng bộ, CMECH rãnh C", "Luxury (Siêu cao cấp)", "Đạt chứng nhận cản nhiệt thụ động Passive House."),
         
        ("Reynaers (Bỉ)", "Hi-Finity", "Cửa lùa slim (Panorama)", "Cửa lùa 3 ray mở rộng tự động bằng motor", 
         "Thiết kế tối giản cực hạn, đố dọc siêu mỏng 26mm, kính hộp chân không cách âm cách nhiệt.", 
         "MC-HF-02", "https://nhomkinhdaiphuc.com/upload/henhom/reynaers-hifinity.webp", 
         "Hệ khóa chốt đa điểm cơ/điện Reynaers", "Luxury (Siêu cao cấp)", "Chịu tải trọng cánh kính nặng lên tới 500kg."),

        ("Reynaers (Bỉ)", "CP 155", "Cửa trượt nâng (Lift & Slide)", "Cửa đi trượt nâng góc 90 độ không cột", 
         "Mở rộng tối đa không gian lối ra hồ bơi biệt thự biển, chống nước bão và ăn mòn muối biển cực tốt.", 
         "MC-CP155-03", "https://nhomkinhdaiphuc.com/upload/henhom/reynaers-cp155.webp", 
         "Bánh xe & khóa nâng hạ Sobinco, CMECH", "Luxury (Siêu cao cấp)", "Nhôm dày 2.2mm gia cường lõi thép đố dọc."),

        # TECHNAL
        ("Technal (Pháp)", "Soleal", "Cửa mở quay", "Cửa đi thông phòng bo viền tròn mềm mại", 
         "Thẩm mỹ thanh lịch kiểu Pháp, cấu trúc gioăng đệm giảm chấn êm ái đóng mở, cách âm tuyệt đối.", 
         "MC-SOL-01", "https://nhomkinhdaiphuc.com/upload/henhom/technal-soleal.webp", 
         "Tay nắm & Bản lề Technal, CMECH", "Cận Cao Cấp", "Bề mặt xử lý anodized satin mịn màng."),

        ("Technal (Pháp)", "Lumeal", "Cửa lùa ẩn cánh", "Cửa đi trượt lùa đố giữa siêu mảnh ẩn khung", 
         "Ngăn nước mưa tuyệt hảo nhờ rãnh thoát nước chìm áp lực, đố dọc mỏng ôm khít góc nhìn rộng thoáng.", 
         "MC-LUM-02", "https://nhomkinhdaiphuc.com/upload/henhom/technal-lumeal.webp", 
         "Bộ phụ kiện khóa âm chìm Technal", "Cận Cao Cấp", "Khung bao cánh ẩn chìm vào tường tạo thẩm mỹ cao."),

        # CORTIZO
        ("Cortizo (Tây Ban Nha)", "Cor Vision", "Cửa lùa slim", "Cửa lùa 2 cánh trượt góc vuông", 
         "Đố dọc trung tâm chỉ 20mm cực mỏng, cho phép ngậm chìm vào sàn trần, view rộng 97% hiện đại.", 
         "MC-CV-01", "https://nhomkinhdaiphuc.com/upload/henhom/cortizo-corvision.webp", 
         "Hộp khóa đa điểm chìm Cortizo, CMECH", "Cận Cao Cấp / Luxury", "Dòng cửa slim ngoại thất bán chạy nhất biệt thự biển."),

        ("Cortizo (Tây Ban Nha)", "4600 Slider", "Cửa trượt nâng (Lift & Slide)", "Cửa đi trượt nâng 3 ray bản cánh dày", 
         "Cách âm lên tới 40dB nhờ gioăng EPDM nén chặt khi hạ cánh, chịu áp lực bão gió cấp 12 ngoài ban công penthouse.", 
         "MC-4600-02", "https://nhomkinhdaiphuc.com/upload/henhom/cortizo-4600.webp", 
         "Phụ kiện khóa tay đòn Roto, CMECH", "Cận Cao Cấp", "Độ dày nhôm lên tới 2.5mm cho khung bao."),

        # XINGFA
        ("Xingfa Quảng Đông", "Xingfa 55", "Cửa sổ/Cửa đi mở quay", "Cửa đi 2 cánh mở ngoài ban công có gân", 
         "Dòng cửa quốc dân siêu bền, gân gia cường trên bề mặt chịu lực va đập tốt, cách âm kín khít cao.", 
         "C3332 / C3303", "https://nhomkinhdaiphuc.com/upload/henhom/xingfa55.webp", 
         "Phụ kiện Kinlong chính hãng, Cogo, Draho", "Trung Cấp / Phổ thông", "Nhôm dày 2.0mm cho cửa đi, 1.4mm cho cửa sổ mở quay."),

        ("Xingfa Quảng Đông", "Xingfa 93", "Cửa đi/sổ trượt lùa", "Cửa lùa 4 cánh trượt 2 ray bản cánh lớn", 
         "Tiết kiệm không gian mở cánh, trượt bánh xe êm ái trên ray inox đệm chống mài mòn.", 
         "MC-XF93-01", "https://nhomkinhdaiphuc.com/upload/henhom/xingfa93.webp", 
         "Bánh xe đơn & Khóa sập Kinlong, Cogo", "Trung Cấp / Phổ thông", "Độ dày nhôm 2.0mm cho cả khung và cánh."),

        ("Xingfa Quảng Đông", "Xingfa 63", "Cửa xếp trượt (Folding)", "Cửa đi xếp trượt 6 cánh xếp về một bên", 
         "Mở rộng tối đa 90% lối ra vào bể bơi hoặc gara xe, xếp gọn cánh chồng khít lên nhau tiết kiệm diện tích.", 
         "MC-XF63-01", "https://nhomkinhdaiphuc.com/upload/henhom/xingfa63.webp", 
         "Bánh xe chịu lực Roto, Hopo, Kinlong", "Trung Cấp", "Hệ thống ray treo trên kết hợp bánh xe dẫn hướng dưới."),

        ("Xingfa Quảng Đông", "Xingfa 65", "Vách mặt dựng (Curtain Wall)", "Vách dựng giấu đố cách nhiệt tòa nhà", 
         "Làm mặt dựng kính khổ lớn cho văn phòng, giấu đố nhôm vào trong tạo độ phẳng mướt kính bên ngoài mặt tiền.", 
         "MC-XF65-01", "https://nhomkinhdaiphuc.com/upload/henhom/xingfa65.webp", 
         "Ke bắt kính chuyên dụng, silicone kết cấu Dow Corning", "Trung Cấp / Dự án B2B", "Nhôm đố đứng dày 2.5mm - 3.0mm chịu uốn xoắn cao."),

        # PMA
        ("PMA (Việt Nam)", "PMA 55", "Cửa sổ/Cửa đi mở quay", "Cửa đi mở quay 1 cánh vát cạnh", 
         "Mặt cắt vát cạnh thẩm mỹ cao giúp chống bám bụi và giảm nước đọng gờ, chi phí kinh tế phù hợp nhà dân.", 
         "MC-PMA55-01", "https://nhomkinhdaiphuc.com/upload/henhom/pma55.webp", 
         "Phụ kiện PMA đồng bộ, Draho, Sigico", "Phổ thông (Kinh tế)", "Nhôm dày 1.2mm - 1.4mm trọng lượng nhẹ dễ sản xuất."),

        ("PMA (Việt Nam)", "PMA Platinum", "Cửa đi mở quay rãnh C", "Cửa đi rãnh C châu Âu tích hợp khóa đa điểm", 
         "Đưa tiêu chuẩn rãnh C lắp phụ kiện cao cấp châu Âu vào dòng nhôm phổ thông để chống xệ cánh và tăng kín khít.", 
         "MC-PLAT-01", "https://nhomkinhdaiphuc.com/upload/henhom/pma-platinum.webp", 
         "Phụ kiện CMECH rãnh C, Roto, Sigico", "Trung Cấp", "Nhôm dày 1.6mm - 1.8mm sơn tĩnh điện cao cấp."),

        ("PMA (Việt Nam)", "PMA Slim", "Cửa lùa Slim nội thất", "Cửa lùa treo 3 cánh trượt liên kết giảm chấn", 
         "Trượt siêu êm không có ray dưới sàn tránh vấp và dễ lau chùi quét nhà, cánh lùa treo nhẹ nhàng đập khít.", 
         "MC-SLIM-IN", "https://nhomkinhdaiphuc.com/upload/henhom/pma-slim.webp", 
         "Bộ phụ kiện lùa treo giảm chấn PMA", "Trung Cấp / Nội thất", "Khung cánh siêu nhỏ gọn bản nhôm rộng chỉ 18mm."),

        # VR
        ("Viralwindow (Việt Nam)", "Viralwindow VRA55 Series", "Cửa đi/sổ mở quay", "Cửa sổ hất bản lề chữ A góc 45 độ", 
         "Kín nước cao, chống va đập gió bão căn hộ tầm trung, sơn tĩnh điện bảo hành mờ mịn vân cát 10 năm.", 
         "MC-Viralwindow VRA55-01", "https://nhomkinhdaiphuc.com/upload/henhom/vr55.webp", 
         "Phụ kiện Draho, Sigico, Sigico, Bogo", "Phổ thông", "Cạnh nhôm vuông phẳng góc cứng cáp dễ cắt lắp."),

        ("Viralwindow (Việt Nam)", "Viralwindow VR Slim", "Cửa lùa Slim ngoại thất", "Cửa lùa 2 ray Slim kính hộp có gioăng nén", 
         "Vách kính phòng khách biệt thự hiện đại view vườn rộng rãi tối giản, đố giữa mảnh rãnh trượt êm khít.", 
         "MC-VRSLIM-OUT", "https://nhomkinhdaiphuc.com/upload/henhom/vr-slim.webp", 
         "Hệ phụ kiện chốt đa điểm âm Viralwindow VR Slim", "Trung Cấp / Cận cao cấp", "Ray inox chịu tải bánh xe kép chống bụi chèn chặt."),

        # TOPAL
        ("Topal (Austdoor)", "Topal Prima", "Cửa đi/sổ mở quay", "Cửa đi 2 cánh mở quay rãnh C tiêu chuẩn", 
         "Được thiết kế rãnh C châu Âu ép góc khít kín tốt, chống xệ cánh nhờ trục bản lề chịu tải lực hướng tâm.", 
         "MC-PRIMA-01", "https://nhomkinhdaiphuc.com/upload/henhom/topal-prima.webp", 
         "Phụ kiện Topal đồng bộ, CMECH, Roto", "Trung Cấp / Cận cao cấp", "Nhôm đùn phôi nhôm sạch 6063-T5 tiêu chuẩn Việt Nam."),

        ("Topal (Austdoor)", "Topal Slim", "Cửa lùa sổ/đi", "Cửa sổ lùa 2 cánh ray inox mảnh", 
         "Thiết kế thanh nhôm mỏng nhẹ tối ưu hóa trọng lượng và diện tích kính, chi phí đầu tư kinh tế.", 
         "MC-TSLIM-01", "https://nhomkinhdaiphuc.com/upload/henhom/topal-slim.webp", 
         "Khóa sập & Bánh xe Topal đồng bộ", "Phổ thông (Kinh tế)", "Thân thiện thợ lắp đặt, lắp nhanh không cần máy dập phức tạp.")
    ]
    
    # Write data rows (Starting from Row 5)
    for row_idx, row_data in enumerate(data, start=5):
        fill_to_apply = fill_zebra if row_idx % 2 == 1 else None
        
        # Check if it's a luxury brand to highlight slightly
        brand_name = row_data[0]
        is_luxury = "Schüco" in brand_name or "Reynaers" in brand_name
        
        for col_idx, cell_value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = cell_value
            cell.font = font_body
            cell.border = box_border
            
            # Apply fill
            if is_luxury and col_idx == 1:
                cell.fill = fill_accent
                cell.font = font_body_bold
            elif fill_to_apply:
                cell.fill = fill_to_apply
                
            # Alignment rules
            if col_idx in [1, 2, 3, 6, 9]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left
                
    # Add a heavy bottom border to the last data row
    last_row_idx = len(data) + 4
    for col_idx in range(1, 11):
        ws.cell(row=last_row_idx, column=col_idx).border = bottom_heavy_border
        
    # 4. Auto-fit column widths with padding
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        # Don't size columns based on the merged title row in A1:J2
        for cell in col:
            if cell.row in [1, 2]:
                continue
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
                
        # Set column width with safety margin
        calculated_width = max(max_len * 1.05 + 4, 12)
        # Cap width to avoid extremely wide columns for URL/description
        ws.column_dimensions[col_letter].width = min(calculated_width, 42)
        
    # Save the Excel Workbook
    output_path = r"d:\Sao Vàng\Website-SaoVang\BaoCao_ThuVien_CuaNhom\ThuVien_HeCuaNhom_SaoVang.xlsx"
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"Excel Catalogue Matrix successfully saved to {output_path}")

if __name__ == "__main__":
    create_catalogue_excel()