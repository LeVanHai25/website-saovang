# -*- coding: utf-8 -*-
"""
Script chuẩn hóa tiêu đề cột trong tất cả các tệp Excel Catalogue_HeCuaNhom_*.xlsx
sang cấu trúc duy nhất:
STT | Nhóm Hệ | Mã Hệ | Tên Hệ | Loại Sản Phẩm | Độ Dày (mm) | Phụ Kiện | Gioăng & Keo | Đặc Điểm Kỹ Thuật
"""
import os, sys, glob
import openpyxl

try:
    sys.stdout.reconfigure(encoding='utf-8')
except AttributeError:
    pass

BASE_DIR = r"d:\Sao Vàng\Website-SaoVang\BaoCao_ThuVien_CuaNhom"
TARGET_HEADERS = [
    "STT", 
    "Nhóm Hệ", 
    "Mã Hệ", 
    "Tên Hệ", 
    "Loại Sản Phẩm", 
    "Độ Dày (mm)", 
    "Phụ Kiện", 
    "Gioăng & Keo", 
    "Đặc Điểm Kỹ Thuật"
]

def standardize_excel_files():
    excel_files = glob.glob(os.path.join(BASE_DIR, "Catalogue_HeCuaNhom_*.xlsx"))
    print(f"Tim thay {len(excel_files)} tep Excel Catalogue.")
    
    for f in excel_files:
        filename = os.path.basename(f)
        try:
            wb = openpyxl.load_workbook(f)
            ws = wb.active
            
            # Quét các dòng đầu tiên (thường từ dòng 1 đến 5) để tìm hàng tiêu đề cột
            header_row_found = None
            for r in range(1, 6):
                row_vals = [str(ws.cell(row=r, column=c).value).strip() if ws.cell(row=r, column=c).value is not None else "" for c in range(1, 11)]
                # Dấu hiệu của hàng tiêu đề: có chứa 'STT' và ít nhất một trong các từ khóa
                if "STT" in row_vals and any(kw in "".join(row_vals).lower() for kw in ["nhóm", "mã", "hệ", "loại", "dày", "phụ kiện", "gioăng", "nhom", "ma", "he", "loai", "day", "phu kien", "gioang"]):
                    header_row_found = r
                    break
            
            if header_row_found:
                print(f"  [+] Tim thay hang tieu de tai dong {header_row_found} cua tep {filename}")
                # Ghi đè các tiêu đề mới nhưng giữ nguyên style hiện tại của ô
                for col_idx, new_h in enumerate(TARGET_HEADERS, start=1):
                    cell = ws.cell(row=header_row_found, column=col_idx)
                    cell.value = new_h
                
                # Lưu lại
                wb.save(f)
                print(f"  >> Da luu thanh cong tep: {filename}")
            else:
                print(f"  [-] Khong tim thay hang tieu de cho tep: {filename}")
                
        except PermissionError:
            print(f"  [!] LOI: Tep {filename} dang bi khoa (mo trong Excel). Vui long dong lai.")
        except Exception as e:
            print(f"  [!] LOI khi xu ly tep {filename}: {str(e)}")

if __name__ == "__main__":
    print("============================================================")
    print("CHUAN HOA TIEU DE COT TRONG CAC FILE EXCEL CATALOGUE")
    print("============================================================")
    standardize_excel_files()
    print("============================================================")
