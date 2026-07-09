import os
import sys
import subprocess

try:
    import openpyxl
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

from openpyxl import load_workbook

path = r"d:\Sao Vàng\Website-SaoVang\BaoCao_ThuVien_CuaNhom\Master_Tracker_HeCuaNhom.xlsx"
output_path = r"d:\Sao Vàng\Website-SaoVang\scratch\master_tracker_summary.txt"

with open(output_path, "w", encoding="utf-8") as out:
    if os.path.exists(path):
        out.write(f"==================== Master_Tracker_HeCuaNhom.xlsx ====================\n")
        wb = load_workbook(path, data_only=True)
        for sheetname in wb.sheetnames:
            out.write(f"\n--- Sheet: {sheetname} ---\n")
            sheet = wb[sheetname]
            # Print first 20 rows
            for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                if r_idx >= 50:
                    break
                row_str = [str(cell).strip() if cell is not None else "" for cell in row]
                out.write(" | ".join(row_str) + "\n")
        out.write("=============================================\n\n")
    else:
        out.write(f"File not found: {path}\n")

print(f"Content written to {output_path}")
