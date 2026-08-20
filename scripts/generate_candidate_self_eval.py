import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.drawing.image import Image as OpenpyxlImage

def create_candidate_self_eval_with_logo():
    wb = openpyxl.Workbook()
    
    # Style configuration
    STEEL_BLUE = "2F5597"
    ICE_BLUE = "D9E1F2"
    LIGHT_GRAY = "F2F2F2"
    BORDER_GRAY = "D9D9D9"
    YELLOW_INPUT = "FFF2CC"
    DARK_GREEN = "385723"
    LIGHT_GREEN = "E2EFDA"
    
    font_company_name = Font(name="Segoe UI", size=12, bold=True, color=STEEL_BLUE)
    font_company_info = Font(name="Segoe UI", size=9, italic=True, color="595959")
    font_title_main = Font(name="Segoe UI", size=15, bold=True, color=STEEL_BLUE)
    font_title_sub = Font(name="Segoe UI", size=10, italic=True, color="7F7F7F")
    font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    font_group_header = Font(name="Segoe UI", size=11, bold=True, color=STEEL_BLUE)
    font_bold = Font(name="Segoe UI", size=10, bold=True)
    font_normal = Font(name="Segoe UI", size=10)
    font_italic = Font(name="Segoe UI", size=9, italic=True)
    font_result = Font(name="Segoe UI", size=11, bold=True, color=DARK_GREEN)
    
    fill_header = PatternFill(start_color=STEEL_BLUE, end_color=STEEL_BLUE, fill_type="solid")
    fill_group = PatternFill(start_color=ICE_BLUE, end_color=ICE_BLUE, fill_type="solid")
    fill_input = PatternFill(start_color=YELLOW_INPUT, end_color=YELLOW_INPUT, fill_type="solid")
    fill_result_box = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    
    thin_side = Side(style="thin", color=BORDER_GRAY)
    border_all_thin = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    border_group = Border(
        left=Side(style="thin", color=STEEL_BLUE),
        right=Side(style="thin", color=STEEL_BLUE),
        top=Side(style="thin", color=STEEL_BLUE),
        bottom=Side(style="thin", color=STEEL_BLUE)
    )

    # -------------------------------------------------------------
    # SHEET 1: PHIẾU TỰ ĐÁNH GIÁ
    # -------------------------------------------------------------
    ws = wb.active
    ws.title = "Phiếu Tự Đánh Giá"
    ws.sheet_view.showGridLines = True
    
    # Column widths
    ws.column_dimensions['A'].width = 6   # STT
    ws.column_dimensions['B'].width = 24  # Nhóm công việc
    ws.column_dimensions['C'].width = 68  # Nội dung công việc (Mục việc cơ bản)
    ws.column_dimensions['D'].width = 26  # Mức độ tự tin (Tự chọn)
    ws.column_dimensions['E'].width = 16  # Số năm kinh nghiệm
    ws.column_dimensions['F'].width = 30  # Ghi chú thêm của ứng viên

    # Company Letterhead Row 2-4
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 15
    ws.row_dimensions[4].height = 15
    
    # Add Company Name & Info
    ws.merge_cells("C2:F2")
    ws["C2"] = "CÔNG TY CỔ PHẦN SẢN XUẤT CƠ KHÍ SAO VÀNG (ASV)"
    ws["C2"].font = font_company_name
    ws["C2"].alignment = Alignment(horizontal="left", vertical="center")
    
    ws.merge_cells("C3:F3")
    ws["C3"] = "Địa chỉ: Tầng 3, TT7-35 Khu đô thị Văn Phú, phường Kiến Hưng, TP Hà Nội, Việt Nam."
    ws["C3"].font = font_company_info
    ws["C3"].alignment = Alignment(horizontal="left", vertical="center")
    
    ws.merge_cells("C4:F4")
    ws["C4"] = "Hotline: 0869 590 279  |  Email: cokhisaovangvn@gmail.com"
    ws["C4"].font = font_company_info
    ws["C4"].alignment = Alignment(horizontal="left", vertical="center")

    # Add Logo image in A2
    try:
        logo_path = "d:/Sao Vàng/Website-SaoVang/Logo/LogoCTY + SV Aluminium/LogoCTY.png"
        img = OpenpyxlImage(logo_path)
        img.width = 110
        img.height = 45
        ws.add_image(img, "A2")
    except Exception as e:
        print(f"Error adding logo: {e}")

    # Title Block (shifted to Row 6-7)
    ws.row_dimensions[6].height = 24
    ws.merge_cells("A6:F6")
    ws["A6"] = "BẢNG TỰ ĐÁNH GIÁ NĂNG LỰC DÀNH CHO ỨNG VIÊN THỢ CƠ KHÍ"
    ws["A6"].font = font_title_main
    ws["A6"].alignment = align_center
    
    ws.row_dimensions[7].height = 18
    ws.merge_cells("A7:F7")
    ws["A7"] = "Kính mong ứng viên tự đánh giá khách quan để công ty bố trí công việc và mức lương phù hợp nhất"
    ws["A7"].font = font_title_sub
    ws["A7"].alignment = align_center

    # Candidate profile block (shifted to Row 9-10)
    ws["B9"] = "Họ và tên ứng viên:"
    ws["B9"].font = font_bold
    ws["C9"] = "[Nhập họ tên]"
    ws["C9"].font = font_normal
    ws["C9"].border = Border(bottom=Side(style="thin", color="000000"))

    ws["E9"] = "Ngày điền phiếu:"
    ws["E9"].font = font_bold
    ws["F9"] = "[Ghi ngày]"
    ws["F9"].font = font_normal
    ws["F9"].border = Border(bottom=Side(style="thin", color="000000"))

    ws["B10"] = "Số điện thoại/CCCD:"
    ws["B10"].font = font_bold
    ws["C10"] = "[Số điện thoại và CCCD]"
    ws["C10"].font = font_normal
    ws["C10"].border = Border(bottom=Side(style="thin", color="000000"))

    ws["E10"] = "Năm sinh/Quê quán:"
    ws["E10"].font = font_bold
    ws["F10"] = "[Ghi năm sinh]"
    ws["F10"].font = font_normal
    ws["F10"].border = Border(bottom=Side(style="thin", color="000000"))

    ws.row_dimensions[11].height = 15

    # Main Table Header (shifted to Row 12)
    headers = [
        ("A12", "STT", align_center),
        ("B12", "NHÓM CÔNG VIỆC", align_left),
        ("C12", "MỤC VIỆC CƠ BẢN VÀ YÊU CẦU KỸ NĂNG", align_left),
        ("D12", "MỨC ĐỘ THÀNH THẠO (ỨNG VIÊN TỰ CHỌN)", align_center),
        ("E12", "NĂM KINH NGHIỆM", align_center),
        ("F12", "GHI CHÚ CHI TIẾT (NẾU CÓ)", align_left)
    ]
    for cell_ref, text, align in headers:
        cell = ws[cell_ref]
        cell.value = text
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align
        cell.border = border_all_thin
    ws.row_dimensions[12].height = 28

    # Helper function for merged headers
    def merge_and_style_header_ws(ws, r_num, title):
        ws.merge_cells(start_row=r_num, start_column=1, end_row=r_num, end_column=6)
        ws.cell(r_num, 1, title).font = font_group_header
        ws.cell(r_num, 1).alignment = align_left
        for col in range(1, 7):
            cell = ws.cell(r_num, col)
            cell.fill = fill_group
            cell.border = border_group

    # 27 Tasks (starts at Row 13)
    tasks_checklist = [
        ("G", "I. TÁC PHONG, AN TOÀN LAO ĐỘNG & 5S"),
        ("1", "Ý thức kỷ luật & Phối hợp", "Đảm bảo sức khỏe dẻo dai làm việc xưởng/công trình; trung thực, nhanh nhẹn, chịu khó học hỏi và phối hợp tốt với đồng nghiệp."),
        ("2", "Vệ sinh xưởng & 5S", "Chủ động dọn dẹp máy móc, quét dọn và vệ sinh sạch sẽ khu vực gia công lắp ráp sau ca làm việc hoặc đi công trình."),
        ("3", "Tư duy ATLĐ xưởng", "Biết tư duy tình huống để phòng ngừa, bảo vệ an toàn cho bản thân, người xung quanh và tài sản trong khu vực thi công."),
        
        ("G", "II. VẬN HÀNH DỤNG CỤ CẦM TAY & KỸ THUẬT MÀI"),
        ("4", "Sử dụng dụng cụ cơ bản", "Sử dụng các loại máy cơ bản: máy cắt tay, máy cắt bàn, máy khoan cầm tay, thước đo khoảng cách, dây điện... an toàn."),
        ("5", "Mài hoàn thiện bề mặt", "Sử dụng máy mài để mài hoàn thiện sản phẩm sắt đen, sắt mạ kẽm, inox; mài góc, mài bề mặt và đánh rỉ sét."),
        ("6", "Mài phẳng tinh mỹ", "Mài tinh mỹ phẳng tuyệt đối hoặc mài lấy góc sắc nét đạt thẩm mỹ cao cho các sản phẩm cơ khí kiến trúc/dân dụng."),
        ("7", "Mài tạo góc kết cấu", "Mài định hình các sản phẩm được tổ hợp từ nguyên liệu tấm hoặc thép hình theo thiết kế kỹ thuật bản vẽ."),
        ("8", "Mài phá mối hàn thô", "Mài phá bavia và làm sạch các vị trí mối hàn trên các bề mặt vật liệu thô hoặc các kết cấu chịu lực hạng nặng."),
        ("9", "Mài đánh bóng inox", "Đánh bóng cơ học cho các sản phẩm sử dụng nguyên liệu inox bóng hoặc làm bóng, mịn bề mặt thép mộc."),
        ("10", "Mài phục hồi mũi khoan", "Kỹ năng tự tinh chỉnh, mài sắc bén các loại mũi khoan thép cầm tay phục vụ công việc hàng ngày."),
        
        ("G", "III. KỸ THUẬT PHUN SƠN HOÀN THIỆN"),
        ("11", "Vận hành súng phun sơn", "Sử dụng thành thạo súng phun sơn khí nén; thực hiện sơn lót chống gỉ và sơn phủ màu hoàn thiện bóng mịn, đều màu."),
        ("12", "Nhận biết chủng loại vật tư", "Nắm rõ quy cách, chủng loại các loại sơn dầu, sơn 2 thành phần và các loại vật tư phụ cơ bản dùng trong xưởng."),
        
        ("G", "IV. KỸ THUẬT CẮT PHÔI & HÀN ĐÍNH GÁ RÁP"),
        ("13", "Sử dụng máy công trình", "Sử dụng cơ bản các máy và phụ kiện lắp ráp đi kèm trong ngành cơ khí dân dụng và xây dựng lắp đặt công trình."),
        ("14", "Đo cắt phôi chính xác", "Nhận biết đơn vị đo lường cơ khí; cắt phôi chính xác từng mm; tự tính toán làm dưỡng gá chuẩn để gia công phôi hàng loạt."),
        ("15", "Hàn đính gá kết cấu", "Sử dụng máy hàn để gá đính chính xác các vị trí kết cấu tấm/hình theo yêu cầu lắp ráp trước khi hàn chết."),
        ("16", "Chuẩn bị công cụ công trình", "Hiểu rõ công năng máy móc; tự tính toán chuẩn bị đầy đủ máy, vật tư phụ cần thiết đi công trình khi được giao đầu việc."),
        ("17", "Hàn hồ quang tay (Que)", "Sử dụng thành thạo máy hàn que (MMA) trên các kết cấu sắt thép chịu lực thông thường."),
        ("18", "Hàn bán tự động MIG", "Sử dụng thành thạo máy hàn bán tự động MIG để hàn các cụm chi tiết hoặc kết cấu thép hộp nhanh, ngấu sâu."),
        ("19", "Hàn khí bảo vệ TIG", "Sử dụng thành thạo máy hàn TIG để hàn inox mỏng hoặc liên kết kim loại yêu cầu mối hàn sạch đẹp, vảy đều."),
        ("20", "Sử dụng máy chuyên dụng", "Vận hành thành thạo máy cân bằng laser, thước đo khoảng cách điện tử và các dụng cụ định vị chuyên dụng."),
        
        ("G", "V. ĐỌC BẢN VẼ, QUẢN LÝ TIẾN ĐỘ & ĐỘI NHÓM"),
        ("21", "Đọc bản vẽ sản xuất", "Đọc thạo bản vẽ kỹ thuật; tự tính toán khai triển kích thước hình học chính xác từ bản vẽ thành sản phẩm thực tế."),
        ("22", "Kỹ năng hoàn thiện & QC", "Tự đánh giá, kiểm tra chất lượng sản phẩm sau khi hoàn thiện trên tinh thần khách quan, chủ động học hỏi sửa lỗi."),
        ("23", "Ước lượng định mức công", "Ước lượng định mức nhân công và khoảng thời gian cần thiết để hoàn thiện sản phẩm hoặc thi công công trình."),
        ("24", "Báo cáo tiến độ cuối ngày", "Báo cáo tiến độ cuối ngày; đề xuất kịp thời các phương án thi công tối ưu hoặc giải quyết nhanh các phát sinh hiện trường."),
        ("25", "Độc lập thi công công việc", "Khả năng độc lập thực hiện công việc được giao; tự chủ động xử lý kỹ thuật thực địa không làm ảnh hưởng cấu trúc chung."),
        ("26", "Kỹ năng làm việc nhóm", "Phối hợp làm việc nhóm nhịp nhàng; đảm bảo sản xuất sản phẩm đúng thiết kế kỹ thuật và đúng tiến độ cam kết."),
        ("27", "Lãnh đạo tổ đội thi công", "Năng lực quản lý, bao quát điều phối công việc cho từng thành viên trong tổ đội thi công khi được giao nhiệm vụ tổ trưởng.")
    ]

    current_row = 13
    eval_rows = []
    
    for row_data in tasks_checklist:
        if row_data[0] == "G":
            ws.row_dimensions[current_row].height = 24
            merge_and_style_header_ws(ws, current_row, row_data[1])
        else:
            stt, group_name, task_desc = row_data
            ws.row_dimensions[current_row].height = 36
            
            # STT
            ws.cell(current_row, 1, stt).font = font_normal
            ws.cell(current_row, 1).alignment = align_center
            
            # Nhóm công việc
            g_title = ""
            if int(stt) <= 3:
                g_title = "I. An toàn & 5S"
            elif int(stt) <= 10:
                g_title = "II. Vận hành & Mài"
            elif int(stt) <= 12:
                g_title = "III. Kỹ thuật Sơn"
            elif int(stt) <= 20:
                g_title = "IV. Cắt & Hàn gá"
            else:
                g_title = "V. Bản vẽ & Đội nhóm"
                
            ws.cell(current_row, 2, g_title).font = font_bold
            ws.cell(current_row, 2).alignment = align_center
            
            # Detailed task
            ws.cell(current_row, 3, task_desc).font = font_normal
            ws.cell(current_row, 3).alignment = align_left
            
            # Dropdown choice: Column D
            cell_choice = ws.cell(current_row, 4, "2 - Làm được (Cần hướng dẫn)")
            cell_choice.font = font_bold
            cell_choice.alignment = align_center
            cell_choice.fill = fill_input
            eval_rows.append(current_row)
            
            # Years experience
            cell_exp = ws.cell(current_row, 5, "[Ghi số năm]")
            cell_exp.font = font_normal
            cell_exp.alignment = align_center
            cell_exp.fill = fill_input
            
            # Notes
            cell_note = ws.cell(current_row, 6, "[Ví dụ: Đã hàn 2 năm]")
            cell_note.font = font_italic
            cell_note.alignment = align_left
            cell_note.fill = fill_input
            
            for col in range(1, 7):
                ws.cell(current_row, col).border = border_all_thin
                
        current_row += 1

    # Place a spacing row
    ws.row_dimensions[current_row].height = 15
    current_row += 1

    # SUMMARY BOX (DASHBOARD AT BOTTOM)
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=6)
    r_hdr = ws.cell(current_row, 2, "BẢNG TỔNG HỢP NĂNG LỰC TỰ ĐÁNH GIÁ")
    r_hdr.font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    r_hdr.fill = PatternFill(start_color=STEEL_BLUE, end_color=STEEL_BLUE, fill_type="solid")
    r_hdr.alignment = align_center
    ws.row_dimensions[current_row].height = 24
    
    for col in range(2, 7):
        ws.cell(current_row, col).border = Border(top=Side(style="medium", color=STEEL_BLUE))
        
    current_row += 1
    
    # 1. Tổng số mục đánh giá
    ws.row_dimensions[current_row].height = 20
    ws.cell(current_row, 2, "Tổng số hạng mục mục việc khảo sát:").font = font_bold
    ws.cell(current_row, 2).alignment = align_left
    
    cell_total = ws.cell(current_row, 3, 27)
    cell_total.font = font_bold
    cell_total.alignment = align_center
    ws.cell(current_row, 4, "Hạng mục công việc").font = font_italic
    ws.cell(current_row, 4).alignment = align_left
    
    for col in range(2, 7):
        cell = ws.cell(current_row, col)
        cell.fill = fill_result_box
        left_s = Side(style="medium", color=STEEL_BLUE) if col == 2 else None
        right_s = Side(style="medium", color=STEEL_BLUE) if col == 6 else None
        cell.border = Border(left=left_s, right=right_s)
        
    total_ref = f"C{current_row}" # C48
    current_row += 1
    
    # 2. Số hạng mục làm tốt độc lập (Score 3)
    ws.row_dimensions[current_row].height = 20
    ws.cell(current_row, 2, "Số hạng mục tự tin LÀM TỐT ĐỘC LẬP:").font = font_bold
    ws.cell(current_row, 2).alignment = align_left
    
    # COUNTIF on column D (from D14 to D45)
    cell_3 = ws.cell(current_row, 3, '=COUNTIF(D14:D45, "*Làm tốt độc lập*")')
    cell_3.font = font_bold
    cell_3.alignment = align_center
    ws.cell(current_row, 4, "Hạng mục đạt mức 3").font = font_italic
    ws.cell(current_row, 4).alignment = align_left
    
    for col in range(2, 7):
        cell = ws.cell(current_row, col)
        cell.fill = fill_result_box
        left_s = Side(style="medium", color=STEEL_BLUE) if col == 2 else None
        right_s = Side(style="medium", color=STEEL_BLUE) if col == 6 else None
        cell.border = Border(left=left_s, right=right_s)
        
    s3_ref = f"C{current_row}" # C49
    current_row += 1

    # 3. Số hạng mục làm được cần hướng dẫn (Score 2)
    ws.row_dimensions[current_row].height = 20
    ws.cell(current_row, 2, "Số hạng mục LÀM ĐƯỢC (cần hướng dẫn):").font = font_bold
    ws.cell(current_row, 2).alignment = align_left
    
    cell_2 = ws.cell(current_row, 3, '=COUNTIF(D14:D45, "*Làm được*")')
    cell_2.font = font_bold
    cell_2.alignment = align_center
    ws.cell(current_row, 4, "Hạng mục đạt mức 2").font = font_italic
    ws.cell(current_row, 4).alignment = align_left
    
    for col in range(2, 7):
        cell = ws.cell(current_row, col)
        cell.fill = fill_result_box
        left_s = Side(style="medium", color=STEEL_BLUE) if col == 2 else None
        right_s = Side(style="medium", color=STEEL_BLUE) if col == 6 else None
        cell.border = Border(left=left_s, right=right_s)
        
    s2_ref = f"C{current_row}" # C50
    current_row += 1

    # 4. Trình độ ứng viên đề xuất tự đánh giá
    ws.row_dimensions[current_row].height = 20
    ws.cell(current_row, 2, "Xếp hạng trình độ tự đề xuất:").font = font_bold
    ws.cell(current_row, 2).alignment = align_left
    
    cell_level = ws.cell(current_row, 3, f'=IF(({s3_ref}+{s2_ref})>=20, "Đề xuất Thợ chính (Cấp A/B)", IF(({s3_ref}+{s2_ref})>=10, "Đề xuất Thợ phụ cứng (Cấp C)", "Đề xuất Thợ học việc"))')
    cell_level.font = font_result
    cell_level.alignment = align_center
    ws.cell(current_row, 4, "(Dựa trên tổng hạng mục biết làm)").font = font_italic
    ws.cell(current_row, 4).alignment = align_left
    
    for col in range(2, 7):
        cell = ws.cell(current_row, col)
        cell.fill = fill_result_box
        left_s = Side(style="medium", color=STEEL_BLUE) if col == 2 else None
        right_s = Side(style="medium", color=STEEL_BLUE) if col == 6 else None
        cell.border = Border(left=left_s, right=right_s, bottom=Side(style="medium", color=STEEL_BLUE))

    current_row += 2

    # Slogan block matching company motto
    ws.row_dimensions[current_row].height = 26
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    slogan_cell = ws.cell(current_row, 1, 'Khẩu hiệu của chúng tôi: "Chất Lượng Tạo Nên Thương Hiệu. Uy Tín Tạo Nên Sự Thành Công"')
    slogan_cell.font = Font(name="Segoe UI", size=10, bold=True, italic=True, color=STEEL_BLUE)
    slogan_cell.alignment = align_center
    slogan_cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    for col in range(1, 7):
        ws.cell(current_row, col).border = Border(
            top=Side(style="thin", color=STEEL_BLUE),
            bottom=Side(style="thin", color=STEEL_BLUE),
            left=Side(style="thin", color=STEEL_BLUE) if col == 1 else None,
            right=Side(style="thin", color=STEEL_BLUE) if col == 6 else None
        )

    current_row += 3
    
    # Signatures
    ws.row_dimensions[current_row].height = 18
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    s1 = ws.cell(current_row, 1, "ỨNG VIÊN TUYỂN DỤNG")
    s1.font = font_bold
    s1.alignment = align_center
    
    ws.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=6)
    s2 = ws.cell(current_row, 4, "PHÒNG NHÂN SỰ XÁC NHẬN")
    s2.font = font_bold
    s2.alignment = align_center
    
    current_row += 1
    ws.row_dimensions[current_row].height = 15
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    sub1 = ws.cell(current_row, 1, "(Ký và ghi rõ họ tên)")
    sub1.font = font_italic
    sub1.alignment = align_center
    
    ws.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=6)
    sub2 = ws.cell(current_row, 4, "(Ký nhận và lưu hồ sơ phỏng vấn)")
    sub2.font = font_italic
    sub2.alignment = align_center

    # Add Data Validation for column D (Dropdown choice)
    dv = DataValidation(
        type="list", 
        formula1='"3 - Làm tốt độc lập, 2 - Làm được (Cần hướng dẫn), 1 - Biết sơ bộ / Đang học, 0 - Chưa làm bao giờ"', 
        allow_blank=True
    )
    dv.prompt = 'Vui lòng chọn mức độ thành thạo của bạn cho công việc này'
    dv.promptTitle = 'Mức độ tự tin'
    ws.add_data_validation(dv)
    
    for r in eval_rows:
        dv.add(ws[f"D{r}"])

    paths = [
        "d:/Sao Vàng/Website-SaoVang/CKSV_Phieu_Tu_Danh_Gia_Tho_Co_Khi.xlsx",
        "d:/Sao Vàng/Website-SaoVang/CKSV_Phieu_Tu_Danh_Gia_Tho_Co_Khi_Co_Logo.xlsx"
    ]
    saved_paths = []
    for path in paths:
        try:
            wb.save(path)
            saved_paths.append(path)
        except PermissionError:
            print(f"Permission denied for {path} (likely open in Excel). Skipping.")
            
    print(f"Self-evaluation sheet with logo saved to: {saved_paths}")

if __name__ == "__main__":
    create_candidate_self_eval_with_logo()
