import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

def create_recruitment_form():
    wb = openpyxl.Workbook()
    
    # -------------------------------------------------------------
    # SHEET 1: FORM TUYỂN DỤNG & THỬ VIỆC
    # -------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Đánh Giá Tuyển Dụng"
    ws1.sheet_view.showGridLines = True
    
    # Colors
    STEEL_BLUE = "2F5597"
    ICE_BLUE = "D9E1F2"
    LIGHT_GRAY = "F2F2F2"
    BORDER_GRAY = "D9D9D9"
    YELLOW_INPUT = "FFF2CC"
    DARK_GREEN = "385723"
    LIGHT_GREEN = "E2EFDA"
    
    # Fonts
    font_title_company = Font(name="Segoe UI", size=11, bold=True, color="595959")
    font_title_main = Font(name="Segoe UI", size=16, bold=True, color=STEEL_BLUE)
    font_title_sub = Font(name="Segoe UI", size=10, italic=True, color="7F7F7F")
    font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    font_group_header = Font(name="Segoe UI", size=11, bold=True, color=STEEL_BLUE)
    font_bold = Font(name="Segoe UI", size=10, bold=True)
    font_normal = Font(name="Segoe UI", size=10)
    font_italic = Font(name="Segoe UI", size=9, italic=True)
    font_salary_result = Font(name="Segoe UI", size=12, bold=True, color=DARK_GREEN)
    
    # Fills
    fill_header = PatternFill(start_color=STEEL_BLUE, end_color=STEEL_BLUE, fill_type="solid")
    fill_group = PatternFill(start_color=ICE_BLUE, end_color=ICE_BLUE, fill_type="solid")
    fill_sum = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
    fill_input = PatternFill(start_color=YELLOW_INPUT, end_color=YELLOW_INPUT, fill_type="solid")
    fill_result_box = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
    
    # Alignments
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    
    # Borders
    thin_side = Side(style="thin", color=BORDER_GRAY)
    border_all_thin = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    border_group = Border(
        left=Side(style="thin", color=STEEL_BLUE),
        right=Side(style="thin", color=STEEL_BLUE),
        top=Side(style="thin", color=STEEL_BLUE),
        bottom=Side(style="thin", color=STEEL_BLUE)
    )
    
    # Column widths
    ws1.column_dimensions['A'].width = 5   # STT
    ws1.column_dimensions['B'].width = 38  # Nhiệm vụ tuyển dụng
    ws1.column_dimensions['C'].width = 62  # Kỹ năng / Kết quả cần đạt được
    ws1.column_dimensions['D'].width = 16  # Ứng viên tự đánh giá (% đáp ứng)
    ws1.column_dimensions['E'].width = 16  # Công ty đánh giá (1: Đạt, 0: Chưa)
    ws1.column_dimensions['F'].width = 25  # Ghi chú / Nhận xét

    # Title Block
    ws1.merge_cells("A2:F2")
    ws1["A2"] = "CÔNG TY CỔ PHẦN SẢN XUẤT CƠ KHÍ SAO VÀNG (ASV)"
    ws1["A2"].font = font_title_company
    ws1["A2"].alignment = Alignment(horizontal="left")

    ws1.merge_cells("A3:F3")
    ws1["A3"] = "FORM TUYỂN DỤNG VÀ ĐÁNH GIÁ NĂNG LỰC THỬ VIỆC THỢ CƠ KHÍ"
    ws1["A3"].font = font_title_main
    ws1["A3"].alignment = align_center

    ws1.merge_cells("A4:F4")
    ws1["A4"] = "Xác định ngạch bậc & lương ngày dựa trên kỹ năng thực tế của ứng viên"
    ws1["A4"].font = font_title_sub
    ws1["A4"].alignment = align_center

    # Candidate Profile Info
    ws1["B6"] = "Họ và tên ứng viên:"
    ws1["B6"].font = font_bold
    ws1["C6"] = "[Nhập tên ứng viên]"
    ws1["C6"].font = font_normal
    ws1["C6"].border = Border(bottom=Side(style="thin", color="000000"))

    ws1["E6"] = "Ngày phỏng vấn:"
    ws1["E6"].font = font_bold
    ws1["F6"] = "[Nhập ngày]"
    ws1["F6"].font = font_normal
    ws1["F6"].border = Border(bottom=Side(style="thin", color="000000"))

    ws1["B7"] = "Số điện thoại/CCCD:"
    ws1["B7"].font = font_bold
    ws1["C7"] = "[Số điện thoại và CCCD]"
    ws1["C7"].font = font_normal
    ws1["C7"].border = Border(bottom=Side(style="thin", color="000000"))

    ws1["E7"] = "Người tuyển dụng:"
    ws1["E7"].font = font_bold
    ws1["F7"] = "[Tên người phỏng vấn]"
    ws1["F7"].font = font_normal
    ws1["F7"].border = Border(bottom=Side(style="thin", color="000000"))

    ws1["B8"] = "Vị trí ứng tuyển:"
    ws1["B8"].font = font_bold
    ws1["C8"] = "Thợ Cơ khí (Học việc/Phụ/Giúp việc/Thợ chính/Tổ trưởng)"
    ws1["C8"].font = font_italic
    ws1["C8"].border = Border(bottom=Side(style="thin", color="000000"))

    ws1["E8"] = "Kỳ thử việc:"
    ws1["E8"].font = font_bold
    ws1["F8"] = "[Ví dụ: 01 tháng thử việc]"
    ws1["F8"].font = font_normal
    ws1["F8"].border = Border(bottom=Side(style="thin", color="000000"))

    # Table Header
    headers = [
        ("A10", "STT", align_center),
        ("B10", "NHIỆM VỤ ĐÁNH GIÁ (VIỆC CẦN LÀM)", align_left),
        ("C10", "YÊU CẦU KỸ NĂNG VÀ KẾT QUẢ BÊN SDLĐ CẦN ĐẠT ĐƯỢC", align_left),
        ("D10", "ỨNG VIÊN TỰ ĐÁNH GIÁ (% Đáp ứng)", align_center),
        ("E10", "CÔNG TY ĐÁNH GIÁ (Đạt: 1, Chưa: 0)", align_center),
        ("F10", "GHI CHÚ / NHẬN XÉT CHI TIẾT", align_left)
    ]
    for cell_ref, text, align in headers:
        cell = ws1[cell_ref]
        cell.value = text
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align
        cell.border = border_all_thin
    ws1.row_dimensions[10].height = 28

    # Criteria data
    # format: (STT, Task, Skill/Outcome, Bậc)
    # Group sections are created dynamically
    criteria_data = [
        # Group I: HỌC VIỆC & PHỤ VIỆC (Bậc 1 - 5)
        ("G", "I. HỌC VIỆC VÀ PHỤ VIỆC (Tiêu chuẩn Bậc 1 - Bậc 5)", "", ""),
        ("1", "Tác phong, sức khỏe & đạo đức nghề nghiệp", "Sức khỏe tốt, nhanh nhẹn, chịu khó học hỏi, trung thực, có trách nhiệm, chấp hành nội quy.", 1),
        ("2", "Vệ sinh và thực hiện 5S tại nơi làm việc", "Chủ động dọn dẹp máy móc thiết bị, quét dọn khu vực thi công sau buổi làm ở xưởng và công trường.", 2),
        ("3", "Sử dụng máy móc cơ bản trong xưởng", "Sử dụng an toàn máy cắt tay, máy cắt bàn, máy khoan cầm tay, máy cân bằng laser, thước đo...", 3),
        ("4", "Kỹ thuật mài hoàn thiện sản phẩm", "Mài tinh mỹ phẳng tuyệt đối/góc sắc nét, mài phá mối hàn thô dầm nặng, mài đánh bóng inox.", 4),
        ("5", "Kỹ thuật phun sơn hoàn thiện", "Sử dụng thành thạo súng phun sơn khí nén, sơn lót, sơn phủ mịn đều không bị chảy sơn.", 5),
        
        # Group II: GIÚP VIỆC (Tiêu chuẩn Bậc 6 - Bậc 7)
        ("G", "II. GIÚP VIỆC (Tiêu chuẩn Bậc 6 - Bậc 7)", "", ""),
        ("6", "Quản lý vật tư và chuẩn bị hiện trường", "Nắm rõ các loại vật tư, phôi cắt sắt/inox; chủ động chuẩn bị đầy đủ vật tư phụ, máy móc công trình.", 6),
        ("7", "Hàn đính gá kết cấu cơ bản", "Đo cắt phôi chính xác theo thước đo, biết làm dưỡng cắt hàng loạt, hàn đính gá chuẩn các góc kết cấu.", 7),
        
        # Group III: THỢ CHÍNH - TỰ THỰC HIỆN (Tiêu chuẩn Bậc 8 - Bậc 10)
        ("G", "III. THỢ CHÍNH - TỰ THỰC HIỆN (Tiêu chuẩn Bậc 8 - Bậc 10)", "", ""),
        ("8", "Hàn kỹ thuật chuyên nghiệp (MIG/TIG/MMA)", "Hàn thành thạo máy hàn que, hàn MIG, hàn TIG. Mối hàn ngấu sâu, đều vảy, không cháy cạnh, không rỗ.", 8),
        ("9", "Đọc bản vẽ sản xuất và khai triển phôi", "Đọc hiểu bản vẽ lắp ráp cơ khí phức tạp, tự bóc tách kích thước phôi để gia công khoa học, tiết kiệm vật tư.", 9),
        ("10", "Kỹ năng tự kiểm soát chất lượng (QC)", "Tự kiểm tra sai lệch kích thước, độ phẳng, độ vuông góc của sản phẩm; tự sửa chữa không giấu dốt.", 10),
        
        # Group IV: TỔ TRƯỞNG & ĐIỀU HÀNH (Tiêu chuẩn Bậc 11 - Bậc 14)
        ("G", "IV. TỔ TRƯỞNG VÀ ĐIỀU HÀNH SẢN XUẤT (Tiêu chuẩn Bậc 11 - Bậc 14)", "", ""),
        ("11", "Ước lượng định mức nhân công & thời gian", "Khả năng đánh giá, ước lượng chuẩn xác số nhân công và khoảng thời gian để hoàn thành một kết cấu.", 11),
        ("12", "Chủ động báo cáo & đề xuất tối ưu", "Báo cáo tiến độ cuối ngày, chủ động đề xuất giải pháp thi công nhanh hơn, năng suất hơn.", 12),
        ("13", "Khả năng làm việc độc lập & Độc quyền quyết định", "Độc lập thi công công việc được giao; tự chủ động điều chỉnh chi tiết nhỏ phù hợp hiện trường mà không ảnh hưởng kết cấu.", 13),
        ("14", "Lãnh đạo đội nhóm và điều phối xưởng", "Thành thạo 90% việc tổ mình và 50% tổ khác. Khả năng phân chia việc, quản lý nhóm 3-5 thợ phụ/chính đạt tiến độ.", 14)
    ]

    current_row = 11
    eval_inputs = [] # to store row numbers of Company Eval for the formula
    
    for row_data in criteria_data:
        stt, task, skill, r_type = row_data[0], row_data[1], row_data[2], row_data[3]
        
        ws1.row_dimensions[current_row].height = 24
        
        if stt == "G":
            # Group Header
            for col in range(1, 7):
                cell = ws1.cell(current_row, col)
                cell.fill = fill_group
                cell.border = border_group
                
            ws1.cell(current_row, 2, task).font = font_group_header
            ws1.cell(current_row, 2).alignment = align_left
            
        else:
            # Item row
            ws1.cell(current_row, 1, stt).font = font_normal
            ws1.cell(current_row, 1).alignment = align_center
            
            ws1.cell(current_row, 2, task).font = font_normal
            ws1.cell(current_row, 2).alignment = align_left
            
            ws1.cell(current_row, 3, skill).font = font_normal
            ws1.cell(current_row, 3).alignment = align_left
            
            # Candidate Self Eval (%) - Input
            cell_self = ws1.cell(current_row, 4)
            cell_self.font = font_bold
            cell_self.alignment = align_center
            cell_self.fill = fill_input
            cell_self.number_format = '0%'
            cell_self.value = 1.0  # default 100%
            
            # Company Eval (1: Đạt, 0: Chưa) - Input
            cell_company = ws1.cell(current_row, 5)
            cell_company.font = font_bold
            cell_company.alignment = align_center
            cell_company.fill = fill_input
            cell_company.value = 0  # default 0
            eval_inputs.append(current_row)
            
            # Note - Input
            cell_note = ws1.cell(current_row, 6)
            cell_note.font = font_normal
            cell_note.alignment = align_left
            cell_note.fill = fill_input
            cell_note.value = "[Ghi chú đánh giá]"
            
            for col in range(1, 7):
                ws1.cell(current_row, col).border = border_all_thin
                
        current_row += 1

    # Place sample ratings for verification
    # Let's say candidate is competent up to Bậc 8 (Hàn chuyên nghiệp)
    ws1["E12"] = 1  # Bậc 1
    ws1["E13"] = 1  # Bậc 2
    ws1["E14"] = 1  # Bậc 3
    ws1["E15"] = 1  # Bậc 4
    ws1["E16"] = 1  # Bậc 5
    ws1["E19"] = 1  # Bậc 6
    ws1["E20"] = 1  # Bậc 7
    ws1["E22"] = 1  # Bậc 8 (hàn chuyên nghiệp)
    ws1["E23"] = 0  # Bậc 9
    ws1["E24"] = 0  # Bậc 10
    ws1["E26"] = 0  # Bậc 11
    ws1["E27"] = 0  # Bậc 12
    ws1["E28"] = 0  # Bậc 13
    ws1["E29"] = 0  # Bậc 14

    ws1.row_dimensions[current_row].height = 15
    current_row += 1

    # RESULTS DASHBOARD BOX
    ws1.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=6)
    r_hdr = ws1.cell(current_row, 2, "KẾT QUẢ PHÂN BẬC LƯƠNG ĐỀ XUẤT SAU THỬ VIỆC")
    r_hdr.font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    r_hdr.fill = PatternFill(start_color=STEEL_BLUE, end_color=STEEL_BLUE, fill_type="solid")
    r_hdr.alignment = align_center
    ws1.row_dimensions[current_row].height = 24
    
    for col in range(2, 7):
        ws1.cell(current_row, col).border = Border(top=Side(style="medium", color=STEEL_BLUE))
        
    current_row += 1
    
    # We construct the MAX formula based on exact row numbers of criteria rows
    # criteria rows are: 
    # Bậc 1: Row 12 (stt 1)
    # Bậc 2: Row 13 (stt 2)
    # Bậc 3: Row 14 (stt 3)
    # Bậc 4: Row 15 (stt 4)
    # Bậc 5: Row 16 (stt 5)
    # Bậc 6: Row 19 (stt 6)
    # Bậc 7: Row 20 (stt 7)
    # Bậc 8: Row 22 (stt 8)
    # Bậc 9: Row 23 (stt 9)
    # Bậc 10: Row 24 (stt 10)
    # Bậc 11: Row 26 (stt 11)
    # Bậc 12: Row 27 (stt 12)
    # Bậc 13: Row 28 (stt 13)
    # Bậc 14: Row 29 (stt 14)
    
    r_map = {
        1: 12, 2: 13, 3: 14, 4: 15, 5: 16, 
        6: 19, 7: 20, 
        8: 22, 9: 23, 10: 24, 
        11: 26, 12: 27, 13: 28, 14: 29
    }
    
    # MAX(IF(E12=1,1,0), IF(E13=1,2,0)...)
    max_parts = []
    for b_num, r_num in r_map.items():
        max_parts.append(f"IF(E{r_num}=1,{b_num},0)")
    max_formula = "=MAX(" + ",".join(max_parts) + ")"
    
    # Write Dashboard Rows
    # 1. Bậc năng lực đạt được
    ws1.row_dimensions[current_row].height = 20
    ws1.cell(current_row, 2, "Bậc năng lực đạt được:").font = font_bold
    ws1.cell(current_row, 2).alignment = align_left
    
    cell_bac = ws1.cell(current_row, 3, max_formula)
    cell_bac.font = font_bold
    cell_bac.alignment = align_center
    
    ws1.cell(current_row, 4, "Bậc / 14").font = font_italic
    ws1.cell(current_row, 4).alignment = align_left
    
    # Style borders
    for col in range(2, 7):
        cell = ws1.cell(current_row, col)
        cell.fill = fill_result_box
        left_s = Side(style="medium", color=STEEL_BLUE) if col == 2 else None
        right_s = Side(style="medium", color=STEEL_BLUE) if col == 6 else None
        cell.border = Border(left=left_s, right=right_s)
        
    bac_cell_ref = f"C{current_row}" # This cell stores the bậc number (e.g. C31)
    current_row += 1
    
    # 2. Chức danh tuyển dụng đề xuất
    ws1.row_dimensions[current_row].height = 20
    ws1.cell(current_row, 2, "Chức danh tuyển dụng đề xuất:").font = font_bold
    ws1.cell(current_row, 2).alignment = align_left
    
    cell_title = ws1.cell(current_row, 3, f"=IF({bac_cell_ref}=0, \"Chưa đạt\", VLOOKUP({bac_cell_ref}, 'Thang Đo & Lương'!A$2:E$15, 2, FALSE))")
    cell_title.font = font_bold
    cell_title.alignment = align_center
    
    for col in range(2, 7):
        cell = ws1.cell(current_row, col)
        cell.fill = fill_result_box
        left_s = Side(style="medium", color=STEEL_BLUE) if col == 2 else None
        right_s = Side(style="medium", color=STEEL_BLUE) if col == 6 else None
        cell.border = Border(left=left_s, right=right_s)
        
    current_row += 1
    
    # 3. Mức lương thực lĩnh (ngày)
    ws1.row_dimensions[current_row].height = 20
    ws1.cell(current_row, 2, "Lương thực lĩnh đề xuất (ngày):").font = font_bold
    ws1.cell(current_row, 2).alignment = align_left
    
    cell_wage_d = ws1.cell(current_row, 3, f"=IF({bac_cell_ref}=0, 0, VLOOKUP({bac_cell_ref}, 'Thang Đo & Lương'!A$2:E$15, 3, FALSE))")
    cell_wage_d.font = font_salary_result
    cell_wage_d.alignment = align_center
    cell_wage_d.number_format = '#,##0" đ"'
    
    ws1.cell(current_row, 4, "(Lương sau thuế/ngày)").font = font_italic
    ws1.cell(current_row, 4).alignment = align_left
    
    for col in range(2, 7):
        cell = ws1.cell(current_row, col)
        cell.fill = fill_result_box
        left_s = Side(style="medium", color=STEEL_BLUE) if col == 2 else None
        right_s = Side(style="medium", color=STEEL_BLUE) if col == 6 else None
        cell.border = Border(left=left_s, right=right_s)
        
    wage_d_ref = f"C{current_row}"
    current_row += 1
    
    # 4. Dự toán thu nhập tháng (26 ngày công)
    ws1.row_dimensions[current_row].height = 20
    ws1.cell(current_row, 2, "Thu nhập tháng dự kiến (26 công):").font = font_bold
    ws1.cell(current_row, 2).alignment = align_left
    
    cell_wage_m = ws1.cell(current_row, 3, f"={wage_d_ref}*26")
    cell_wage_m.font = font_salary_result
    cell_wage_m.alignment = align_center
    cell_wage_m.number_format = '#,##0" đ"'
    
    ws1.cell(current_row, 4, "(Tương ứng 15-22 triệu/tháng)").font = font_italic
    ws1.cell(current_row, 4).alignment = align_left
    
    for col in range(2, 7):
        cell = ws1.cell(current_row, col)
        cell.fill = fill_result_box
        left_s = Side(style="medium", color=STEEL_BLUE) if col == 2 else None
        right_s = Side(style="medium", color=STEEL_BLUE) if col == 6 else None
        cell.border = Border(left=left_s, right=right_s)
        
    current_row += 1
    
    # 5. Khuyến nghị và Kế hoạch sử dụng
    ws1.row_dimensions[current_row].height = 24
    ws1.cell(current_row, 2, "Khuyến nghị & Phân công:").font = font_bold
    ws1.cell(current_row, 2).alignment = align_left
    
    ws1.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=6)
    rec_cell = ws1.cell(current_row, 3, f"=IF({bac_cell_ref}=0, \"Yêu cầu đào tạo thêm hoặc từ chối thử việc\", VLOOKUP({bac_cell_ref}, 'Thang Đo & Lương'!A$2:E$15, 5, FALSE))")
    rec_cell.font = font_bold
    rec_cell.alignment = align_left
    
    for col in range(2, 7):
        cell = ws1.cell(current_row, col)
        cell.fill = fill_result_box
        left_s = Side(style="medium", color=STEEL_BLUE) if col == 2 else None
        right_s = Side(style="medium", color=STEEL_BLUE) if col == 6 else None
        cell.border = Border(left=left_s, right=right_s, bottom=Side(style="medium", color=STEEL_BLUE))

    # Data validations
    # Limit self-eval to decimals between 0 and 1 (0% and 100%)
    dv_self = DataValidation(type="decimal", operator="between", formula1="0", formula2="1", allow_blank=True)
    dv_self.error = 'Tỷ lệ đáp ứng phải từ 0% đến 100%!'
    dv_self.errorTitle = 'Lỗi nhập liệu'
    dv_self.prompt = 'Nhập tỷ lệ % tự đánh giá (Ví dụ: 80%)'
    dv_self.promptTitle = 'Tự đánh giá'
    ws1.add_data_validation(dv_self)
    
    # Limit company-eval to integer between 0 and 1
    dv_company = DataValidation(type="whole", operator="between", formula1="0", formula2="1", allow_blank=True)
    dv_company.error = 'Điểm đánh giá phải là 1 (Đạt) hoặc 0 (Chưa đạt)!'
    dv_company.errorTitle = 'Lỗi nhập điểm'
    dv_company.prompt = 'Nhập 1 nếu đạt tiêu chuẩn, nhập 0 nếu chưa đạt'
    dv_company.promptTitle = 'Chấm điểm đạt'
    ws1.add_data_validation(dv_company)

    # Apply validations to columns D and E
    for r_num in r_map.values():
        dv_self.add(ws1[f"D{r_num}"])
        dv_company.add(ws1[f"E{r_num}"])

    current_row += 3
    
    # Signatures
    ws1.row_dimensions[current_row].height = 18
    ws1.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    s1 = ws1.cell(current_row, 1, "ỨNG VIÊN TUYỂN DỤNG")
    s1.font = font_bold
    s1.alignment = align_center
    
    ws1.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=6)
    s2 = ws1.cell(current_row, 4, "TỔ TRƯỞNG / QUẢN ĐỐC DUYỆT")
    s2.font = font_bold
    s2.alignment = align_center
    
    current_row += 1
    ws1.row_dimensions[current_row].height = 15
    ws1.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    sub1 = ws1.cell(current_row, 1, "(Ký và ghi rõ họ tên)")
    sub1.font = font_italic
    sub1.alignment = align_center
    
    ws1.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=6)
    sub2 = ws1.cell(current_row, 4, "(Phê duyệt ngạch bậc và lương)")
    sub2.font = font_italic
    sub2.alignment = align_center
    
    current_row += 4
    ws1.row_dimensions[current_row].height = 18
    ws1.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    n1 = ws1.cell(current_row, 1, "[Họ tên ứng viên]")
    n1.font = font_normal
    n1.alignment = align_center
    
    ws1.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=6)
    n2 = ws1.cell(current_row, 4, "[Họ tên người đánh giá]")
    n2.font = font_normal
    n2.alignment = align_center

    # -------------------------------------------------------------
    # SHEET 2: THANG ĐO & LƯƠNG
    # -------------------------------------------------------------
    ws2 = wb.create_sheet(title="Thang Đo & Lương")
    ws2.sheet_view.showGridLines = True
    
    ws2.column_dimensions['A'].width = 10
    ws2.column_dimensions['B'].width = 24
    ws2.column_dimensions['C'].width = 24
    ws2.column_dimensions['D'].width = 24
    ws2.column_dimensions['E'].width = 65
    
    # Headers
    ws2["A1"] = "Bậc"
    ws2["B1"] = "Chức danh đề xuất"
    ws2["C1"] = "Lương thực lĩnh (ngày)"
    ws2["D1"] = "Tổng chi phí công ty (ngày)"
    ws2["E1"] = "Phân công / Vai trò chính trong xưởng"
    
    for col in ["A", "B", "C", "D", "E"]:
        cell = ws2[f"{col}1"]
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_all_thin
        
    scale_data = [
        (1, "Thợ học việc (Cấp 1)", 220000, 300000, "Phụ việc vệ sinh, di chuyển phôi mài mộc dưới sự hướng dẫn"),
        (2, "Thợ học việc (Cấp 2)", 220000, 300000, "Thực hiện tốt 5S, dọn dẹp mặt bằng và chuẩn bị an toàn ATLĐ"),
        (3, "Thợ phụ cơ bản (Cấp 1)", 250000, 330000, "Sử dụng máy cắt cầm tay, khoan bê tông, thước đo chính xác"),
        (4, "Thợ phụ mài tinh (Cấp 2)", 290000, 370000, "Mài hoàn thiện sản phẩm đạt phẳng tuyệt đối/góc sắc cạnh"),
        (5, "Thợ sơn hoàn thiện", 330000, 410000, "Đứng buồng sơn phun khí nén, sơn chống rỉ và sơn phủ thẩm mỹ"),
        (6, "Thợ gá lắp phôi", 380000, 460000, "Nắm rõ các loại vật tư, phôi cắt sắt/inox, chuẩn bị công trình"),
        (7, "Thợ gá hàn liên hợp", 400000, 480000, "Gá kết cấu cơ bản, hàn đính gá chuẩn góc trước khi hàn chính"),
        (8, "Thợ hàn chính chuyên nghiệp", 430000, 510000, "Độc lập hàn MIG/TIG kết cấu thép hoặc inox thẩm mỹ cao"),
        (9, "Thợ lắp ráp lành nghề", 440000, 520000, "Đọc bản vẽ sản xuất phức tạp, tự bóc tách khai triển phôi gá"),
        (10, "Thợ hoàn thiện độc lập", 470000, 550000, "Tự kiểm soát chất lượng hoàn thiện sản phẩm không cần giám sát"),
        (11, "Thợ dự toán & tiến độ", 480000, 560000, "Ước lượng thời gian/nhân công, lập tiến độ dự án nhỏ"),
        (12, "Tổ phó kỹ thuật", 490000, 570000, "Tự chủ động điều chỉnh chi tiết nhỏ ở hiện trường, thay thế tổ trưởng"),
        (13, "Trưởng nhóm lắp dựng", 500000, 580000, "Điều phối nhóm 2-3 người lắp dựng kết cấu thép/inox ngoài công trường"),
        (14, "Tổ trưởng sản xuất", 550000, 630000, "Phân việc toàn xưởng, chịu trách nhiệm tiến độ & an toàn toàn xưởng")
    ]
    
    for r_idx, row_val in enumerate(scale_data, start=2):
        ws2.cell(r_idx, 1, row_val[0]).alignment = align_center
        ws2.cell(r_idx, 1).font = font_bold
        
        ws2.cell(r_idx, 2, row_val[1]).alignment = align_left
        ws2.cell(r_idx, 2).font = font_bold
        
        ws2.cell(r_idx, 3, row_val[2]).number_format = '#,##0" đ"'
        ws2.cell(r_idx, 3).alignment = align_right
        ws2.cell(r_idx, 3).font = font_bold
        
        ws2.cell(r_idx, 4, row_val[3]).number_format = '#,##0" đ"'
        ws2.cell(r_idx, 4).alignment = align_right
        ws2.cell(r_idx, 4).font = font_normal
        
        ws2.cell(r_idx, 5, row_val[4]).alignment = align_left
        ws2.cell(r_idx, 5).font = font_normal
        
        for col in range(1, 6):
            cell = ws2.cell(r_idx, col)
            cell.border = border_all_thin
            if r_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

    # Save workbook
    output_path = "d:/Sao Vàng/Website-SaoVang/CKSV_Form_Tuyen_Dung_Tho_Co_Khi.xlsx"
    wb.save(output_path)
    print(f"Recruitment form created successfully at {output_path}")

if __name__ == "__main__":
    create_recruitment_form()
