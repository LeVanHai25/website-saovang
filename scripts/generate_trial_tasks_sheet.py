import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.drawing.image import Image as OpenpyxlImage

def create_trial_tasks_workbook():
    wb = openpyxl.Workbook()
    
    # -------------------------------------------------------------
    # STYLES CONFIGURATION
    # -------------------------------------------------------------
    STEEL_BLUE = "2F5597"
    ICE_BLUE = "D9E1F2"
    LIGHT_GRAY = "F2F2F2"
    BORDER_GRAY = "D9D9D9"
    YELLOW_INPUT = "FFF2CC"
    DARK_GREEN = "385723"
    LIGHT_GREEN = "E2EFDA"
    
    # Fonts
    font_company_name = Font(name="Segoe UI", size=12, bold=True, color=STEEL_BLUE)
    font_company_info = Font(name="Segoe UI", size=9, italic=True, color="595959")
    font_title_main = Font(name="Segoe UI", size=15, bold=True, color=STEEL_BLUE)
    font_title_sub = Font(name="Segoe UI", size=10, italic=True, color="7F7F7F")
    font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    font_group_header = Font(name="Segoe UI", size=11, bold=True, color=STEEL_BLUE)
    font_bold = Font(name="Segoe UI", size=10, bold=True)
    font_normal = Font(name="Segoe UI", size=10)
    font_italic = Font(name="Segoe UI", size=9, italic=True)
    font_result = Font(name="Segoe UI", size=12, bold=True, color=DARK_GREEN)
    
    # Fills
    fill_header = PatternFill(start_color=STEEL_BLUE, end_color=STEEL_BLUE, fill_type="solid")
    fill_group = PatternFill(start_color=ICE_BLUE, end_color=ICE_BLUE, fill_type="solid")
    fill_input = PatternFill(start_color=YELLOW_INPUT, end_color=YELLOW_INPUT, fill_type="solid")
    fill_result_box = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
    
    # Alignments
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    
    # Borders
    thin_side = Side(style="thin", color=BORDER_GRAY)
    border_all_thin = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    border_double_bottom = Border(
        left=thin_side, right=thin_side, 
        top=thin_side, 
        bottom=Side(style="double", color="000000")
    )

    logo_path = "d:/Sao Vàng/Website-SaoVang/Logo/LogoCTY + SV Aluminium/LogoCTY.png"

    # Active sheet
    ws = wb.active
    ws.title = "Bảng Đánh Giá Thực Chiến"
    ws.sheet_view.showGridLines = True
    
    # Column widths
    ws.column_dimensions['A'].width = 6   # STT
    ws.column_dimensions['B'].width = 24  # Nhóm chuyên môn
    ws.column_dimensions['C'].width = 30  # Bài thi thực hành
    ws.column_dimensions['D'].width = 48  # Công việc cần làm
    ws.column_dimensions['E'].width = 45  # Kết quả đạt được
    ws.column_dimensions['F'].width = 18  # Thời gian kiểm tra
    ws.column_dimensions['G'].width = 16  # KẾT QUẢ (1: Đạt, 0: Chưa)
    ws.column_dimensions['H'].width = 25  # Ghi chú nhận xét

    # 1. Company Letterhead
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 15
    ws.row_dimensions[4].height = 15
    
    ws.merge_cells("C2:H2")
    ws["C2"] = "CÔNG TY CỔ PHẦN SẢN XUẤT CƠ KHÍ SAO VÀNG (ASV)"
    ws["C2"].font = font_company_name
    ws["C2"].alignment = Alignment(horizontal="left", vertical="center")
    
    ws.merge_cells("C3:H3")
    ws["C3"] = "Địa chỉ: Tầng 3, TT7-35 Khu đô thị Văn Phú, phường Kiến Hưng, TP Hà Nội, Việt Nam."
    ws["C3"].font = font_company_info
    ws["C3"].alignment = Alignment(horizontal="left", vertical="center")
    
    ws.merge_cells("C4:H4")
    ws["C4"] = "Hotline: 0869 590 279  |  Email: cokhisaovangvn@gmail.com"
    ws["C4"].font = font_company_info
    ws["C4"].alignment = Alignment(horizontal="left", vertical="center")

    try:
        img = OpenpyxlImage(logo_path)
        img.width = 110
        img.height = 45
        ws.add_image(img, "A2")
    except Exception as e:
        print(f"Error adding logo: {e}")

    # Document Title Block (shifted to Row 6-7)
    ws.row_dimensions[6].height = 24
    ws.merge_cells("A6:H6")
    ws["A6"] = "BẢNG CHẤM ĐIỂM CÁC BÀI THI THỬ TAY NGHỀ THỰC TẾ"
    ws["A6"].font = font_title_main
    ws["A6"].alignment = align_center
    
    ws.row_dimensions[7].height = 18
    ws.merge_cells("A7:H7")
    ws["A7"] = "Liệt kê danh sách các hạng mục việc thực tế tự chọn tại xưởng để đánh giá độ thành thạo thợ cơ khí"
    ws["A7"].font = font_title_sub
    ws["A7"].alignment = align_center

    # Candidate profile block (shifted to Row 9-10)
    ws["B9"] = "Họ và tên ứng viên:"
    ws["B9"].font = font_bold
    ws["C9"] = "[Nhập họ tên]"
    ws["C9"].font = font_normal
    ws["C9"].border = Border(bottom=Side(style="thin", color="000000"))

    ws["F9"] = "Người chấm thi (Quản đốc):"
    ws["F9"].font = font_bold
    ws["G9"] = "[Nhập tên]"
    ws["G9"].font = font_normal
    ws["G9"].border = Border(bottom=Side(style="thin", color="000000"))

    ws["B10"] = "Vị trí ứng tuyển:"
    ws["B10"].font = font_bold
    ws["C10"] = "Thợ Cơ khí (Học việc / Phụ / Giúp việc / Thợ chính)"
    ws["C10"].font = font_italic
    ws["C10"].border = Border(bottom=Side(style="thin", color="000000"))

    ws["F10"] = "Ngày kiểm tra:"
    ws["F10"].font = font_bold
    ws["G10"] = "[Ghi ngày]"
    ws["G10"].font = font_normal
    ws["G10"].border = Border(bottom=Side(style="thin", color="000000"))

    ws.row_dimensions[11].height = 15

    # Main Table Header (Row 12)
    ws.row_dimensions[12].height = 28
    headers = [
        ("A12", "STT", align_center),
        ("B12", "NHÓM CHUYÊN MÔN", align_left),
        ("C12", "TÊN BÀI THI THỰC HÀNH", align_left),
        ("D12", "CÔNG VIỆC CẦN LÀM CHI TIẾT", align_left),
        ("E12", "KẾT QUẢ ĐẠT ĐƯỢC YÊU CẦU", align_left),
        ("F12", "THỜI GIAN KIỂM TRA", align_center),
        ("G12", "KẾT QUẢ CHẤM (Đạt: 1 | 0)", align_center),
        ("H12", "GHI CHÚ CHI TIẾT", align_left)
    ]
    for cell_ref, text, align in headers:
        cell = ws[cell_ref]
        cell.value = text
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align
        cell.border = border_all_thin

    # 12 practical trial tasks data
    trials_data = [
        ("1", "Cơ bản / Học việc", "Đo cắt phôi sắt hộp thẳng (90°)", "Đo đạc và vận hành máy cắt bàn cắt phôi sắt hộp thẳng góc theo bản vẽ.", "Phôi đúng kích thước bản vẽ (sai số <= 1mm), mép cắt thẳng đứng.", "5 - 10 phút"),
        ("2", "Cơ bản / Học việc", "Cắt vát góc mỹ thuật (45°)", "Cắt chéo vát góc phôi sắt hộp mạ kẽm ghép góc khung bao.", "Cắt vát góc chính xác ghép khít góc vuông 90 độ, không hở khe.", "5 - 10 phút"),
        ("3", "Cơ bản / Phụ việc", "Mài phẳng mối hàn sắt", "Sử dụng máy mài góc cầm tay mài phẳng bavia mối hàn thô sắt đen.", "Mối hàn phẳng mịn, sờ tay không gợn, không mài lẹm mỏng phôi.", "5 - 10 phút"),
        ("4", "Thợ Inox chuyên biệt", "Mài đánh bóng inox bóng gương", "Mài xử lý bavia thô mối hàn và đánh bóng inox đạt bề mặt láng mịn gương.", "Bề mặt sáng bóng gương (mirror finish), không xước vân mài cũ.", "10 - 15 phút"),
        ("5", "Giúp việc / Thợ chính", "Gá ráp khung thép hộp ke góc", "Gá lắp, lấy ke vuông góc và hàn đính gá góc khung bao dầm chính.", "Góc ke vuông phẳng, khung phẳng đét không bị vênh chéo góc hình học.", "10 - 15 phút"),
        ("6", "Thợ hàn kết cấu dày", "Hàn hồ quang tay (Hàn que)", "Hàn kéo dải hàn que liên kết thép tấm dày kết cấu dầm chính chịu lực dày.", "Mối hàn ngấu sâu mép kim loại, vảy hàn đều, không nứt rỗ, xỉ dễ bong.", "10 - 15 phút"),
        ("7", "Thợ hàn sắt mỏng", "Hàn bán tự động MIG sắt hộp", "Hàn kéo liên kết sắt hộp mạ kẽm mỏng 1.2mm dầm phụ mái kính.", "Mối hàn ngấu đẹp chắc chắn, không bị quá nhiệt làm chảy thủng phôi.", "10 - 15 phút"),
        ("8", "Thợ hàn Inox mỏng", "Hàn TIG inox khí bảo vệ", "Hàn TIG liên kết inox tấm mỏng 0.8 - 1.2mm kết cấu hộp trang trí.", "Mối hàn sáng bóng màu vàng rơm hoặc trắng, vảy cá đều đẹp tăm tắp.", "10 - 15 phút"),
        ("9", "Hàn Công nghệ cao", "Hàn máy Laser chuyên dụng", "Sử dụng máy hàn laser để liên kết tấm phôi nhôm/inox mỏng.", "Mối hàn cực mịn bóng đẹp, không cong vênh vật liệu do nhiệt.", "5 - 10 phút"),
        ("10", "Thợ Sơn chuyên biệt", "Vận hành phun sơn khí nén", "Pha chế sơn theo tỷ lệ và sơn lót, sơn phủ màu hoàn thiện mối hàn dầm.", "Nước sơn mịn màu bóng láng, sơn phủ kín, không chảy giọt sơn đọng lại.", "5 - 10 phút"),
        ("11", "Thợ Lắp dựng công trình", "Khoan cấy bulong hóa chất", "Khoan bê tông dầm sàn cấy bulong hóa chất liên kết chân bản mã cột đỡ.", "Đường khoan chuẩn đứng, bulong cấy chắc nịch, hóa chất điền đầy.", "10 - 15 phút"),
        ("12", "Thợ chính / Tổ trưởng", "Đọc bản vẽ & Khai triển phôi", "Đọc bản vẽ thiết kế, tự tính toán chiều dài các thanh phôi tối ưu để cắt.", "Bóc tách phôi chuẩn xác từng mm, phương án cắt khoa học tiết kiệm thép.", "15 - 20 phút")
    ]

    current_row = 13
    eval_rows = []
    
    for row in trials_data:
        stt, group, name, desc, target, duration = row
        ws.row_dimensions[current_row].height = 42
        
        ws.cell(current_row, 1, int(stt)).alignment = align_center
        ws.cell(current_row, 1).font = font_normal
        
        ws.cell(current_row, 2, group).font = font_bold
        ws.cell(current_row, 2).alignment = align_left
        
        ws.cell(current_row, 3, name).font = font_bold
        ws.cell(current_row, 3).alignment = align_left
        
        ws.cell(current_row, 4, desc).font = font_normal
        ws.cell(current_row, 4).alignment = align_left
        
        ws.cell(current_row, 5, target).font = font_normal
        ws.cell(current_row, 5).alignment = align_left
        
        ws.cell(current_row, 6, duration).font = font_italic
        ws.cell(current_row, 6).alignment = align_center
        
        # Result (1/0 input)
        cell_res = ws.cell(current_row, 7, 0)
        cell_res.font = font_bold
        cell_res.alignment = align_center
        cell_res.fill = fill_input
        eval_rows.append(current_row)
        
        # Note
        cell_note = ws.cell(current_row, 8, "[Nhận xét nhanh]")
        cell_note.font = font_italic
        cell_note.alignment = align_left
        cell_note.fill = fill_input
        
        for col in range(1, 9):
            ws.cell(current_row, col).border = border_all_thin
            if current_row % 2 == 0:
                if col not in [7, 8]:
                    ws.cell(current_row, col).fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
                    
        current_row += 1

    # Insert sample scores for mock test (e.g. tested tasks 1, 2, 3, 5, 7)
    ws["G13"] = 1 # Task 1: Đạt
    ws["G14"] = 1 # Task 2: Đạt
    ws["G15"] = 1 # Task 3: Đạt
    ws["G16"] = "" # Task 4: Không chọn thi
    ws["G17"] = 1 # Task 5: Đạt
    ws["G18"] = "" # Task 6: Không chọn thi
    ws["G19"] = 1 # Task 7: Đạt
    ws["G20"] = ""
    ws["G21"] = ""
    ws["G22"] = ""
    ws["G23"] = ""
    ws["G24"] = ""

    # Spacing
    ws.row_dimensions[current_row].height = 15
    current_row += 1

    # SUMMARY DASHBOARD PANEL AT BOTTOM
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=8)
    r_hdr = ws.cell(current_row, 2, "BẢNG TỔNG HỢP ĐÁNH GIÁ KẾT QUẢ THI THỰC CHIẾN TẠI XƯỞNG")
    r_hdr.font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    r_hdr.fill = PatternFill(start_color=STEEL_BLUE, end_color=STEEL_BLUE, fill_type="solid")
    r_hdr.alignment = align_center
    ws.row_dimensions[current_row].height = 24
    
    for col in range(2, 9):
        ws.cell(current_row, col).border = Border(top=Side(style="medium", color=STEEL_BLUE))
        
    current_row += 1
    
    # 1. Số bài chọn thi thực tế
    ws.row_dimensions[current_row].height = 20
    ws.cell(current_row, 2, "Tổng số bài ứng viên lựa chọn thi thực tế:").font = font_bold
    ws.cell(current_row, 2).alignment = align_left
    
    cell_total_trials = ws.cell(current_row, 3, "=COUNT(G13:G24)")
    cell_total_trials.font = font_bold
    cell_total_trials.alignment = align_center
    ws.cell(current_row, 4, "Bài thi").font = font_italic
    ws.cell(current_row, 4).alignment = align_left
    
    for col in range(2, 9):
        cell = ws.cell(current_row, col)
        cell.fill = fill_result_box
        left_s = Side(style="medium", color=STEEL_BLUE) if col == 2 else None
        right_s = Side(style="medium", color=STEEL_BLUE) if col == 8 else None
        cell.border = Border(left=left_s, right=right_s)
        
    total_ref = f"C{current_row}" # C27
    current_row += 1
    
    # 2. Số bài thi Đạt
    ws.row_dimensions[current_row].height = 20
    ws.cell(current_row, 2, "Tổng số bài thi thực tế ĐẠT YÊU CẦU:").font = font_bold
    ws.cell(current_row, 2).alignment = align_left
    
    cell_passed = ws.cell(current_row, 3, "=COUNTIF(G13:G24, 1)")
    cell_passed.font = font_bold
    cell_passed.alignment = align_center
    ws.cell(current_row, 4, "Bài thi đạt").font = font_italic
    ws.cell(current_row, 4).alignment = align_left
    
    for col in range(2, 9):
        cell = ws.cell(current_row, col)
        cell.fill = fill_result_box
        left_s = Side(style="medium", color=STEEL_BLUE) if col == 2 else None
        right_s = Side(style="medium", color=STEEL_BLUE) if col == 8 else None
        cell.border = Border(left=left_s, right=right_s)
        
    passed_ref = f"C{current_row}" # C28
    current_row += 1
    
    # 3. Tỷ lệ đạt
    ws.row_dimensions[current_row].height = 20
    ws.cell(current_row, 2, "Tỷ lệ hoàn thành bài thi đạt yêu cầu:").font = font_bold
    ws.cell(current_row, 2).alignment = align_left
    
    cell_rate = ws.cell(current_row, 3, f"=IF({total_ref}>0, {passed_ref}/{total_ref}, 0)")
    cell_rate.font = font_bold
    cell_rate.alignment = align_center
    cell_rate.number_format = '0.0%'
    ws.cell(current_row, 4, "(Tỷ lệ đạt chuẩn)").font = font_italic
    ws.cell(current_row, 4).alignment = align_left
    
    for col in range(2, 9):
        cell = ws.cell(current_row, col)
        cell.fill = fill_result_box
        left_s = Side(style="medium", color=STEEL_BLUE) if col == 2 else None
        right_s = Side(style="medium", color=STEEL_BLUE) if col == 8 else None
        cell.border = Border(left=left_s, right=right_s)
        
    rate_ref = f"C{current_row}" # C29
    current_row += 1

    # 4. Đánh giá xếp loại tay nghề đề xuất
    ws.row_dimensions[current_row].height = 20
    ws.cell(current_row, 2, "Xếp loại trình độ tay nghề đề xuất:").font = font_bold
    ws.cell(current_row, 2).alignment = align_left
    
    cell_level = ws.cell(current_row, 3, f'=IF({rate_ref}>=0.9, "Đạt tay nghề xuất sắc (Thợ cứng)", IF({rate_ref}>=0.75, "Đạt chuẩn tay nghề khá (Thợ chính)", IF({rate_ref}>=0.5, "Đạt mức độ trung bình (Thợ phụ)", "Chưa đạt yêu cầu thử tay nghề xưởng")))')
    cell_level.font = font_result
    cell_level.alignment = align_center
    ws.cell(current_row, 4, "(Tự động chấm dựa trên tỷ lệ đạt)").font = font_italic
    ws.cell(current_row, 4).alignment = align_left
    
    for col in range(2, 9):
        cell = ws.cell(current_row, col)
        cell.fill = fill_result_box
        left_s = Side(style="medium", color=STEEL_BLUE) if col == 2 else None
        right_s = Side(style="medium", color=STEEL_BLUE) if col == 8 else None
        cell.border = Border(left=left_s, right=right_s, bottom=Side(style="medium", color=STEEL_BLUE))

    current_row += 2

    # Slogan block matching company motto
    ws.row_dimensions[current_row].height = 26
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
    slogan_cell = ws.cell(current_row, 1, 'Khẩu hiệu của chúng tôi: "Chất Lượng Tạo Nên Thương Hiệu. Uy Tín Tạo Nên Sự Thành Công"')
    slogan_cell.font = Font(name="Segoe UI", size=10, bold=True, italic=True, color=STEEL_BLUE)
    slogan_cell.alignment = align_center
    slogan_cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    for col in range(1, 9):
        ws.cell(current_row, col).border = Border(
            top=Side(style="thin", color=STEEL_BLUE),
            bottom=Side(style="thin", color=STEEL_BLUE),
            left=Side(style="thin", color=STEEL_BLUE) if col == 1 else None,
            right=Side(style="thin", color=STEEL_BLUE) if col == 8 else None
        )

    current_row += 3
    
    # Signatures
    ws.row_dimensions[current_row].height = 18
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    s1 = ws.cell(current_row, 1, "ỨNG VIÊN TUYỂN DỤNG")
    s1.font = font_bold
    s1.alignment = align_center
    
    ws.merge_cells(start_row=current_row, start_column=5, end_row=current_row, end_column=8)
    s2 = ws.cell(current_row, 5, "QUẢN ĐỐC / FOREMAN GIÁM KHẢO CHẤM")
    s2.font = font_bold
    s2.alignment = align_center
    
    current_row += 1
    ws.row_dimensions[current_row].height = 15
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    sub1 = ws.cell(current_row, 1, "(Ký và ghi rõ họ tên)")
    sub1.font = font_italic
    sub1.alignment = align_center
    
    ws.merge_cells(start_row=current_row, start_column=5, end_row=current_row, end_column=8)
    sub2 = ws.cell(current_row, 5, "(Xác nhận kết quả thi thực tế tại xưởng)")
    sub2.font = font_italic
    sub2.alignment = align_center

    # Add validations for Column G (Result)
    dv = DataValidation(
        type="whole", 
        operator="between", 
        formula1="0", 
        formula2="1", 
        allow_blank=True
    )
    dv.prompt = 'Vui lòng nhập 1: Đạt hoặc 0: Chưa đạt. Bỏ trống nếu không thi hạng mục này.'
    dv.promptTitle = 'Điểm thi thực tế'
    ws.add_data_validation(dv)
    
    for r in eval_rows:
        dv.add(ws[f"G{r}"])

    paths = [
        "d:/Sao Vàng/Website-SaoVang/CKSV_Danh_Sach_Bai_Thi_Thuc_Chien_Tho_Co_Khi.xlsx",
        "d:/Sao Vàng/Website-SaoVang/CKSV_Danh_Sach_Bai_Thi_Thuc_Chien_Tho_Co_Khi_Co_Logo.xlsx"
    ]
    saved_paths = []
    for path in paths:
        try:
            wb.save(path)
            saved_paths.append("Success")
        except PermissionError:
            print("Permission denied for path. Skipping.")
            
    print("Trial tasks workbook generated successfully.")

if __name__ == "__main__":
    create_trial_tasks_workbook()
