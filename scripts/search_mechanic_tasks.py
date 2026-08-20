import os
import openpyxl

def search_mechanic_info():
    folder_path = "d:/Sao Vàng/Website-SaoVang/CKSV"
    keywords = ["thợ", "cơ khí", "hàn", "ráp", "mài", "sơn", "thi công", "lắp dựng", "tổ đội"]
    results = []
    
    for root, dirs, files in os.walk(folder_path):
        for file in files:
            if file.endswith((".xlsx", ".xlsm")):
                path = os.path.join(root, file)
                try:
                    wb = openpyxl.load_workbook(path, data_only=True)
                    for sheetname in wb.sheetnames:
                        ws = wb[sheetname]
                        # Scan first 100 rows and 20 columns for keyword matches
                        matches = []
                        max_r = min(ws.max_row, 150)
                        max_c = min(ws.max_column, 25)
                        for r in range(1, max_r + 1):
                            row_vals = [str(ws.cell(r, c).value or '') for c in range(1, max_c + 1)]
                            row_str = " | ".join(row_vals).lower()
                            if any(kw in row_str for kw in keywords):
                                # Clean up row values for reporting
                                clean_vals = [v.strip() for v in row_vals if v.strip()]
                                if clean_vals:
                                    matches.append((r, clean_vals))
                        if matches:
                            results.append((path, sheetname, matches))
                except Exception as e:
                    results.append((path, f"ERROR: {repr(e)}", []))
                    
    # Write report
    report_path = "d:/Sao Vàng/Website-SaoVang/scripts/mechanic_search_results.txt"
    with open(report_path, "w", encoding="utf-8") as f:
        f.write("BÁO CÁO TÌM KIẾM THÔNG TIN MỤC VIỆC THỢ CƠ KHÍ\n")
        f.write("="*80 + "\n\n")
        for path, sheet, matches in results:
            f.write(f"FILE: {os.path.basename(path)} ({path})\n")
            f.write(f"SHEET: {sheet}\n")
            f.write(f"Số dòng khớp từ khóa: {len(matches)}\n")
            f.write("-" * 40 + "\n")
            for r, vals in matches[:15]: # Show first 15 matches of each sheet
                f.write(f"  Dòng {r:03d}: {vals[:5]} ...\n")
            if len(matches) > 15:
                f.write(f"  ... và {len(matches) - 15} dòng khác ...\n")
            f.write("\n" + "="*80 + "\n\n")
            
    print(f"Search complete. Results written to {report_path}")

if __name__ == "__main__":
    search_mechanic_info()
