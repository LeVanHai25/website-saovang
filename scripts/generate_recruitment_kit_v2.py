import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.drawing.image import Image as OpenpyxlImage

def create_complete_recruitment_kit():
    wb = openpyxl.Workbook()
    
    # -------------------------------------------------------------
    # STYLES CONFIGURATION
    # -------------------------------------------------------------
    STEEL_BLUE = "2F5597"
    ICE_BLUE = "D9E1F2"
    LIGHT_GRAY = "F2F2F2"
    BORDER_GRAY = "D9D9D9"
    YELLOW_INPUT = "FFF2CC"
    DARK_GREEN = "385723"
    LIGHT_GREEN = "E2EFDA"
    
    # Fonts
    font_company_name = Font(name="Segoe UI", size=12, bold=True, color=STEEL_BLUE)
    font_company_info = Font(name="Segoe UI", size=9, italic=True, color="595959")
    font_title_main = Font(name="Segoe UI", size=15, bold=True, color=STEEL_BLUE)
    font_title_sub = Font(name="Segoe UI", size=10, italic=True, color="7F7F7F")
    font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    font_group_header = Font(name="Segoe UI", size=11, bold=True, color=STEEL_BLUE)
    font_bold = Font(name="Segoe UI", size=10, bold=True)
    font_normal = Font(name="Segoe UI", size=10)
    font_italic = Font(name="Segoe UI", size=9, italic=True)
    font_result = Font(name="Segoe UI", size=11, bold=True, color=DARK_GREEN)
    font_salary_result = Font(name="Segoe UI", size=12, bold=True, color=DARK_GREEN)
    
    # Fills
    fill_header = PatternFill(start_color=STEEL_BLUE, end_color=STEEL_BLUE, fill_type="solid")
    fill_group = PatternFill(start_color=ICE_BLUE, end_color=ICE_BLUE, fill_type="solid")
    fill_input = PatternFill(start_color=YELLOW_INPUT, end_color=YELLOW_INPUT, fill_type="solid")
    fill_result_box = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
    fill_sum = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
    
    # Alignments
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    
    # Borders
    thin_side = Side(style="thin", color=BORDER_GRAY)
    border_all_thin = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    border_group = Border(
        left=Side(style="thin", color=STEEL_BLUE),
        right=Side(style="thin", color=STEEL_BLUE),
        top=Side(style="thin", color=STEEL_BLUE),
        bottom=Side(style="thin", color=STEEL_BLUE)
    )
    border_double_bottom = Border(
        left=thin_side, right=thin_side, 
        top=thin_side, 
        bottom=Side(style="double", color="000000")
    )
    def merge_and_style_header_ws1(ws, r_num, title):
        ws.merge_cells(start_row=r_num, start_column=1, end_row=r_num, end_column=7)
        ws.cell(r_num, 1, title).font = font_group_header
        ws.cell(r_num, 1).alignment = align_left
        for col in range(1, 8):
            cell = ws.cell(r_num, col)
            cell.fill = fill_group
            cell.border = border_group

    logo_path = "d:/Sao Vàng/Website-SaoVang/Logo/LogoCTY + SV Aluminium/LogoCTY.png"

    def apply_letterhead(ws, max_col_letter):
        # Merge cells for company info
        ws.row_dimensions[2].height = 20
        ws.row_dimensions[3].height = 15
        ws.row_dimensions[4].height = 15
        
        ws.merge_cells(f"C2:{max_col_letter}2")
        ws["C2"] = "CÔNG TY CỔ PHẦN SẢN XUẤT CƠ KHÍ SAO VÀNG (ASV)"
        ws["C2"].font = font_company_name
        ws["C2"].alignment = Alignment(horizontal="left", vertical="center")
        
        ws.merge_cells(f"C3:{max_col_letter}3")
        ws["C3"] = "Địa chỉ: Tầng 3, TT7-35 Khu đô thị Văn Phú, phường Kiến Hưng, TP Hà Nội, Việt Nam."
        ws["C3"].font = font_company_info
        ws["C3"].alignment = Alignment(horizontal="left", vertical="center")
        
        ws.merge_cells(f"C4:{max_col_letter}4")
        ws["C4"] = "Hotline: 0869 590 279  |  Email: cokhisaovangvn@gmail.com"
        ws["C4"].font = font_company_info
        ws["C4"].alignment = Alignment(horizontal="left", vertical="center")

        try:
            img = OpenpyxlImage(logo_path)
            img.width = 110
            img.height = 45
            ws.add_image(img, "A2")
        except Exception as e:
            print(f"Error adding logo: {e}")

    # -------------------------------------------------------------
    # SHEET 1: QUY TRÌNH PHỎNG VẤN (USER GUIDE)
    # -------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Quy Trình Hướng Dẫn"
    ws1.sheet_view.showGridLines = True
    
    ws1.column_dimensions['A'].width = 12
    ws1.column_dimensions['B'].width = 25
    ws1.column_dimensions['C'].width = 75
    
    apply_letterhead(ws1, "C")
    
    ws1.row_dimensions[6].height = 26
    ws1.merge_cells("A6:C6")
    ws1["A6"] = "QUY TRÌNH PHỎNG VẤN TUYỂN DỤNG & THỬ TAY NGHỀ THỢ CƠ KHÍ"
    ws1["A6"].font = font_title_main
    ws1["A6"].alignment = align_center

    ws1.row_dimensions[7].height = 18
    ws1.merge_cells("A7:C7")
    ws1["A7"] = "Hướng dẫn phối hợp đánh giá giữa Phòng Nhân Sự (Kế toán/HR) và Quản Đốc Xưởng"
    ws1["A7"].font = font_title_sub
    ws1["A7"].alignment = align_center
    
    ws1["A9"] = "Bước"
    ws1["B9"] = "Tên hoạt động"
    ws1["C9"] = "Nội dung chi tiết quy trình thực hiện"
    for col in ["A", "B", "C"]:
        ws1[f"{col}9"].font = font_header
        ws1[f"{col}9"].fill = fill_header
        ws1[f"{col}9"].alignment = align_center
        ws1[f"{col}9"].border = border_all_thin
    ws1.row_dimensions[9].height = 24

    guide_steps = [
        ("Bước 1", "Phát Phiếu Tự Khai", "HR phát Phiếu tự khai cho Ứng viên tự tích chọn các đầu việc biết làm tại Sheet '1. Phỏng Vấn Lý Thuyết' (Cột E). Việc này giúp ứng viên làm quen và sàng lọc nhanh hồ sơ."),
        ("Bước 2", "Phỏng Vấn Lý Thuyết", "Kế toán hoặc HR tiến hành hỏi ứng viên các câu hỏi gợi ý có sẵn ở Cột D tại Sheet '1. Phỏng Vấn Lý Thuyết'. Việc này giúp đánh giá tư duy kỹ thuật lý thuyết mà không cần người hỏi có chuyên môn sâu. Tích Đạt (1) hoặc Chưa (0) vào Cột F."),
        ("Bước 3", "Thử Tay Nghề Thực Chiến", "Quản đốc xưởng đưa ứng viên xuống thực tế xưởng làm thử 5 bài kiểm tra thực tế cấu kiện của công trình (Sheet '2. Bài Thi Thực Chiến'). Quản đốc ghi điểm Đạt (1) hoặc Chưa (0) trực tiếp vào máy tính."),
        ("Bước 4", "Xem Kết Quả Đề Xuất", "Kế toán hoặc HR kiểm tra kết quả tính tự động ở Sheet '3. Kết Quả & Đề Xuất Lương'. Excel tự động tra cứu cấp bậc, chức danh và mức lương đề xuất ngày/tháng để làm tờ trình Sếp duyệt."),
        ("Lưu ý cho Sếp", "Thay Đổi Cấu Hình Lương", "Sếp có toàn quyền thay đổi định mức lương ngày và chức danh ở Sheet 'Cấu Hình Lương & Cấp Bậc'. Bảng tính chính sẽ tự động cập nhật theo giá trị Sếp thay đổi mà không làm hỏng công thức.")
    ]
    
    for r_idx, (step, act, desc) in enumerate(guide_steps, start=10):
        ws1.cell(r_idx, 1, step).alignment = align_center
        ws1.cell(r_idx, 1).font = font_bold
        ws1.cell(r_idx, 1).border = border_all_thin
        
        ws1.cell(r_idx, 2, act).alignment = align_left
        ws1.cell(r_idx, 2).font = font_bold
        ws1.cell(r_idx, 2).border = border_all_thin
        
        ws1.cell(r_idx, 3, desc).alignment = align_left
        ws1.cell(r_idx, 3).font = font_normal
        ws1.cell(r_idx, 3).border = border_all_thin
        ws1.row_dimensions[r_idx].height = 42

    # -------------------------------------------------------------
    # SHEET 2: 1. PHỎNG VẤN LÝ THUYẾT (THE 27 TASKS)
    # -------------------------------------------------------------
    ws2 = wb.create_sheet(title="1. Phỏng Vấn Lý Thuyết")
    ws2.sheet_view.showGridLines = True
    
    ws2.column_dimensions['A'].width = 6   # STT
    ws2.column_dimensions['B'].width = 24  # Nhóm công việc
    ws2.column_dimensions['C'].width = 50  # Nhiệm vụ
    ws2.column_dimensions['D'].width = 54  # Câu hỏi phỏng vấn gợi ý
    ws2.column_dimensions['E'].width = 16  # Ứng viên tự khai
    ws2.column_dimensions['F'].width = 16  # Phỏng vấn đạt (1/0)
    ws2.column_dimensions['G'].width = 25  # Ghi chú chi tiết

    apply_letterhead(ws2, "G")

    ws2.row_dimensions[6].height = 24
    ws2.merge_cells("A6:G6")
    ws2["A6"] = "BẢNG PHỎNG VẤN KỸ NĂNG VÀ MỤC VIỆC CƠ BẢN"
    ws2["A6"].font = font_title_main
    ws2["A6"].alignment = align_center

    ws2.row_dimensions[7].height = 18
    ws2.merge_cells("A7:G7")
    ws2["A7"] = "Hệ thống 27 câu hỏi gợi ý giúp HR & Kế toán phỏng vấn nhanh trình độ thợ cơ khí"
    ws2["A7"].font = font_title_sub
    ws2["A7"].alignment = align_center

    # Header Row (Row 9)
    ws2.row_dimensions[9].height = 28
    headers_ws2 = [
        ("A9", "STT", align_center),
        ("B9", "NHÓM CÔNG VIỆC", align_left),
        ("C9", "NHIỆM VỤ CHI TIẾT (YÊU CẦU CỦA SẾP)", align_left),
        ("D9", "CÂU HỎI PHỎNG VẤN GỢI Ý DÀNH CHO HR/KẾ TOÁN", align_left),
        ("E9", "ỨNG VIÊN TỰ KHAI (Đạt: 1 | 0)", align_center),
        ("F9", "PHỎNG VẤN ĐẠT (Đạt: 1 | 0)", align_center),
        ("G9", "GHI CHÚ KỸ THUẬT / NHẬN XÉT NHANH", align_left)
    ]
    for cell_ref, text, align in headers_ws2:
        cell = ws2[cell_ref]
        cell.value = text
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align
        cell.border = border_all_thin

    # 27 Tasks with Interview Questions
    tasks_data = [
        ("G", "I. TÁC PHONG, AN TOÀN LAO ĐỘNG & 5S"),
        ("1", "Ý thức kỷ luật & Sức khỏe", "Đảm bảo sức khỏe dẻo dai làm việc xưởng/công trình; thái độ trung thực, nhanh nhẹn, phối hợp tốt với đồng nghiệp.", 
         "Em có bệnh mãn tính hay đau khớp gì không? Có ngại tiếng ồn, bụi bặm của xưởng cơ khí không?"),
        ("2", "Vệ sinh xưởng & 5S", "Chủ động dọn dẹp máy móc, quét dọn và vệ sinh sạch sẽ khu vực gia công lắp dựng sau ca làm hoặc khi đi công trình.", 
         "Cuối buổi làm em thường dọn dẹp khu vực hàn/cắt thế nào để đảm bảo an toàn phòng chống cháy nổ?"),
        ("3", "Tư duy ATLĐ xưởng", "Biết tư duy tình huống để có biện pháp nhắc nhở, bảo vệ an toàn cho bản thân, người xung quanh và tài sản thi công.", 
         "Nếu thấy đồng nghiệp mài thép mà không đeo kính bảo hộ, em sẽ xử lý như thế nào?"),
         
        ("G", "II. VẬN HÀNH DỤNG CỤ CẦM TAY & KỸ THUẬT MÀI"),
        ("4", "Sử dụng dụng cụ cầm tay", "Sử dụng các loại máy cơ bản trong xưởng: máy cắt tay, máy cắt bàn, máy khoan, thước đo khoảng cách, dây điện... an toàn.", 
         "Em hãy chỉ và cho biết cách sử dụng máy cắt tay, máy khoan cầm tay và cách cắm dây nguồn thế nào là an toàn?"),
        ("5", "Mài hoàn thiện sản phẩm", "Sử dụng máy mài để mài hoàn thiện sản phẩm sắt đen, sắt mạ kẽm, inox; mài bề mặt, mài góc, đánh bóng, đánh rỉ sét.", 
         "Làm thế nào để mài sạch bavia mối hàn thô của inox mà không làm trầy xước bề mặt xung quanh?"),
        ("6", "Mài tinh mỹ phẳng góc", "Mài tinh mỹ phẳng tuyệt đối hoặc mài lấy góc sắc nét đạt thẩm mỹ cao (yêu cầu chi tiết được gá hàn đạt chuẩn trước).", 
         "Để mài lấy một mặt phẳng tuyệt đối sau khi hàn đính, em sử dụng kỹ thuật mài như thế nào và sờ tay kiểm tra thế nào?"),
        ("7", "Mài định hình tạo góc", "Mài lấy góc yêu cầu cho các sản phẩm được tổ hợp từ nguyên liệu tấm hoặc thép hình theo thiết kế bản vẽ.", 
         "Khi mài góc của sản phẩm tổ hợp từ thép tấm, làm thế nào để đảm bảo góc sắc nét, không bị tròn đầu?"),
        ("8", "Mài phá kết cấu nặng", "Mài phá bavia mối hàn thô trên các bề mặt vật liệu thô hoặc các cấu kiện chịu lực nặng.", 
         "Khi mài phá mối hàn của kết cấu chịu lực nặng, em cần chú ý điều gì để không làm giảm khả năng chịu lực của kết cấu?"),
        ("9", "Mài đánh bóng inox", "Đánh bóng cơ học cho các sản phẩm sử dụng nguyên liệu inox bóng hoặc cần làm mịn, làm bóng bề mặt kim loại khác.", 
         "Em sử dụng các bước nhám xếp và nỉ đánh bóng như thế nào để xử lý bề mặt inox đạt độ bóng gương?"),
        ("10", "Mài phục hồi mũi khoan", "Biết tinh chỉnh lại các vật dụng phục vụ cho công việc, tự mài sắc phục hồi các mũi khoan thép bị cùn.", 
         "Góc mài chuẩn của mũi khoan kim loại thường là bao nhiêu độ? Em mài thử một mũi khoan xem nào?"),
         
        ("G", "III. KỸ THUẬT PHUN SƠN HOÀN THIỆN"),
        ("11", "Vận hành súng phun sơn", "Sử dụng thành thạo súng phun sơn khí nén, sơn chống rỉ và sơn phủ hoàn thiện màu bóng mịn đều tay.", 
         "Khi phun sơn bị hiện tượng chảy sệt giọt sơn xuống dưới, nguyên nhân là do đâu và cách khắc phục thế nào?"),
        ("12", "Nhận biết chủng loại vật tư", "Nắm rõ cơ bản các loại vật tư, chủng loại vật liệu sơn phủ chống rỉ và dung môi pha chế.", 
         "Làm thế nào để phân biệt nhanh giữa sắt đen, sắt mạ kẽm và inox 304 khi nhìn bằng mắt thường hoặc dùng nam châm?"),
         
        ("G", "IV. KỸ THUẬT CẮT PHÔI & HÀN ĐÍNH GÁ RÁP"),
        ("13", "Sử dụng máy công trình", "Biết sử dụng cơ bản các loại máy móc và phụ kiện đi kèm trong nghề cơ khí dân dụng, xây dựng hiện trường.", 
         "Em đã từng vận hành máy khoan bê tông rút lõi hoặc máy khoan từ chưa? Sử dụng chúng cần lưu ý ATLĐ gì?"),
        ("14", "Đo cắt phôi làm dưỡng", "Nhận biết đơn vị đo lường cơ khí, cắt phôi chính xác; biết cách làm dưỡng chuẩn để cắt hàng loạt phôi thép.", 
         "Nếu cần cắt 50 đoạn sắt hộp cùng kích thước 450mm, em sẽ cắt từng thanh hay làm thế nào để nhanh và chuẩn nhất?"),
        ("15", "Hàn đính gá kết cấu", "Có thể sử dụng máy hàn để hàn đính gá chuẩn vuông góc các vị trí cần thiết khi được giao phôi lắp ráp.", 
         "Mục đích của việc hàn đính gá ráp trước khi hàn chết là gì? Khoảng cách giữa các mối đính thế nào là vừa?"),
        ("16", "Chuẩn bị thiết bị công trình", "Hiểu rõ công năng máy; chủ động chuẩn bị đầy đủ dụng cụ, vật tư phụ đi công trình khi được giao hạng mục thi công.", 
         "Khi chuẩn bị đi lắp đặt kết cấu ngoài công trình, em cần kiểm tra và mang theo những loại máy móc, bảo hộ gì?"),
        ("17", "Hàn hồ quang tay (Que)", "Sử dụng thành thạo máy hàn que (MMA) trên các kết cấu sắt thép dày chịu lực thông thường.", 
         "Khi hàn que bị dính que hàn vào phôi, em xử lý thế nào? Làm thế nào để điều chỉnh dòng điện hàn phù hợp với độ dày thép?"),
        ("18", "Hàn bán tự động MIG", "Sử dụng thành thạo máy hàn bán tự động MIG/MAG trên kết cấu thép hộp, thép hình phổ thông.", 
         "Hàn MIG dùng khí bảo vệ gì? Nếu mối hàn MIG bị rỗ khí hoặc sủi bọt, nguyên nhân thường do đâu?"),
        ("19", "Hàn khí bảo vệ TIG", "Sử dụng thành thạo máy hàn TIG để hàn inox mỏng hoặc liên kết yêu cầu mối hàn trắng sáng đều vảy cá.", 
         "Khi hàn TIG inox mỏng làm thế nào để không bị thủng phôi và mối hàn có màu vàng rơm hoặc trắng sáng?"),
        ("20", "Sử dụng thiết bị chuyên dụng", "Vận hành thành thạo máy cân bằng laser để lấy cốt lắp dựng và các thước đo khoảng cách điện tử chuyên dụng.", 
         "Em hãy chỉ cách sử dụng máy cân bằng laser để lấy cốt cao độ khi lắp dựng cột?"),
         
        ("G", "V. ĐỌC BẢN VẼ, TIẾN ĐỘ & QUẢN LÝ ĐỘI NHÓM"),
        ("22", "Đọc bản vẽ sản xuất", "Đọc hiểu bản vẽ thiết kế chế tạo; biết cách tính toán khai triển phôi thép tấm/hình một cách khoa học, hiệu quả.", 
         "Đưa bản vẽ dầm mái Monaco: Em hãy bóc tách chiều dài cắt phôi của dầm chính và dầm phụ trên bản vẽ này?"),
        ("23", "Kỹ năng hoàn thiện & QC", "Kỹ năng tự kiểm tra chất lượng sản phẩm (Self-QC) sau gia công; tinh thần khách quan, chủ động khắc phục lỗi kỹ thuật.", 
         "Nếu phát hiện sản phẩm mình vừa hoàn thiện bị sai lệch kích thước 5mm so với bản vẽ, em xử lý thế nào?"),
        ("24", "Ước lượng định mức công", "Khả năng đánh giá bản vẽ; ước lượng chuẩn định mức công thợ và thời gian hoàn thành dầm kết cấu hoặc công trình mái.", 
         "Để ráp và hàn hoàn thiện 1 bộ dầm đỡ kính mái Monaco này, em ước lượng cần bao nhiêu công thợ và làm trong bao lâu?"),
        ("25", "Báo cáo & Đề xuất tối ưu", "Lập báo cáo tiến độ cuối ngày; đề xuất các phương án thi công tối ưu tại chỗ cho bộ phận khác phối hợp xử lý nhanh.", 
         "Khi gặp vướng mắc kỹ thuật tại công trình mà không có thiết kế ở đó, em báo cáo và đề xuất phương án giải quyết thế nào?"),
        ("26", "Độc lập thi công hiện trường", "Khả năng độc lập thi công; tự đưa ra quyết định điều chỉnh kỹ thuật nhỏ tại hiện trường không làm ảnh hưởng kết cấu chung.", 
         "Khi được giao độc lập thi công một hạng mục tại công trường, em tự quyết định thay đổi chi tiết nhỏ thế nào để không sai kết cấu?"),
        ("27", "Kỹ năng làm việc nhóm", "Khả năng làm việc nhóm nhịp nhàng; phối hợp đẩy nhanh tiến độ gia công sản xuất đúng bản vẽ thiết kế.", 
         "Em làm thế nào để phối hợp ăn ý với thợ phụ mài và cắt phôi để đẩy nhanh tiến độ ráp khung của mình?"),
        ("28", "Lãnh đạo tổ đội thi công", "Năng lực quản lý và lãnh đạo tổ đội 3-5 thợ phụ/chính; bao quát tiến độ và chịu trách nhiệm chất lượng của tổ.", 
         "Nếu được giao làm tổ trưởng nhóm 3 người lắp dựng mái kính, em sẽ phân chia công việc thế nào dựa trên tay nghề của từng người?")
    ]

    current_row = 10
    eval_rows = []
    row_bac_map = {}
    
    for row_data in tasks_data:
        if row_data[0] == "G":
            ws2.row_dimensions[current_row].height = 24
            merge_and_style_header_ws1(ws2, current_row, row_data[1])
        else:
            stt, task, standard, question = row_data
            ws2.row_dimensions[current_row].height = 42
            
            ws2.cell(current_row, 1, stt).font = font_normal
            ws2.cell(current_row, 1).alignment = align_center
            
            # Group Column
            g_title = ""
            if int(stt) <= 3:
                g_title = "I. An toàn & 5S"
            elif int(stt) <= 10:
                g_title = "II. Vận hành & Mài"
            elif int(stt) <= 12:
                g_title = "III. Kỹ thuật Sơn"
            elif int(stt) <= 20:
                g_title = "IV. Cắt & Hàn gá"
            else:
                g_title = "V. Bản vẽ & Đội nhóm"
            ws2.cell(current_row, 2, g_title).font = font_bold
            ws2.cell(current_row, 2).alignment = align_center
            
            ws2.cell(current_row, 3, task).font = font_bold
            ws2.cell(current_row, 3).alignment = align_left
            
            ws2.cell(current_row, 4, question).font = font_italic
            ws2.cell(current_row, 4).alignment = align_left
            
            # Candidate self eval (1/0)
            cell_self = ws2.cell(current_row, 5, 1)
            cell_self.font = font_bold
            cell_self.alignment = align_center
            cell_self.fill = fill_input
            
            # Interviewer evaluation (1/0)
            cell_comp = ws2.cell(current_row, 6, 0)
            cell_comp.font = font_bold
            cell_comp.alignment = align_center
            cell_comp.fill = fill_input
            eval_rows.append(current_row)
            
            # Note
            cell_note = ws2.cell(current_row, 7, "[Nhận xét câu hỏi]")
            cell_note.font = font_italic
            cell_note.alignment = align_left
            cell_note.fill = fill_input
            
            for col in range(1, 8):
                ws2.cell(current_row, col).border = border_all_thin
                
            w_map = {
                11: 1, 12: 2, 13: 3,          # I
                15: 4, 16: 5, 17: 6, 18: 7, 19: 8, 20: 9, 21: 10, # II
                23: 11, 24: 12,               # III
                26: 13, 27: 14, 28: 15, 29: 16, 30: 17, 31: 18, 32: 19, 33: 20, # IV
                35: 21, 36: 22, 37: 23, 38: 24, 39: 25, 40: 26, 41: 27  # V (27 tasks total)
            }
            # Note: tasks are mapped to levels for auto evaluation
            # If they pass the task, they accumulate level capability
            
        current_row += 1

    # Place sample test scores (probation pass up to Bậc 7 welding)
    ws2["E11"] = 1; ws2["F11"] = 1
    ws2["E12"] = 1; ws2["F12"] = 1
    ws2["E13"] = 1; ws2["F13"] = 1
    ws2["E15"] = 1; ws2["F15"] = 1
    ws2["E16"] = 1; ws2["F16"] = 1
    ws2["E17"] = 1; ws2["F17"] = 1
    ws2["E18"] = 1; ws2["F18"] = 1
    ws2["E19"] = 1; ws2["F19"] = 1
    ws2["E20"] = 1; ws2["F20"] = 1
    ws2["E21"] = 1; ws2["F21"] = 1
    ws2["E23"] = 1; ws2["F23"] = 1
    ws2["E24"] = 1; ws2["F24"] = 1
    ws2["E26"] = 1; ws2["F26"] = 1
    ws2["E27"] = 1; ws2["F27"] = 1
    ws2["E28"] = 1; ws2["F28"] = 1
    ws2["E29"] = 1; ws2["F29"] = 1
    ws2["E30"] = 1; ws2["F30"] = 1
    ws2["E31"] = 1; ws2["F31"] = 1  # Passed MIG/TIG up to Bậc 7 (equivalent to task 19)
    ws2["E32"] = 0; ws2["F32"] = 0
    ws2["E33"] = 0; ws2["F33"] = 0
    ws2["E35"] = 0; ws2["F35"] = 0
    ws2["E36"] = 0; ws2["F36"] = 0
    ws2["E37"] = 0; ws2["F37"] = 0
    ws2["E38"] = 0; ws2["F38"] = 0
    ws2["E39"] = 0; ws2["F39"] = 0
    ws2["E40"] = 0; ws2["F40"] = 0
    ws2["E41"] = 0; ws2["F41"] = 0

    # Add validations for Column E & F in Sheet 2
    dv_comp = DataValidation(type="whole", operator="between", formula1="0", formula2="1", allow_blank=True)
    dv_comp.prompt = 'Nhập 1: Đạt, 0: Chưa đạt'
    dv_comp.promptTitle = 'Đánh giá'
    ws2.add_data_validation(dv_comp)
    for r in eval_rows:
        dv_comp.add(ws2[f"E{r}"])
        dv_comp.add(ws2[f"F{r}"])

    # -------------------------------------------------------------
    # SHEET 3: 2. BÀI THI THỰC CHIẾN (MONACO PROJECT BASE)
    # -------------------------------------------------------------
    ws3 = wb.create_sheet(title="2. Bài Thi Thực Chiến")
    ws3.sheet_view.showGridLines = True
    
    ws3.column_dimensions['A'].width = 6
    ws3.column_dimensions['B'].width = 25
    ws3.column_dimensions['C'].width = 30
    ws3.column_dimensions['D'].width = 48
    ws3.column_dimensions['E'].width = 45
    ws3.column_dimensions['F'].width = 16
    ws3.column_dimensions['G'].width = 25

    apply_letterhead(ws3, "G")

    ws3.row_dimensions[6].height = 24
    ws3.merge_cells("A6:G6")
    ws3["A6"] = "BÀI THI THỬ TAY NGHỀ THỰC CHIẾN TẠI XƯỞNG"
    ws3["A6"].font = font_title_main
    ws3["A6"].alignment = align_center

    ws3.row_dimensions[7].height = 18
    ws3.merge_cells("A7:G7")
    ws3["A7"] = "Đánh giá trực tiếp khi cho ứng viên làm thử sản phẩm thực tế của dự án Mái kính Monaco"
    ws3["A7"].font = font_title_sub
    ws3["A7"].alignment = align_center

    # Header Row (Row 9)
    ws3.row_dimensions[9].height = 28
    headers_ws3 = [
        ("A9", "STT", align_center),
        ("B9", "BÀI THI THỰC HÀNH", align_left),
        ("C9", "CẤU KIỆN MẪU DỰ ÁN MONACO", align_left),
        ("D9", "MÔ TẢ CHI TIẾT ĐẦU VIỆC YÊU CẦU", align_left),
        ("E9", "TIÊU CHUẨN ĐÁNH GIÁ (BẮT BUỘC ĐẠT)", align_left),
        ("F9", "KẾT QUẢ (Đạt: 1 | 0)", align_center),
        ("G9", "NHẬN XÉT CỦA FOREMAN CHẤM THI", align_left)
    ]
    for cell_ref, text, align in headers_ws3:
        cell = ws3[cell_ref]
        cell.value = text
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align
        cell.border = border_all_thin

    practical_trials = [
        ("T1", "Đo cắt phôi dầm", "Dầm chính thép U100x200x100x6mm", 
         "Đo đạc và vận hành máy cắt đĩa để cắt 1 đoạn dầm chính dài đúng 1200mm theo bản vẽ.", 
         "Đo thước mét kiểm tra lại: Sai số chiều dài không vượt quá ±1mm; góc cắt ke vuông phẳng tắp."),
        
        ("T2", "Mài phẳng & vệ sinh dầm", "Dầm chịu lực chính mái kính", 
         "Mài vát mép (beveling) đầu dầm thép U và mài sạch lớp mạ kẽm bavia tại khu vực chuẩn bị liên kết hàn.", 
         "Góc vát đều từ 30-45 độ, bề mặt mài sạch hết lớp màng oxit sáng bóng để đảm bảo mối hàn ngấu ngấu lực."),
        
        ("T3", "Ráp kết cấu cột", "Cột nâng đỡ kết cấu phụ đỡ kính", 
         "Gá lắp và hàn đính (bon) bản mã thép tấm SS400 240x300x8mm vào thân cột thép hộp 200x100x3mm ke góc vuông.", 
         "Kiểm tra ke góc 90 độ khít sát cả 4 cạnh; cột phẳng tắp không bị vênh, xoắn vặn góc hình học."),
        
        ("T4", "Hàn dầm liên kết", "Mối nối liên kết dầm U chịu lực", 
         "Hàn kéo chịu lực ngấu đều nối 2 đầu dầm chính thép U đúc dày 6mm bằng máy hàn que hoặc máy hàn MIG.", 
         "Mối hàn ngấu sâu, không thủng thép dầm, vảy hàn đều và chắc chắn, gõ nhẹ xỉ hàn phải tự bong sạch."),
        
        ("T5", "Sơn lót hoàn thiện", "Cụm cấu kiện dầm cột mẫu Monaco", 
         "Làm sạch bề mặt mối hàn, dùng súng phun sơn khí nén phun sơn lót chống gỉ đều lên cấu kiện mẫu vừa gia công.", 
         "Nước sơn láng mịn, phủ kín mối hàn và góc ke cột; không có lỗi chảy sệ sơn thành dòng hoặc nổi bong bóng.")
    ]

    ws3_eval_rows = []
    for idx, row in enumerate(practical_trials, start=10):
        ws3.row_dimensions[idx].height = 42
        ws3.cell(idx, 1, row[0]).alignment = align_center
        ws3.cell(idx, 2, row[1]).font = font_bold
        ws3.cell(idx, 3, row[2]).font = font_bold
        ws3.cell(idx, 4, row[3]).alignment = align_left
        ws3.cell(idx, 5, row[4]).alignment = align_left
        
        # Result (1/0)
        cell_res = ws3.cell(idx, 6, 1)
        cell_res.font = font_bold
        cell_res.alignment = align_center
        cell_res.fill = fill_input
        ws3_eval_rows.append(idx)
        
        # Notes
        cell_note = ws3.cell(idx, 7, "[Nhận xét thực hành]")
        cell_note.font = font_italic
        cell_note.alignment = align_left
        cell_note.fill = fill_input
        
        for col in range(1, 8):
            ws3.cell(idx, col).border = border_all_thin
            if idx % 2 == 0:
                ws3.cell(idx, col).fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid") if col != 6 and col != 7 else fill_input

    # Sample results (passed T1-T4, failed T5 painting?)
    ws3["F10"] = 1
    ws3["F11"] = 1
    ws3["F12"] = 1
    ws3["F13"] = 1
    ws3["F14"] = 0  # Painting failed in trial

    for r in ws3_eval_rows:
        dv_comp.add(ws3[f"F{r}"])

    # -------------------------------------------------------------
    # SHEET 4: 3. KẾT QUẢ & ĐỀ XUẤT LƯƠNG (DASHBOARD)
    # -------------------------------------------------------------
    ws4 = wb.create_sheet(title="3. Kết Quả & Đề Xuất Lương")
    ws4.sheet_view.showGridLines = True
    
    ws4.column_dimensions['A'].width = 6
    ws4.column_dimensions['B'].width = 38
    ws4.column_dimensions['C'].width = 24
    ws4.column_dimensions['D'].width = 18
    ws4.column_dimensions['E'].width = 24
    ws4.column_dimensions['F'].width = 25

    apply_letterhead(ws4, "F")

    ws4.row_dimensions[6].height = 24
    ws4.merge_cells("A6:F6")
    ws4["A6"] = "BÁO CÁO KẾT QUẢ ĐÁNH GIÁ VÀ ĐỀ XUẤT TUYỂN DỤNG"
    ws4["A6"].font = font_title_main
    ws4["A6"].alignment = align_center

    ws4.row_dimensions[7].height = 18
    ws4.merge_cells("A7:F7")
    ws4["A7"] = "Kết quả tổng hợp tự động năng lực ứng viên và đề xuất mức lương ngày thử việc"
    ws4["A7"].font = font_title_sub
    ws4["A7"].alignment = align_center

    # Candidate info block
    ws4["B9"] = "Họ và tên ứng viên:"
    ws4["B9"].font = font_bold
    ws4["C9"] = "='1. Phỏng Vấn Lý Thuyết'!C9"
    ws4["C9"].font = font_bold
    ws4["C9"].border = Border(bottom=Side(style="thin", color="000000"))

    ws4["E9"] = "Ngày phỏng vấn:"
    ws4["E9"].font = font_bold
    ws4["F9"] = "='1. Phỏng Vấn Lý Thuyết'!F9"
    ws4["F9"].font = font_normal
    ws4["F9"].border = Border(bottom=Side(style="thin", color="000000"))

    ws4["B10"] = "Số điện thoại/CCCD:"
    ws4["B10"].font = font_bold
    ws4["C10"] = "='1. Phỏng Vấn Lý Thuyết'!C10"
    ws4["C10"].font = font_normal
    ws4["C10"].border = Border(bottom=Side(style="thin", color="000000"))

    ws4["E10"] = "Người phỏng vấn:"
    ws4["E10"].font = font_bold
    ws4["F10"] = "='1. Phỏng Vấn Lý Thuyết'!F10"
    ws4["F10"].font = font_normal
    ws4["F10"].border = Border(bottom=Side(style="thin", color="000000"))

    ws4.row_dimensions[11].height = 15

    # Summary Panel
    ws4.merge_cells("B12:F12")
    ws4["B12"] = "KẾT QUẢ ĐÁNH GIÁ NĂNG LỰC TOÀN DIỆN"
    ws4["B12"].font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    ws4["B12"].fill = PatternFill(start_color=STEEL_BLUE, end_color=STEEL_BLUE, fill_type="solid")
    ws4["B12"].alignment = align_center
    ws4.row_dimensions[12].height = 24
    
    for col in range(2, 7):
        ws4.cell(12, col).border = Border(top=Side(style="medium", color=STEEL_BLUE))

    # 1. Điểm lý thuyết
    ws4.row_dimensions[13].height = 20
    ws4.cell(13, 2, "Hạng mục phỏng vấn lý thuyết ĐẠT:").font = font_bold
    ws4.cell(13, 2).alignment = align_left
    
    # Range of tasks rows in Sheet 2: 11,12,13,15,16,17,18,19,20,21,23,24,26,27,28,29,30,31,32,33,35,36,37,38,39,40,41
    sh2_task_rows = [11,12,13,15,16,17,18,19,20,21,23,24,26,27,28,29,30,31,32,33,35,36,37,38,39,40,41]
    sum_sh2_f = "+".join(f"'1. Phỏng Vấn Lý Thuyết'!F{r}" for r in sh2_task_rows)
    cell_p1 = ws4.cell(13, 3, f"={sum_sh2_f}")
    cell_p1.font = font_bold
    cell_p1.alignment = align_center
    ws4.cell(13, 4, f"/ {len(sh2_task_rows)} mục việc").font = font_italic
    ws4.cell(13, 4).alignment = align_left
    
    # 2. Điểm thực chiến
    ws4.row_dimensions[14].height = 20
    ws4.cell(14, 2, "Bài thi Thực chiến tại xưởng ĐẠT:").font = font_bold
    ws4.cell(14, 2).alignment = align_left
    
    cell_p2 = ws4.cell(14, 3, "=SUM('2. Bài Thi Thực Chiến'!F10:F14)")
    cell_p2.font = font_bold
    cell_p2.alignment = align_center
    ws4.cell(14, 4, "/ 5 bài thi").font = font_italic
    ws4.cell(14, 4).alignment = align_left

    # 3. Tỷ lệ thành thạo quy đổi
    ws4.row_dimensions[15].height = 20
    ws4.cell(15, 2, "Tỷ lệ thành thạo kỹ năng quy đổi:").font = font_bold
    ws4.cell(15, 2).alignment = align_left
    
    # Combined rate = (passed theory + passed practical) / (total theory + total practical)
    cell_p3 = ws4.cell(15, 3, f"=(C13+C14)/({len(sh2_task_rows)}+5)")
    cell_p3.font = font_result
    cell_p3.alignment = align_center
    cell_p3.number_format = '0.0%'
    ws4.cell(15, 4, "(Tỷ lệ năng lực thực tế)").font = font_italic
    ws4.cell(15, 4).alignment = align_left

    # 4. Phân bậc tự động
    ws4.row_dimensions[16].height = 20
    ws4.cell(16, 2, "Bậc năng lực đề xuất:").font = font_bold
    ws4.cell(16, 2).alignment = align_left
    
    # Map the percentage to Level 1 to 14
    cell_p4 = ws4.cell(16, 3, "=IF(C15>=0.95, 14, IF(C15>=0.88, 12, IF(C15>=0.8, 10, IF(C15>=0.72, 8, IF(C15>=0.64, 7, IF(C15>=0.55, 5, IF(C15>=0.45, 4, IF(C15>=0.3, 3, IF(C15>=0.15, 2, 1)))))))))")
    cell_p4.font = font_bold
    cell_p4.alignment = align_center
    ws4.cell(16, 4, "Bậc / 14").font = font_italic
    ws4.cell(16, 4).alignment = align_left

    # Style borders for evaluation box
    for r in range(13, 17):
        for col in range(2, 7):
            cell = ws4.cell(r, col)
            cell.fill = fill_result_box
            left_s = Side(style="medium", color=STEEL_BLUE) if col == 2 else None
            right_s = Side(style="medium", color=STEEL_BLUE) if col == 6 else None
            cell.border = Border(left=left_s, right=right_s)

    # -------------------------------------------------------------
    # SALARY RECOMMENDATION SUB-BOX
    # -------------------------------------------------------------
    # Spacing and border bottom
    for col in range(2, 7):
        ws4.cell(16, col).border = Border(left=Side(style="medium", color=STEEL_BLUE) if col==2 else None,
                                           right=Side(style="medium", color=STEEL_BLUE) if col==6 else None,
                                           bottom=Side(style="thin", color=STEEL_BLUE))

    bac_ref = "C16"

    # 5. Chức danh đề xuất
    ws4.row_dimensions[17].height = 20
    ws4.cell(17, 2, "Chức danh tuyển dụng đề xuất:").font = font_bold
    ws4.cell(17, 2).alignment = align_left
    
    cell_title = ws4.cell(17, 3, f"=VLOOKUP({bac_ref}, 'Cấu Hình Lương & Cấp Bậc'!A$2:E$15, 2, FALSE)")
    cell_title.font = font_bold
    cell_title.alignment = align_center
    
    # 6. Mức lương ngày thực lĩnh
    ws4.row_dimensions[18].height = 20
    ws4.cell(18, 2, "Lương thực lĩnh ngày đề xuất:").font = font_bold
    ws4.cell(18, 2).alignment = align_left
    
    cell_wage = ws4.cell(18, 3, f"=VLOOKUP({bac_ref}, 'Cấu Hình Lương & Cấp Bậc'!A$2:E$15, 3, FALSE)")
    cell_wage.font = font_salary_result
    cell_wage.alignment = align_center
    cell_wage.number_format = '#,##0" đ"'
    ws4.cell(18, 4, "(Lương sau thuế)").font = font_italic
    ws4.cell(18, 4).alignment = align_left

    # 7. Tổng chi phí ngày công
    ws4.row_dimensions[19].height = 20
    ws4.cell(19, 2, "Tổng chi phí ngày công (ước tính):").font = font_bold
    ws4.cell(19, 2).alignment = align_left
    
    cell_cost = ws4.cell(19, 3, f"=VLOOKUP({bac_ref}, 'Cấu Hình Lương & Cấp Bậc'!A$2:E$15, 4, FALSE)")
    cell_cost.font = font_bold
    cell_cost.alignment = align_center
    cell_cost.number_format = '#,##0" đ"'
    ws4.cell(19, 4, "(Đã gồm bảo hiểm/phụ cấp)").font = font_italic
    ws4.cell(19, 4).alignment = align_left

    # 8. Dự toán lương tháng 26 công
    ws4.row_dimensions[20].height = 20
    ws4.cell(20, 2, "Dự toán thu nhập tháng (26 công):").font = font_bold
    ws4.cell(20, 2).alignment = align_left
    
    cell_month = ws4.cell(20, 3, "=C18*26")
    cell_month.font = font_salary_result
    cell_month.alignment = align_center
    cell_month.number_format = '#,##0" đ"'
    ws4.cell(20, 4, "(Lương thử việc đề xuất)").font = font_italic
    ws4.cell(20, 4).alignment = align_left

    # 9. Khuyến nghị sử dụng
    ws4.row_dimensions[21].height = 24
    ws4.cell(21, 2, "Khuyến nghị phân công:").font = font_bold
    ws4.cell(21, 2).alignment = align_left
    
    ws4.merge_cells("C21:F21")
    cell_rec = ws4.cell(21, 3, f"=VLOOKUP({bac_ref}, 'Cấu Hình Lương & Cấp Bậc'!A$2:E$15, 5, FALSE)")
    cell_rec.font = font_bold
    cell_rec.alignment = align_left

    # Style borders for salary recommendation box
    for r in range(17, 22):
        for col in range(2, 7):
            cell = ws4.cell(r, col)
            cell.fill = fill_result_box
            left_s = Side(style="medium", color=STEEL_BLUE) if col == 2 else None
            right_s = Side(style="medium", color=STEEL_BLUE) if col == 6 else None
            bottom_s = Side(style="medium", color=STEEL_BLUE) if r == 21 else None
            cell.border = Border(left=left_s, right=right_s, bottom=bottom_s)

    current_row = 23
    ws4.row_dimensions[current_row].height = 15
    current_row += 1

    # Slogan block matching company motto
    ws4.row_dimensions[current_row].height = 26
    ws4.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    slogan_cell = ws4.cell(current_row, 1, 'Khẩu hiệu của chúng tôi: "Chất Lượng Tạo Nên Thương Hiệu. Uy Tín Tạo Nên Sự Thành Công"')
    slogan_cell.font = Font(name="Segoe UI", size=10, bold=True, italic=True, color=STEEL_BLUE)
    slogan_cell.alignment = align_center
    slogan_cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    for col in range(1, 7):
        ws4.cell(current_row, col).border = Border(
            top=Side(style="thin", color=STEEL_BLUE),
            bottom=Side(style="thin", color=STEEL_BLUE),
            left=Side(style="thin", color=STEEL_BLUE) if col == 1 else None,
            right=Side(style="thin", color=STEEL_BLUE) if col == 6 else None
        )

    current_row += 3
    
    # Signatures
    ws4.row_dimensions[current_row].height = 18
    ws4.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    s1 = ws4.cell(current_row, 1, "ỨNG VIÊN TUYỂN DỤNG")
    s1.font = font_bold
    s1.alignment = align_center
    
    ws4.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=6)
    s2 = ws4.cell(current_row, 4, "QUẢN ĐỐC / BAN GIÁM ĐỐC DUYỆT")
    s2.font = font_bold
    s2.alignment = align_center
    
    current_row += 1
    ws4.row_dimensions[current_row].height = 15
    ws4.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    sub1 = ws4.cell(current_row, 1, "(Ký và ghi rõ họ tên)")
    sub1.font = font_italic
    sub1.alignment = align_center
    
    ws4.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=6)
    sub2 = ws4.cell(current_row, 4, "(Ký nhận phê duyệt ngạch lương ngày)")
    sub2.font = font_italic
    sub2.alignment = align_center
    
    current_row += 4
    ws4.row_dimensions[current_row].height = 18
    ws4.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    n1 = ws4.cell(current_row, 1, "='1. Phỏng Vấn Lý Thuyết'!C9")
    n1.font = font_normal
    n1.alignment = align_center
    
    ws4.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=6)
    n2 = ws4.cell(current_row, 4, "[Họ tên Ban Giám Đốc]")
    n2.font = font_normal
    n2.alignment = align_center

    # -------------------------------------------------------------
    # SHEET 5: CẤU HÌNH LƯƠNG & CẤP BẬC (SẾP TỰ ĐIỀU CHỈNH)
    # -------------------------------------------------------------
    ws5 = wb.create_sheet(title="Cấu Hình Lương & Cấp Bậc")
    ws5.sheet_view.showGridLines = True
    
    ws5.column_dimensions['A'].width = 10
    ws5.column_dimensions['B'].width = 24
    ws5.column_dimensions['C'].width = 24
    ws5.column_dimensions['D'].width = 24
    ws5.column_dimensions['E'].width = 65
    
    # Headers
    ws5["A1"] = "Bậc"
    ws5["B1"] = "Chức danh đề xuất"
    ws5["C1"] = "Lương thực lĩnh (ngày)"
    ws5["D1"] = "Tổng chi phí công ty (ngày)"
    ws5["E1"] = "Phân công / Vai trò chính trong xưởng"
    
    for col in ["A", "B", "C", "D", "E"]:
        cell = ws5[f"{col}1"]
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_all_thin
        
    scale_data = [
        (1, "Thợ học việc (Cấp 1)", 220000, 300000, "Phụ việc vệ sinh, mài thô, nâng hạ phôi dưới sự chỉ dẫn"),
        (2, "Thợ học việc (Cấp 2)", 220000, 300000, "Thực hiện tốt 5S khu vực hàn gá, đảm bảo an toàn ATLĐ xưởng"),
        (3, "Thợ phụ cơ bản (Cấp 1)", 255000, 335000, "Biết dùng máy cắt cầm tay, khoan tay, thước đo chính xác"),
        (4, "Thợ phụ mài tinh (Cấp 2)", 290000, 370000, "Mài tinh mỹ bề mặt, mài lấy góc sắc nét, mài phá bavia"),
        (5, "Thợ sơn & gá phụ", 350000, 430000, "Đứng buồng sơn phun khí nén; nắm vật tư và hỗ trợ hàn đính gá"),
        (6, "Thợ gá lắp phôi", 380000, 460000, "Nắm rõ các loại vật tư, phôi cắt sắt/inox, chuẩn bị công trình"),
        (7, "Thợ gá hàn liên hợp", 405000, 485000, "Sử dụng thành thạo máy hàn que/MIG/TIG để đính gá kết cấu"),
        (8, "Thợ hàn chính chuyên nghiệp", 430000, 510000, "Đọc bản vẽ sản xuất và tự hàn độc lập sản phẩm chất lượng cao"),
        (9, "Thợ lắp ráp lành nghề", 440000, 520000, "Đọc bản vẽ cơ khí phức tạp, tự bóc tách khai triển phôi gá ráp"),
        (10, "Thợ hoàn thiện độc lập", 470000, 550000, "Tự kiểm soát chất lượng hoàn thiện sản phẩm không cần giám sát"),
        (11, "Thợ dự toán & tiến độ", 480000, 560000, "Ước lượng thời gian/nhân công, lập tiến độ thi công dự án nhỏ"),
        (12, "Tổ phó kỹ thuật", 490000, 570000, "Tự chủ động điều chỉnh chi tiết nhỏ ở hiện trường, thay thế tổ trưởng"),
        (13, "Trưởng nhóm lắp dựng", 500000, 580000, "Điều phối nhóm 2-3 người lắp dựng kết cấu thép/inox ngoài công trường"),
        (14, "Tổ trưởng sản xuất", 550000, 630000, "Phân việc toàn xưởng, chịu trách nhiệm tiến độ & chất lượng tổ đội")
    ]
    
    for r_idx, row_val in enumerate(scale_data, start=2):
        ws5.cell(r_idx, 1, row_val[0]).alignment = align_center
        ws5.cell(r_idx, 1).font = font_bold
        ws5.cell(r_idx, 1).border = border_all_thin
        
        ws5.cell(r_idx, 2, row_val[1]).alignment = align_left
        ws5.cell(r_idx, 2).font = font_bold
        ws5.cell(r_idx, 2).border = border_all_thin
        
        ws5.cell(r_idx, 3, row_val[2]).number_format = '#,##0" đ"'
        ws5.cell(r_idx, 3).alignment = align_right
        ws5.cell(r_idx, 3).font = font_bold
        ws5.cell(r_idx, 3).border = border_all_thin
        
        ws5.cell(r_idx, 4, row_val[3]).number_format = '#,##0" đ"'
        ws5.cell(r_idx, 4).alignment = align_right
        ws5.cell(r_idx, 4).font = font_normal
        ws5.cell(r_idx, 4).border = border_all_thin
        
        ws5.cell(r_idx, 5, row_val[4]).alignment = align_left
        ws5.cell(r_idx, 5).font = font_normal
        ws5.cell(r_idx, 5).border = border_all_thin
        
        for col in range(1, 6):
            cell = ws5.cell(r_idx, col)
            if r_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

    # -------------------------------------------------------------
    # SAVE & EXPORT
    # -------------------------------------------------------------
    paths = [
        "d:/Sao Vàng/Website-SaoVang/CKSV_Bo_Cong_Cu_Tuyen_Dung_Tho_Co_Khi.xlsx",
        "d:/Sao Vàng/Website-SaoVang/CKSV_Bo_Cong_Cu_Tuyen_Dung_Tho_Co_Khi_Co_Logo.xlsx"
    ]
    saved_paths = []
    for path in paths:
        try:
            wb.save(path)
            saved_paths.append(path)
        except PermissionError:
            print(f"Permission denied for {path}. Skipping.")
            
    print(f"Unified Recruitment Kit created successfully. Saved to: {saved_paths}")

if __name__ == "__main__":
    create_complete_recruitment_kit()
