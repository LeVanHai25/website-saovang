import openpyxl

def inspect_file():
    path = "d:/Sao Vàng/Website-SaoVang/MẪU FOM TUYỂN DỤNG 2024.xlsx"
    wb = openpyxl.load_workbook(path, data_only=True)
    out_path = "d:/Sao Vàng/Website-SaoVang/scripts/inspect_recruitment_2024.txt"
    
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(f"Sheets in workbook: {wb.sheetnames}\n")
        
        for sheet_name in wb.sheetnames:
            f.write(f"\n--- Sheet: {sheet_name} ---\n")
            ws = wb[sheet_name]
            for r in range(1, 100):
                row_vals = [ws.cell(r, c).value for c in range(1, 15)]
                if any(v is not None for v in row_vals):
                    f.write(f"Row {r:02d}: {row_vals}\n")
    print("Done inspecting, written to inspect_recruitment_2024.txt")

if __name__ == "__main__":
    inspect_file()
