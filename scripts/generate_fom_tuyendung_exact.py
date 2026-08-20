import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

def create_exact_recruitment_form():
    wb = openpyxl.Workbook()
    
    # -------------------------------------------------------------
    # SHEET 1: ĐÁNH GIÁ TUYỂN DỤNG (EXACT COPY OF SHEET1 DATA)
    # -------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Đánh Giá Tuyển Dụng"
    ws1.sheet_view.showGridLines = True
    
    # Theme: Steel Blue (Cơ khí lắp ráp chuyên nghiệp)
    STEEL_BLUE = "2F5597"
    ICE_BLUE = "D9E1F2"
    LIGHT_GRAY = "F2F2F2"
    BORDER_GRAY = "D9D9D9"
    YELLOW_INPUT = "FFF2CC"
    DARK_GREEN = "385723"
    LIGHT_GREEN = "E2EFDA"
    
    # Fonts
    font_title_company = Font(name="Segoe UI", size=11, bold=True, color="595959")
    font_title_main = Font(name="Segoe UI", size=16, bold=True, color=STEEL_BLUE)
    font_title_sub = Font(name="Segoe UI", size=10, italic=True, color="7F7F7F")
    font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    font_group_header = Font(name="Segoe UI", size=11, bold=True, color=STEEL_BLUE)
    font_bold = Font(name="Segoe UI", size=10, bold=True)
    font_normal = Font(name="Segoe UI", size=10)
    font_italic = Font(name="Segoe UI", size=9, italic=True)
    font_salary_result = Font(name="Segoe UI", size=12, bold=True, color=DARK_GREEN)
    
    # Fills
    fill_header = PatternFill(start_color=STEEL_BLUE, end_color=STEEL_BLUE, fill_type="solid")
    fill_group = PatternFill(start_color=ICE_BLUE, end_color=ICE_BLUE, fill_type="solid")
    fill_sum = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
    fill_input = PatternFill(start_color=YELLOW_INPUT, end_color=YELLOW_INPUT, fill_type="solid")
    fill_result_box = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
    
    # Alignments
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    
    # Borders
    thin_side = Side(style="thin", color=BORDER_GRAY)
    border_all_thin = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    border_group = Border(
        left=Side(style="thin", color=STEEL_BLUE),
        right=Side(style="thin", color=STEEL_BLUE),
        top=Side(style="thin", color=STEEL_BLUE),
        bottom=Side(style="thin", color=STEEL_BLUE)
    )
    
    # Column widths
    ws1.column_dimensions['A'].width = 6   # STT
    ws1.column_dimensions['B'].width = 38  # NHIỆM VỤ (VIỆC CẦN LÀM)
    ws1.column_dimensions['C'].width = 58  # KẾT QUẢ BÊN SDLĐ (CẦN ĐẠT ĐƯỢC)
    ws1.column_dimensions['D'].width = 16  # ỨNG VIÊN TỰ ĐÁNH GIÁ (% Đáp ứng)
    ws1.column_dimensions['E'].width = 16  # CÔNG TY ĐÁNH GIÁ (Đạt: 1, Chưa: 0)
    ws1.column_dimensions['F'].width = 12  # CUNG BẬC LƯƠNG
    ws1.column_dimensions['G'].width = 16  # LƯƠNG THỰC LĨNH
    ws1.column_dimensions['H'].width = 18  # TỔNG LƯƠNG / NGÀY CÔNG
    ws1.column_dimensions['I'].width = 25  # GHI CHÚ / NHẬN XÉT

    # Title Block
    ws1.merge_cells("A2:I2")
    ws1["A2"] = "CÔNG TY CỔ PHẦN SẢN XUẤT CƠ KHÍ SAO VÀNG (ASV)"
    ws1["A2"].font = font_title_company
    ws1["A2"].alignment = Alignment(horizontal="left")

    ws1.merge_cells("A3:I3")
    ws1["A3"] = "BẢNG TUYỂN DỤNG THỢ PHỤ - THỢ CHÍNH NGÀNH CƠ KHÍ DÂN DỤNG"
    ws1["A3"].font = font_title_main
    ws1["A3"].alignment = align_center

    ws1.merge_cells("A4:I4")
    ws1["A4"] = "Đánh giá chi tiết năng lực thử việc & xác định cung bậc lương thực tế"
    ws1["A4"].font = font_title_sub
    ws1["A4"].alignment = align_center

    # Candidate Profile Info
    ws1["B6"] = "Họ và tên ứng viên:"
    ws1["B6"].font = font_bold
    ws1["C6"] = "[Nhập tên ứng viên]"
    ws1["C6"].font = font_normal
    ws1["C6"].border = Border(bottom=Side(style="thin", color="000000"))

    ws1["E6"] = "Ngày phỏng vấn:"
    ws1["E6"].font = font_bold
    ws1["F6"] = "[Nhập ngày]"
    ws1["F6"].font = font_normal
    ws1["F6"].border = Border(bottom=Side(style="thin", color="000000"))

    ws1["H6"] = "Người chấm điểm:"
    ws1["H6"].font = font_bold
    ws1["I6"] = "[Tên Quản đốc/Tổ trưởng]"
    ws1["I6"].font = font_normal
    ws1["I6"].border = Border(bottom=Side(style="thin", color="000000"))

    ws1["B7"] = "Số điện thoại/CCCD:"
    ws1["B7"].font = font_bold
    ws1["C7"] = "[Số điện thoại và CCCD]"
    ws1["C7"].font = font_normal
    ws1["C7"].border = Border(bottom=Side(style="thin", color="000000"))

    ws1["E7"] = "Vị trí ứng tuyển:"
    ws1["E7"].font = font_bold
    ws1["F7"] = "Thợ Cơ khí Dân dụng"
    ws1["F7"].font = font_italic
    ws1["F7"].border = Border(bottom=Side(style="thin", color="000000"))

    ws1["H7"] = "Kỳ thử việc:"
    ws1["H7"].font = font_bold
    ws1["I7"] = "01 tháng thử việc"
    ws1["I7"].font = font_normal
    ws1["I7"].border = Border(bottom=Side(style="thin", color="000000"))

    ws1.row_dimensions[8].height = 15

    # Table Header
    headers = [
        ("A9", "STT", align_center),
        ("B9", "NHIỆM VỤ (VIỆC CẦN LÀM)", align_left),
        ("C9", "KẾT QUẢ BÊN SDLĐ (CẦN ĐẠT ĐƯỢC)", align_left),
        ("D9", "ỨNG VIÊN TỰ ĐÁNH GIÁ (% Đáp ứng)", align_center),
        ("E9", "CÔNG TY ĐÁNH GIÁ (Đạt: 1, Chưa: 0)", align_center),
        ("F9", "CUNG BẬC LƯƠNG", align_center),
        ("G9", "LƯƠNG THỰC LĨNH", align_center),
        ("H9", "TỔNG LƯƠNG / NGÀY CÔNG", align_center),
        ("I9", "GHI CHÚ / NHẬN XÉT CHI TIẾT", align_left)
    ]
    for cell_ref, text, align in headers:
        cell = ws1[cell_ref]
        cell.value = text
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align
        cell.border = border_all_thin
    ws1.row_dimensions[9].height = 28

    # Exact Tasks from CKSV file (rows 7 to 33 in Sheet1)
    exact_tasks = [
        # (STT, Nhiệm vụ, Kết quả, Cung bậc, Lương lĩnh, Tổng lương)
        ("1", "_Sức khỏe tốt, nhanh nhẹn, chịu khó học hỏi, có trách nhiệm, phối hợp được với các vị trí trong xưởng", "", 1, 220000, 300000),
        ("", "Chủ động thu dọn máy móc dụng cụ làm việc, quét dọn vệ sinh khu vực thi công sau mỗi buổi làm và khi kết thúc công việc ở xưởng cũng như đi công trường", "", 2, 220000, 300000),
        ("3", "_Biết sử dụng các loại máy móc và dụng cụ cơ bản trong xưởng sản xuất: máy cắt tay, máy cắt bàn, máy khoan, thước đo khoảng cách, dây điện…vv", "", 3, 250000, 330000),
        ("", "Biết tư duy tình huống để có biện pháp nhắc nhở, bảo vệ an toàn cho bản thân và những người xung quanh, cũng như tài sản trong khu vực thi công", "", 3, 250000, 330000),
        ("", "Có thể sử dụng máy mài để mài hoàn thiện sản phẩm theo yêu cầu về độ tinh mỹ của sản phẩm sắt đen, sắt mạ, inox và vệ sinh bề mặt sản phẩm sau khi hoàn thành phần thô (mài bề mặt, mài góc, đánh bóng, đánh rỉ...)", "", 3, 260000, 340000),
        ("4", "Mài tinh mỹ (yêu cầu cho sp có độ thẩm mỹ cao, cần mài lấy mặt phẳng tuyệt đối hoặc mài lấy góc sắc nét nếu chi tiết được gá, hàn đạt tiêu chuẩn)", "", 4, 270000, 350000),
        ("", "Mài lấy góc (yêu cầu cho các sản phẩm được tổ hợp từ nguyên liệu tấm, hình vv, tổ hợp thành sản phẩm theo thiết kế)", "", 4, 280000, 360000),
        ("", "Mài phá (yêu cầu cho các vị trí mối hàn trên các bề mặt vật liệu thô, kết cấu hạng nặng, hoặc các mối hàn không đạt tiêu chuẩn về thẩm mỹ, cần tinh chỉnh lại)", "", 4, 290000, 370000),
        ("", "Mài đánh bóng (yêu cầu cho các sản phẩm sử dụng nguyên liệu inox bóng, hoặc cần làm bóng, mịn cho sp ngoài inox)", "", 4, 300000, 380000),
        ("", "Biết tinh chỉnh lại các vật dụng phục vụ cho cv (mài mũi khoan)", "", 4, 320000, 400000),
        ("", "Sử dụng thành thạo súng phun sơn khí nén, sơn phủ và sơn hoàn thiện sản phẩm sau quá trình hàn, mài tinh mỹ", "", 5, 330000, 410000),
        ("5", "Nắm rõ cơ bản các loại vật tư, chủng loại vật liệu", "", 5, 340000, 420000),
        ("", "Sử dụng được các loại máy cơ bản trong công việc: máy hàn, máy cắt tay, máy cắt bàn, máy khoan, máy cân bằng laze, thước đo khoảng cách...vv", "", 5, 350000, 430000),
        ("", "Nhận biết được đơn vị đo lường, có thể cắt phôi chính xác theo kích thước được giao, biết cách tính toán làm dưỡng chuẩn để cắt hàng loạt nếu vật liệu cần cắt cùng một kích thước và với số lượng trên 2", "", 5, 360000, 440000),
        ("", "Có thể sử dụng máy hàn để hàn đính gá chuẩn các vị trí cần thiết, khi được giao việc", "", 5, 370000, 450000),
        ("6", "Nắm bắt và hiểu rõ hết công năng cho từng loại máy móc, chủ động tính toán chuẩn bị đầy đủ vt phụ, máy móc dùng cho công trình khi đã được thông báo về nội dung các hạng mục công việc của công trình sắp triển khai", "", 6, 380000, 460000),
        ("7", "Sử dụng thành thạo máy hàn que", "", 7, 390000, 470000),
        ("", "Sử dụng thành thạo máy hàn mic", "", 7, 400000, 480000),
        ("", "Sử dụng thành thạo máy hàn tic", "", 7, 410000, 490000),
        ("", "Biết sử dụng cơ bản các loại máy và phụ kiện đi kèm trong nghề cơ khí dân dụng, xây dựng", "", 7, 420000, 500000),
        ("8", "Đọc hiểu bản vẽ sản xuất và có thể triển khai công việc trên bản vẽ khi được giao, biết cách tính toán khai triển công việc từ bản vẽ thành sp thực tế một cách khoa học, nhanh và hiệu quả nhất, nghiên cứu, đóng góp ý kiến kịp thời nếu thấy có phương án thi công hiệu quả, năng suất hơn", "", 8, 430000, 510000),
        ("25", "Kỹ năng hoàn thiện và đánh giá sản phẩm sau khi hoàn thiện trên tinh thần khách quan, không giấu dốt, luôn luôn nâng cấp, hoàn thiện bản thân trong công việc", "", 9, 440000, 520000),
        ("9", "Khả năng đánh giá, ước lượng số nhân công và khoảng thời gian hoàn thành sản phẩm hoặc công trình, Chịu trách nhiệm cho những phần việc được giao khi đã được hướng dẫn cụ thể", "", 10, 470000, 550000),
        ("", "Báo cáo tình hình tiến độ công việc vào cuối ngày, và trao đổi ngay với những phần việc phát sinh hoặc đề xuất các phương án thi công tối ưu để các bộ phận khác nắm rõ thông tin", "", 11, 480000, 560000),
        ("10", "Có thể độc lập trong công việc khi giao, có thể tự quyết định điều chỉnh cách thức làm hoặc chủng loại vật tư mà không ảnh hưởng đến cấu trúc tổng thể của sản phẩm, đồng thời tăng tính thẩm mỹ, hiệu quả", "", 12, 490000, 570000),
        ("11", "Khả năng làm việc nhóm, đảm bảo sản xuất các sản phẩm đúng thiết kế, đúng tiến độ", "", 13, 500000, 580000),
        ("12", "Khả năng lãnh đạo đội nhóm, bao quát, điều phối, bố trí công việc phù hợp cho từng thành viên trong nhóm khi được giao nhiệm vụ lãnh đạo tổ đội thi công các hạng mục công việc đã trao đổi", "", 14, 550000, 630000)
    ]

    current_row = 11
    eval_rows = []
    
    for row_data in exact_tasks:
        stt, task, r_result, b_num, wage_l, wage_t = row_data
        
        ws1.row_dimensions[current_row].height = 24
        
        ws1.cell(current_row, 1, stt).font = font_normal
        ws1.cell(current_row, 1).alignment = align_center
        
        ws1.cell(current_row, 2, task).font = font_normal
        ws1.cell(current_row, 2).alignment = align_left
        
        ws1.cell(current_row, 3, r_result).font = font_normal
        ws1.cell(current_row, 3).alignment = align_left
        
        # Candidate self eval (%) - Input
        cell_self = ws1.cell(current_row, 4, 1.0)
        cell_self.font = font_bold
        cell_self.alignment = align_center
        cell_self.fill = fill_input
        cell_self.number_format = '0%'
        
        # Company eval (1: Đạt, 0: Chưa) - Input
        cell_company = ws1.cell(current_row, 5, 0)
        cell_company.font = font_bold
        cell_company.alignment = align_center
        cell_company.fill = fill_input
        eval_rows.append(current_row)
        
        # Cung bậc lương
        cell_lvl = ws1.cell(current_row, 6, b_num)
        cell_lvl.font = font_bold
        cell_lvl.alignment = align_center
        
        # Lương thực lĩnh (ngày)
        cell_wl = ws1.cell(current_row, 7, wage_l)
        cell_wl.font = font_bold
        cell_wl.alignment = align_right
        cell_wl.number_format = '#,##0" đ"'
        
        # Tổng lương / ngày công
        cell_wt = ws1.cell(current_row, 8, wage_t)
        cell_wt.font = font_normal
        cell_wt.alignment = align_right
        cell_wt.number_format = '#,##0" đ"'
        
        # Ghi chú
        cell_note = ws1.cell(current_row, 9, "[Ghi chú đánh giá]")
        cell_note.font = font_normal
        cell_note.alignment = align_left
        cell_note.fill = fill_input
        
        # Apply borders to all columns
        for col in range(1, 10):
            ws1.cell(current_row, col).border = border_all_thin
            
        current_row += 1

    # Add sample evaluations to match real-world test
    ws1["E11"] = 1  # Đạt Bậc 1
    ws1["E12"] = 1  # Đạt Bậc 2
    ws1["E13"] = 1  # Đạt Bậc 3
    ws1["E14"] = 1  
    ws1["E15"] = 1  
    ws1["E16"] = 1  # Đạt Bậc 4
    ws1["E17"] = 1  
    ws1["E18"] = 1  
    ws1["E19"] = 1  
    ws1["E20"] = 1  
    ws1["E21"] = 1  # Đạt Bậc 5
    ws1["E22"] = 1  # Đạt Bậc 6
    ws1["E23"] = 1  # Đạt Bậc 7
    ws1["E24"] = 1  
    ws1["E25"] = 0  # Chưa đạt tiếp
    
    ws1.row_dimensions[current_row].height = 15
    current_row += 1

    # RESULTS DASHBOARD BOX
    ws1.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=9)
    r_hdr = ws1.cell(current_row, 2, "KẾT QUẢ ĐÁNH GIÁ NĂNG LỰC & XÁC ĐỊNH BẬC LƯƠNG SAU THỬ VIỆC")
    r_hdr.font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    r_hdr.fill = PatternFill(start_color=STEEL_BLUE, end_color=STEEL_BLUE, fill_type="solid")
    r_hdr.alignment = align_center
    ws1.row_dimensions[current_row].height = 24
    
    for col in range(2, 10):
        ws1.cell(current_row, col).border = Border(top=Side(style="medium", color=STEEL_BLUE))
        
    current_row += 1
    
    # Formulas for wage extraction
    # Wage is determined as the MAX of Column G (Lương thực lĩnh) where Column E is 1.
    # Formula format: `=MAX(IF(E11=1,G11,0), IF(E12=1,G12,0), ...)`
    wage_parts_l = []
    wage_parts_t = []
    bac_parts = []
    
    for r_num in eval_rows:
        wage_parts_l.append(f"IF(E{r_num}=1,G{r_num},0)")
        wage_parts_t.append(f"IF(E{r_num}=1,H{r_num},0)")
        bac_parts.append(f"IF(E{r_num}=1,F{r_num},0)")
        
    formula_wage_l = "=MAX(" + ",".join(wage_parts_l) + ")"
    formula_wage_t = "=MAX(" + ",".join(wage_parts_t) + ")"
    formula_bac = "=MAX(" + ",".join(bac_parts) + ")"
    
    # 1. Bậc năng lực đạt được
    ws1.row_dimensions[current_row].height = 20
    ws1.cell(current_row, 2, "Bậc năng lực cao nhất đạt được:").font = font_bold
    ws1.cell(current_row, 2).alignment = align_left
    
    cell_bac = ws1.cell(current_row, 3, formula_bac)
    cell_bac.font = font_bold
    cell_bac.alignment = align_center
    
    ws1.cell(current_row, 4, "Cung bậc / 14").font = font_italic
    ws1.cell(current_row, 4).alignment = align_left
    
    for col in range(2, 10):
        cell = ws1.cell(current_row, col)
        cell.fill = fill_result_box
        left_s = Side(style="medium", color=STEEL_BLUE) if col == 2 else None
        right_s = Side(style="medium", color=STEEL_BLUE) if col == 9 else None
        cell.border = Border(left=left_s, right=right_s)
        
    bac_cell_ref = f"C{current_row}"
    current_row += 1
    
    # 2. Chức danh tuyển dụng đề xuất
    ws1.row_dimensions[current_row].height = 20
    ws1.cell(current_row, 2, "Chức danh tuyển dụng đề xuất:").font = font_bold
    ws1.cell(current_row, 2).alignment = align_left
    
    cell_title = ws1.cell(current_row, 3, f"=IF({bac_cell_ref}=0, \"Chưa đạt tiêu chuẩn\", VLOOKUP({bac_cell_ref}, 'Thang Đo & Lương'!A$2:E$15, 2, FALSE))")
    cell_title.font = font_bold
    cell_title.alignment = align_center
    
    for col in range(2, 10):
        cell = ws1.cell(current_row, col)
        cell.fill = fill_result_box
        left_s = Side(style="medium", color=STEEL_BLUE) if col == 2 else None
        right_s = Side(style="medium", color=STEEL_BLUE) if col == 9 else None
        cell.border = Border(left=left_s, right=right_s)
        
    current_row += 1
    
    # 3. Lương thực lĩnh ngày công
    ws1.row_dimensions[current_row].height = 20
    ws1.cell(current_row, 2, "Mức lương thực lĩnh đề xuất (ngày):").font = font_bold
    ws1.cell(current_row, 2).alignment = align_left
    
    cell_wage_l = ws1.cell(current_row, 3, formula_wage_l)
    cell_wage_l.font = font_salary_result
    cell_wage_l.alignment = align_center
    cell_wage_l.number_format = '#,##0" đ"'
    
    ws1.cell(current_row, 4, "(Tiền thực lĩnh về tay)").font = font_italic
    ws1.cell(current_row, 4).alignment = align_left
    
    for col in range(2, 10):
        cell = ws1.cell(current_row, col)
        cell.fill = fill_result_box
        left_s = Side(style="medium", color=STEEL_BLUE) if col == 2 else None
        right_s = Side(style="medium", color=STEEL_BLUE) if col == 9 else None
        cell.border = Border(left=left_s, right=right_s)
        
    wage_l_ref = f"C{current_row}"
    current_row += 1
    
    # 4. Tổng lương ngày công (Chi phí công ty)
    ws1.row_dimensions[current_row].height = 20
    ws1.cell(current_row, 2, "Tổng chi phí công ty chi trả (ngày):").font = font_bold
    ws1.cell(current_row, 2).alignment = align_left
    
    cell_wage_t = ws1.cell(current_row, 3, formula_wage_t)
    cell_wage_t.font = font_bold
    cell_wage_t.alignment = align_center
    cell_wage_t.number_format = '#,##0" đ"'
    
    ws1.cell(current_row, 4, "(Lương gốc + Phụ cấp ngày công)").font = font_italic
    ws1.cell(current_row, 4).alignment = align_left
    
    for col in range(2, 10):
        cell = ws1.cell(current_row, col)
        cell.fill = fill_result_box
        left_s = Side(style="medium", color=STEEL_BLUE) if col == 2 else None
        right_s = Side(style="medium", color=STEEL_BLUE) if col == 9 else None
        cell.border = Border(left=left_s, right=right_s)
        
    current_row += 1
    
    # 5. Dự toán thu nhập tháng (26 ngày công)
    ws1.row_dimensions[current_row].height = 20
    ws1.cell(current_row, 2, "Dự toán thu nhập thực lĩnh (26 công):").font = font_bold
    ws1.cell(current_row, 2).alignment = align_left
    
    cell_wage_m = ws1.cell(current_row, 3, f"={wage_l_ref}*26")
    cell_wage_m.font = font_salary_result
    cell_wage_m.alignment = align_center
    cell_wage_m.number_format = '#,##0" đ"'
    
    ws1.cell(current_row, 4, "(Dải lương theo poster 15 - 22 triệu)").font = font_italic
    ws1.cell(current_row, 4).alignment = align_left
    
    for col in range(2, 10):
        cell = ws1.cell(current_row, col)
        cell.fill = fill_result_box
        left_s = Side(style="medium", color=STEEL_BLUE) if col == 2 else None
        right_s = Side(style="medium", color=STEEL_BLUE) if col == 9 else None
        cell.border = Border(left=left_s, right=right_s)
        
    current_row += 1
    
    # 6. Khuyến nghị và Kế hoạch sử dụng nhân sự
    ws1.row_dimensions[current_row].height = 24
    ws1.cell(current_row, 2, "Kế hoạch sử dụng nhân sự:").font = font_bold
    ws1.cell(current_row, 2).alignment = align_left
    
    ws1.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=9)
    rec_cell = ws1.cell(current_row, 3, f"=IF({bac_cell_ref}=0, \"Yêu cầu đào tạo thêm tay nghề hoặc từ chối thử việc\", VLOOKUP({bac_cell_ref}, 'Thang Đo & Lương'!A$2:E$15, 5, FALSE))")
    rec_cell.font = font_bold
    rec_cell.alignment = align_left
    
    for col in range(2, 10):
        cell = ws1.cell(current_row, col)
        cell.fill = fill_result_box
        left_s = Side(style="medium", color=STEEL_BLUE) if col == 2 else None
        right_s = Side(style="medium", color=STEEL_BLUE) if col == 9 else None
        cell.border = Border(left=left_s, right=right_s, bottom=Side(style="medium", color=STEEL_BLUE))

    # Data Validations
    dv_self = DataValidation(type="decimal", operator="between", formula1="0", formula2="1", allow_blank=True)
    dv_self.error = 'Tỷ lệ đáp ứng phải từ 0% đến 100%!'
    dv_self.errorTitle = 'Lỗi nhập liệu'
    dv_self.prompt = 'Nhập % tự đánh giá (Ví dụ: 80%)'
    dv_self.promptTitle = 'Ứng viên tự đánh giá'
    ws1.add_data_validation(dv_self)
    
    dv_company = DataValidation(type="whole", operator="between", formula1="0", formula2="1", allow_blank=True)
    dv_company.error = 'Chỉ được nhập 1 (Đạt tiêu chuẩn) hoặc 0 (Chưa đạt)!'
    dv_company.errorTitle = 'Lỗi chấm điểm'
    dv_company.prompt = 'Nhập 1: Đạt, 0: Chưa đạt'
    dv_company.promptTitle = 'Quản đốc đánh giá'
    ws1.add_data_validation(dv_company)

    for r_num in eval_rows:
        dv_self.add(ws1[f"D{r_num}"])
        dv_company.add(ws1[f"E{r_num}"])

    current_row += 3
    
    # Signatures
    ws1.row_dimensions[current_row].height = 18
    ws1.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    s1 = ws1.cell(current_row, 1, "ỨNG VIÊN TUYỂN DỤNG")
    s1.font = font_bold
    s1.alignment = align_center
    
    ws1.merge_cells(start_row=current_row, start_column=5, end_row=current_row, end_column=9)
    s2 = ws1.cell(current_row, 5, "QUẢN ĐỐC / TỔ TRƯỞNG DUYỆT")
    s2.font = font_bold
    s2.alignment = align_center
    
    current_row += 1
    ws1.row_dimensions[current_row].height = 15
    ws1.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    sub1 = ws1.cell(current_row, 1, "(Ký và ghi rõ họ tên)")
    sub1.font = font_italic
    sub1.alignment = align_center
    
    ws1.merge_cells(start_row=current_row, start_column=5, end_row=current_row, end_column=9)
    sub2 = ws1.cell(current_row, 5, "(Xác nhận ngạch bậc lương ngày)")
    sub2.font = font_italic
    sub2.alignment = align_center
    
    current_row += 4
    ws1.row_dimensions[current_row].height = 18
    ws1.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    n1 = ws1.cell(current_row, 1, "[Họ tên ứng viên]")
    n1.font = font_normal
    n1.alignment = align_center
    
    ws1.merge_cells(start_row=current_row, start_column=5, end_row=current_row, end_column=9)
    n2 = ws1.cell(current_row, 5, "[Họ tên người đánh giá]")
    n2.font = font_normal
    n2.alignment = align_center

    # -------------------------------------------------------------
    # SHEET 2: THANG ĐO & LƯƠNG
    # -------------------------------------------------------------
    ws2 = wb.create_sheet(title="Thang Đo & Lương")
    ws2.sheet_view.showGridLines = True
    
    ws2.column_dimensions['A'].width = 10
    ws2.column_dimensions['B'].width = 24
    ws2.column_dimensions['C'].width = 24
    ws2.column_dimensions['D'].width = 24
    ws2.column_dimensions['E'].width = 65
    
    # Headers
    ws2["A1"] = "Bậc"
    ws2["B1"] = "Chức danh đề xuất"
    ws2["C1"] = "Lương thực lĩnh (ngày)"
    ws2["D1"] = "Tổng chi phí công ty (ngày)"
    ws2["E1"] = "Phân công / Vai trò chính trong xưởng"
    
    for col in ["A", "B", "C", "D", "E"]:
        cell = ws2[f"{col}1"]
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_all_thin
        
    scale_data = [
        (1, "Thợ học việc (Cấp 1)", 220000, 300000, "Phụ việc vệ sinh, mài mộc, nâng hạ phôi dưới sự chỉ dẫn"),
        (2, "Thợ học việc (Cấp 2)", 220000, 300000, "Thực hiện tốt 5S khu vực hàn gá, đảm bảo an toàn ATLĐ xưởng"),
        (3, "Thợ phụ cơ bản (Cấp 1)", 250000, 330000, "Biết dùng máy cắt cầm tay, khoan tay, thước đo chính xác"),
        (4, "Thợ phụ mài tinh (Cấp 2)", 290000, 370000, "Mài tinh mỹ bề mặt, mài lấy góc sắc nét, mài phá bavia"),
        (5, "Thợ sơn hoàn thiện", 330000, 410000, "Đứng buồng sơn phun khí nén, sơn chống rỉ và sơn phủ thẩm mỹ"),
        (6, "Thợ gá lắp phôi", 380000, 460000, "Nắm rõ các loại vật tư, phôi cắt sắt/inox, chuẩn bị công trình"),
        (7, "Thợ gá hàn liên hợp", 400000, 480000, "Sử dụng thành thạo máy hàn que/MIG/TIG để đính gá kết cấu"),
        (8, "Thợ hàn chính chuyên nghiệp", 430000, 510000, "Đọc bản vẽ sản xuất và tự hàn độc lập sản phẩm chất lượng cao"),
        (9, "Thợ lắp ráp lành nghề", 440000, 520000, "Đọc bản vẽ cơ khí phức tạp, tự bóc tách khai triển phôi gá ráp"),
        (10, "Thợ hoàn thiện độc lập", 470000, 550000, "Tự kiểm soát chất lượng hoàn thiện sản phẩm không cần giám sát"),
        (11, "Thợ dự toán & tiến độ", 480000, 560000, "Ước lượng thời gian/nhân công, lập tiến độ thi công dự án nhỏ"),
        (12, "Tổ phó kỹ thuật", 490000, 570000, "Tự chủ động điều chỉnh chi tiết nhỏ ở hiện trường, thay thế tổ trưởng"),
        (13, "Trưởng nhóm lắp dựng", 500000, 580000, "Điều phối nhóm 2-3 người lắp dựng kết cấu thép/inox ngoài công trường"),
        (14, "Tổ trưởng sản xuất", 550000, 630000, "Phân việc toàn xưởng, chịu trách nhiệm tiến độ & chất lượng tổ đội")
    ]
    
    for r_idx, row_val in enumerate(scale_data, start=2):
        ws2.cell(r_idx, 1, row_val[0]).alignment = align_center
        ws2.cell(r_idx, 1).font = font_bold
        
        ws2.cell(r_idx, 2, row_val[1]).alignment = align_left
        ws2.cell(r_idx, 2).font = font_bold
        
        ws2.cell(r_idx, 3, row_val[2]).number_format = '#,##0" đ"'
        ws2.cell(r_idx, 3).alignment = align_right
        ws2.cell(r_idx, 3).font = font_bold
        
        ws2.cell(r_idx, 4, row_val[3]).number_format = '#,##0" đ"'
        ws2.cell(r_idx, 4).alignment = align_right
        ws2.cell(r_idx, 4).font = font_normal
        
        ws2.cell(r_idx, 5, row_val[4]).alignment = align_left
        ws2.cell(r_idx, 5).font = font_normal
        
        for col in range(1, 6):
            cell = ws2.cell(r_idx, col)
            cell.border = border_all_thin
            if r_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

    # Save workbook
    output_path = "d:/Sao Vàng/Website-SaoVang/CKSV_Form_Tuyen_Dung_Tho_Co_Khi_Chinh_Xac.xlsx"
    wb.save(output_path)
    print(f"Exact recruitment form created successfully at {output_path}")

if __name__ == "__main__":
    create_exact_recruitment_form()
