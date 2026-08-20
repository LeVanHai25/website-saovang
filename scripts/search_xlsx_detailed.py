import os
import openpyxl

def search_files_detailed():
    base_dir = "d:/Sao Vàng/Website-SaoVang"
    files_checked = []
    found = []
    
    for root, dirs, files in os.walk(base_dir):
        if ".git" in root or "node_modules" in root:
            continue
        for file in files:
            if file.endswith((".xlsx", ".xlsm", ".xls")):
                path = os.path.join(root, file)
                files_checked.append(path)
                try:
                    wb = openpyxl.load_workbook(path)
                    sheetnames = wb.sheetnames
                    if "Bảng dự toán" in sheetnames or "Bảng khối lượng" in sheetnames or "Bảng dự toán" in "".join(sheetnames):
                        found.append((path, sheetnames))
                except Exception as e:
                    # Log error
                    files_checked.append(f"{path} -> ERROR: {repr(e)}")
                    
    with open("d:/Sao Vàng/Website-SaoVang/scripts/search_xlsx_detailed.txt", "w", encoding="utf-8") as f:
        f.write("CHECKED FILES:\n")
        f.write("\n".join(files_checked))
        f.write("\n\nFOUND FILES:\n")
        for path, sheets in found:
            f.write(f"{path} -> {sheets}\n")
    print(f"Done, found {len(found)} files")

if __name__ == "__main__":
    search_files_detailed()
