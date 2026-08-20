import os
import openpyxl

def search_files():
    base_dir = "d:/Sao Vàng/Website-SaoVang"
    found = []
    
    for root, dirs, files in os.walk(base_dir):
        # Skip .git and node_modules
        if ".git" in root or "node_modules" in root:
            continue
        for file in files:
            if file.endswith((".xlsx", ".xlsm", ".xls")):
                path = os.path.join(root, file)
                try:
                    # check if we can read sheetnames
                    wb = openpyxl.load_workbook(path, read_only=True)
                    sheetnames = wb.sheetnames
                    if "Bảng dự toán" in sheetnames or "Bảng khối lượng" in sheetnames:
                        found.append((path, sheetnames))
                except Exception as e:
                    # Maybe it's old xls or locked
                    pass
                    
    with open("d:/Sao Vàng/Website-SaoVang/scripts/search_xlsx_results.txt", "w", encoding="utf-8") as f:
        for path, sheets in found:
            f.write(f"{path} -> {sheets}\n")
    print(f"Done, found {len(found)} files")

if __name__ == "__main__":
    search_files()
