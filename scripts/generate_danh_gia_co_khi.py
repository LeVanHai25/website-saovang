import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

def create_evaluation_matrix():
    wb = openpyxl.Workbook()
    
    # -------------------------------------------------------------
    # SHEET 1: BẢNG ĐÁNH GIÁ NĂNG LỰC
    # -------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Đánh Giá Năng Lực"
    ws1.sheet_view.showGridLines = True
    
    # Define color palette (Navy Blue & Steel Blue)
    NAVY_BLUE = "1B365D"
    ICE_BLUE = "DDEBF7"
    LIGHT_GRAY = "F2F2F2"
    BORDER_GRAY = "D9D9D9"
    YELLOW_INPUT = "FFF2CC"
    DARK_GREEN = "385723"
    LIGHT_GREEN = "E2EFDA"
    
    # Fonts
    font_title_company = Font(name="Segoe UI", size=11, bold=True, color="595959")
    font_title_main = Font(name="Segoe UI", size=16, bold=True, color=NAVY_BLUE)
    font_title_sub = Font(name="Segoe UI", size=10, italic=True, color="7F7F7F")
    font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    font_group_header = Font(name="Segoe UI", size=11, bold=True, color=NAVY_BLUE)
    font_bold = Font(name="Segoe UI", size=10, bold=True)
    font_normal = Font(name="Segoe UI", size=10)
    font_italic = Font(name="Segoe UI", size=9, italic=True)
    font_salary_result = Font(name="Segoe UI", size=12, bold=True, color=DARK_GREEN)
    
    # Fills
    fill_header = PatternFill(start_color=NAVY_BLUE, end_color=NAVY_BLUE, fill_type="solid")
    fill_group = PatternFill(start_color=ICE_BLUE, end_color=ICE_BLUE, fill_type="solid")
    fill_sum = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
    fill_input = PatternFill(start_color=YELLOW_INPUT, end_color=YELLOW_INPUT, fill_type="solid")
    fill_result_header = PatternFill(start_color="8FAADC", end_color="8FAADC", fill_type="solid")
    fill_result_box = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
    
    # Alignments
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    align_title = Alignment(horizontal="center", vertical="center")
    
    # Borders
    thin_border_side = Side(style="thin", color=BORDER_GRAY)
    border_all_thin = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    border_group_header = Border(
        left=Side(style="thin", color=NAVY_BLUE),
        right=Side(style="thin", color=NAVY_BLUE),
        top=Side(style="medium", color=NAVY_BLUE),
        bottom=Side(style="medium", color=NAVY_BLUE)
    )
    
    border_sum_row = Border(
        left=thin_border_side,
        right=thin_border_side,
        top=Side(style="thin", color=NAVY_BLUE),
        bottom=Side(style="thin", color=NAVY_BLUE)
    )
    
    border_total_row = Border(
        left=Side(style="thin", color=NAVY_BLUE),
        right=Side(style="thin", color=NAVY_BLUE),
        top=Side(style="thin", color=NAVY_BLUE),
        bottom=Side(style="double", color=NAVY_BLUE)
    )
    
    border_result_box = Border(
        left=Side(style="medium", color=DARK_GREEN),
        right=Side(style="medium", color=DARK_GREEN),
        top=Side(style="medium", color=DARK_GREEN),
        bottom=Side(style="medium", color=DARK_GREEN)
    )

    # Column widths
    ws1.column_dimensions['A'].width = 5   # STT
    ws1.column_dimensions['B'].width = 42  # Tiêu chí
    ws1.column_dimensions['C'].width = 13  # Điểm tối đa
    ws1.column_dimensions['D'].width = 14  # Điểm đánh giá
    ws1.column_dimensions['E'].width = 13  # Tỷ lệ %
    ws1.column_dimensions['F'].width = 62  # Hướng dẫn chấm điểm

    # 1. Title Block
    ws1.merge_cells("A2:F2")
    ws1["A2"] = "CÔNG TY CỔ PHẦN SẢN XUẤT CƠ KHÍ SAO VÀNG (ASV)"
    ws1["A2"].font = font_title_company
    ws1["A2"].alignment = Alignment(horizontal="left")

    ws1.merge_cells("A3:F3")
    ws1["A3"] = "BẢNG ĐÁNH GIÁ NĂNG LỰC & XÁC ĐỊNH BẬC LƯƠNG"
    ws1["A3"].font = font_title_main
    ws1["A3"].alignment = align_center

    ws1.merge_cells("A4:F4")
    ws1["A4"] = "Áp dụng cho: Tổ sản xuất / Thợ Cơ khí (Hàn, Lắp ráp, Gia công) - Chu kỳ: 3 tháng/lần"
    ws1["A4"].font = font_title_sub
    ws1["A4"].alignment = align_center

    # 2. Info Block
    ws1["B6"] = "Họ và tên nhân viên:"
    ws1["B6"].font = font_bold
    ws1["C6"] = "[Nhập tên nhân viên]"
    ws1["C6"].font = font_normal
    ws1["C6"].border = Border(bottom=Side(style="thin", color="000000"))

    ws1["E6"] = "Ngày đánh giá:"
    ws1["E6"].font = font_bold
    ws1["F6"] = "[Nhập ngày]"
    ws1["F6"].font = font_normal
    ws1["F6"].border = Border(bottom=Side(style="thin", color="000000"))

    ws1["B7"] = "Người đánh giá:"
    ws1["B7"].font = font_bold
    ws1["C7"] = "[Họ tên Quản lý/Tổ trưởng]"
    ws1["C7"].font = font_normal
    ws1["C7"].border = Border(bottom=Side(style="thin", color="000000"))

    ws1["E7"] = "Kỳ đánh giá:"
    ws1["E7"].font = font_bold
    ws1["F7"] = "[Ví dụ: Định kỳ Q3/2026]"
    ws1["F7"].font = font_normal
    ws1["F7"].border = Border(bottom=Side(style="thin", color="000000"))

    # Table Header
    headers = [
        ("A9", "STT", align_center),
        ("B9", "TIÊU CHÍ ĐÁNH GIÁ NĂNG LỰC", align_left),
        ("C9", "ĐIỂM TỐI ĐA", align_center),
        ("D9", "ĐIỂM ĐÁNH GIÁ", align_center),
        ("E9", "TỶ LỆ (%) đạt", align_center),
        ("F9", "HƯỚNG DẪN ĐÁNH GIÁ CHẤT LƯỢNG THỰC TẾ", align_left)
    ]
    for cell_ref, text, align in headers:
        cell = ws1[cell_ref]
        cell.value = text
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align
        cell.border = border_all_thin
    ws1.row_dimensions[9].height = 28

    # Data Rows
    # Structure: (STT, Criteria, Max Score, Description, Type)
    # Type: 'G' for Group Header, 'I' for Item, 'S' for Subtotal
    data = [
        # Group I
        ("I", "TAY NGHỀ CHUYÊN MÔN (Trọng số 40%)", None, "Nhóm năng lực cốt lõi về kỹ thuật trực tiếp", "G"),
        ("1", "Hàn MIG/MAG", 8, "Hàn kết cấu thép (sắt), yêu cầu mối hàn ngấu, đều tay, không rỗ khí, khuyết tật.", "I"),
        ("2", "Hàn TIG Inox", 8, "Hàn ống, bồn, tấm Inox mỏng; mối hàn sáng bóng (trắng/vàng nhạt), không đen/thủng.", "I"),
        ("3", "Đọc bản vẽ kỹ thuật cơ khí", 6, "Hiểu ký hiệu mối hàn, dung sai bản vẽ; tự bóc tách kích thước phôi để chế tạo độc lập.", "I"),
        ("4", "Gia công & Lắp ráp sản phẩm", 6, "Lấy dấu chính xác, gá dựng kết cấu khung đúng góc, biết tự chế/sử dụng jig gá đặt.", "I"),
        ("5", "Sử dụng thiết bị cầm tay hiệu quả", 4, "Sử dụng máy cắt bàn, máy mài góc, khoan tay an toàn, đúng góc, giữ gìn máy móc.", "I"),
        ("6", "Độ chính xác & Thẩm mỹ sản phẩm", 8, "Kích thước sai số nằm trong dung sai cho phép, mài sạch xỉ hàn, sản phẩm láng mịn.", "I"),
        ("", "Cộng nhóm I: Tay nghề chuyên môn", "=SUM(C11:C16)", "S"),

        # Group II
        ("II", "CHẤT LƯỢNG CÔNG VIỆC (Trọng số 20%)", None, "Đo lường sai hỏng và tuân thủ kỹ thuật", "G"),
        ("7", "Tỷ lệ sản phẩm đạt ngay từ lần đầu (Rework)", 8, "Không phải sửa lại lỗi sau hàn/gá; không gây lãng phí vật tư làm lại.", "I"),
        ("8", "Hoàn thành đúng yêu cầu kỹ thuật bản vẽ", 6, "Đảm bảo đúng biên dạng hình học, độ phẳng mặt phẳng, độ vuông góc của kết cấu.", "I"),
        ("9", "Tuân thủ quy trình hàn & lắp ráp", 6, "Đi đúng thứ tự hàn để chống co rút cong vênh; cài đặt đúng dòng điện/khí bảo vệ.", "I"),
        ("", "Cộng nhóm II: Chất lượng công việc", "=SUM(C19:C21)", "S"),

        # Group III
        ("III", "TIẾN ĐỘ & HIỆU SUẤT (Trọng số 15%)", None, "Tốc độ và tính chủ động trong sản xuất", "G"),
        ("10", "Hoàn thành công việc đúng tiến độ", 5, "Đạt định mức thời gian sản xuất công ty đưa ra; sẵn sàng tăng ca khi có tiến độ gấp.", "I"),
        ("11", "Chủ động trong công việc", 5, "Tự sắp xếp khu vực làm việc gọn gàng, chủ động chuẩn bị phôi và dụng cụ trước ca.", "I"),
        ("12", "Khả năng làm việc độc lập", 5, "Nhận bản vẽ & phôi là tự triển khai ra thành phẩm mà không cần giám sát liên tục.", "I"),
        ("", "Cộng nhóm III: Tiến độ & Hiệu suất", "=SUM(C24:C26)", "S"),

        # Group IV
        ("IV", "KỶ LUẬT & THÁI ĐỘ (Trọng số 15%)", None, "Tuân thủ nội quy và an toàn nhà xưởng", "G"),
        ("13", "Đi làm đúng giờ & Tác phong công nghiệp", 3, "Đúng giờ làm việc; mặc đầy đủ quần áo bảo hộ, giày, mũ bảo hiểm khi hàn/mài.", "I"),
        ("14", "Tuân thủ nội quy & Thực hành 5S xưởng", 3, "Vệ sinh máy móc, quét dọn khu vực làm việc trước và sau ca; sắp xếp công cụ gọn gàng.", "I"),
        ("15", "Tuân thủ An toàn lao động (ATLĐ)", 5, "Tuyệt đối không bỏ kính bảo hộ khi mài; che chắn khi hàn tránh ảnh hưởng người khác.", "I"),
        ("16", "Tinh thần trách nhiệm với tài sản xưởng", 4, "Bảo quản máy hàn, dụng cụ cầm tay tốt; báo cáo ngay khi phát hiện máy hỏng/rò khí.", "I"),
        ("", "Cộng nhóm IV: Kỷ luật & Thái độ", "=SUM(C29:C32)", "S"),

        # Group V
        ("V", "KHẢ NĂNG PHÁT TRIỂN & HỖ TRỢ (Trọng số 10%)", None, "Đóng góp vào sự phát triển chung và đội ngũ", "G"),
        ("17", "Tinh thần hỗ trợ đồng đội (Teamwork)", 2, "Hỗ trợ thợ khác khi nâng hạ phôi nặng, gá ghép kết cấu lớn; hòa nhã với đồng nghiệp.", "I"),
        ("18", "Hướng dẫn thợ phụ & thợ mới thử việc", 2, "Nhiệt tình truyền đạt mẹo gá đặt, chỉnh dòng điện cho thợ phụ nâng cao tay nghề.", "I"),
        ("19", "Đề xuất cải tiến & Sáng kiến sản xuất", 2, "Đóng góp ý tưởng làm đồ gá xoay nhanh, hoặc cách phân chia phôi tiết kiệm vật tư.", "I"),
        ("20", "Khả năng quản lý & Lãnh đạo tổ đội", 4, "Tư duy bao quát tốt; có thể đứng ra điều phối, chỉ dẫn công việc cho tổ 2-3 người.", "I"),
        ("", "Cộng nhóm V: Khả năng phát triển", "=SUM(C35:C38)", "S"),
    ]

    current_row = 10
    input_cells = []
    
    for row_data in data:
        stt, name, max_score, desc_or_type, r_type = row_data[0], row_data[1], row_data[2], row_data[3], row_data[4] if len(row_data) > 4 else row_data[3]
        
        ws1.row_dimensions[current_row].height = 22
        
        if r_type == "G":
            # Group Header - DO NOT MERGE TO AVOID READ-ONLY MERGED CELL ERROR
            for col in range(1, 7):
                cell = ws1.cell(current_row, col)
                cell.fill = fill_group
                cell.border = border_group_header
                
            ws1.cell(current_row, 1, stt).font = font_group_header
            ws1.cell(current_row, 1).alignment = align_center
            
            ws1.cell(current_row, 2, name).font = font_group_header
            ws1.cell(current_row, 2).alignment = align_left
            
            ws1.cell(current_row, 6, desc_or_type).font = font_italic
            ws1.cell(current_row, 6).alignment = align_left
            
        elif r_type == "I":
            # Item Row
            ws1.cell(current_row, 1, stt).font = font_normal
            ws1.cell(current_row, 1).alignment = align_center
            
            ws1.cell(current_row, 2, name).font = font_normal
            ws1.cell(current_row, 2).alignment = align_left
            
            ws1.cell(current_row, 3, max_score).font = font_bold
            ws1.cell(current_row, 3).alignment = align_center
            
            # Evaluated score (input)
            cell_eval = ws1.cell(current_row, 4)
            cell_eval.font = font_bold
            cell_eval.alignment = align_center
            cell_eval.fill = fill_input
            input_cells.append((current_row, max_score))
            
            # Tỷ lệ %
            cell_pct = ws1.cell(current_row, 5, f"=D{current_row}/C{current_row}")
            cell_pct.font = font_normal
            cell_pct.alignment = align_center
            cell_pct.number_format = '0.0%'
            
            # Description
            ws1.cell(current_row, 6, desc_or_type).font = font_normal
            ws1.cell(current_row, 6).alignment = align_left
            
            # Borders
            for col in range(1, 7):
                ws1.cell(current_row, col).border = border_all_thin
                
        elif r_type == "S":
            # Subtotal Row
            ws1.cell(current_row, 2, name).font = font_bold
            ws1.cell(current_row, 2).alignment = align_left
            
            ws1.cell(current_row, 3, max_score).font = font_bold
            ws1.cell(current_row, 3).alignment = align_center
            
            # Subtotal score formula
            sub_d = f"=SUM(D{current_row-1-int(current_row in [17,22,27,33,39])}:D{current_row-1})"
            # Actually, let's write out specific SUM formulas for robustness:
            if current_row == 17:  # Tay nghề
                ws1.cell(current_row, 4, "=SUM(D11:D16)").font = font_bold
            elif current_row == 22: # Chất lượng
                ws1.cell(current_row, 4, "=SUM(D19:D21)").font = font_bold
            elif current_row == 27: # Hiệu suất
                ws1.cell(current_row, 4, "=SUM(D24:D26)").font = font_bold
            elif current_row == 33: # Kỷ luật
                ws1.cell(current_row, 4, "=SUM(D29:D32)").font = font_bold
            elif current_row == 39: # Khả năng phát triển
                ws1.cell(current_row, 4, "=SUM(D35:D38)").font = font_bold
                
            ws1.cell(current_row, 4).alignment = align_center
            
            # Subtotal pct formula
            ws1.cell(current_row, 5, f"=D{current_row}/C{current_row}").font = font_bold
            ws1.cell(current_row, 5).alignment = align_center
            ws1.cell(current_row, 5).number_format = '0.0%'
            
            for col in range(1, 7):
                cell = ws1.cell(current_row, col)
                cell.fill = fill_sum
                cell.border = border_sum_row
                
        current_row += 1

    # Row 40 is empty
    ws1.row_dimensions[current_row].height = 15
    current_row += 1
    
    # Row 41: TOTAL ROW
    ws1.row_dimensions[current_row].height = 25
    ws1.cell(current_row, 2, "TỔNG ĐIỂM ĐÁNH GIÁ NĂNG LỰC").font = font_bold
    ws1.cell(current_row, 2).alignment = align_left
    
    ws1.cell(current_row, 3, "=C17+C22+C27+C33+C39").font = font_bold
    ws1.cell(current_row, 3).alignment = align_center
    
    ws1.cell(current_row, 4, "=D17+D22+D27+D33+D39").font = font_bold
    ws1.cell(current_row, 4).alignment = align_center
    ws1.cell(current_row, 4).fill = fill_sum
    
    ws1.cell(current_row, 5, f"=D{current_row}/C{current_row}").font = font_bold
    ws1.cell(current_row, 5).alignment = align_center
    ws1.cell(current_row, 5).number_format = '0.0%'
    
    for col in range(1, 7):
        ws1.cell(current_row, col).border = border_total_row
        if col in [2,3,4,5]:
            ws1.cell(current_row, col).font = Font(name="Segoe UI", size=11, bold=True)
            
    total_score_row = current_row # Save Row 41
    current_row += 2  # Skip a row
    
    # -------------------------------------------------------------
    # RESULTS DASHBOARD BOX (Rows 43-48)
    # -------------------------------------------------------------
    ws1.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=6)
    r_hdr = ws1.cell(current_row, 2, "KẾT QUẢ ĐÁNH GIÁ & XÁC ĐỊNH BẬC LƯƠNG ĐỀ XUẤT")
    r_hdr.font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    r_hdr.fill = PatternFill(start_color=NAVY_BLUE, end_color=NAVY_BLUE, fill_type="solid")
    r_hdr.alignment = align_center
    ws1.row_dimensions[current_row].height = 24
    
    # Outer box border applying
    for col in range(2, 7):
        ws1.cell(current_row, col).border = Border(top=Side(style="medium", color=NAVY_BLUE))
    
    current_row += 1
    
    dashboard_fields = [
        ("Tổng điểm đạt được:", f"=D{total_score_row}", " Điểm / 100", '0.0'),
        ("Tỷ lệ năng lực đạt được:", f"=E{total_score_row}", "", '0.0%'),
        ("Xếp loại năng lực & Bậc thợ:", f"=VLOOKUP(C{current_row+1}, 'Thang Đo'!A$2:E$8, 4, TRUE)", "", None),
        ("Mức lương đề xuất (VND):", f'=IF(VLOOKUP(C{current_row+1}, \'Thang Đo\'!A$2:E$8, 3, TRUE)=0, "Đào tạo lại / Chưa đạt", VLOOKUP(C{current_row+1}, \'Thang Đo\'!A$2:E$8, 3, TRUE))', "", '#,##0" đ"'),
    ]
    
    # For writing calculations properly, we need to know the cell address:
    # Row 44: label in B44, value in C44, note in D44.
    # We will write percent value in C45, so VLOOKUP refers to C45 (which is C{current_row_pct}).
    pct_cell_ref = f"C{current_row + 1}" # This will be C45
    
    for idx, (label, formula_or_val, suffix, num_format) in enumerate(dashboard_fields):
        ws1.row_dimensions[current_row].height = 20
        
        ws1.cell(current_row, 2, label).font = font_bold
        ws1.cell(current_row, 2).alignment = align_left
        
        # We replace the placeholder C{current_row+1} in formula with the actual pct_cell_ref
        actual_formula = formula_or_val.replace(f"C{current_row+1}", pct_cell_ref) if "C" in formula_or_val else formula_or_val
        
        cell_val = ws1.cell(current_row, 3, actual_formula)
        cell_val.alignment = align_center
        
        if label == "Mức lương đề xuất (VND):":
            cell_val.font = font_salary_result
        else:
            cell_val.font = font_bold
            
        if num_format:
            cell_val.number_format = num_format
            
        if suffix:
            ws1.cell(current_row, 4, suffix).font = font_italic
            ws1.cell(current_row, 4).alignment = Alignment(horizontal="left", vertical="center")
            
        # Style borders for the result box
        for col in range(2, 7):
            cell = ws1.cell(current_row, col)
            cell.fill = fill_result_box
            left_s = Side(style="medium", color=NAVY_BLUE) if col == 2 else None
            right_s = Side(style="medium", color=NAVY_BLUE) if col == 6 else None
            cell.border = Border(left=left_s, right=right_s)
            
        current_row += 1
        
    # Khuyến nghị nhân sự row (spans C to F)
    ws1.row_dimensions[current_row].height = 24
    ws1.cell(current_row, 2, "Khuyến nghị nhân sự:").font = font_bold
    ws1.cell(current_row, 2).alignment = align_left
    
    ws1.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=6)
    rec_cell = ws1.cell(current_row, 3, f"=VLOOKUP({pct_cell_ref}, 'Thang Đo'!A$2:E$8, 5, TRUE)")
    rec_cell.font = font_bold
    rec_cell.alignment = align_left
    
    for col in range(2, 7):
        cell = ws1.cell(current_row, col)
        cell.fill = fill_result_box
        left_s = Side(style="medium", color=NAVY_BLUE) if col == 2 else None
        right_s = Side(style="medium", color=NAVY_BLUE) if col == 6 else None
        cell.border = Border(left=left_s, right=right_s, bottom=Side(style="medium", color=NAVY_BLUE))

    current_row += 3  # Gap for signature
    
    # 3. Signature Block
    ws1.row_dimensions[current_row].height = 18
    ws1.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    s1 = ws1.cell(current_row, 1, "NHÂN VIÊN ĐƯỢC ĐÁNH GIÁ")
    s1.font = font_bold
    s1.alignment = align_center
    
    ws1.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=6)
    s2 = ws1.cell(current_row, 4, "TỔ TRƯỞNG / QUẢN ĐỐC XƯỞNG")
    s2.font = font_bold
    s2.alignment = align_center
    
    current_row += 1
    ws1.row_dimensions[current_row].height = 15
    ws1.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    sub1 = ws1.cell(current_row, 1, "(Ký và ghi rõ họ tên)")
    sub1.font = font_italic
    sub1.alignment = align_center
    
    ws1.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=6)
    sub2 = ws1.cell(current_row, 4, "(Ký và phê duyệt)")
    sub2.font = font_italic
    sub2.alignment = align_center
    
    current_row += 4
    ws1.row_dimensions[current_row].height = 18
    ws1.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    n1 = ws1.cell(current_row, 1, "[Họ tên nhân viên]")
    n1.font = font_normal
    n1.alignment = align_center
    
    ws1.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=6)
    n2 = ws1.cell(current_row, 4, "[Họ tên người đánh giá]")
    n2.font = font_normal
    n2.alignment = align_center

    # 4. Data Validation for Scores
    # We restrict inputs to a decimal between 0 and the max score specified in Column C.
    for r, m_score in input_cells:
        dv = DataValidation(
            type="decimal", 
            operator="between", 
            formula1="0", 
            formula2=f"C{r}", 
            allow_blank=True
        )
        dv.error = f'Điểm nhập vào phải từ 0 đến tối đa {m_score}!'
        dv.errorTitle = 'Lỗi nhập điểm'
        dv.prompt = f'Nhập số điểm từ 0 đến {m_score} điểm'
        dv.promptTitle = 'Nhập điểm đánh giá'
        
        ws1.add_data_validation(dv)
        dv.add(ws1[f"D{r}"])

    # Place default sample scores to verify look (or leave empty - let's set some typical values for display)
    ws1["D11"] = 7.0  # MIG/MAG
    ws1["D12"] = 6.5  # TIG Inox
    ws1["D13"] = 5.0  # Đọc bản vẽ
    ws1["D14"] = 5.0  # Gia công
    ws1["D15"] = 3.5  # Thiết bị cầm tay
    ws1["D16"] = 7.0  # Độ chính xác
    ws1["D19"] = 7.0  # Rework
    ws1["D20"] = 5.5  # Đúng yêu cầu
    ws1["D21"] = 5.5  # Quy trình
    ws1["D24"] = 4.5  # Tiến độ
    ws1["D25"] = 4.0  # Chủ động
    ws1["D26"] = 4.0  # Độc lập
    ws1["D29"] = 3.0  # Đúng giờ
    ws1["D30"] = 2.5  # 5S
    ws1["D31"] = 4.5  # ATLĐ
    ws1["D32"] = 3.5  # Trách nhiệm
    ws1["D35"] = 1.5  # Teamwork
    ws1["D36"] = 1.5  # Hướng dẫn
    ws1["D37"] = 1.0  # Sáng kiến
    ws1["D38"] = 2.0  # Quản lý

    # -------------------------------------------------------------
    # SHEET 2: BẢNG CẤU HÌNH THANG ĐO (THANG ĐO)
    # -------------------------------------------------------------
    ws2 = wb.create_sheet(title="Thang Đo")
    ws2.sheet_view.showGridLines = True
    
    ws2.column_dimensions['A'].width = 15
    ws2.column_dimensions['B'].width = 15
    ws2.column_dimensions['C'].width = 24
    ws2.column_dimensions['D'].width = 30
    ws2.column_dimensions['E'].width = 65
    
    # Headers for reference table
    ws2["A1"] = "Tỷ lệ tối thiểu"
    ws2["B1"] = "Khoảng năng lực"
    ws2["C1"] = "Lương đề xuất (VND)"
    ws2["D1"] = "Xếp loại năng lực"
    ws2["E1"] = "Đề xuất chi tiết về nhân sự"
    
    for col in ["A", "B", "C", "D", "E"]:
        ws2[f"{col}1"].font = font_header
        ws2[f"{col}1"].fill = fill_header
        ws2[f"{col}1"].alignment = align_center
        ws2[f"{col}1"].border = border_all_thin
        
    scale_data = [
        (0.00, "Dưới 70%", 0, "Chưa Đạt / Thợ học việc", "Cần đào tạo thêm, chưa đạt tiêu chuẩn thợ chính độc lập"),
        (0.70, "70% - 74%", 15000000, "Thợ Cơ Khí - Cấp 1", "Đạt chuẩn thợ chính bậc cơ bản (Lương 15.0 triệu)"),
        (0.75, "75% - 79%", 16000000, "Thợ Cơ Khí - Cấp 2", "Đạt chuẩn thợ chính bậc trung (Lương 16.0 triệu)"),
        (0.80, "80% - 84%", 17500000, "Thợ Cơ Khí - Cấp 3 (Khá)", "Thợ chính khá, độc lập giải quyết công việc phổ thông (Lương 17.5 triệu)"),
        (0.85, "85% - 89%", 19000000, "Thợ Cơ Khí - Cấp 4 (Lành nghề)", "Thợ lành nghề, gia công gá đặt và hàn độ khó cao (Lương 19.0 triệu)"),
        (0.90, "90% - 94%", 20500000, "Thợ Cơ Khí - Cấp 5 (Bậc cao)", "Thợ bậc cao, có thể quản lý nhóm và hướng dẫn thợ phụ (Lương 20.5 triệu)"),
        (0.95, "95% - 100%", 22000000, "Thợ Cơ Khí - Cấp 6 (Xuất sắc / Tổ trưởng)", "Thợ cơ khí xuất sắc, kỹ năng toàn diện, năng lực Tổ trưởng (Lương 22.0 triệu)")
    ]
    
    for r_idx, row_val in enumerate(scale_data, start=2):
        ws2.cell(r_idx, 1, row_val[0]).number_format = '0.0%'
        ws2.cell(r_idx, 1).alignment = align_center
        ws2.cell(r_idx, 1).font = font_bold
        
        ws2.cell(r_idx, 2, row_val[1]).alignment = align_center
        ws2.cell(r_idx, 2).font = font_normal
        
        ws2.cell(r_idx, 3, row_val[2]).number_format = '#,##0" đ"'
        ws2.cell(r_idx, 3).alignment = align_right
        ws2.cell(r_idx, 3).font = font_bold
        
        ws2.cell(r_idx, 4, row_val[3]).alignment = align_left
        ws2.cell(r_idx, 4).font = font_bold
        
        ws2.cell(r_idx, 5, row_val[4]).alignment = align_left
        ws2.cell(r_idx, 5).font = font_normal
        
        for col in range(1, 6):
            cell = ws2.cell(r_idx, col)
            cell.border = border_all_thin
            if r_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

    # -------------------------------------------------------------
    # SHEET 3: LỘ TRÌNH KHUNG NĂNG LỰC 5 CẤP
    # -------------------------------------------------------------
    ws3 = wb.create_sheet(title="Khung Năng Lực Chi Tiết")
    ws3.sheet_view.showGridLines = True
    
    ws3.column_dimensions['A'].width = 12
    ws3.column_dimensions['B'].width = 20
    ws3.column_dimensions['C'].width = 18
    ws3.column_dimensions['D'].width = 38
    ws3.column_dimensions['E'].width = 38
    ws3.column_dimensions['F'].width = 38
    
    ws3.merge_cells("A2:F2")
    ws3["A2"] = "KHUNG TIÊU CHUẨN NĂNG LỰC ĐỊNH HƯỚNG PHÁT TRIỂN NGHỀ NGHIỆP"
    ws3["A2"].font = Font(name="Segoe UI", size=14, bold=True, color=NAVY_BLUE)
    ws3["A2"].alignment = align_center
    
    ws3.row_dimensions[4].height = 26
    headers3 = [
        ("A4", "Bậc Lộ Trình", align_center),
        ("B4", "Chức Danh Vị Trí", align_center),
        ("C4", "Mức Lương Tham Chiếu", align_center),
        ("D4", "Tiêu Chuẩn Kỹ Thuật Chuyên Môn", align_left),
        ("E4", "Tiêu Chuẩn Chất Lượng & Tiến Độ", align_left),
        ("F4", "Tiêu Chuẩn Thái Độ & Quản Lý", align_left)
    ]
    
    for cell_ref, text, align in headers3:
        cell = ws3[cell_ref]
        cell.value = text
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align
        cell.border = border_all_thin
        
    levels_info = [
        ("Level 1", "Thợ Phụ / Học Việc", "7 - 12 triệu", 
         "• Sử dụng được máy mài, cắt cầm tay độc lập.\n• Hàn đính/hàn gá cơ bản.\n• Phụ trợ thợ chính gá lắp kết cấu.", 
         "• Mối hàn thô, cần sửa lỗi nhiều.\n• Làm việc theo chỉ đạo trực tiếp của tổ trưởng.\n• Tiến độ chậm.", 
         "• Chăm chỉ, chịu khó học hỏi nâng tay nghề.\n• Tuân thủ nghiêm ngặt ATLĐ (kính mài, găng tay).\n• Thực hiện tốt 5S khu vực phụ trách."),
        
        ("Level 2", "Thợ Chính Bậc Trung", "15 - 16 triệu", 
         "• Thành thạo hàn MIG/MAG hoặc TIG cơ bản.\n• Đọc hiểu bản vẽ lắp dựng đơn giản.\n• Gia công các kết cấu thép định hình thông dụng.", 
         "• Tỷ lệ sửa lỗi (rework) dưới 10%.\n• Hoàn thành công việc đúng tiến độ được phân công.\n• Đạt độ chính xác góc, phẳng cơ bản.", 
         "• Tác phong công nghiệp, đúng giờ.\n• Tự giác thực hiện 5S tại khu vực hàn.\n• Có tinh thần hợp tác, làm việc nhóm tốt."),
        
        ("Level 3", "Thợ Lành Nghề", "17.5 - 19 triệu", 
         "• Thành thạo đồng thời cả hàn MIG/MAG và TIG.\n• Đọc hiểu bản vẽ lắp ráp cơ khí phức tạp.\n• Tự chế tạo jig gá đặt cơ bản; gá kết cấu chính xác.", 
         "• Tỷ lệ rework dưới 3%.\n• Tiến độ bàn giao nhanh, ổn định.\n• Sản phẩm hàn xong sạch xỉ, thẩm mỹ cao.", 
         "• Chủ động cao trong công việc.\n• Biết tự bảo dưỡng thiết bị máy móc được giao.\n• Hướng dẫn được cho thợ phụ và thợ mới."),
        
        ("Level 4", "Thợ Bậc Cao / Tổ Phó Kỹ Thuật", "20.5 triệu", 
         "• Hàn điêu luyện các vật liệu khó (inox mỏng, ống áp lực).\n• Đọc, bóc tách bản vẽ cơ khí thành thạo.\n• Thiết kế/chế tạo đồ gá lắp ráp phức tạp.", 
         "• Hầu như không xảy ra sai hỏng (Rework < 1%).\n• Tốc độ gia công cực nhanh, tối ưu hóa nguyên vật liệu tránh hao phí.", 
         "• Có khả năng đứng ra quản lý nhóm 3-5 thợ chính/phụ.\n• Chủ động điều phối công việc khi tổ trưởng vắng mặt.\n• Tích cực đề xuất Kaizen."),
        
        ("Level 5", "Tổ Trưởng Sản Xuất", "22 triệu + Phụ cấp", 
         "• Trình độ chuyên môn cơ khí toàn diện.\n• Có kỹ năng kiểm tra và kiểm soát chất lượng (QC) toàn xưởng.\n• Xử lý triệt để sự cố kỹ thuật phát sinh.", 
         "• Chịu trách nhiệm chung về tiến độ & chất lượng của cả tổ sản xuất.\n• Quản lý tốt tỷ lệ hao hụt vật tư của tổ sản xuất.", 
         "• Kỹ năng quản lý, phân chia công việc xuất sắc.\n• Đánh giá năng lực nhân sự khách quan.\n• Đại diện tổ làm việc với Ban Giám đốc.")
    ]
    
    for r_idx, row_val in enumerate(levels_info, start=5):
        ws3.row_dimensions[r_idx].height = 90
        
        ws3.cell(r_idx, 1, row_val[0]).alignment = align_center
        ws3.cell(r_idx, 1).font = font_bold
        
        ws3.cell(r_idx, 2, row_val[1]).alignment = align_left
        ws3.cell(r_idx, 2).font = font_bold
        
        ws3.cell(r_idx, 3, row_val[2]).alignment = align_center
        ws3.cell(r_idx, 3).font = font_bold
        ws3.cell(r_idx, 3).font = Font(name="Segoe UI", size=10, bold=True, color=DARK_GREEN)
        
        for col_idx, val in enumerate(row_val[3:], start=4):
            cell = ws3.cell(r_idx, col_idx, val)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.font = font_normal
            
        for col in range(1, 7):
            cell = ws3.cell(r_idx, col)
            cell.border = border_all_thin
            if r_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

    # Save Workbook
    filepath = "d:/Sao Vàng/Website-SaoVang/Bang_Danh_Gia_Nang_Luc_Tho_Co_Khi_ASV.xlsx"
    wb.save(filepath)
    print(f"Excel workbook created successfully at {filepath}")

if __name__ == "__main__":
    create_evaluation_matrix()
