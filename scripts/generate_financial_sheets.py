import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_financial_workbook():
    wb = openpyxl.Workbook()
    
    # ----------------------------------------------------
    # Styles Definition
    # ----------------------------------------------------
    title_font = Font(name='Calibri', size=16, bold=True, color='1B365D')
    subtitle_font = Font(name='Calibri', size=11, italic=True, color='4A5568')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    section_font = Font(name='Calibri', size=11, bold=True, color='1B365D')
    subtotal_font = Font(name='Calibri', size=11, bold=True, color='000000')
    bold_font = Font(name='Calibri', size=11, bold=True)
    normal_font = Font(name='Calibri', size=11)
    note_font = Font(name='Calibri', size=10, italic=True, color='718096')

    header_fill = PatternFill(start_color='1B365D', end_color='1B365D', fill_type='solid')
    sub_header_fill = PatternFill(start_color='2B6CB0', end_color='2B6CB0', fill_type='solid')
    section_fill = PatternFill(start_color='EBF8FF', end_color='EBF8FF', fill_type='solid')
    subtotal_fill = PatternFill(start_color='EDF2F7', end_color='EDF2F7', fill_type='solid')
    highlight_fill = PatternFill(start_color='FEFCBF', end_color='FEFCBF', fill_type='solid')
    kpi_fill = PatternFill(start_color='C6F6D5', end_color='C6F6D5', fill_type='solid')

    thin_border_side = Side(border_style='thin', color='CBD5E0')
    thick_bottom_side = Side(border_style='double', color='1B365D')
    top_thin_side = Side(border_style='thin', color='1B365D')
    
    cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    subtotal_border = Border(top=top_thin_side, bottom=thick_bottom_side)
    
    currency_fmt = '#,##0 "VNĐ"'
    percent_fmt = '0.00%'

    # =========================================================================
    # SHEET 1: 01_Bao_Cao_KQKD_PnL
    # =========================================================================
    ws1 = wb.active
    ws1.title = "01_Bao_Cao_KQKD_PnL"
    ws1.views.sheetView[0].showGridLines = True

    ws1.merge_cells('A1:F1')
    ws1['A1'] = "CÔNG TY CỔ PHẦN CƠ KHÍ & NHÔM SAO VÀNG (CKSV)"
    ws1['A1'].font = title_font

    ws1.merge_cells('A2:F2')
    ws1['A2'] = "BÁO CÁO KẾT QUẢ HOẠT ĐỘNG KINH DOANH QUẢN TRỊ (P&L STATEMENT)"
    ws1['A2'].font = Font(name='Calibri', size=14, bold=True, color='2B6CB0')

    ws1.merge_cells('A3:F3')
    ws1['A3'] = "Kỳ báo cáo: Năm 2026 (Mô hình hợp nhất & Phân tích chi phí từng mảng) | Đơn vị tính: VNĐ"
    ws1['A3'].font = subtitle_font

    pnl_headers = ["STT", "Chỉ Tiêu Báo Cáo KQKD", "Mã Số", "Thuyết Minh / TK", "Số Tiền (VNĐ)", "Tỷ Trọng / Ghi Chú"]
    for col_idx, h in enumerate(pnl_headers, 1):
        cell = ws1.cell(row=5, column=col_idx)
        cell.value = h
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    pnl_rows = [
        ("I", "DOANH THU THUẦN BÁN HÀNG & CUNG CẤP DỊCH VỤ", "01", "TK 511", 12500000000, "100.00%"),
        ("1", "  1. Doanh thu Mảng OEM & Bán buôn Thanh nhôm", "01.1", "TK 5111", 7200000000, "=E7/E6"),
        ("2", "  2. Doanh thu Mảng Gia công & Sản xuất Cửa nhôm", "01.2", "TK 5112", 3800000000, "=E8/E6"),
        ("3", "  3. Doanh thu Mảng Dự án Thi công & Công trình", "01.3", "TK 5113", 1500000000, "=E9/E6"),
        ("II", "GIÁ VỐN HÀNG BÁN (COGS)", "11", "TK 632", 9150000000, "73.20%"),
        ("1", "  1. Chi phí nguyên vật liệu trực tiếp (NVLTT)", "11.1", "TK 621", 7350000000, "Phôi nhôm, kính, phụ kiện, sơn"),
        ("2", "  2. Chi phí nhân công trực tiếp (NCTT)", "11.2", "TK 622", 1100000000, "Lương thợ đùn, thợ cơ khí, thợ ráp"),
        ("3", "  3. Chi phí sản xuất chung (SXC)", "11.3", "TK 627", 700000000, "Khấu hao máy, điện xưởng, khuôn"),
        ("III", "LỢI NHUẬN GỘP VỀ BÁN HÀNG & CUNG CẤP DỊCH VỤ", "20", "20 = 01 - 11", "=E6-E10", "=E14/E6"),
        ("IV", "DOANH THU HOẠT ĐỘNG TÀI CHÍNH", "21", "TK 515", 35000000, "Lãi tiền gửi, chiết khấu thanh toán"),
        ("V", "CHI PHÍ TÀI CHÍNH", "22", "TK 635", 140000000, "1.12%"),
        ("1", "  - Trong đó: Chi phí lãi vay ngân hàng", "23", "TK 6351", 120000000, "Vay vốn lưu động nhập vật tư"),
        ("VI", "CHI PHÍ BÁN HÀNG", "25", "TK 641", 850000000, "6.80%"),
        ("1", "  1. Lương & hoa hồng bán hàng (Sales)", "25.1", "TK 6411", 380000000, "Kinh doanh OEM & cửa"),
        ("2", "  2. Chi phí Vận chuyển & Logistics", "25.2", "TK 6412", 220000000, "Cước xe giao nhôm & công trình"),
        ("3", "  3. Marketing, Quảng cáo & Mẫu Catalogue", "25.3", "TK 6413", 160000000, "In catalogue, website, chạy ads"),
        ("4", "  4. Chi phí đóng gói, màng PE & bọc bảo vệ", "25.4", "TK 6414", 90000000, "Bọc màng chống xước thanh nhôm"),
        ("VII", "CHI PHÍ QUẢN LÝ DOANH NGHIỆP", "26", "TK 642", 620000000, "4.96%"),
        ("1", "  1. Lương bộ phận Quản lý & Văn phòng", "26.1", "TK 6421", 360000000, "Ban giám đốc, Kế toán, HR, Admin"),
        ("2", "  2. Khấu hao tài sản cố định văn phòng", "26.2", "TK 6422", 50000000, "Máy tính, bàn ghế, thiết bị VP"),
        ("3", "  3. Chi phí dịch vụ mua ngoài (Điện, Nước, Net)", "26.3", "TK 6423", 70000000, "Văn phòng điều hành"),
        ("4", "  4. Phí kiểm định Quatest, CQ & Bảo hộ nhãn hiệu", "26.4", "TK 6424", 80000000, "Đăng ký nhãn hiệu & chứng nhận Quatest"),
        ("5", "  5. Chi phí QLDN khác (VPP, Tiếp khách, Dự phòng)", "26.5", "TK 6428", 60000000, "Chi phí giao tế & quản lý chung"),
        ("VIII", "LỢI NHUẬN THUẦN TỪ HOẠT ĐỘNG KINH DOANH (EBIT)", "30", "30 = 20 + 21 - (22+25+26)", "=E14+E15-(E16+E18+E23)", "=E29/E6"),
        ("IX", "THU NHẬP KHÁC", "31", "TK 711", 20000000, "Thanh lý phế liệu, nhôm vụn"),
        ("X", "CHI PHÍ KHÁC", "32", "TK 811", 10000000, "Chi phí xử lý môi trường/phát sinh"),
        ("XI", "TỔNG LỢI NHUẬN KẾ TOÁN TRƯỚC THUẾ (EBT)", "50", "50 = 30 + 31 - 32", "=E29+E30-E31", "=E32/E6"),
        ("XII", "CHI PHÍ THUẾ TNDN HÀNH CHÍNH (20%)", "51", "TK 821", "=E32*0.2", "Thuế TNDN 20%"),
        ("XIII", "LỢI NHUẬN SAU THUẾ TNDN (NET PROFIT)", "60", "60 = 50 - 51", "=E32-E33", "=E34/E6")
    ]

    for row_idx, rdata in enumerate(pnl_rows, 6):
        stt, chitiieu, maso, thuk, sotien, ghichu = rdata
        ws1.cell(row=row_idx, column=1, value=stt).alignment = Alignment(horizontal='center')
        ws1.cell(row=row_idx, column=2, value=chitiieu)
        ws1.cell(row=row_idx, column=3, value=maso).alignment = Alignment(horizontal='center')
        ws1.cell(row=row_idx, column=4, value=thuk).alignment = Alignment(horizontal='center')
        
        c5 = ws1.cell(row=row_idx, column=5, value=sotien)
        c6 = ws1.cell(row=row_idx, column=6, value=ghichu)

        is_major = stt in ["I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X", "XI", "XII", "XIII"]
        
        if is_major:
            for c in range(1, 7):
                ws1.cell(row=row_idx, column=c).font = bold_font
                ws1.cell(row=row_idx, column=c).fill = section_fill if stt not in ["III", "VIII", "XIII"] else kpi_fill
            c5.number_format = currency_fmt
            if isinstance(ghichu, str) and ghichu.startswith('='):
                c6.number_format = percent_fmt
        else:
            for c in range(1, 7):
                ws1.cell(row=row_idx, column=c).font = normal_font
            c5.number_format = currency_fmt
            if isinstance(ghichu, str) and ghichu.startswith('='):
                c6.number_format = percent_fmt
        
        for c in range(1, 7):
            ws1.cell(row=row_idx, column=c).border = cell_border

    # Highlight Net profit row
    for c in range(1, 7):
        ws1.cell(row=34, column=c).fill = PatternFill(start_color='C6F6D5', end_color='C6F6D5', fill_type='solid')
        ws1.cell(row=34, column=c).font = Font(name='Calibri', size=11, bold=True, color='22543D')
        ws1.cell(row=34, column=c).border = Border(top=Side(border_style='thin', color='22543D'), bottom=Side(border_style='double', color='22543D'))

    # =========================================================================
    # SHEET 2: 02_Bang_Phan_Bo_Chi_Phi
    # =========================================================================
    ws2 = wb.create_sheet(title="02_Bang_Phan_Bo_Chi_Phi")
    ws2.views.sheetView[0].showGridLines = True

    ws2.merge_cells('A1:J1')
    ws2['A1'] = "BẢNG PHÂN BỔ CHI PHÍ CHI TIẾT THEO TRUNG TÂM CHI PHÍ (COST CENTERS)"
    ws2['A1'].font = title_font

    ws2.merge_cells('A2:J2')
    ws2['A2'] = "Phân bổ chi tiết cho 3 Mảng: (1) OEM Nhôm Bán Buôn | (2) Gia Công Cửa Nhôm | (3) Thi Công Dự Án Monaco"
    ws2['A2'].font = subtitle_font

    alloc_headers = [
        "Mã CP", "Tên Loại Chi Phí Chi Tiết", "Mã TK", "Tổng Chi Phí (VNĐ)",
        "Tiêu Thức Phân Bổ", "% OEM Nhôm", "Tiền OEM Nhôm (VNĐ)",
        "% Cửa Nhôm", "Tiền Cửa Nhôm (VNĐ)", "% Dự Án", "Tiền Dự Án (VNĐ)", "Ghi Chú Quản Trị"
    ]

    for col_idx, h in enumerate(alloc_headers, 1):
        cell = ws2.cell(row=4, column=col_idx)
        cell.value = h
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    cost_items = [
        # Group 1: 621
        ("CP-NVL01", "Phôi nhôm billet AL6063 & Thanh nhôm thô", "621", 5200000000, "Trực tiếp sản lượng kg đùn OEM & Cửa", 0.75, 0.20, 0.05, "Tập hợp trực tiếp theo PO sản xuất"),
        ("CP-NVL02", "Bột sơn tĩnh điện cao cấp (Jotun/Akzo)", "621", 650000000, "Diện tích bề mặt sơn m2", 0.70, 0.20, 0.10, "Bảo hành 10-15 năm theo chứng thư"),
        ("CP-NVL03", "Kính cường lực, kính hộp, kính dán an toàn", "621", 850000000, "Trực tiếp mảng Cửa & Dự án", 0.00, 0.65, 0.35, "OEM nhôm không dùng kính"),
        ("CP-NVL04", "Bộ phụ kiện kim khí (Bản lề, khóa, tay nắm)", "621", 450000000, "Trực tiếp số bộ cửa sản xuất", 0.00, 0.70, 0.30, "Đồng bộ hệ nhôm rãnh C"),
        ("CP-NVL05", "Vật tư phụ (Gioăng EPDM, keo silicone, ốc vít)", "621", 200000000, "Theo mét dài cửa & khung", 0.10, 0.60, 0.30, "Gioăng chèn thanh nhôm & lắp đặt"),
        
        # Group 2: 622
        ("CP-NC01", "Lương thợ vận hành máy đùn & đùn nhôm", "622", 420000000, "Sản lượng kg đùn OEM", 0.85, 0.15, 0.00, "Chi phí nhân công xưởng OEM"),
        ("CP-NC02", "Lương thợ cơ khí cắt, ép ke, gia công cửa", "622", 480000000, "Giờ công thợ gia công cửa", 0.00, 0.80, 0.20, "Xưởng gia công CKSV"),
        ("CP-NC03", "Lương thợ thi công lắp đặt tại công trình", "622", 200000000, "Trực tiếp công trình dự án", 0.00, 0.10, 0.90, "Lắp đặt biệt thự Monaco"),

        # Group 3: 627
        ("CP-SXC01", "Điện năng sản xuất xưởng & máy móc", "627", 220000000, "Công suất KW máy & giờ chạy", 0.60, 0.30, 0.10, "Máy đùn nhôm tiêu thụ điện lớn"),
        ("CP-SXC02", "Khấu hao dây chuyền đùn & máy cơ khí", "627", 250000000, "Nguyên giá & tỷ lệ khai thác máy", 0.55, 0.35, 0.10, "Khấu hao tài sản cố định nhà xưởng"),
        ("CP-SXC03", "Khấu hao khuôn đùn nhôm OEM Sao Vàng", "627", 110000000, "Sản lượng sản xuất từng bộ khuôn", 0.90, 0.10, 0.00, "Khuôn đùn độc quyền tem Sao Vàng"),
        ("CP-SXC04", "Chi phí bảo trì, sửa chữa xưởng & máy móc", "627", 70000000, "Tỷ lệ diện tích & giờ máy", 0.50, 0.40, 0.10, "Bảo dưỡng định kỳ"),
        ("CP-SXC05", "Lương quản đốc xưởng & kỹ thuật KCS", "627", 50000000, "Tỷ lệ thời gian giám sát", 0.50, 0.35, 0.15, "Kiểm soát chất lượng CQ xuất xưởng"),

        # Group 4: 641
        ("CP-BH01", "Lương & hoa hồng phòng kinh doanh (Sales)", "641", 380000000, "Doanh thu từng mảng", 0.58, 0.30, 0.12, "Tỷ lệ theo doanh thu bán hàng"),
        ("CP-BH02", "Chi phí xe tải & logistics giao hàng", "641", 220000000, "Số chuyến xe & cự ly km", 0.65, 0.20, 0.15, "Vận chuyển nhôm thanh & cửa hoàn thiện"),
        ("CP-BH03", "Marketing, Website & In Catalogue nhôm", "641", 160000000, "Định hướng phát triển thương hiệu", 0.50, 0.30, 0.20, "Hồ sơ năng lực & mẫu góc cửa"),
        ("CP-BH04", "Màng PE bọc bảo vệ thanh nhôm & đóng gói", "641", 90000000, "Số mét thanh nhôm xuất kho", 0.80, 0.15, 0.05, "Tem nhãn & màng chống xước"),

        # Group 5: 642
        ("CP-QL01", "Lương Ban Giám Đốc, Kế toán, Nhân sự", "642", 360000000, "Tỷ lệ Doanh thu hợp nhất", 0.58, 0.30, 0.12, "Bộ phận quản lý điều hành"),
        ("CP-QL02", "Phí kiểm định Quatest & Bảo hộ thương hiệu", "642", 80000000, "Trực tiếp mảng OEM & Hồ sơ thầu", 0.70, 0.10, 0.20, "Phí Quatest 1 / IBST & Đăng ký sở hữu trí tuệ"),
        ("CP-QL03", "Khấu hao thiết bị văn phòng & Công cụ", "642", 50000000, "Số lượng nhân sự văn phòng", 0.50, 0.30, 0.20, "Máy tính, máy in, phần mềm"),
        ("CP-QL04", "Chi phí văn phòng (Điện, Nước, Net, VPP)", "642", 130000000, "Tỷ lệ Doanh thu hợp nhất", 0.58, 0.30, 0.12, "Văn phòng điều hành CKSV"),

        # Group 6: 635
        ("CP-TC01", "Chi phí lãi vay ngân hàng vốn lưu động", "635", 120000000, "Giá trị vốn lưu động sử dụng", 0.70, 0.20, 0.10, "Vay nhập phôi nhôm Billet số lượng lớn"),
        ("CP-TC02", "Phí bảo lãnh hợp đồng & phí ngân hàng", "635", 20000000, "Trực tiếp hợp đồng dự án/OEM", 0.30, 0.20, 0.50, "Bảo lãnh thực hiện hợp đồng Monaco")
    ]

    r_start = 5
    for idx, item in enumerate(cost_items, r_start):
        macp, tencp, matk, tongtien, tieuthuc, p1, p2, p3, ghichu = item
        
        ws2.cell(row=idx, column=1, value=macp).alignment = Alignment(horizontal='center')
        ws2.cell(row=idx, column=2, value=tencp)
        ws2.cell(row=idx, column=3, value=matk).alignment = Alignment(horizontal='center')
        
        c4 = ws2.cell(row=idx, column=4, value=tongtien)
        c4.number_format = currency_fmt
        
        ws2.cell(row=idx, column=5, value=tieuthuc)
        
        # % and Amount OEM
        c6 = ws2.cell(row=idx, column=6, value=p1)
        c6.number_format = percent_fmt
        c7 = ws2.cell(row=idx, column=7, value=f"=D{idx}*F{idx}")
        c7.number_format = currency_fmt
        
        # % and Amount Cua Nhom
        c8 = ws2.cell(row=idx, column=8, value=p2)
        c8.number_format = percent_fmt
        c9 = ws2.cell(row=idx, column=9, value=f"=D{idx}*H{idx}")
        c9.number_format = currency_fmt
        
        # % and Amount Du An
        c10 = ws2.cell(row=idx, column=10, value=p3)
        c10.number_format = percent_fmt
        c11 = ws2.cell(row=idx, column=11, value=f"=D{idx}*J{idx}")
        c11.number_format = currency_fmt
        
        ws2.cell(row=idx, column=12, value=ghichu)
        
        for col in range(1, 13):
            ws2.cell(row=idx, column=col).border = cell_border
            ws2.cell(row=idx, column=col).font = normal_font

    # Total Row for Allocation Table
    tot_row = len(cost_items) + r_start
    ws2.cell(row=tot_row, column=1, value="TỔNG").alignment = Alignment(horizontal='center')
    ws2.cell(row=tot_row, column=2, value="TỔNG CỘNG CHI PHÍ PHÂN BỔ")
    ws2.cell(row=tot_row, column=4, value=f"=SUM(D{r_start}:D{tot_row-1})").number_format = currency_fmt
    ws2.cell(row=tot_row, column=7, value=f"=SUM(G{r_start}:G{tot_row-1})").number_format = currency_fmt
    ws2.cell(row=tot_row, column=9, value=f"=SUM(I{r_start}:I{tot_row-1})").number_format = currency_fmt
    ws2.cell(row=tot_row, column=11, value=f"=SUM(K{r_start}:K{tot_row-1})").number_format = currency_fmt

    for col in range(1, 13):
        cell = ws2.cell(row=tot_row, column=col)
        cell.font = bold_font
        cell.fill = subtotal_fill
        cell.border = subtotal_border

    # =========================================================================
    # SHEET 3: 03_Tieu_Thuc_&_Quy_Trinh
    # =========================================================================
    ws3 = wb.create_sheet(title="03_Tieu_Thuc_&_Quy_Trinh")
    ws3.views.sheetView[0].showGridLines = True

    ws3.merge_cells('A1:E1')
    ws3['A1'] = "HƯỚNG DẪN TIÊU THỨC PHÂN BỔ & PHƯƠNG PHÁP QUẢN TRỊ CHI PHÍ SAO VÀNG"
    ws3['A1'].font = title_font

    ws3.merge_cells('A2:E2')
    ws3['A2'] = "Nguyên tắc kế toán quản trị áp dụng cho Doanh nghiệp Cơ khí & Sản xuất Nhôm OEM"
    ws3['A2'].font = subtitle_font

    guides = [
        ("STT", "Phương Pháp / Tiêu Thức Phân Bổ", "Công Thức / Cách Tính Tỷ Lệ", "Loại Chi Phí Áp Dụng", "Ưu Điểm & Lưu Ý Quản Trị"),
        ("1", "Trực tiếp (Direct Tracing)", "Tập hợp trực tiếp 100% cho sản phẩm/mảng đó", "NVL chính (Phôi nhôm, Kính, Phụ kiện theo PO), Khấu hao khuôn OEM", "Chính xác tuyệt đối 100%, không cần tiêu thức ước tính."),
        ("2", "Theo Sản lượng / Khối lượng (Weight/Volume Driver)", "Tỷ lệ % = Khối lượng mảng (kg, m2) / Tổng khối lượng hợp nhất", "Bột sơn tĩnh điện, Điện xưởng đùn nhôm, Màng PE bọc nhôm, Vận chuyển nhôm", "Phù hợp nhất cho mảng sản xuất nhôm đùn và sơn tĩnh điện."),
        ("3", "Theo Giờ công lao động (Man-hour Driver)", "Tỷ lệ % = Số giờ công thợ mảng A / Tổng số giờ công", "Chi phí nhân công gia công cửa, Lương thợ cơ khí, Sửa chữa bảo trì", "Phản ánh đúng mức độ tiêu tốn sức lao động và năng suất xưởng."),
        ("4", "Theo Doanh thu (Revenue-based Allocation)", "Tỷ lệ % = Doanh thu mảng A / Tổng Doanh thu Hợp nhất", "Chi phí bán hàng, Hoa hồng Sales, Lương BOD/Kế toán, Văn phòng phẩm", "Đơn giản, phản ánh khả năng gánh vác chi phí quản lý của từng mảng."),
        ("5", "Theo Công suất & Thiết bị (Machine-hour / Power Driver)", "Tỷ lệ % = (Công suất máy kW x Số giờ chạy) / Tổng KWh", "Chi phí điện sản xuất tiêu thụ lớn, Khấu hao máy móc cơ khí, Bảo dưỡng", "Giúp tính đúng giá thành cửa m2 và nhôm thanh kg.")
    ]

    for r_idx, g in enumerate(guides, 4):
        stt, phuongphap, congthuc, loaicp, uudiem = g
        ws3.cell(row=r_idx, column=1, value=stt).alignment = Alignment(horizontal='center')
        ws3.cell(row=r_idx, column=2, value=phuongphap)
        ws3.cell(row=r_idx, column=3, value=congthuc)
        ws3.cell(row=r_idx, column=4, value=loaicp)
        ws3.cell(row=r_idx, column=5, value=uudiem)

        is_head = r_idx == 4
        for c in range(1, 6):
            cell = ws3.cell(row=r_idx, column=c)
            cell.border = cell_border
            if is_head:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            else:
                cell.font = normal_font
                if c == 2:
                    cell.font = bold_font

    # Adjusting Column Widths
    for ws in [ws1, ws2, ws3]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if cell.number_format == currency_fmt and isinstance(cell.value, (int, float)):
                    val_str = f"{cell.value:,.0f} VNĐ"
                if len(val_str) > max_len and not cell.coordinate in ['A1', 'A2', 'A3']:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # Manual tweaking for better display
    ws1.column_dimensions['A'].width = 8
    ws1.column_dimensions['B'].width = 48
    ws1.column_dimensions['C'].width = 10
    ws1.column_dimensions['D'].width = 22
    ws1.column_dimensions['E'].width = 22
    ws1.column_dimensions['F'].width = 38

    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 42
    ws2.column_dimensions['C'].width = 10
    ws2.column_dimensions['D'].width = 22
    ws2.column_dimensions['E'].width = 35
    ws2.column_dimensions['F'].width = 14
    ws2.column_dimensions['G'].width = 22
    ws2.column_dimensions['H'].width = 14
    ws2.column_dimensions['I'].width = 22
    ws2.column_dimensions['J'].width = 14
    ws2.column_dimensions['K'].width = 22
    ws2.column_dimensions['L'].width = 38

    ws3.column_dimensions['A'].width = 8
    ws3.column_dimensions['B'].width = 30
    ws3.column_dimensions['C'].width = 35
    ws3.column_dimensions['D'].width = 35
    ws3.column_dimensions['E'].width = 45

    output_path = "Bang_Phan_Bo_Va_Ket_Qua_Kinh_Doanh_Sao_Vang.xlsx"
    wb.save(output_path)
    print(f"Successfully generated financial allocation and P&L excel at: {output_path}")

if __name__ == '__main__':
    create_financial_workbook()
