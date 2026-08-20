import os
import openpyxl

def inspect_detail():
    base_dir = "d:/Sao Vàng/Website-SaoVang/CKSV"
    sub_dir = os.path.join(base_dir, "bảng duyển dụng")
    
    inspect_targets = [
        # File path, sheet name, max rows
        (os.path.join(base_dir, "1, CKSV, Nhiệm vụ phòng Thiết kế - Dự toán - 010.9.24 (1).xlsx"), "ĐÁNH GIÁ CHUNG ", 25),
        (os.path.join(base_dir, "1, CKSV, Nhiệm vụ phòng Thiết kế - Dự toán - 010.9.24 (1).xlsx"), "TP kỹ thuật 09.8.24", 25),
        (os.path.join(base_dir, "3, CKSV, Nhiệm vụ Tổng đội thi công - 10.9.24.xlsx"), "Quản đốc sx 09.8.24", 25),
        (os.path.join(base_dir, "3, CKSV, Nhiệm vụ Tổng đội thi công - 10.9.24.xlsx"), "Sheet1", 30),
        (os.path.join(sub_dir, "CKSV THANG BẢNG LƯƠNG & FOM TUYỂN DỤNG.xlsx"), "Ngạch bậc 09.9.24", 30),
        (os.path.join(sub_dir, "CKSV THANG BẢNG LƯƠNG & FOM TUYỂN DỤNG.xlsx"), "Bảng lương 09.9.24", 30),
        (os.path.join(sub_dir, "CKSV THANG BẢNG LƯƠNG & FOM TUYỂN DỤNG.xlsx"), "CÔNG NHÂN, TỔ TRƯỞNG", 30),
        (os.path.join(sub_dir, "CKSV THANG BẢNG LƯƠNG & FOM TUYỂN DỤNG.xlsx"), "N-VỤ PHÒNG THIẾT KẾ, KỸ THUẬT", 30),
        (os.path.join(sub_dir, "CKSV, Tự đánh gia bản thân - 31.7.24.xlsx"), "Sheet1", 30),
        (os.path.join(sub_dir, "CKSV,_Sơ_đồ_tổ_chức_công_ty_và_quỹ_lương,_Gốc.xlsx"), "Sheet1", 40) # let's see what sheet names it has first
    ]
    
    # We will first check sheets of CKSV,_Sơ_đồ_tổ_chức_công_ty_và_quỹ_lương,_Gốc.xlsx
    company_org_file = os.path.join(sub_dir, "CKSV,_Sơ_đồ_tổ_chức_công_ty_và_quỹ_lương,_Gốc.xlsx")
    org_sheets = []
    if os.path.exists(company_org_file):
        try:
            wb = openpyxl.load_workbook(company_org_file, read_only=True)
            org_sheets = wb.sheetnames
            print(f"Sheets in Org File: {org_sheets}")
        except Exception as e:
            print(f"Error checking org file sheets: {e}")
            
    output_lines = []
    output_lines.append("DETAIL INSPECTION OF CKSV WORKBOOK SHEETS:")
    
    for filepath, sheetname, max_r in inspect_targets:
        if not os.path.exists(filepath):
            output_lines.append(f"\nFile NOT found: {filepath}")
            continue
            
        output_lines.append("\n" + "="*80)
        output_lines.append(f"FILE: {os.path.basename(filepath)} | SHEET: {sheetname}")
        output_lines.append("="*80)
        
        try:
            # We open without read_only because some sheets might have issues, but let's check
            wb = openpyxl.load_workbook(filepath, data_only=True)
            if sheetname not in wb.sheetnames:
                output_lines.append(f"Sheet {sheetname} not found. Available sheets: {wb.sheetnames}")
                continue
            ws = wb[sheetname]
            rows = list(ws.iter_rows(max_row=max_r, max_col=12, values_only=True))
            for idx, r in enumerate(rows):
                if any(x is not None for x in r):
                    clean_row = []
                    for val in r:
                        if val is not None:
                            # format floats to look clean
                            if isinstance(val, float):
                                if val.is_integer():
                                    clean_row.append(str(int(val)))
                                else:
                                    clean_row.append(f"{val:.2f}")
                            else:
                                clean_row.append(str(val))
                        else:
                            clean_row.append("")
                    # remove trailing empties
                    while clean_row and clean_row[-1] == "":
                        clean_row.pop()
                    if clean_row:
                        output_lines.append(f"  Row {idx+1}: {clean_row}")
        except Exception as e:
            output_lines.append(f"Error reading sheet {sheetname} from {os.path.basename(filepath)}: {e}")
            
    # Also inspect sheets of org file dynamically if we found sheets
    if os.path.exists(company_org_file) and org_sheets:
        for sname in org_sheets[:3]: # inspect first 3 sheets of org file
            output_lines.append("\n" + "="*80)
            output_lines.append(f"FILE: CKSV,_Sơ_đồ_tổ_chức_công_ty_và_quỹ_lương,_Gốc.xlsx | SHEET: {sname}")
            output_lines.append("="*80)
            try:
                wb = openpyxl.load_workbook(company_org_file, data_only=True)
                ws = wb[sname]
                rows = list(ws.iter_rows(max_row=30, max_col=12, values_only=True))
                for idx, r in enumerate(rows):
                    if any(x is not None for x in r):
                        clean_row = []
                        for val in r:
                            if val is not None:
                                clean_row.append(str(val))
                            else:
                                clean_row.append("")
                        while clean_row and clean_row[-1] == "":
                            clean_row.pop()
                        if clean_row:
                            output_lines.append(f"  Row {idx+1}: {clean_row}")
            except Exception as e:
                output_lines.append(f"Error reading sheet {sname}: {e}")

    output_path = "d:/Sao Vàng/Website-SaoVang/scripts/inspect_detail_results.txt"
    with open(output_path, "w", encoding="utf-8") as f_out:
        f_out.write("\n".join(output_lines))
    print(f"Detailed inspection written to {output_path}")

if __name__ == "__main__":
    inspect_detail()
