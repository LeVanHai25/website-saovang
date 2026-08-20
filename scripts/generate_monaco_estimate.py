import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage

def generate_monaco_estimate_with_logo():
    wb = openpyxl.Workbook()
    
    # Define colors
    STEEL_BLUE = "2F5597"
    ICE_BLUE = "D9E1F2"
    LIGHT_GRAY = "F2F2F2"
    BORDER_GRAY = "D9D9D9"
    YELLOW_INPUT = "FFF2CC"
    DARK_GREEN = "385723"
    LIGHT_GREEN = "E2EFDA"
    
    # Fonts
    font_company_name = Font(name="Segoe UI", size=12, bold=True, color=STEEL_BLUE)
    font_company_info = Font(name="Segoe UI", size=9, italic=True, color="595959")
    font_title_main = Font(name="Segoe UI", size=15, bold=True, color=STEEL_BLUE)
    font_title_sub = Font(name="Segoe UI", size=10, italic=True, color="7F7F7F")
    font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    font_group_header = Font(name="Segoe UI", size=11, bold=True, color="000000")
    font_bold = Font(name="Segoe UI", size=10, bold=True)
    font_normal = Font(name="Segoe UI", size=10)
    font_italic = Font(name="Segoe UI", size=9, italic=True)
    font_total = Font(name="Segoe UI", size=11, bold=True, color=DARK_GREEN)
    
    # Fills
    fill_header = PatternFill(start_color=STEEL_BLUE, end_color=STEEL_BLUE, fill_type="solid")
    fill_group = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    fill_input = PatternFill(start_color=YELLOW_INPUT, end_color=YELLOW_INPUT, fill_type="solid")
    fill_total = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
    
    # Alignments
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    
    # Borders
    thin_side = Side(style="thin", color=BORDER_GRAY)
    border_all_thin = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    border_double_bottom = Border(
        left=thin_side, right=thin_side, 
        top=thin_side, 
        bottom=Side(style="double", color="000000")
    )
    
    # -------------------------------------------------------------
    # SHEET 1: TỔNG HỢP VẬT TƯ KIỂM KHO
    # -------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Tổng Hợp Kiểm Kho"
    ws1.sheet_view.showGridLines = True
    
    # Column widths
    ws1.column_dimensions['A'].width = 6
    ws1.column_dimensions['B'].width = 25
    ws1.column_dimensions['C'].width = 30
    ws1.column_dimensions['D'].width = 16
    ws1.column_dimensions['E'].width = 16
    ws1.column_dimensions['F'].width = 18
    ws1.column_dimensions['G'].width = 18
    ws1.column_dimensions['H'].width = 28
    ws1.column_dimensions['I'].width = 25

    # Company Letterhead Row 2-4
    ws1.row_dimensions[2].height = 20
    ws1.row_dimensions[3].height = 15
    ws1.row_dimensions[4].height = 15
    
    # Add Company Name & Info
    ws1.merge_cells("C2:I2")
    ws1["C2"] = "CÔNG TY CỔ PHẦN SẢN XUẤT CƠ KHÍ SAO VÀNG (ASV)"
    ws1["C2"].font = font_company_name
    ws1["C2"].alignment = Alignment(horizontal="left", vertical="center")
    
    ws1.merge_cells("C3:I3")
    ws1["C3"] = "Địa chỉ: Tầng 3, TT7-35 Khu đô thị Văn Phú, phường Kiến Hưng, TP Hà Nội, Việt Nam."
    ws1["C3"].font = font_company_info
    ws1["C3"].alignment = Alignment(horizontal="left", vertical="center")
    
    ws1.merge_cells("C4:I4")
    ws1["C4"] = "Hotline: 0869 590 279  |  Email: cokhisaovangvn@gmail.com"
    ws1["C4"].font = font_company_info
    ws1["C4"].alignment = Alignment(horizontal="left", vertical="center")

    # Add Logo image in A2
    try:
        logo_path = "d:/Sao Vàng/Website-SaoVang/Logo/LogoCTY + SV Aluminium/LogoCTY.png"
        img = OpenpyxlImage(logo_path)
        img.width = 110
        img.height = 45
        ws1.add_image(img, "A2")
    except Exception as e:
        print(f"Error adding logo to sheet 1: {e}")

    # Title Block (shifted to Row 6-7)
    ws1.row_dimensions[6].height = 24
    ws1.merge_cells("A6:I6")
    ws1["A6"] = "DANH SÁCH THÉP U & THÉP HỘP CẦN KIỂM KHO DỰ ÁN MONACO"
    ws1["A6"].font = font_title_main
    ws1["A6"].alignment = align_center
    
    ws1.row_dimensions[7].height = 18
    ws1.merge_cells("A7:I7")
    ws1["A7"] = "Báo cáo tổng hợp số lượng cây quy chuẩn phục vụ công tác chuẩn bị vật tư hiện trường"
    ws1["A7"].font = font_title_sub
    ws1["A7"].alignment = align_center

    # Metadata (shifted to Row 9)
    ws1["B9"] = "Dự án:"
    ws1["B9"].font = font_bold
    ws1["C9"] = "Nối mái kính MONACO Hạ Long"
    ws1["C9"].font = font_normal
    ws1["C9"].border = Border(bottom=Side(style="thin", color="000000"))

    ws1["E9"] = "Ngày lập:"
    ws1["E9"].font = font_bold
    ws1["F9"] = "15/07/2026"
    ws1["F9"].font = font_normal
    ws1["F9"].border = Border(bottom=Side(style="thin", color="000000"))

    ws1["H9"] = "Người đề xuất:"
    ws1["H9"].font = font_bold
    ws1["I9"] = "Tổ kỹ thuật / Thiết kế"
    ws1["I9"].font = font_normal
    ws1["I9"].border = Border(bottom=Side(style="thin", color="000000"))

    ws1.row_dimensions[10].height = 15

    # Header row (shifted to Row 11)
    ws1.row_dimensions[11].height = 28
    headers_ws1 = [
        ("A11", "STT", align_center),
        ("B11", "LOẠI VẬT TƯ", align_left),
        ("C11", "QUY CÁCH CHI TIẾT", align_left),
        ("D11", "CHIỀU DÀI TK (M)", align_center),
        ("E11", "TỶ TRỌNG (KG/M)", align_center),
        ("F11", "SỐ LƯỢNG YÊU CẦU", align_center),
        ("G11", "ĐƠN VỊ TÍNH", align_center),
        ("H11", "MỤC ĐÍCH SỬ DỤNG TRONG MÁI KÍNH", align_left),
        ("I11", "TRẠNG THÁI KHO (CÒN / THIẾU)", align_center)
    ]
    for cell_ref, text, align in headers_ws1:
        cell = ws1[cell_ref]
        cell.value = text
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align
        cell.border = border_all_thin

    # Materials (starts at row 12)
    materials_data = [
        ("1", "Thép U đúc SS400", "U100 x 200 x 100 x 6 mm", 19.8, 18.7, 4, "Cây 6m", "Chế tạo hệ khung chính nối mái kính", "[  ] Còn tồn  /  [  ] Cần mua"),
        ("2", "Thép U chấn SS400", "200 x 100 x 4 mm (đáy 200, cánh 100)", 5.3, 15.7, 1, "Cây 6m", "Khung liên kết phụ đầu mái kính", "[  ] Còn tồn  /  [  ] Cần mua"),
        ("3", "Thép hộp vuông", "SS400 100 x 100 x 3 mm", 14.8, 9.14, 3, "Cây 6m", "Lắp ráp hệ xương phụ đỡ kính", "[  ] Còn tồn  /  [  ] Cần mua"),
        ("4", "Thép hộp chữ nhật", "SS400 200 x 100 x 3 mm", 4.93, 13.85, 1, "Cây 6m", "Dựng cột chịu lực chính cho mái", "[  ] Còn tồn  /  [  ] Cần mua")
    ]

    for idx, row in enumerate(materials_data, start=12):
        ws1.row_dimensions[idx].height = 36
        ws1.cell(idx, 1, row[0]).alignment = align_center
        ws1.cell(idx, 2, row[1]).font = font_bold
        ws1.cell(idx, 3, row[2]).font = font_bold
        
        ws1.cell(idx, 4, row[3]).alignment = align_right
        ws1.cell(idx, 4).number_format = '0.00'
        
        ws1.cell(idx, 5, row[4]).alignment = align_right
        ws1.cell(idx, 5).number_format = '0.00'
        
        ws1.cell(idx, 6, row[5]).font = font_bold
        ws1.cell(idx, 6).alignment = align_center
        
        ws1.cell(idx, 7, row[6]).alignment = align_center
        ws1.cell(idx, 8, row[7]).alignment = align_left
        
        cell_status = ws1.cell(idx, 9, row[8])
        cell_status.alignment = align_center
        cell_status.fill = fill_input
        cell_status.font = font_italic
        
        for col in range(1, 10):
            cell = ws1.cell(idx, col)
            cell.border = border_all_thin
            if not cell.fill.start_color.rgb:
                cell.fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid") if idx % 2 == 0 else PatternFill(fill_type=None)

    note_row = 17
    ws1.merge_cells(f"A{note_row}:I{note_row}")
    ws1[f"A{note_row}"] = "Ghi chú xưởng: Bộ phận kho tích chọn trạng thái của từng loại thép trên để phòng Vật tư chủ động đặt mua các loại còn thiếu."
    ws1[f"A{note_row}"].font = font_italic
    ws1[f"A{note_row}"].alignment = align_left

    # -------------------------------------------------------------
    # SHEET 2: BẢNG DỰ TOÁN CHI TIẾT MONACO
    # -------------------------------------------------------------
    ws2 = wb.create_sheet(title="Dự Toán Chi Tiết")
    ws2.sheet_view.showGridLines = True
    
    ws2.column_dimensions['A'].width = 6   # STT
    ws2.column_dimensions['B'].width = 38  # Tên vật tư
    ws2.column_dimensions['C'].width = 18  # Chiều dài thiết kế
    ws2.column_dimensions['D'].width = 22  # Trọng lượng đơn vị
    ws2.column_dimensions['E'].width = 18  # Quy cách cấp
    ws2.column_dimensions['F'].width = 25  # Diễn giải
    ws2.column_dimensions['G'].width = 10  # Đơn vị tính
    ws2.column_dimensions['H'].width = 18  # Khối lượng thanh toán
    ws2.column_dimensions['I'].width = 18  # Đơn giá
    ws2.column_dimensions['J'].width = 22  # Thành tiền
    
    # Company Letterhead Row 2-4
    ws2.row_dimensions[2].height = 20
    ws2.row_dimensions[3].height = 15
    ws2.row_dimensions[4].height = 15
    
    ws2.merge_cells("C2:J2")
    ws2["C2"] = "CÔNG TY CỔ PHẦN SẢN XUẤT CƠ KHÍ SAO VÀNG (ASV)"
    ws2["C2"].font = font_company_name
    ws2["C2"].alignment = Alignment(horizontal="left", vertical="center")
    
    ws2.merge_cells("C3:J3")
    ws2["C3"] = "Địa chỉ: Tầng 3, TT7-35 Khu đô thị Văn Phú, phường Kiến Hưng, TP Hà Nội, Việt Nam."
    ws2["C3"].font = font_company_info
    ws2["C3"].alignment = Alignment(horizontal="left", vertical="center")
    
    ws2.merge_cells("C4:J4")
    ws2["C4"] = "Hotline: 0869 590 279  |  Email: cokhisaovangvn@gmail.com"
    ws2["C4"].font = font_company_info
    ws2["C4"].alignment = Alignment(horizontal="left", vertical="center")

    # Add Logo image in A2
    try:
        logo_path = "d:/Sao Vàng/Website-SaoVang/Logo/LogoCTY + SV Aluminium/LogoCTY.png"
        img = OpenpyxlImage(logo_path)
        img.width = 110
        img.height = 45
        ws2.add_image(img, "A2")
    except Exception as e:
        print(f"Error adding logo to sheet 2: {e}")

    # Title Block (shifted to Row 6-7)
    ws2.row_dimensions[6].height = 24
    ws2.merge_cells("A6:J6")
    ws2["A6"] = "BẢNG DỰ TOÁN KHỐI LƯỢNG VÀ CHI PHÍ THỰC TẾ"
    ws2["A6"].font = Font(name="Segoe UI", size=14, bold=True, color=STEEL_BLUE)
    ws2["A6"].alignment = align_center

    ws2.row_dimensions[7].height = 18
    ws2.merge_cells("A7:J7")
    ws2["A7"] = "Hạng mục: Nối mái kính MONACO (Diện tích: 29 m2)"
    ws2["A7"].font = font_title_sub
    ws2["A7"].alignment = align_center

    # Table Header (shifted to Row 9)
    headers_ws2 = [
        ("A9", "STT", align_center),
        ("B9", "TÊN VẬT TƯ / QUY CÁCH CHẤT LIỆU", align_left),
        ("C9", "C.DÀI T.KẾ (m/m2)", align_center),
        ("D9", "TRỌNG LƯỢNG Đ.VỊ", align_center),
        ("E9", "QUY CÁCH CẤP", align_center),
        ("F9", "DIỄN GIẢI KHỐI LƯỢNG", align_left),
        ("G9", "ĐVT", align_center),
        ("H9", "K.LƯỢNG T.TOÁN", align_center),
        ("I9", "ĐƠN GIÁ (đ)", align_center),
        ("J9", "THÀNH TIỀN (đ)", align_center)
    ]
    for cell_ref, text, align in headers_ws2:
        cell = ws2[cell_ref]
        cell.value = text
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align
        cell.border = border_all_thin
    ws2.row_dimensions[9].height = 24

    # Data Rows (starts at row 10)
    estimate_data = [
        ("1", "Kính CL dán 13.52mm", "29 m2", "Trung bình 2.37m2/tấm", "12 tấm", "Khối lượng tính thành tiền = Diện tích m2 kính", "m2", 29.0, 750000),
        ("2", "SS400 U100 x 200 x 100 x 6 mm", "19.8m", 18.7, "4 (cây/6m)", "18.70 x 24m = 448.80 kg", "kg", "=18.7*24", 27500),
        ("3", "SS400 U chấn 200 x 100 x 4 mm (đáy 200, cánh 100, dày 4)", "5.3m", 15.7, "1 (cây/6m)", "15.70 x 6m = 94.20 kg", "kg", "=15.7*6", 27500),
        ("4", "Xương phụ SS400 100x100x3mm", "14.8m", 9.14, "3 (cây/6m)", "9.14 x 18m = 164.52 kg", "kg", "=9.14*18", 20000),
        ("5", "Cột SS400 200x100x3mm", "4.93m", 13.85, "1 (cây/6m)", "13.85 x 6m = 83.10 kg", "kg", "=13.85*6", 20000),
        ("6", "Bản mã cột SS400 240x300x8mm", "3 cái", 4.52, "3 cái", "3 cái x 4.52 kg = 13.56 kg", "kg", "=4.52*3", 25000),
        ("7", "Bản mã nối khung chính SS400 200x200x8mm", "2 cái", 2.52, "2 cái", "2 cái x 2.52 kg = 5.04 kg", "kg", "=2.52*2", 25000),
        ("8", "V úp che chân cột 310x250x40x1.2 mm", "6 cái", 1.5, "6 cái", "6 cái x 1.5 kg = 9.00 kg", "kg", "=1.5*6", 35000),
        ("9", "Máng nước inox 1.2mm", "12m (gồm 5.3m mái mới và mái cũ)", "", "12m", "Đo lại theo chiều dài thực tế thi công ngoài công trường", "md", 12.0, 590000),
        ("10", "Sơn Chống rỉ", "Diện tích sơn 35m2", "", "2kg", "Định mức phủ chống rỉ bề mặt kết cấu", "kg", 2.0, 350000),
        ("11", "Sơn Hoa Việt", "Diện tích sơn 35m2", "", "4kg", "Sơn phủ màu hoàn thiện kết cấu sắt thép", "kg", 4.0, 350000),
        ("12", "Dung môi", "", "", "2kg", "Dung môi pha chế sơn đạt độ nhớt tiêu chuẩn", "kg", 2.0, 50000),
        ("13", "Kẹp giữ kính inox 304", "", "", "20 cái", "Kẹp cố định tấm kính vào hệ khung xương", "Cái", 20.0, 10000),
        ("14", "Vít inox khoan giữ kính", "", "", "40 cái", "Vít gia cố các điểm nối kẹp kính", "Cái", 40.0, 5000),
        ("15", "Buông zamset M16x160", "", "", "20 cái", "Bulong hóa chất cấy liên kết dầm bê tông", "Cái", 20.0, 185000),
        ("16", "Que hàn + dây CO2", "", "", "1.00 gói", "Vật tư phụ phục vụ công tác hàn liên kết", "Gói", 1.0, 500000),
        ("17", "Đá cắt + đá mài", "", "", "1.00 gói", "Đá cắt phôi và đá mài bavia hoàn thiện", "Gói", 1.0, 250000),
        ("18", "Silicone + Keo", "", "", "1.00 gói", "Keo kết cấu chống thấm mép kính và máng nước", "Gói", 1.0, 1200000),
        ("19", "Bulong + phụ kiện lắp", "", "", "1.00 gói", "Các loại bulong cường độ cao và long đen phụ", "Gói", 1.0, 100000)
    ]

    r_start = 10
    for idx, row in enumerate(estimate_data, start=r_start):
        ws2.row_dimensions[idx].height = 24
        
        ws2.cell(idx, 1, row[0]).alignment = align_center
        ws2.cell(idx, 2, row[1]).alignment = align_left
        ws2.cell(idx, 2).font = font_bold if idx in [11, 12, 13, 14] else font_normal
        
        ws2.cell(idx, 3, row[2]).alignment = align_center
        ws2.cell(idx, 4, row[3]).alignment = align_center if isinstance(row[3], str) else align_right
        if isinstance(row[3], (int, float)):
            ws2.cell(idx, 4).number_format = '0.00'
            
        ws2.cell(idx, 5, row[4]).alignment = align_center
        ws2.cell(idx, 6, row[5]).alignment = align_left
        ws2.cell(idx, 7, row[6]).alignment = align_center
        
        # Quantity column H
        cell_kl = ws2.cell(idx, 8, row[7])
        cell_kl.alignment = align_right
        if isinstance(row[7], (int, float)):
            cell_kl.number_format = '0.00'
            
        # Price column I
        cell_price = ws2.cell(idx, 9, row[8])
        cell_price.alignment = align_right
        cell_price.number_format = '#,##0'
        cell_price.fill = fill_input
        
        # Subtotal column J
        cell_subtotal = ws2.cell(idx, 10, f"=H{idx}*I{idx}")
        cell_subtotal.alignment = align_right
        cell_subtotal.number_format = '#,##0" đ"'
        cell_subtotal.font = font_bold
        
        for col in range(1, 11):
            ws2.cell(idx, col).border = border_all_thin
            if idx % 2 == 0:
                ws2.cell(idx, col).fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

    # Insert Group Header row at Row 10 (Yellow Accent like in screenshot)
    ws2.insert_rows(10, 1)
    ws2.row_dimensions[10].height = 24
    ws2.merge_cells("B10:F10")
    ws2["B10"] = "I. Nối mái kính MONACO"
    ws2["B10"].font = font_group_header
    ws2["B10"].alignment = align_left
    
    ws2["G10"] = "m2"
    ws2["G10"].font = font_bold
    ws2["G10"].alignment = align_center
    
    ws2["H10"] = 29.0
    ws2["H10"].font = font_bold
    ws2["H10"].alignment = align_right
    ws2["H10"].number_format = '0.00'
    
    for col in range(1, 11):
        cell = ws2.cell(10, col)
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        cell.border = border_all_thin
        
    # Re-adjust formulas in column J due to inserted row:
    # Now our data rows are row 11 to 29.
    for r in range(11, 30):
        ws2.cell(r, 10, f"=H{r}*I{r}")
        
    # Total row at bottom (Row 30)
    total_row = 30
    ws2.row_dimensions[total_row].height = 28
    ws2.merge_cells(f"A{total_row}:I{total_row}")
    ws2.cell(total_row, 1, "TỔNG CHI PHÍ VẬT TƯ THỰC TẾ DỰ TOÁN:").font = font_bold
    ws2.cell(total_row, 1).alignment = align_right
    
    cell_total_sum = ws2.cell(total_row, 10, f"=SUM(J11:J29)")
    cell_total_sum.font = font_total
    cell_total_sum.alignment = align_right
    cell_total_sum.number_format = '#,##0" đ"'
    
    for col in range(1, 11):
        cell = ws2.cell(total_row, col)
        cell.fill = fill_total
        cell.border = border_double_bottom

    # Adjust widths
    ws1.column_dimensions['A'].width = 6
    ws1.column_dimensions['B'].width = 24
    ws1.column_dimensions['C'].width = 38
    ws1.column_dimensions['D'].width = 18
    ws1.column_dimensions['E'].width = 18
    ws1.column_dimensions['F'].width = 18
    ws1.column_dimensions['G'].width = 14
    ws1.column_dimensions['H'].width = 42
    ws1.column_dimensions['I'].width = 28

    ws2.column_dimensions['A'].width = 6
    ws2.column_dimensions['B'].width = 48
    ws2.column_dimensions['C'].width = 18
    ws2.column_dimensions['D'].width = 18
    ws2.column_dimensions['E'].width = 18
    ws2.column_dimensions['F'].width = 38
    ws2.column_dimensions['G'].width = 10
    ws2.column_dimensions['H'].width = 18
    ws2.column_dimensions['I'].width = 18
    ws2.column_dimensions['J'].width = 24

    paths = [
        "d:/Sao Vàng/Website-SaoVang/CKSV_Du_Toan_Mai_Kinh_Monaco.xlsx",
        "d:/Sao Vàng/Website-SaoVang/CKSV_Du_Toan_Mai_Kinh_Monaco_Co_Logo.xlsx"
    ]
    saved_paths = []
    for path in paths:
        try:
            wb.save(path)
            saved_paths.append(path)
        except PermissionError:
            print(f"Permission denied for {path}. Skipping.")
            
    print(f"Monaco estimate workbook with logo saved to: {saved_paths}")

if __name__ == "__main__":
    generate_monaco_estimate_with_logo()
