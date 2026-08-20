import os
import openpyxl

def search_text_in_all_xlsx():
    base_dir = "d:/Sao Vàng/Website-SaoVang"
    query = "monaco"
    found = []
    
    for root, dirs, files in os.walk(base_dir):
        if ".git" in root or "node_modules" in root:
            continue
        for file in files:
            if file.endswith((".xlsx", ".xlsm")):
                path = os.path.join(root, file)
                try:
                    wb = openpyxl.load_workbook(path, data_only=True)
                    for sheetname in wb.sheetnames:
                        ws = wb[sheetname]
                        for r in range(1, ws.max_row + 1):
                            for c in range(1, ws.max_column + 1):
                                val = ws.cell(r, c).value
                                if val and query in str(val).lower():
                                    found.append((path, sheetname, r, c, val))
                except Exception as e:
                    pass
                    
    with open("d:/Sao Vàng/Website-SaoVang/scripts/monaco_search_results.txt", "w", encoding="utf-8") as f:
        for path, sheet, r, c, val in found:
            f.write(f"File: {path}\nSheet: {sheet}\nCell: {openpyxl.utils.get_column_letter(c)}{r}\nValue: {val}\n{'-'*60}\n")
    print(f"Done, found {len(found)} instances")

if __name__ == "__main__":
    search_text_in_all_xlsx()
