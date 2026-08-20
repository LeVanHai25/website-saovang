import os
import openpyxl

def inspect_xlsx_files():
    base_dir = "d:/Sao Vàng/Website-SaoVang/CKSV"
    sub_dir = os.path.join(base_dir, "bảng duyển dụng")
    
    files = []
    # Scan root directory
    for f in os.listdir(base_dir):
        if f.endswith(".xlsx"):
            files.append(os.path.join(base_dir, f))
            
    # Scan sub directory
    if os.path.exists(sub_dir):
        for f in os.listdir(sub_dir):
            if f.endswith(".xlsx"):
                files.append(os.path.join(sub_dir, f))
                
    output_lines = []
    output_lines.append(f"Found {len(files)} .xlsx files to inspect:")
    for filepath in files:
        output_lines.append("\n" + "="*80)
        output_lines.append(f"FILE: {os.path.basename(filepath)}")
        output_lines.append("="*80)
        try:
            wb = openpyxl.load_workbook(filepath, read_only=True)
            output_lines.append(f"Sheets: {wb.sheetnames}")
            for name in wb.sheetnames:
                ws = wb[name]
                output_lines.append(f"\n  Sheet: {name}")
                rows = list(ws.iter_rows(max_row=20, max_col=10, values_only=True))
                for idx, r in enumerate(rows):
                    if any(x is not None for x in r):
                        # Filter out empty trailing elements in row
                        clean_row = []
                        for val in r:
                            if val is not None:
                                clean_row.append(str(val))
                            else:
                                clean_row.append("")
                        # remove empty cells from the right side of the row to keep it short
                        while clean_row and clean_row[-1] == "":
                            clean_row.pop()
                        if clean_row:
                            output_lines.append(f"    Row {idx+1}: {clean_row}")
        except Exception as e:
            output_lines.append(f"Error reading {os.path.basename(filepath)}: {e}")
            
    # Write to a text file using utf-8 encoding
    output_path = "d:/Sao Vàng/Website-SaoVang/scripts/inspect_results.txt"
    with open(output_path, "w", encoding="utf-8") as f_out:
        f_out.write("\n".join(output_lines))
    print(f"Results written to {output_path}")

if __name__ == "__main__":
    inspect_xlsx_files()
