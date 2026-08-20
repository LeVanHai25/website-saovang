import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

def create_exact_6col_recruitment_form():
    wb = openpyxl.Workbook()
    
    # -------------------------------------------------------------
    # SHEET 1: ĐÁNH GIÁ TUYỂN DỤNG (6-COLUMN PREMIUM FORMAT)
    # -------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Đánh Giá Tuyển Dụng"
    ws1.sheet_view.showGridLines = True
    
    # Theme: Steel Blue
    STEEL_BLUE = "2F5597"
    ICE_BLUE = "D9E1F2"
    LIGHT_GRAY = "F2F2F2"
    BORDER_GRAY = "D9D9D9"
    YELLOW_INPUT = "FFF2CC"
    DARK_GREEN = "385723"
    LIGHT_GREEN = "E2EFDA"
    
    # Fonts
    font_title_company = Font(name="Segoe UI", size=11, bold=True, color="595959")
    font_title_main = Font(name="Segoe UI", size=16, bold=True, color=STEEL_BLUE)
    font_title_sub = Font(name="Segoe UI", size=10, italic=True, color="7F7F7F")
    font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    font_group_header = Font(name="Segoe UI", size=11, bold=True, color=STEEL_BLUE)
    font_bold = Font(name="Segoe UI", size=10, bold=True)
    font_normal = Font(name="Segoe UI", size=10)
    font_italic = Font(name="Segoe UI", size=9, italic=True)
    font_salary_result = Font(name="Segoe UI", size=12, bold=True, color=DARK_GREEN)
    
    # Fills
    fill_header = PatternFill(start_color=STEEL_BLUE, end_color=STEEL_BLUE, fill_type="solid")
    fill_group = PatternFill(start_color=ICE_BLUE, end_color=ICE_BLUE, fill_type="solid")
    fill_sum = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
    fill_input = PatternFill(start_color=YELLOW_INPUT, end_color=YELLOW_INPUT, fill_type="solid")
    fill_result_box = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
    
    # Alignments
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    
    # Borders
    thin_side = Side(style="thin", color=BORDER_GRAY)
    border_all_thin = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    border_group = Border(
        left=Side(style="thin", color=STEEL_BLUE),
        right=Side(style="thin", color=STEEL_BLUE),
        top=Side(style="thin", color=STEEL_BLUE),
        bottom=Side(style="thin", color=STEEL_BLUE)
    )
    
    # Column widths (6 Columns layout from screenshot)
    ws1.column_dimensions['A'].width = 6   # STT
    ws1.column_dimensions['B'].width = 38  # NHIỆM VỤ ĐÁNH GIÁ (VIỆC CẦN LÀM)
    ws1.column_dimensions['C'].width = 58  # YÊU CẦU KỸ NĂNG VÀ KẾT QUẢ BÊN SDLĐ CẦN ĐẠT ĐƯỢC
    ws1.column_dimensions['D'].width = 16  # ỨNG VIÊN TỰ ĐÁNH GIÁ (% Đáp ứng)
    ws1.column_dimensions['E'].width = 16  # CÔNG TY ĐÁNH GIÁ (Đạt: 1, Chưa: 0)
    ws1.column_dimensions['F'].width = 25  # GHI CHÚ / NHẬN XÉT CHI TIẾT

    # Title Block
    ws1.merge_cells("A2:F2")
    ws1["A2"] = "CÔNG TY CỔ PHẦN SẢN XUẤT CƠ KHÍ SAO VÀNG (ASV)"
    ws1["A2"].font = font_title_company
    ws1["A2"].alignment = Alignment(horizontal="left")

    ws1.merge_cells("A3:F3")
    ws1["A3"] = "FORM TUYỂN DỤNG VÀ ĐÁNH GIÁ NĂNG LỰC THỬ VIỆC THỢ CƠ KHÍ"
    ws1["A3"].font = font_title_main
    ws1["A3"].alignment = align_center

    ws1.merge_cells("A4:F4")
    ws1["A4"] = "Xác định ngạch bậc & lương ngày dựa trên kỹ năng thực tế của ứng viên"
    ws1["A4"].font = font_title_sub
    ws1["A4"].alignment = align_center

    # Candidate Profile Info
    ws1["B6"] = "Họ và tên ứng viên:"
    ws1["B6"].font = font_bold
    ws1["C6"] = "[Nhập tên ứng viên]"
    ws1["C6"].font = font_normal
    ws1["C6"].border = Border(bottom=Side(style="thin", color="000000"))

    ws1["E6"] = "Ngày phỏng vấn:"
    ws1["E6"].font = font_bold
    ws1["F6"] = "[Nhập ngày]"
    ws1["F6"].font = font_normal
    ws1["F6"].border = Border(bottom=Side(style="thin", color="000000"))

    ws1["B7"] = "Số điện thoại/CCCD:"
    ws1["B7"].font = font_bold
    ws1["C7"] = "[Số điện thoại và CCCD]"
    ws1["C7"].font = font_normal
    ws1["C7"].border = Border(bottom=Side(style="thin", color="000000"))

    ws1["E7"] = "Người chấm điểm:"
    ws1["E7"].font = font_bold
    ws1["F7"] = "[Tên Quản đốc/Tổ trưởng]"
    ws1["F7"].font = font_normal
    ws1["F7"].border = Border(bottom=Side(style="thin", color="000000"))

    ws1["B8"] = "Vị trí ứng tuyển:"
    ws1["B8"].font = font_bold
    ws1["C8"] = "Thợ Cơ khí Dân dụng (Học việc/Phụ/Giúp việc/Thợ chính/Tổ trưởng)"
    ws1["C8"].font = font_italic
    ws1["C8"].border = Border(bottom=Side(style="thin", color="000000"))

    ws1["E8"] = "Kỳ thử việc:"
    ws1["E8"].font = font_bold
    ws1["F8"] = "01 tháng thử việc"
    ws1["F8"].font = font_normal
    ws1["F8"].border = Border(bottom=Side(style="thin", color="000000"))

    ws1.row_dimensions[9].height = 15

    # Table Header
    headers = [
        ("A10", "STT", align_center),
        ("B10", "NHIỆM VỤ ĐÁNH GIÁ (VIỆC CẦN LÀM)", align_left),
        ("C10", "YÊU CẦU KỸ NĂNG VÀ KẾT QUẢ BÊN SDLĐ CẦN ĐẠT ĐƯỢC", align_left),
        ("D10", "ỨNG VIÊN TỰ ĐÁNH GIÁ (% Đáp ứng)", align_center),
        ("E10", "CÔNG TY ĐÁNH GIÁ (Đạt: 1, Chưa: 0)", align_center),
        ("F10", "GHI CHÚ / NHẬN XÉT CHI TIẾT", align_left)
    ]
    for cell_ref, text, align in headers:
        cell = ws1[cell_ref]
        cell.value = text
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align
        cell.border = border_all_thin
    ws1.row_dimensions[10].height = 28

    # Helper function to write merged headers
    def merge_and_style_header(ws, r_num, title):
        ws.merge_cells(start_row=r_num, start_column=1, end_row=r_num, end_column=6)
        ws.cell(r_num, 1, title).font = font_group_header
        ws.cell(r_num, 1).alignment = align_left
        for col in range(1, 7):
            cell = ws.cell(r_num, col)
            cell.fill = fill_group
            cell.border = border_group

    # Exact Tasks details
    exact_tasks = [
        # (Group Header Title) OR (STT, Task, Skill/Outcome)
        ("G", "I. HỌC VIỆC VÀ PHỤ VIỆC (Tiêu chuẩn Bậc 1 - Bậc 5)"),
        ("1", "_Sức khỏe tốt, nhanh nhẹn, chịu khó học hỏi, có trách nhiệm, phối hợp được với các vị trí trong xưởng", "Sức khỏe dẻo dai, bền bỉ; chấp hành tốt nội quy; làm việc hòa đồng với đồng đội."),
        ("2", "Chủ động thu dọn máy móc dụng cụ làm việc, quét dọn vệ sinh khu vực thi công sau mỗi buổi làm và khi kết thúc công việc ở xưởng cũng như đi công trường", "Giữ gìn vệ sinh chung, thực hành tốt 5S tại khu vực phân công trước và sau ca làm."),
        ("3a", "_Biết sử dụng các loại máy móc và dụng cụ cơ bản trong xưởng sản xuất: máy cắt tay, máy cắt bàn, máy khoan, thước đo khoảng cách, dây điện…vv", "Sử dụng đúng kỹ thuật, an toàn các dụng cụ cơ bản phục vụ công việc hàng ngày."),
        ("3b", "Biết tư duy tình huống để có biện pháp nhắc nhở, bảo vệ an toàn cho bản thân và những người xung quanh, cũng như tài sản trong khu vực thi công", "Có ý thức phòng ngừa rủi ro mất an toàn; cảnh báo đồng nghiệp khi thấy nguy hiểm."),
        ("3c", "Có thể sử dụng máy mài để mài hoàn thiện sản phẩm theo yêu cầu về độ tinh mỹ của sản phẩm sắt đen, sắt mạ, inox và vệ sinh bề mặt sản phẩm sau khi hoàn thành phần thô (mài bề mặt, mài góc, đánh bóng, đánh rỉ...)", "Vệ sinh sạch bavia mối hàn kết cấu cơ bản phẳng phiu."),
        ("4a", "Mài tinh mỹ (yêu cầu cho sp có độ thẩm mỹ cao, cần mài lấy mặt phẳng tuyệt đối hoặc mài lấy góc sắc nét nếu chi tiết được gá, hàn đạt tiêu chuẩn)", "Mài đạt độ phẳng tuyệt đối, góc vuông sắc cạnh cho sản phẩm trưng bày, trang trí."),
        ("4b", "Mài lấy góc (yêu cầu cho các sản phẩm được tổ hợp từ nguyên liệu tấm, hình vv, tổ hợp thành sản phẩm theo thiết kế)", "Mài tổ hợp các góc ghép của phôi tấm mỏng chính xác theo thiết kế bản vẽ."),
        ("4c", "Mài phá (yêu cầu cho các vị trí mối hàn trên các bề mặt vật liệu thô, kết cấu hạng nặng, hoặc các mối hàn không đạt tiêu chuẩn về thẩm mỹ, cần tinh chỉnh lại)", "Sử dụng đĩa mài phá xử lý các mối hàn thô của kết cấu chịu lực, dầm nặng."),
        ("4d", "Mài đánh bóng (yêu cầu cho các sản phẩm sử dụng nguyên liệu inox bóng, hoặc cần làm bóng, mịn cho sp ngoài inox)", "Đánh bóng inox đạt độ gương sáng loáng, mịn, không còn vết xước bề mặt."),
        ("4e", "Biết tinh chỉnh lại các vật dụng phục vụ cho cv (mài mũi khoan)", "Tự mài sắc lại mũi khoan thép bén ngọt, mài các dụng cụ gá đặt gọn gàng."),
        ("5a", "Sử dụng thành thạo súng phun sơn khí nén, sơn phủ và sơn hoàn thiện sản phẩm sau quá trình hàn, mài tinh mỹ", "Sơn lót chống rỉ và sơn màu phủ đều tay, bề mặt bóng mịn, không chảy sơn, không bọt khí."),
        ("5b", "Nắm rõ cơ bản các loại vật tư, chủng loại vật liệu", "Nhận biết chính xác các độ dày tấm, ống sắt hộp, hộp mạ kẽm, inox 201, 304."),
        ("5c", "Sử dụng được các loại máy cơ bản trong công việc: máy hàn, máy cắt tay, máy cắt bàn, máy khoan, máy cân bằng laze, thước đo khoảng cách...vv", "Sử dụng linh hoạt các máy móc đo đạc định vị để thực hiện công việc được giao."),
        ("5d", "Nhận biết được đơn vị đo lường, có thể cắt phôi chính xác theo kích thước được giao, biết cách tính toán làm dưỡng chuẩn để cắt hàng loạt nếu vật liệu cần cắt cùng một kích thước và với số lượng trên 2", "Đọc số đo thước mét chính xác từng mm; làm cữ cắt loạt phôi tiết kiệm vật tư."),
        ("5e", "Có thể sử dụng máy hàn để hàn đính gá chuẩn các vị trí cần thiết, khi được giao việc", "Hàn đính gá kết cấu vững chãi, chính xác theo cữ lắp dựng."),
        
        ("G", "II. GIÚP VIỆC (Tiêu chuẩn Bậc 6 - Bậc 7)"),
        ("6", "Nắm bắt và hiểu rõ hết công năng cho từng loại máy móc, chủ động tính toán chuẩn bị đầy đủ vt phụ, máy móc dùng cho công trình khi đã được thông báo về nội dung các hạng mục công việc của công trình sắp triển khai", "Chủ động chuẩn bị đầy đủ dụng cụ đi lắp dựng công trình xa, không để thiếu đồ nghề."),
        ("7a", "Sử dụng thành thạo máy hàn que", "Hàn que (MMA) điêu luyện, ngấu sâu, chịu lực cao, mối hàn đều vảy."),
        ("7b", "Sử dụng thành thạo máy hàn mic", "Hàn MIG/MAG đều tay, không bọt khí, tốc độ đắp nhanh trên kết cấu sắt thép."),
        ("7c", "Sử dụng thành thạo máy hàn tic", "Hàn TIG Inox trắng sáng, mối hàn kết vảy cá đều đẹp, không bị đen mối hàn."),
        ("7d", "Biết sử dụng cơ bản các loại máy và phụ kiện đi kèm trong nghề cơ khí dân dụng, xây dựng", "Thao tác thuần thục tất cả máy chuyên dụng trong ngành cơ khí dân dụng."),
        
        ("G", "III. THỢ CHÍNH - TỰ THỰC HIỆN (Tiêu chuẩn Bậc 8 - Bậc 10)"),
        ("8", "Đọc hiểu bản vẽ sản xuất và có thể triển khai công việc trên bản vẽ khi được giao, biết cách tính toán khai triển công việc từ bản vẽ thành sp thực tế một cách khoa học, nhanh và hiệu quả nhất, nghiên cứu, đóng góp ý kiến kịp thời nếu thấy có phương án thi công hiệu quả, năng suất hơn", "Tự triển khai sản phẩm chuẩn từ bản vẽ kỹ thuật 2D/3D mà không cần hướng dẫn."),
        ("25", "Kỹ năng hoàn thiện và đánh giá sản phẩm sau khi hoàn thiện trên tinh thần khách quan, không giấu dốt, luôn luôn nâng cấp, hoàn thiện bản thân trong công việc", "Tự kiểm định chất lượng thẩm mỹ trước khi QC kiểm tra; nâng cao tay nghề định kỳ."),
        ("9a", "Khả năng đánh giá, ước lượng số nhân công và khoảng thời gian hoàn thành sản phẩm hoặc công trình, Chịu trách nhiệm cho những phần việc được giao khi đã được hướng dẫn cụ thể", "Dự toán chuẩn tiến độ và lượng công thợ cần thiết để hoàn thành hạng mục."),
        ("9b", "Báo cáo tình hình tiến độ công việc vào cuối ngày, và trao đổi ngay với những phần việc phát sinh hoặc đề xuất các phương án thi công tối ưu để các bộ phận khác nắm rõ thông tin", "Thông tin tiến độ thông suốt, đề xuất giải pháp thi công tại công trình kịp thời."),
        ("10", "Có thể độc lập trong công việc khi giao, có thể tự quyết định điều chỉnh cách thức làm hoặc chủng loại vật tư mà không ảnh hưởng đến cấu trúc tổng thể của sản phẩm, đồng thời tăng tính thẩm mỹ, hiệu quả", "Tự giải quyết sự cố kỹ thuật nhỏ ngoài thực địa, nâng cao năng suất chung."),
        
        ("G", "IV. TỔ TRƯỞNG VÀ ĐIỀU HÀNH SẢN XUẤT (Tiêu chuẩn Bậc 11 - Bậc 14)"),
        ("11", "Khả năng làm việc nhóm, đảm bảo sản xuất các sản phẩm đúng thiết kế, đúng tiến độ", "Phối hợp điều phối nhóm thợ hoàn thành cụm kết cấu lớn nhịp nhàng, đúng tiến độ."),
        ("12", "Khả năng lãnh đạo đội nhóm, bao quát, điều phối, bố trí công việc phù hợp cho từng thành viên trong nhóm khi được giao nhiệm vụ lãnh đạo tổ đội thi công các hạng mục công việc đã trao đổi", "Phân chia công việc hợp lý; giám sát chất lượng và tiến độ toàn đội sản xuất.")
    ]

    current_row = 11
    eval_rows = []
    
    # Mapping to evaluate wages based on company data row-by-row
    row_wage_map = {}
    
    for row_data in exact_tasks:
        if row_data[0] == "G":
            ws1.row_dimensions[current_row].height = 24
            merge_and_style_header(ws1, current_row, row_data[1])
        else:
            stt, task, skill = row_data
            ws1.row_dimensions[current_row].height = 24
            
            ws1.cell(current_row, 1, stt).font = font_normal
            ws1.cell(current_row, 1).alignment = align_center
            
            ws1.cell(current_row, 2, task).font = font_normal
            ws1.cell(current_row, 2).alignment = align_left
            
            ws1.cell(current_row, 3, skill).font = font_normal
            ws1.cell(current_row, 3).alignment = align_left
            
            # Candidate Self Eval (%) - Input
            cell_self = ws1.cell(current_row, 4, 1.0)
            cell_self.font = font_bold
            cell_self.alignment = align_center
            cell_self.fill = fill_input
            cell_self.number_format = '0%'
            
            # Company Eval (1: Đạt, 0: Chưa) - Input
            cell_company = ws1.cell(current_row, 5, 0)
            cell_company.font = font_bold
            cell_company.alignment = align_center
            cell_company.fill = fill_input
            eval_rows.append(current_row)
            
            # Note - Input
            cell_note = ws1.cell(current_row, 6, "[Nhận xét nhanh]")
            cell_note.font = font_italic
            cell_note.alignment = align_left
            cell_note.fill = fill_input
            
            for col in range(1, 7):
                ws1.cell(current_row, col).border = border_all_thin
                
            # Define wages for each specific row based on the CKSV files
            # Mapping row number to (Bậc, Lương thực lĩnh, Lương tổng chi phí)
            # Row 12 (stt 1) -> Bậc 1 (220k)
            # Row 13 (stt 2) -> Bậc 2 (220k)
            # Row 14 (stt 3a) -> Bậc 3 (250k)
            # Row 15 (stt 3b) -> Bậc 3 (250k)
            # Row 16 (stt 3c) -> Bậc 3 (260k)
            # Row 17 (stt 4a) -> Bậc 4 (270k)
            # Row 18 (stt 4b) -> Bậc 4 (280k)
            # Row 19 (stt 4c) -> Bậc 4 (290k)
            # Row 20 (stt 4d) -> Bậc 4 (300k)
            # Row 21 (stt 4e) -> Bậc 4 (320k)
            # Row 22 (stt 5a) -> Bậc 5 (330k)
            # Row 23 (stt 5b) -> Bậc 5 (340k)
            # Row 24 (stt 5c) -> Bậc 5 (350k)
            # Row 25 (stt 5d) -> Bậc 5 (360k)
            # Row 26 (stt 5e) -> Bậc 5 (370k)
            # Row 28 (stt 6) -> Bậc 6 (380k)
            # Row 29 (stt 7a) -> Bậc 7 (390k)
            # Row 30 (stt 7b) -> Bậc 7 (400k)
            # Row 31 (stt 7c) -> Bậc 7 (410k)
            # Row 32 (stt 7d) -> Bậc 7 (420k)
            # Row 34 (stt 8) -> Bậc 8 (430k)
            # Row 35 (stt 25) -> Bậc 9 (440k)
            # Row 36 (stt 9a) -> Bậc 10 (470k)
            # Row 37 (stt 9b) -> Bậc 11 (480k)
            # Row 38 (stt 10) -> Bậc 12 (490k)
            # Row 40 (stt 11) -> Bậc 13 (500k)
            # Row 41 (stt 12) -> Bậc 14 (550k)
            
            w_map = {
                12: (1, 220000, 300000), 13: (2, 220000, 300000),
                14: (3, 250000, 330000), 15: (3, 250000, 330000), 16: (3, 260000, 340000),
                17: (4, 270000, 350000), 18: (4, 280000, 360000), 19: (4, 290000, 370000), 20: (4, 300000, 380000), 21: (4, 320000, 400000),
                22: (5, 330000, 410000), 23: (5, 340000, 420000), 24: (5, 350000, 430000), 25: (5, 360000, 440000), 26: (5, 370000, 450000),
                28: (6, 380000, 460000),
                29: (7, 390000, 470000), 30: (7, 400000, 480000), 31: (7, 410000, 490000), 32: (7, 420000, 500000),
                34: (8, 430000, 510000), 35: (9, 440000, 520000), 36: (10, 470000, 550000), 37: (11, 480000, 560000), 38: (12, 490000, 570000),
                40: (13, 500000, 580000), 41: (14, 550000, 630000)
            }
            row_wage_map[current_row] = w_map[current_row]
            
        current_row += 1

    # Place sample test scores (probation pass up to Bậc 7 welding)
    ws1["E12"] = 1  # Bậc 1
    ws1["E13"] = 1  # Bậc 2
    ws1["E14"] = 1  # Bậc 3
    ws1["E15"] = 1
    ws1["E16"] = 1
    ws1["E17"] = 1  # Bậc 4
    ws1["E18"] = 1
    ws1["E19"] = 1
    ws1["E20"] = 1
    ws1["E21"] = 1
    ws1["E22"] = 1  # Bậc 5
    ws1["E23"] = 1
    ws1["E24"] = 1
    ws1["E25"] = 1
    ws1["E26"] = 1
    ws1["E28"] = 1  # Bậc 6
    ws1["E29"] = 1  # Bậc 7
    ws1["E30"] = 1
    ws1["E31"] = 1
    ws1["E32"] = 1
    ws1["E34"] = 0  # Bậc 8 (Chưa đạt đọc bản vẽ)

    ws1.row_dimensions[current_row].height = 15
    current_row += 1

    # RESULTS DASHBOARD BOX (6-COLUMN FORMAT BELOW THE TABLE)
    ws1.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=6)
    r_hdr = ws1.cell(current_row, 2, "KẾT QUẢ PHÂN BẬC LƯƠNG ĐỀ XUẤT SAU THỬ VIỆC")
    r_hdr.font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    r_hdr.fill = PatternFill(start_color=STEEL_BLUE, end_color=STEEL_BLUE, fill_type="solid")
    r_hdr.alignment = align_center
    ws1.row_dimensions[current_row].height = 24
    
    for col in range(2, 7):
        ws1.cell(current_row, col).border = Border(top=Side(style="medium", color=STEEL_BLUE))
        
    current_row += 1
    
    # We construct the MAX formula based on exact row numbers of criteria rows
    wage_parts_l = []
    wage_parts_t = []
    bac_parts = []
    
    for r_num in eval_rows:
        b_num, wage_l, wage_t = row_wage_map[r_num]
        wage_parts_l.append(f"IF(E{r_num}=1,{wage_l},0)")
        wage_parts_t.append(f"IF(E{r_num}=1,{wage_t},0)")
        bac_parts.append(f"IF(E{r_num}=1,{b_num},0)")
        
    formula_wage_l = "=MAX(" + ",".join(wage_parts_l) + ")"
    formula_wage_t = "=MAX(" + ",".join(wage_parts_t) + ")"
    formula_bac = "=MAX(" + ",".join(bac_parts) + ")"
    
    # 1. Bậc năng lực đạt được
    ws1.row_dimensions[current_row].height = 20
    ws1.cell(current_row, 2, "Bậc năng lực đạt được:").font = font_bold
    ws1.cell(current_row, 2).alignment = align_left
    
    cell_bac = ws1.cell(current_row, 3, formula_bac)
    cell_bac.font = font_bold
    cell_bac.alignment = align_center
    
    ws1.cell(current_row, 4, "Bậc / 14").font = font_italic
    ws1.cell(current_row, 4).alignment = align_left
    
    # Style borders
    for col in range(2, 7):
        cell = ws1.cell(current_row, col)
        cell.fill = fill_result_box
        left_s = Side(style="medium", color=STEEL_BLUE) if col == 2 else None
        right_s = Side(style="medium", color=STEEL_BLUE) if col == 6 else None
        cell.border = Border(left=left_s, right=right_s)
        
    bac_cell_ref = f"C{current_row}" # This cell stores the bậc number (e.g. C31)
    current_row += 1
    
    # 2. Chức danh tuyển dụng đề xuất
    ws1.row_dimensions[current_row].height = 20
    ws1.cell(current_row, 2, "Chức danh tuyển dụng đề xuất:").font = font_bold
    ws1.cell(current_row, 2).alignment = align_left
    
    cell_title = ws1.cell(current_row, 3, f"=IF({bac_cell_ref}=0, \"Chưa đạt\", VLOOKUP({bac_cell_ref}, 'Thang Đo & Lương'!A$2:E$15, 2, FALSE))")
    cell_title.font = font_bold
    cell_title.alignment = align_center
    
    for col in range(2, 7):
        cell = ws1.cell(current_row, col)
        cell.fill = fill_result_box
        left_s = Side(style="medium", color=STEEL_BLUE) if col == 2 else None
        right_s = Side(style="medium", color=STEEL_BLUE) if col == 6 else None
        cell.border = Border(left=left_s, right=right_s)
        
    current_row += 1
    
    # 3. Mức lương thực lĩnh (ngày)
    ws1.row_dimensions[current_row].height = 20
    ws1.cell(current_row, 2, "Lương thực lĩnh đề xuất (ngày):").font = font_bold
    ws1.cell(current_row, 2).alignment = align_left
    
    cell_wage_d = ws1.cell(current_row, 3, formula_wage_l)
    cell_wage_d.font = font_salary_result
    cell_wage_d.alignment = align_center
    cell_wage_d.number_format = '#,##0" đ"'
    
    ws1.cell(current_row, 4, "(Lương sau thuế/ngày)").font = font_italic
    ws1.cell(current_row, 4).alignment = align_left
    
    for col in range(2, 7):
        cell = ws1.cell(current_row, col)
        cell.fill = fill_result_box
        left_s = Side(style="medium", color=STEEL_BLUE) if col == 2 else None
        right_s = Side(style="medium", color=STEEL_BLUE) if col == 6 else None
        cell.border = Border(left=left_s, right=right_s)
        
    wage_d_ref = f"C{current_row}"
    current_row += 1
    
    # 4. Tổng chi phí công ty chi trả (ngày)
    ws1.row_dimensions[current_row].height = 20
    ws1.cell(current_row, 2, "Tổng chi phí chi trả (ngày):").font = font_bold
    ws1.cell(current_row, 2).alignment = align_left
    
    cell_wage_t = ws1.cell(current_row, 3, formula_wage_t)
    cell_wage_t.font = font_bold
    cell_wage_t.alignment = align_center
    cell_wage_t.number_format = '#,##0" đ"'
    
    ws1.cell(current_row, 4, "(Chi phí công ty chi trả/ngày)").font = font_italic
    ws1.cell(current_row, 4).alignment = align_left
    
    for col in range(2, 7):
        cell = ws1.cell(current_row, col)
        cell.fill = fill_result_box
        left_s = Side(style="medium", color=STEEL_BLUE) if col == 2 else None
        right_s = Side(style="medium", color=STEEL_BLUE) if col == 6 else None
        cell.border = Border(left=left_s, right=right_s)
        
    current_row += 1
    
    # 5. Dự toán thu nhập tháng (26 ngày công)
    ws1.row_dimensions[current_row].height = 20
    ws1.cell(current_row, 2, "Thu nhập tháng dự kiến (26 công):").font = font_bold
    ws1.cell(current_row, 2).alignment = align_left
    
    cell_wage_m = ws1.cell(current_row, 3, f"={wage_d_ref}*26")
    cell_wage_m.font = font_salary_result
    cell_wage_m.alignment = align_center
    cell_wage_m.number_format = '#,##0" đ"'
    
    ws1.cell(current_row, 4, "(Tương ứng 15-22 triệu/tháng)").font = font_italic
    ws1.cell(current_row, 4).alignment = align_left
    
    for col in range(2, 7):
        cell = ws1.cell(current_row, col)
        cell.fill = fill_result_box
        left_s = Side(style="medium", color=STEEL_BLUE) if col == 2 else None
        right_s = Side(style="medium", color=STEEL_BLUE) if col == 6 else None
        cell.border = Border(left=left_s, right=right_s)
        
    current_row += 1
    
    # 6. Khuyến nghị và Kế hoạch sử dụng
    ws1.row_dimensions[current_row].height = 24
    ws1.cell(current_row, 2, "Khuyến nghị & Phân công:").font = font_bold
    ws1.cell(current_row, 2).alignment = align_left
    
    ws1.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=6)
    rec_cell = ws1.cell(current_row, 3, f"=IF({bac_cell_ref}=0, \"Yêu cầu đào tạo thêm tay nghề hoặc từ chối thử việc\", VLOOKUP({bac_cell_ref}, 'Thang Đo & Lương'!A$2:E$15, 5, FALSE))")
    rec_cell.font = font_bold
    rec_cell.alignment = align_left
    
    for col in range(2, 7):
        cell = ws1.cell(current_row, col)
        cell.fill = fill_result_box
        left_s = Side(style="medium", color=STEEL_BLUE) if col == 2 else None
        right_s = Side(style="medium", color=STEEL_BLUE) if col == 6 else None
        cell.border = Border(left=left_s, right=right_s, bottom=Side(style="medium", color=STEEL_BLUE))

    # Data validations
    dv_self = DataValidation(type="decimal", operator="between", formula1="0", formula2="1", allow_blank=True)
    dv_self.error = 'Tỷ lệ đáp ứng phải từ 0% đến 100%!'
    dv_self.errorTitle = 'Lỗi nhập liệu'
    dv_self.prompt = 'Nhập % tự đánh giá (Ví dụ: 80%)'
    dv_self.promptTitle = 'Ứng viên tự đánh giá'
    ws1.add_data_validation(dv_self)
    
    dv_company = DataValidation(type="whole", operator="between", formula1="0", formula2="1", allow_blank=True)
    dv_company.error = 'Chỉ được nhập 1 (Đạt tiêu chuẩn) hoặc 0 (Chưa đạt)!'
    dv_company.errorTitle = 'Lỗi chấm điểm'
    dv_company.prompt = 'Nhập 1: Đạt, 0: Chưa đạt'
    dv_company.promptTitle = 'Quản đốc đánh giá'
    ws1.add_data_validation(dv_company)

    for r_num in eval_rows:
        dv_self.add(ws1[f"D{r_num}"])
        dv_company.add(ws1[f"E{r_num}"])

    current_row += 3
    
    # Signatures
    ws1.row_dimensions[current_row].height = 18
    ws1.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    s1 = ws1.cell(current_row, 1, "ỨNG VIÊN TUYỂN DỤNG")
    s1.font = font_bold
    s1.alignment = align_center
    
    ws1.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=6)
    s2 = ws1.cell(current_row, 4, "QUẢN ĐỐC / TỔ TRƯỞNG DUYỆT")
    s2.font = font_bold
    s2.alignment = align_center
    
    current_row += 1
    ws1.row_dimensions[current_row].height = 15
    ws1.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    sub1 = ws1.cell(current_row, 1, "(Ký và ghi rõ họ tên)")
    sub1.font = font_italic
    sub1.alignment = align_center
    
    ws1.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=6)
    sub2 = ws1.cell(current_row, 4, "(Xác nhận ngạch bậc lương ngày)")
    sub2.font = font_italic
    sub2.alignment = align_center
    
    current_row += 4
    ws1.row_dimensions[current_row].height = 18
    ws1.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    n1 = ws1.cell(current_row, 1, "[Họ tên ứng viên]")
    n1.font = font_normal
    n1.alignment = align_center
    
    ws1.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=6)
    n2 = ws1.cell(current_row, 4, "[Họ tên người đánh giá]")
    n2.font = font_normal
    n2.alignment = align_center

    # -------------------------------------------------------------
    # SHEET 2: THANG ĐO & LƯƠNG
    # -------------------------------------------------------------
    ws2 = wb.create_sheet(title="Thang Đo & Lương")
    ws2.sheet_view.showGridLines = True
    
    ws2.column_dimensions['A'].width = 10
    ws2.column_dimensions['B'].width = 24
    ws2.column_dimensions['C'].width = 24
    ws2.column_dimensions['D'].width = 24
    ws2.column_dimensions['E'].width = 65
    
    # Headers
    ws2["A1"] = "Bậc"
    ws2["B1"] = "Chức danh đề xuất"
    ws2["C1"] = "Lương thực lĩnh (ngày)"
    ws2["D1"] = "Tổng chi phí công ty (ngày)"
    ws2["E1"] = "Phân công / Vai trò chính trong xưởng"
    
    for col in ["A", "B", "C", "D", "E"]:
        cell = ws2[f"{col}1"]
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_all_thin
        
    scale_data = [
        (1, "Thợ học việc (Cấp 1)", 220000, 300000, "Phụ việc vệ sinh, mài mộc, nâng hạ phôi dưới sự chỉ dẫn"),
        (2, "Thợ học việc (Cấp 2)", 220000, 300000, "Thực hiện tốt 5S khu vực hàn gá, đảm bảo an toàn ATLĐ xưởng"),
        (3, "Thợ phụ cơ bản (Cấp 1)", 250000, 330000, "Biết dùng máy cắt cầm tay, khoan tay, thước đo chính xác"),
        (4, "Thợ phụ mài tinh (Cấp 2)", 290000, 370000, "Mài tinh mỹ bề mặt, mài lấy góc sắc nét, mài phá bavia"),
        (5, "Thợ sơn hoàn thiện", 330000, 410000, "Đứng buồng sơn phun khí nén, sơn chống rỉ và sơn phủ thẩm mỹ"),
        (6, "Thợ gá lắp phôi", 380000, 460000, "Nắm rõ các loại vật tư, phôi cắt sắt/inox, chuẩn bị công trình"),
        (7, "Thợ gá hàn liên hợp", 400000, 480000, "Sử dụng thành thạo máy hàn que/MIG/TIG để đính gá kết cấu"),
        (8, "Thợ hàn chính chuyên nghiệp", 430000, 510000, "Đọc bản vẽ sản xuất và tự hàn độc lập sản phẩm chất lượng cao"),
        (9, "Thợ lắp ráp lành nghề", 440000, 520000, "Đọc bản vẽ cơ khí phức tạp, tự bóc tách khai triển phôi gá ráp"),
        (10, "Thợ hoàn thiện độc lập", 470000, 550000, "Tự kiểm soát chất lượng hoàn thiện sản phẩm không cần giám sát"),
        (11, "Thợ dự toán & tiến độ", 480000, 560000, "Ước lượng thời gian/nhân công, lập tiến độ thi công dự án nhỏ"),
        (12, "Tổ phó kỹ thuật", 490000, 570000, "Tự chủ động điều chỉnh chi tiết nhỏ ở hiện trường, thay thế tổ trưởng"),
        (13, "Trưởng nhóm lắp dựng", 500000, 580000, "Điều phối nhóm 2-3 người lắp dựng kết cấu thép/inox ngoài công trường"),
        (14, "Tổ trưởng sản xuất", 550000, 630000, "Phân việc toàn xưởng, chịu trách nhiệm tiến độ & chất lượng tổ đội")
    ]
    
    for r_idx, row_val in enumerate(scale_data, start=2):
        ws2.cell(r_idx, 1, row_val[0]).alignment = align_center
        ws2.cell(r_idx, 1).font = font_bold
        
        ws2.cell(r_idx, 2, row_val[1]).alignment = align_left
        ws2.cell(r_idx, 2).font = font_bold
        
        ws2.cell(r_idx, 3, row_val[2]).number_format = '#,##0" đ"'
        ws2.cell(r_idx, 3).alignment = align_right
        ws2.cell(r_idx, 3).font = font_bold
        
        ws2.cell(r_idx, 4, row_val[3]).number_format = '#,##0" đ"'
        ws2.cell(r_idx, 4).alignment = align_right
        ws2.cell(r_idx, 4).font = font_normal
        
        ws2.cell(r_idx, 5, row_val[4]).alignment = align_left
        ws2.cell(r_idx, 5).font = font_normal
        
        for col in range(1, 6):
            cell = ws2.cell(r_idx, col)
            cell.border = border_all_thin
            if r_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

    # Save workbook to CKSV_Form_Tuyen_Dung_Tho_Co_Khi_Hoan_Thien.xlsx
    output_path = "d:/Sao Vàng/Website-SaoVang/CKSV_Form_Tuyen_Dung_Tho_Co_Khi_Hoan_Thien.xlsx"
    wb.save(output_path)
    print(f"Final 6col recruitment form created successfully at {output_path}")

if __name__ == "__main__":
    create_exact_6col_recruitment_form()
