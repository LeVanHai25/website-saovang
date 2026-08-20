import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.drawing.image import Image as OpenpyxlImage

def create_skills_test_sheet_with_logo():
    wb = openpyxl.Workbook()
    
    # Define Styles
    STEEL_BLUE = "2F5597"
    ICE_BLUE = "D9E1F2"
    LIGHT_GRAY = "F2F2F2"
    BORDER_GRAY = "D9D9D9"
    YELLOW_INPUT = "FFF2CC"
    DARK_GREEN = "385723"
    LIGHT_GREEN = "E2EFDA"
    
    # Fonts
    font_company_name = Font(name="Segoe UI", size=12, bold=True, color=STEEL_BLUE)
    font_company_info = Font(name="Segoe UI", size=9, italic=True, color="595959")
    font_title_main = Font(name="Segoe UI", size=15, bold=True, color=STEEL_BLUE)
    font_title_sub = Font(name="Segoe UI", size=10, italic=True, color="7F7F7F")
    font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    font_group_header = Font(name="Segoe UI", size=11, bold=True, color=STEEL_BLUE)
    font_bold = Font(name="Segoe UI", size=10, bold=True)
    font_normal = Font(name="Segoe UI", size=10)
    font_italic = Font(name="Segoe UI", size=9, italic=True)
    font_result = Font(name="Segoe UI", size=11, bold=True, color=DARK_GREEN)
    
    # Fills
    fill_header = PatternFill(start_color=STEEL_BLUE, end_color=STEEL_BLUE, fill_type="solid")
    fill_group = PatternFill(start_color=ICE_BLUE, end_color=ICE_BLUE, fill_type="solid")
    fill_input = PatternFill(start_color=YELLOW_INPUT, end_color=YELLOW_INPUT, fill_type="solid")
    fill_result_box = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
    
    # Alignments
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    
    # Borders
    thin_side = Side(style="thin", color=BORDER_GRAY)
    border_all_thin = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    border_group = Border(
        left=Side(style="thin", color=STEEL_BLUE),
        right=Side(style="thin", color=STEEL_BLUE),
        top=Side(style="thin", color=STEEL_BLUE),
        bottom=Side(style="thin", color=STEEL_BLUE)
    )

    # -------------------------------------------------------------
    # SHEET 1: ĐÁNH GIÁ TAY NGHỀ
    # -------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Đánh Giá Tay Nghề"
    ws1.sheet_view.showGridLines = True
    
    # Column widths
    ws1.column_dimensions['A'].width = 6   # STT
    ws1.column_dimensions['B'].width = 30  # HẠNG MỤC KỸ NĂNG
    ws1.column_dimensions['C'].width = 50  # TIÊU CHUẨN ĐÁNH GIÁ CHI TIẾT
    ws1.column_dimensions['D'].width = 54  # HƯỚNG DẪN KIỂM TRA (DÀNH CHO KẾ TOÁN/HR)
    ws1.column_dimensions['E'].width = 16  # ỨNG VIÊN TỰ KHAI (1/0)
    ws1.column_dimensions['F'].width = 16  # THỰC TẾ ĐẠT (1/0)
    ws1.column_dimensions['G'].width = 25  # GHI CHÚ CHI TIẾT

    # 1. Company Letterhead Row 2-4
    ws1.row_dimensions[2].height = 20
    ws1.row_dimensions[3].height = 15
    ws1.row_dimensions[4].height = 15
    
    # Add Company Name & Info
    ws1.merge_cells("C2:G2")
    ws1["C2"] = "CÔNG TY CỔ PHẦN SẢN XUẤT CƠ KHÍ SAO VÀNG (ASV)"
    ws1["C2"].font = font_company_name
    ws1["C2"].alignment = Alignment(horizontal="left", vertical="center")
    
    ws1.merge_cells("C3:G3")
    ws1["C3"] = "Địa chỉ: Tầng 3, TT7-35 Khu đô thị Văn Phú, phường Kiến Hưng, TP Hà Nội, Việt Nam."
    ws1["C3"].font = font_company_info
    ws1["C3"].alignment = Alignment(horizontal="left", vertical="center")
    
    ws1.merge_cells("C4:G4")
    ws1["C4"] = "Hotline: 0869 590 279  |  Email: cokhisaovangvn@gmail.com"
    ws1["C4"].font = font_company_info
    ws1["C4"].alignment = Alignment(horizontal="left", vertical="center")

    # Add Logo image in A2 (spanning A2:B4, width=110px, height=45px)
    try:
        logo_path = "d:/Sao Vàng/Website-SaoVang/Logo/LogoCTY + SV Aluminium/LogoCTY.png"
        img = OpenpyxlImage(logo_path)
        img.width = 110
        img.height = 45
        ws1.add_image(img, "A2")
    except Exception as e:
        print(f"Error adding logo: {e}")

    # Title Block (shifted to Row 6-7)
    ws1.row_dimensions[6].height = 24
    ws1.merge_cells("A6:G6")
    ws1["A6"] = "PHIẾU KIỂM TRA & ĐÁNH GIÁ TRÌNH ĐỘ TAY NGHỀ THỢ CƠ KHÍ"
    ws1["A6"].font = font_title_main
    ws1["A6"].alignment = align_center

    ws1.row_dimensions[7].height = 18
    ws1.merge_cells("A7:G7")
    ws1["A7"] = "Biểu mẫu thử tay nghề thực chiến xưởng - Sát thực tế - Nhân sự/Kế toán dễ dàng đánh giá"
    ws1["A7"].font = font_title_sub
    ws1["A7"].alignment = align_center

    # Profile Info (shifted to Row 9-10)
    ws1["B9"] = "Họ và tên thợ:"
    ws1["B9"].font = font_bold
    ws1["C9"] = "[Nhập tên ứng viên]"
    ws1["C9"].font = font_normal
    ws1["C9"].border = Border(bottom=Side(style="thin", color="000000"))

    ws1["E9"] = "Ngày kiểm tra:"
    ws1["E9"].font = font_bold
    ws1["F9"] = "[Nhập ngày]"
    ws1["F9"].font = font_normal
    ws1["F9"].border = Border(bottom=Side(style="thin", color="000000"))

    ws1["B10"] = "Số điện thoại/CCCD:"
    ws1["B10"].font = font_bold
    ws1["C10"] = "[Số điện thoại và CCCD]"
    ws1["C10"].font = font_normal
    ws1["C10"].border = Border(bottom=Side(style="thin", color="000000"))

    ws1["E10"] = "Người chấm điểm:"
    ws1["E10"].font = font_bold
    ws1["F10"] = "[Tên Quản đốc/Tổ trưởng/Kế toán]"
    ws1["F10"].font = font_normal
    ws1["F10"].border = Border(bottom=Side(style="thin", color="000000"))

    ws1.row_dimensions[11].height = 15

    # Table Header (shifted to Row 13)
    headers = [
        ("A13", "STT", align_center),
        ("B13", "HẠNG MỤC KỸ NĂNG KIỂM TRA", align_left),
        ("C13", "YÊU CẦU ĐẠT TIÊU CHUẨN (SẾP DUYỆT)", align_left),
        ("D13", "HƯỚNG DẪN KIỂM TRA (KẾ TOÁN/HR NHÌN THỰC TẾ)", align_left),
        ("E13", "ỨNG VIÊN TỰ KHAI (Đạt: 1 | 0)", align_center),
        ("F13", "THỰC TẾ ĐẠT (Đạt: 1 | 0)", align_center),
        ("G13", "GHI CHÚ / NHẬN XÉT CỦA NGƯỜI CHẤM", align_left)
    ]
    for cell_ref, text, align in headers:
        cell = ws1[cell_ref]
        cell.value = text
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align
        cell.border = border_all_thin
    ws1.row_dimensions[13].height = 28

    # Helper function to write merged headers
    def merge_and_style_header_ws1(ws, r_num, title):
        ws.merge_cells(start_row=r_num, start_column=1, end_row=r_num, end_column=7)
        ws.cell(r_num, 1, title).font = font_group_header
        ws.cell(r_num, 1).alignment = align_left
        for col in range(1, 8):
            cell = ws.cell(r_num, col)
            cell.fill = fill_group
            cell.border = border_group

    # Technical checklist data (shifted start to row 14)
    technical_tasks = [
        ("G", "I. KỸ THUẬT HÀN & RÁP KHUNG"),
        ("1", "Kỹ năng hàn que (MMA)", 
         "Hàn liên tục được đường hàn dài, ngấu đều, không bị thủng phôi hộp, xỉ hàn bong sạch và vảy đều.",
         "Yêu cầu hàn thử 10cm trên sắt hộp. Kế toán/HR kiểm tra: Mối hàn liền mạch, gõ xỉ ra thấy sáng đều, không có lỗ kim li ti."),
        
        ("2", "Kỹ năng hàn MIG/MAG", 
         "Vận hành máy hàn bán tự động MIG, đi mối hàn nhanh, ngấu sâu kết cấu thép chịu lực.",
         "Yêu cầu hàn thử góc chữ T. Kế toán/HR kiểm tra: Đường hàn mượt, đều, không có vết cháy đen/thủng phôi."),
        
        ("3", "Kỹ năng hàn TIG Inox", 
         "Hàn TIG phôi inox mỏng từ 0.8mm - 1.2mm; mối hàn vảy cá đều đặn, trắng sáng bóng, không bị đen/thủng.",
         "Yêu cầu hàn thử inox mỏng. Kế toán/HR kiểm tra mối hàn: Phải có màu vàng rơm hoặc trắng sáng, nếu đen thui là chưa đạt."),
        
        ("4", "Kỹ thuật đính ráp kết cấu", 
         "Đo đạc gá đính ráp chính xác góc vuông 90 độ, song song của khung bao trước khi hàn chết kết cấu.",
         "Yêu cầu đính gá góc khung. Dùng thước ê-ke đặt vào góc xem có hở khít không, góc vuông khít 90 độ là ĐẠT."),
        
        ("G", "II. KỸ THUẬT MÀI & XỬ LÝ BỀ MẶT"),
        ("5", "Mài phẳng mối hàn kết cấu", 
         "Dùng máy mài góc làm phẳng mối hàn thô phẳng phiu, không lồi lõm, sạch bavia bám xung quanh.",
         "Yêu cầu mài phẳng mối hàn thô. Kế toán dùng tay sờ lên mối mài: nếu thấy phẳng mịn như bề mặt ống/hộp bên cạnh là ĐẠT."),
        
        ("6", "Mài bo góc cạnh mỹ thuật", 
         "Gia công bo tròn góc hàn mịn đẹp, vuông góc sắc nét đối với các sản phẩm dân dụng trang trí kiến trúc.",
         "Quan sát thực tế góc cạnh: Phải được mài bo đều tay, không có góc nhọn bén đứt tay, góc vuông khít."),
        
        ("7", "Đánh bóng gương & Xước Hairline", 
         "Xử lý bề mặt inox đạt độ bóng gương (mirror finish) sáng loáng hoặc đánh xước hairline đồng đều thẩm mỹ.",
         "Quan sát sản phẩm sau đánh bóng: Phải soi gương thấy ảnh rõ nét, không còn vết xước mờ."),
        
        ("G", "III. ĐỌC BẢN VẼ & CẮT PHÔI THÉP"),
        ("8", "Đọc hiểu bản vẽ kỹ thuật", 
         "Đọc hiểu bản vẽ lắp ráp cơ khí 2D/3D; nhận diện đúng kích thước phủ bì và các góc ghép kết cấu.",
         "Đưa một bản vẽ cửa cổng sắt mẫu. Hỏi ứng viên: 'Kích thước thanh đố đứng và đố ngang dài bao nhiêu?'"),
        
        ("9", "Bóc tách phôi & Khai triển", 
         "Bóc tách bản vẽ thành danh sách phôi cần cắt chi tiết; tính toán sao cho tiết kiệm phôi, hao hụt ít nhất.",
         "Yêu cầu ứng viên liệt kê kích thước cắt cho 10 thanh sắt hộp từ cây nguyên bản 6 mét để hao hụt ít nhất."),
        
        ("10", "Vận hành máy cắt phôi", 
         "Cắt phôi bằng máy cắt bàn, cắt đĩa chính xác từng mm; mép cắt phẳng, không bị chéo góc.",
         "Yêu cầu cắt 1 đoạn sắt hộp dài đúng 500mm. Kế toán dùng thước mét đo lại: sai số tối đa ±1mm là ĐẠT."),
        
        ("G", "IV. KỸ THUẬT SƠN HOÀN THIỆN"),
        ("11", "Vận hành súng phun sơn", 
         "Sử dụng thành thạo súng phun sơn khí nén, sơn chống rỉ và sơn phủ mịn, bóng đều màng sơn.",
         "Yêu cầu sơn thử 1 mặt thép. Quan sát màng sơn: Phải mịn, đều màu, không có vệt chảy xệ sơn xuống dưới."),
        
        ("12", "Pha chế sơn dầu & Sơn 2 thành phần", 
         "Pha sơn dầu, sơn Epoxy 2 thành phần đúng tỷ lệ dung môi chỉ định để sơn đạt độ cứng bám tốt.",
         "Hỏi: 'Sơn Epoxy 2 thành phần pha tỷ lệ chất đóng rắn và dung môi thế nào? Pha xong để được bao lâu?'"),
        
        ("G", "V. TÁC PHONG AN TOÀN & 5S"),
        ("13", "Chấp hành an toàn lao động (ATLĐ)", 
         "Đeo kính bảo hộ khi mài, găng tay da khi hàn, mặt nạ hàn và chấp hành an toàn cháy nổ xưởng.",
         "Quan sát thực tế khi thợ thực hành: có tự giác đeo kính mài, đeo mặt nạ hàn và đi giày bảo hộ không."),
        
        ("14", "Dọn dẹp vệ sinh xưởng (5S)", 
         "Chủ động dọn dẹp máy móc, cuộn dây nguồn gọn gàng, quét sạch bavia sắt và bụi sơn khu vực mình làm.",
         "Sau khi thợ làm xong bài thực hành, quan sát xem họ có tự giác thu dọn máy cắt/hàn và quét dọn bãi làm không.")
    ]

    current_row = 14
    eval_rows = []
    row_bac_map = {}
    
    for row_data in technical_tasks:
        if row_data[0] == "G":
            ws1.row_dimensions[current_row].height = 24
            merge_and_style_header_ws1(ws1, current_row, row_data[1])
        else:
            stt, task, standard, instruction = row_data
            ws1.row_dimensions[current_row].height = 42
            
            ws1.cell(current_row, 1, stt).font = font_normal
            ws1.cell(current_row, 1).alignment = align_center
            
            ws1.cell(current_row, 2, task).font = font_bold
            ws1.cell(current_row, 2).alignment = align_left
            
            ws1.cell(current_row, 3, standard).font = font_normal
            ws1.cell(current_row, 3).alignment = align_left
            
            ws1.cell(current_row, 4, instruction).font = font_italic
            ws1.cell(current_row, 4).alignment = align_left
            
            # Candidate Self-Declaration (1/0)
            cell_self = ws1.cell(current_row, 5, 1)
            cell_self.font = font_bold
            cell_self.alignment = align_center
            cell_self.fill = fill_input
            
            # Company Eval / Practical results (1/0)
            cell_company = ws1.cell(current_row, 6, 0)
            cell_company.font = font_bold
            cell_company.alignment = align_center
            cell_company.fill = fill_input
            eval_rows.append(current_row)
            
            # Note
            cell_note = ws1.cell(current_row, 7, "[Nhận xét chi tiết]")
            cell_note.font = font_italic
            cell_note.alignment = align_left
            cell_note.fill = fill_input
            
            for col in range(1, 8):
                ws1.cell(current_row, col).border = border_all_thin
                
            w_map = {
                15: 1, 16: 2, 17: 3, 18: 4,  # Hàn
                20: 5, 21: 6, 22: 7,          # Mài
                24: 8, 25: 9, 26: 10,         # Bản vẽ
                28: 11, 29: 12,               # Sơn
                31: 13, 32: 14                # 5S
            }
            row_bac_map[current_row] = w_map[current_row]
            
        current_row += 1

    # Place sample test values
    ws1["E15"] = 1  # Hàn que
    ws1["E16"] = 1  # Hàn MIG
    ws1["E17"] = 0
    ws1["E18"] = 1
    ws1["E20"] = 1
    ws1["E21"] = 1
    ws1["E22"] = 0
    ws1["E24"] = 1
    ws1["E25"] = 1
    ws1["E26"] = 1
    ws1["E28"] = 1
    ws1["E29"] = 1
    ws1["E31"] = 1
    ws1["E32"] = 1

    # Company Eval marks
    ws1["F15"] = 1
    ws1["F16"] = 1
    ws1["F17"] = 0
    ws1["F18"] = 1
    ws1["F20"] = 1
    ws1["F21"] = 1
    ws1["F22"] = 0
    ws1["F24"] = 1
    ws1["F25"] = 1
    ws1["F26"] = 1
    ws1["F28"] = 1
    ws1["F29"] = 1
    ws1["F31"] = 1
    ws1["F32"] = 1

    ws1.row_dimensions[current_row].height = 15
    current_row += 1

    # TECHNICAL RATING DASHBOARD BOX
    ws1.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=7)
    r_hdr = ws1.cell(current_row, 2, "KẾT QUẢ ĐÁNH GIÁ TRÌNH ĐỘ TAY NGHỀ THỰC TẾ")
    r_hdr.font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    r_hdr.fill = PatternFill(start_color=STEEL_BLUE, end_color=STEEL_BLUE, fill_type="solid")
    r_hdr.alignment = align_center
    ws1.row_dimensions[current_row].height = 24
    
    for col in range(2, 8):
        ws1.cell(current_row, col).border = Border(top=Side(style="medium", color=STEEL_BLUE))
        
    current_row += 1
    
    sum_e_str = "+".join(f"E{r}" for r in eval_rows)
    sum_f_str = "+".join(f"F{r}" for r in eval_rows)
    
    # 1. Tổng số tiêu chí đánh giá
    ws1.row_dimensions[current_row].height = 20
    ws1.cell(current_row, 2, "Tổng số hạng mục kiểm tra:").font = font_bold
    ws1.cell(current_row, 2).alignment = align_left
    
    cell_total = ws1.cell(current_row, 3, len(eval_rows))
    cell_total.font = font_bold
    cell_total.alignment = align_center
    
    ws1.cell(current_row, 4, "Hạng mục chuyên môn").font = font_italic
    ws1.cell(current_row, 4).alignment = align_left
    
    for col in range(2, 8):
        cell = ws1.cell(current_row, col)
        cell.fill = fill_result_box
        left_s = Side(style="medium", color=STEEL_BLUE) if col == 2 else None
        right_s = Side(style="medium", color=STEEL_BLUE) if col == 7 else None
        cell.border = Border(left=left_s, right=right_s)
        
    total_ref = f"C{current_row}" # C35
    current_row += 1
    
    # 2. Số hạng mục ứng viên TỰ KHAI
    ws1.row_dimensions[current_row].height = 20
    ws1.cell(current_row, 2, "Hạng mục ứng viên tự khai biết làm:").font = font_bold
    ws1.cell(current_row, 2).alignment = align_left
    
    cell_self_sum = ws1.cell(current_row, 3, f"={sum_e_str}")
    cell_self_sum.font = font_bold
    cell_self_sum.alignment = align_center
    
    ws1.cell(current_row, 4, "Đạt yêu cầu tự khai").font = font_italic
    ws1.cell(current_row, 4).alignment = align_left
    
    for col in range(2, 8):
        cell = ws1.cell(current_row, col)
        cell.fill = fill_result_box
        left_s = Side(style="medium", color=STEEL_BLUE) if col == 2 else None
        right_s = Side(style="medium", color=STEEL_BLUE) if col == 7 else None
        cell.border = Border(left=left_s, right=right_s)
        
    current_row += 1
    
    # 3. Số hạng mục KIỂM TRA THỰC TẾ ĐẠT
    ws1.row_dimensions[current_row].height = 20
    ws1.cell(current_row, 2, "Hạng mục kiểm tra thực tế đạt:").font = font_bold
    ws1.cell(current_row, 2).alignment = align_left
    
    cell_actual_sum = ws1.cell(current_row, 3, f"={sum_f_str}")
    cell_actual_sum.font = font_bold
    cell_actual_sum.alignment = align_center
    
    ws1.cell(current_row, 4, "Đạt yêu cầu kiểm tra thực chiến").font = font_italic
    ws1.cell(current_row, 4).alignment = align_left
    
    for col in range(2, 8):
        cell = ws1.cell(current_row, col)
        cell.fill = fill_result_box
        left_s = Side(style="medium", color=STEEL_BLUE) if col == 2 else None
        right_s = Side(style="medium", color=STEEL_BLUE) if col == 7 else None
        cell.border = Border(left=left_s, right=right_s)
        
    actual_sum_ref = f"C{current_row}" # C37
    current_row += 1
    
    # 4. Tỷ lệ thành thạo kỹ năng (%)
    ws1.row_dimensions[current_row].height = 20
    ws1.cell(current_row, 2, "Tỷ lệ thành thạo kỹ năng thực tế:").font = font_bold
    ws1.cell(current_row, 2).alignment = align_left
    
    cell_rate = ws1.cell(current_row, 3, f"={actual_sum_ref}/{total_ref}")
    cell_rate.font = font_result
    cell_rate.alignment = align_center
    cell_rate.number_format = '0.0%'
    
    ws1.cell(current_row, 4, "(Thực tế đạt / Tổng hạng mục)").font = font_italic
    ws1.cell(current_row, 4).alignment = align_left
    
    for col in range(2, 8):
        cell = ws1.cell(current_row, col)
        cell.fill = fill_result_box
        left_s = Side(style="medium", color=STEEL_BLUE) if col == 2 else None
        right_s = Side(style="medium", color=STEEL_BLUE) if col == 7 else None
        cell.border = Border(left=left_s, right=right_s)
        
    rate_ref = f"C{current_row}" # C38
    current_row += 1
    
    # 5. Phân loại trình độ đề xuất
    ws1.row_dimensions[current_row].height = 20
    ws1.cell(current_row, 2, "Xếp hạng trình độ tay nghề:").font = font_bold
    ws1.cell(current_row, 2).alignment = align_left
    
    cell_level = ws1.cell(current_row, 3, f'=IF({rate_ref}>=0.85, "Thợ chính xuất sắc (Bậc A)", IF({rate_ref}>=0.7, "Thợ cứng/Thợ khá (Bậc B)", IF({rate_ref}>=0.5, "Thợ phụ cơ bản (Bậc C)", "Thợ học việc/Học nghề"))')
    cell_level.font = font_result
    cell_level.alignment = align_center
    
    ws1.cell(current_row, 4, "(Phân nhóm thợ tham chiếu)").font = font_italic
    ws1.cell(current_row, 4).alignment = align_left
    
    for col in range(2, 8):
        cell = ws1.cell(current_row, col)
        cell.fill = fill_result_box
        left_s = Side(style="medium", color=STEEL_BLUE) if col == 2 else None
        right_s = Side(style="medium", color=STEEL_BLUE) if col == 7 else None
        cell.border = Border(left=left_s, right=right_s)
        
    current_row += 1
    
    # 6. Khuyến nghị phân công
    ws1.row_dimensions[current_row].height = 24
    ws1.cell(current_row, 2, "Khuyến nghị phân công công việc:").font = font_bold
    ws1.cell(current_row, 2).alignment = align_left
    
    ws1.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=7)
    cell_rec = ws1.cell(current_row, 3, f'=IF({rate_ref}>=0.85, "Độc lập chế tạo kết cấu phức tạp, đọc bản vẽ tốt, gá ráp hoàn thiện.", IF({rate_ref}>=0.7, "Độc lập thi công cụm kết cấu thép/inox thông dụng theo bản vẽ.", IF({rate_ref}>=0.5, "Giao phụ việc, cắt phôi, mài bavia và hỗ trợ thợ chính gá ráp.", "Bố trí phụ việc dọn vệ sinh xưởng, thực hành 5S và học việc tại xưởng.")))')
    cell_rec.font = font_bold
    cell_rec.alignment = align_left
    
    for col in range(2, 8):
        cell = ws1.cell(current_row, col)
        cell.fill = fill_result_box
        left_s = Side(style="medium", color=STEEL_BLUE) if col == 2 else None
        right_s = Side(style="medium", color=STEEL_BLUE) if col == 7 else None
        cell.border = Border(left=left_s, right=right_s, bottom=Side(style="medium", color=STEEL_BLUE))

    current_row += 2

    # Slogan block matching company motto
    ws1.row_dimensions[current_row].height = 26
    ws1.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
    slogan_cell = ws1.cell(current_row, 1, 'Khẩu hiệu của chúng tôi: "Chất Lượng Tạo Nên Thương Hiệu. Uy Tín Tạo Nên Sự Thành Công"')
    slogan_cell.font = Font(name="Segoe UI", size=10, bold=True, italic=True, color=STEEL_BLUE)
    slogan_cell.alignment = align_center
    slogan_cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Border for slogan
    for col in range(1, 8):
        ws1.cell(current_row, col).border = Border(
            top=Side(style="thin", color=STEEL_BLUE),
            bottom=Side(style="thin", color=STEEL_BLUE),
            left=Side(style="thin", color=STEEL_BLUE) if col == 1 else None,
            right=Side(style="thin", color=STEEL_BLUE) if col == 7 else None
        )

    current_row += 3
    
    # Signatures
    ws1.row_dimensions[current_row].height = 18
    ws1.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    s1 = ws1.cell(current_row, 1, "ỨNG VIÊN TUYỂN DỤNG")
    s1.font = font_bold
    s1.alignment = align_center
    
    ws1.merge_cells(start_row=current_row, start_column=5, end_row=current_row, end_column=7)
    s2 = ws1.cell(current_row, 5, "QUẢN ĐỐC / TỔ TRƯỞNG DUYỆT")
    s2.font = font_bold
    s2.alignment = align_center
    
    current_row += 1
    ws1.row_dimensions[current_row].height = 15
    ws1.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    sub1 = ws1.cell(current_row, 1, "(Ký và ghi rõ họ tên)")
    sub1.font = font_italic
    sub1.alignment = align_center
    
    ws1.merge_cells(start_row=current_row, start_column=5, end_row=current_row, end_column=7)
    sub2 = ws1.cell(current_row, 5, "(Xác nhận kết quả tay nghề)")
    sub2.font = font_italic
    sub2.alignment = align_center
    
    current_row += 4
    ws1.row_dimensions[current_row].height = 18
    ws1.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    n1 = ws1.cell(current_row, 1, "[Họ tên ứng viên]")
    n1.font = font_normal
    n1.alignment = align_center
    
    ws1.merge_cells(start_row=current_row, start_column=5, end_row=current_row, end_column=7)
    n2 = ws1.cell(current_row, 5, "[Họ tên người chấm điểm]")
    n2.font = font_normal
    n2.alignment = align_center

    # Data validations to input columns E & F
    dv_input = DataValidation(type="whole", operator="between", formula1="0", formula2="1", allow_blank=True)
    dv_input.prompt = 'Nhập 1: Đạt, 0: Chưa đạt'
    dv_input.promptTitle = 'Chấm điểm'
    ws1.add_data_validation(dv_input)

    for r_num in eval_rows:
        dv_input.add(ws1[f"E{r_num}"])
        dv_input.add(ws1[f"F{r_num}"])

    # -------------------------------------------------------------
    # SHEET 2: PHÂN LOẠI TRÌNH ĐỘ
    # -------------------------------------------------------------
    ws2 = wb.create_sheet(title="Phân Loại Trình Độ")
    ws2.sheet_view.showGridLines = True
    
    ws2.column_dimensions['A'].width = 24
    ws2.column_dimensions['B'].width = 15
    ws2.column_dimensions['C'].width = 24
    ws2.column_dimensions['D'].width = 65
    
    ws2["A1"] = "Xếp hạng trình độ"
    ws2["B1"] = "Tỷ lệ đạt (%)"
    ws2["C1"] = "Chức danh tương thích"
    ws2["D1"] = "Mô tả năng lực kỹ thuật thực tế"
    
    for col in ["A", "B", "C", "D"]:
        cell = ws2[f"{col}1"]
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_all_thin
        
    class_data = [
        ("Thợ chính xuất sắc (Bậc A)", ">= 85%", "Thợ cơ khí chính", "Đọc thạo bản vẽ kỹ thuật, độc lập bóc tách phôi và gá ráp thành thạo, tay nghề hàn que/MIG/TIG ngấu đẹp thẩm mỹ cao."),
        ("Thợ cứng/Thợ khá (Bậc B)", "70% - 84%", "Thợ cơ khí cứng", "Hàn tốt các phương pháp hàn phổ thông chịu lực; tự gá ráp khung cơ bản, mài hoàn thiện thẩm mỹ đạt chuẩn."),
        ("Thợ phụ cơ bản (Bậc C)", "50% - 69%", "Thợ phụ", "Sử dụng tốt dụng cụ cầm tay (máy cắt, khoan...); mài bavia sạch sẽ; chuẩn bị phôi và hỗ trợ gá ráp theo yêu cầu của thợ chính."),
        ("Thợ học việc/Học nghề", "< 50%", "Học việc", "Chưa đáp ứng được tay nghề độc lập; cần đào tạo thêm; chịu khó học hỏi, dọn vệ sinh 5S an toàn xưởng.")
    ]
    
    for r_idx, row_val in enumerate(class_data, start=2):
        ws2.cell(r_idx, 1, row_val[0]).alignment = align_left
        ws2.cell(r_idx, 1).font = font_bold
        ws2.cell(r_idx, 1).border = border_all_thin
        
        ws2.cell(r_idx, 2, row_val[1]).alignment = align_center
        ws2.cell(r_idx, 2).font = font_normal
        ws2.cell(r_idx, 2).border = border_all_thin
        
        ws2.cell(r_idx, 3, row_val[2]).alignment = align_center
        ws2.cell(r_idx, 3).font = font_bold
        ws2.cell(r_idx, 3).border = border_all_thin
        
        ws2.cell(r_idx, 4, row_val[3]).alignment = align_left
        ws2.cell(r_idx, 4).font = font_normal
        ws2.cell(r_idx, 4).border = border_all_thin
        
        for col in range(1, 5):
            cell = ws2.cell(r_idx, col)
            if r_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

    paths = [
        "d:/Sao Vàng/Website-SaoVang/CKSV_Bo_Cong_Cu_Tuyen_Dung_Tho_Co_Khi.xlsx",
        "d:/Sao Vàng/Website-SaoVang/CKSV_Bo_Cong_Cu_Tuyen_Dung_Tho_Co_Khi_Co_Logo.xlsx"
    ]
    saved_paths = []
    for path in paths:
        try:
            wb.save(path)
            saved_paths.append(path)
        except PermissionError:
            print(f"Permission denied for {path} (likely open in Excel). Skipping.")
            
    print(f"Skills evaluation workbook with logo saved to: {saved_paths}")

if __name__ == "__main__":
    create_skills_test_sheet_with_logo()
