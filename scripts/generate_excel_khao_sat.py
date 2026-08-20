import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Create workbook and worksheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Bảng Khảo Sát"

# Ensure grid lines are visible
ws.views.sheetView[0].showGridLines = True

# Colors
TITLE_COLOR = "1F4E78"       # Navy Blue
HEADER_BG_COLOR = "1F4E78"   # Dark Navy
HEADER_FG_COLOR = "FFFFFF"   # White
ALT_ROW_BG = "F2F5F9"        # Very light blue-grey
BORDER_COLOR = "D9D9D9"      # Light grey

# Fonts
font_title = Font(name="Segoe UI", size=16, bold=True, color=TITLE_COLOR)
font_meta_label = Font(name="Segoe UI", size=10, bold=True, color="333333")
font_meta_val = Font(name="Segoe UI", size=10, italic=False, color="333333")
font_header = Font(name="Segoe UI", size=11, bold=True, color=HEADER_FG_COLOR)
font_body = Font(name="Segoe UI", size=10, color="000000")
font_stt = Font(name="Segoe UI", size=10, bold=True, color="333333")

# Fills
fill_header = PatternFill(start_color=HEADER_BG_COLOR, end_color=HEADER_BG_COLOR, fill_type="solid")
fill_alt = PatternFill(start_color=ALT_ROW_BG, end_color=ALT_ROW_BG, fill_type="solid")

# Borders
thin_side = Side(border_style="thin", color=BORDER_COLOR)
thick_bottom_side = Side(border_style="medium", color="1F4E78")
border_cell = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
border_header = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thick_bottom_side)

# Alignments
align_center_top = Alignment(horizontal="center", vertical="top")
align_left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)
align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Write Title Section
ws.cell(row=1, column=1, value="BẢNG KHẢO SÁT HẠNG MỤC CẤU KIỆN & QUY CÁCH VẬT TƯ").font = font_title
ws.merge_cells("A1:D1")

# Write Metadata
meta_info = [
    ("Công trình:", "C02L07 – An Vượng"),
    ("Nguồn dữ liệu khảo sát:", "Mô hình SketchUp (SKP)"),
    ("Mục đích:", "Khảo sát cấu tạo sơ bộ các hạng mục phục vụ triển khai thiết kế Shopdrawing, bóc tách vật tư và lập dự toán.")
]

for idx, (label, val) in enumerate(meta_info, start=3):
    cell_lbl = ws.cell(row=idx, column=1, value=label)
    cell_lbl.font = font_meta_label
    cell_lbl.alignment = Alignment(horizontal="left", vertical="center")
    
    cell_val = ws.cell(row=idx, column=2, value=val)
    cell_val.font = font_meta_val
    cell_val.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=4)

# Table Header Row
headers = ["STT", "Vị trí", "Hạng mục", "Nội dung khảo sát kỹ thuật"]
header_row = 7

ws.row_dimensions[header_row].height = 28
for col_idx, text in enumerate(headers, start=1):
    cell = ws.cell(row=header_row, column=col_idx, value=text)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_header
    cell.border = border_header

# Table Data
data = [
    (
        1, "Tầng 3", "Lan can",
        "• Tay vịn đo được hộp 40×80 mm.\n"
        "• Nan đứng đo được hộp 20×40 mm.\n"
        "• Chưa bố trí thanh đáy (sẽ xác định khi triển khai bản vẽ sản xuất)."
    ),
    (
        2, "Tầng 3", "Mái che",
        "• Chiều cao tổng phủ bì hệ khung: 136 mm.\n"
        "• Khung bao ngoài đo được hộp 50×100 mm.\n"
        "• Thanh hộp trang trí trên và dưới đo được hộp 20×50 mm.\n"
        "• Chưa bố trí hệ khung xương bên trong.\n"
        "• Vật liệu lợp mái: Tấm Aluminium (Alu)."
    ),
    (
        3, "Tầng 2", "Mái che",
        "• Chiều cao tổng hệ khung: 200 mm.\n"
        "• Khung bao mái đo được hộp 140×100 mm.\n"
        "• Khung xương chính đo được hộp 100×200 mm.\n"
        "• Thanh hộp trang trí trên và dưới đo được hộp 30×50 mm.\n"
        "• Có bố trí nan đứng trang trí.\n"
        "• Vật liệu lợp mái: Kính."
    ),
    (
        4, "Tầng 1", "Mái che",
        "• Chiều cao tổng hệ khung: 200 mm.\n"
        "• Khung bao ngoài mái đo được hộp 140×100 mm.\n"
        "• Khung xương chính đo được hộp 100×200 mm.\n"
        "• Thanh hộp trang trí trên và dưới đo được hộp 30×50 mm.\n"
        "• Có bố trí nan đứng trang trí.\n"
        "• Vật liệu hoàn thiện mái: Kính kết hợp Aluminium (Alu)."
    ),
    (
        5, "Tầng 1", "Khung hàng rào",
        "• Khung bao hàng rào đo được hộp 47×80 mm.\n"
        "• Thanh ngang giữa đo được hộp 23×80 mm.\n"
        "• Hệ lam trang trí sử dụng lập là/bản dày 11 mm."
    ),
    (
        6, "Tầng 1", "Cổng",
        "• Thanh đứng hai bên cổng đo được hộp 21×62 mm.\n"
        "• Khung bao cổng đo được hộp 35×37 mm.\n"
        "• Nan đứng sử dụng bản dày 5 mm và 12 mm.\n"
        "• Nan trang trí sử dụng tấm/bản dày 3 mm."
    )
]

start_row = 8
for i, (stt, vi_tri, hang_muc, noi_dung) in enumerate(data):
    current_row = start_row + i
    
    c_stt = ws.cell(row=current_row, column=1, value=stt)
    c_vi_tri = ws.cell(row=current_row, column=2, value=vi_tri)
    c_hang_muc = ws.cell(row=current_row, column=3, value=hang_muc)
    c_noi_dung = ws.cell(row=current_row, column=4, value=noi_dung)

    c_stt.font = font_stt
    c_vi_tri.font = font_body
    c_hang_muc.font = Font(name="Segoe UI", size=10, bold=True, color="000000")
    c_noi_dung.font = font_body

    c_stt.alignment = align_center_top
    c_vi_tri.alignment = align_center_top
    c_hang_muc.alignment = align_left_top
    c_noi_dung.alignment = align_left_top

    # Apply borders and zebra pattern
    is_even = (i % 2 == 1)
    for col in range(1, 5):
        cell = ws.cell(row=current_row, column=col)
        cell.border = border_cell
        if is_even:
            cell.fill = fill_alt

# Column Widths
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 14
ws.column_dimensions['C'].width = 22
ws.column_dimensions['D'].width = 75

# Save workbook
output_path = "Bang_Khao_Sat_C02L07_An_Vuong.xlsx"
wb.save(output_path)
print(f"File created successfully: {output_path}")
