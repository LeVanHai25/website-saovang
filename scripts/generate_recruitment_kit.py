import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

def create_recruitment_kit():
    wb = openpyxl.Workbook()
    
    # Define Styles
    STEEL_BLUE = "2F5597"
    ICE_BLUE = "D9E1F2"
    LIGHT_GRAY = "F2F2F2"
    BORDER_GRAY = "D9D9D9"
    YELLOW_INPUT = "FFF2CC"
    DARK_GREEN = "385723"
    LIGHT_GREEN = "E2EFDA"
    
    # Fonts
    font_title_company = Font(name="Segoe UI", size=11, bold=True, color="595959")
    font_title_main = Font(name="Segoe UI", size=16, bold=True, color=STEEL_BLUE)
    font_title_sub = Font(name="Segoe UI", size=10, italic=True, color="7F7F7F")
    font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    font_group_header = Font(name="Segoe UI", size=11, bold=True, color=STEEL_BLUE)
    font_bold = Font(name="Segoe UI", size=10, bold=True)
    font_normal = Font(name="Segoe UI", size=10)
    font_italic = Font(name="Segoe UI", size=9, italic=True)
    font_salary_result = Font(name="Segoe UI", size=12, bold=True, color=DARK_GREEN)
    
    # Fills
    fill_header = PatternFill(start_color=STEEL_BLUE, end_color=STEEL_BLUE, fill_type="solid")
    fill_group = PatternFill(start_color=ICE_BLUE, end_color=ICE_BLUE, fill_type="solid")
    fill_sum = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
    fill_input = PatternFill(start_color=YELLOW_INPUT, end_color=YELLOW_INPUT, fill_type="solid")
    fill_result_box = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
    
    # Alignments
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    
    # Borders
    thin_side = Side(style="thin", color=BORDER_GRAY)
    border_all_thin = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    border_group = Border(
        left=Side(style="thin", color=STEEL_BLUE),
        right=Side(style="thin", color=STEEL_BLUE),
        top=Side(style="thin", color=STEEL_BLUE),
        bottom=Side(style="thin", color=STEEL_BLUE)
    )

    # -------------------------------------------------------------
    # SHEET 1: ĐÁNH GIÁ TUYỂN DỤNG
    # -------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Đánh Giá Tuyển Dụng"
    ws1.sheet_view.showGridLines = True
    
    # Column widths
    ws1.column_dimensions['A'].width = 6   # STT
    ws1.column_dimensions['B'].width = 30  # NHIỆM VỤ ĐÁNH GIÁ
    ws1.column_dimensions['C'].width = 50  # YÊU CẦU KỸ NĂNG CHI TIẾT
    ws1.column_dimensions['D'].width = 50  # HƯỚNG DẪN KIỂM TRA CHO KẾ THỪA/KẾ TOÁN
    ws1.column_dimensions['E'].width = 16  # ỨNG VIÊN TỰ KHAI (1/0)
    ws1.column_dimensions['F'].width = 16  # THỰC CHIẾN ĐẠT (1/0)
    ws1.column_dimensions['G'].width = 20  # GHI CHÚ CHI TIẾT

    # Title Block
    ws1.merge_cells("A2:G2")
    ws1["A2"] = "CÔNG TY CỔ PHẦN SẢN XUẤT CƠ KHÍ SAO VÀNG (ASV)"
    ws1["A2"].font = font_title_company
    ws1["A2"].alignment = Alignment(horizontal="left")

    ws1.merge_cells("A3:G3")
    ws1["A3"] = "BỘ ĐÁNH GIÁ NĂNG LỰC TUYỂN DỤNG THỢ CƠ KHÍ (DÀNH CHO MỌI NHÂN SỰ)"
    ws1["A3"].font = font_title_main
    ws1["A3"].alignment = align_center

    ws1.merge_cells("A4:G4")
    ws1["A4"] = "Tích hợp hướng dẫn đánh giá cho nhân sự không chuyên (Kế toán/HR) & Điểm thực chiến tay nghề"
    ws1["A4"].font = font_title_sub
    ws1["A4"].alignment = align_center

    # Candidate Profile Info
    ws1["B6"] = "Họ và tên ứng viên:"
    ws1["B6"].font = font_bold
    ws1["C6"] = "[Nhập tên ứng viên]"
    ws1["C6"].font = font_normal
    ws1["C6"].border = Border(bottom=Side(style="thin", color="000000"))

    ws1["E6"] = "Ngày phỏng vấn:"
    ws1["E6"].font = font_bold
    ws1["F6"] = "[Nhập ngày]"
    ws1["F6"].font = font_normal
    ws1["F6"].border = Border(bottom=Side(style="thin", color="000000"))

    ws1["B7"] = "Số điện thoại/CCCD:"
    ws1["B7"].font = font_bold
    ws1["C7"] = "[Số điện thoại và CCCD]"
    ws1["C7"].font = font_normal
    ws1["C7"].border = Border(bottom=Side(style="thin", color="000000"))

    ws1["E7"] = "Người phỏng vấn:"
    ws1["E7"].font = font_bold
    ws1["F7"] = "[Kế toán/HR/Quản đốc]"
    ws1["F7"].font = font_normal
    ws1["F7"].border = Border(bottom=Side(style="thin", color="000000"))

    ws1["B8"] = "Vị trí ứng tuyển:"
    ws1["B8"].font = font_bold
    ws1["C8"] = "Thợ Cơ khí Dân dụng (Học việc/Phụ/Giúp việc/Thợ chính/Tổ trưởng)"
    ws1["C8"].font = font_italic
    ws1["C8"].border = Border(bottom=Side(style="thin", color="000000"))

    ws1["E8"] = "Kỳ thử việc:"
    ws1["E8"].font = font_bold
    ws1["F8"] = "01 tháng thử việc"
    ws1["F8"].font = font_normal
    ws1["F8"].border = Border(bottom=Side(style="thin", color="000000"))

    ws1.row_dimensions[9].height = 15

    # Main Table Header
    headers = [
        ("A10", "STT", align_center),
        ("B10", "NHIỆM VỤ ĐÁNH GIÁ (VIỆC CẦN LÀM)", align_left),
        ("C10", "YÊU CẦU KỸ NĂNG CHI TIẾT (SẾP DUYỆT)", align_left),
        ("D10", "HƯỚNG DẪN KIỂM TRA (DÀNH CHO KẾ TOÁN/HR)", align_left),
        ("E10", "ỨNG VIÊN TỰ KHAI (Đạt: 1 | 0)", align_center),
        ("F10", "THỰC CHIẾN ĐẠT (Đạt: 1 | 0)", align_center),
        ("G10", "GHI CHÚ / NHẬN XÉT CHI TIẾT", align_left)
    ]
    for cell_ref, text, align in headers:
        cell = ws1[cell_ref]
        cell.value = text
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align
        cell.border = border_all_thin
    ws1.row_dimensions[10].height = 28

    # Helper function to write merged headers
    def merge_and_style_header_ws1(ws, r_num, title):
        ws.merge_cells(start_row=r_num, start_column=1, end_row=r_num, end_column=7)
        ws.cell(r_num, 1, title).font = font_group_header
        ws.cell(r_num, 1).alignment = align_left
        for col in range(1, 8):
            cell = ws.cell(r_num, col)
            cell.fill = fill_group
            cell.border = border_group

    # Core Competency Evaluation Table with HR Testing Instructions
    competency_tasks = [
        ("G", "I. HỌC VIỆC VÀ PHỤ VIỆC (Tiêu chuẩn Bậc 1 - Bậc 5)"),
        ("1", "Tác phong, Sức khỏe & Thái độ (Bậc 1)", 
         "Đảm bảo sức khỏe dẻo dai, bền bỉ để làm việc trong môi trường nhà xưởng cơ khí; thái độ trung thực, nhanh nhẹn, chịu khó học hỏi, có ý thức trách nhiệm.",
         "Quan sát nhanh nhẹn khi đi đứng, chào hỏi. Hỏi: 'Em có bệnh nền hay đau nhức xương khớp gì không? Có ngại làm việc bụi bặm, tiếng ồn không?'"),
        
        ("2", "Thực hành Vệ sinh & 5S (Bậc 2)", 
         "Chủ động dọn dẹp máy móc, thiết bị cầm tay gọn gàng sau ca làm việc; quét dọn và vệ sinh sạch sẽ khu vực gia công lắp dựng tại xưởng cũng như khi đi công trường.",
         "Hỏi ứng viên: 'Sau mỗi ca làm việc, quy trình dọn dẹp bàn giao khu vực làm việc của em thường gồm những bước nào?'"),
        
        ("3", "Vận hành dụng cụ cầm tay & Mài thô (Bậc 3)", 
         "Sử dụng đúng kỹ thuật, an toàn các dụng cụ cơ bản (máy cắt tay, máy cắt bàn, máy khoan cầm tay, thước đo...); mài hoàn thiện bề mặt kết cấu cơ bản (sắt đen, sắt mạ kẽm, inox).",
         "Hỏi: 'Khi đá của máy cắt tay bị mòn hoặc cần thay, em làm thế nào để đảm bảo an toàn?' Nhìn thực tế xem họ có đeo kính bảo hộ khi mài không."),
        
        ("4", "Kỹ thuật mài tinh mỹ & Đánh bóng (Bậc 4)", 
         "Mài phẳng tuyệt đối, mài bo góc cạnh sắc nét đạt độ thẩm mỹ cao đối với sản phẩm hoàn thiện; mài xử lý mối hàn thô; mài đánh bóng inox đạt độ bóng gương (mirror finish); biết mài sắc mũi khoan.",
         "Yêu cầu mài phẳng một mối hàn thử. Đưa tay sờ bề mặt mối mài: nếu phẳng mịn, không lồi lõm, không lẹm phôi là ĐẠT."),
        
        ("5", "Sơn phủ hoàn thiện & Hàn đính gá phụ (Bậc 5)", 
         "Vận hành hệ thống phun sơn khí nén, sơn lót chống gỉ và sơn phủ màu bóng mịn; đo cắt phôi chính xác và làm dưỡng loạt phôi; hàn đính gá cơ bản.",
         "Yêu cầu sơn thử một góc sắt hộp. Quan sát: bề mặt sơn phải láng mịn, không chảy giọt sơn, không bọt khí là ĐẠT."),
        
        ("G", "II. GIÚP VIỆC (Tiêu chuẩn Bậc 6 - Bậc 7)"),
        ("6", "Chuẩn bị hiện trường & Thiết bị (Bậc 6)", 
         "Nắm rõ công năng các loại thiết bị cơ khí cầm tay/chuyên dụng; chủ động tính toán và chuẩn bị đầy đủ đồ nghề thi công cần thiết khi đi lắp ráp công trình ngoài hiện trường.",
         "Đưa tình huống: 'Sắp tới đi lắp tôn mái nhà xưởng 100m2, em cần chuẩn bị những loại máy móc, đồ bảo hộ và vật tư phụ nào?'"),
        
        ("7", "Vận hành máy hàn chuyên nghiệp (Bậc 7)", 
         "Hàn thành thạo các phương pháp hàn que (MMA), hàn bán tự động MIG trên kết cấu sắt thép chịu lực và hàn TIG trên inox mỏng (mối hàn ngấu sâu, vảy cá đều đẹp, trắng sáng).",
         "Yêu cầu hàn thử một đường que/MIG/TIG liên tục dài 10cm. Quan sát: mối hàn đều đặn, không thủng phôi, xỉ bong dễ dàng, mối hàn TIG trắng sáng là ĐẠT."),
        
        ("G", "III. THỢ CHÍNH - TỰ THỰC HIỆN (Tiêu chuẩn Bậc 8 - Bậc 10)"),
        ("8", "Đọc bản vẽ sản xuất & Bóc tách (Bậc 8)", 
         "Đọc hiểu bản vẽ thiết kế kỹ thuật cơ khí gia công chế tạo; tự tính toán bóc tách phôi và khai triển kích thước hình học chính xác để gia công sản phẩm thực tế.",
         "Đưa một bản vẽ cửa cổng sắt hộp mẫu. Hỏi: 'Với bản vẽ này, em cần cắt bao nhiêu thanh sắt dài bao nhiêu để gá ráp?' Nhập số cắt đúng là ĐẠT."),
        
        ("9", "Hoàn thiện & Tự kiểm soát QC (Bậc 9)", 
         "Kỹ năng tự kiểm tra chất lượng sản phẩm (Self-QC) sau gia công (độ phẳng, độ vuông góc, tính thẩm mỹ); tinh thần cầu tiến, không giấu dốt, tự khắc phục lỗi kỹ thuật phát sinh.",
         "Hỏi: 'Nếu em phát hiện sản phẩm của mình bị hàn lệch góc 2 độ sau khi đã nguội, em sẽ xử lý như thế nào?'"),
        
        ("10", "Dự toán định mức nhân công & Tiến độ (Bậc 10)", 
         "Khả năng đánh giá độ phức tạp của kết cấu bản vẽ; ước lượng chuẩn xác định mức giờ công thợ và lượng nhân lực cần thiết để hoàn thành sản phẩm.",
         "Hỏi: 'Ráp hoàn thiện 1 cái cửa đi sắt hộp 4 cánh 3x3m cần bao nhiêu công thợ và khoảng mấy ngày?' Nhìn đáp án so với định mức công ty."),
        
        ("G", "IV. TỔ TRƯỞNG VÀ ĐIỀU HÀNH SẢN XUẤT (Tiêu chuẩn Bậc 11 - Bậc 14)"),
        ("11", "Báo cáo tiến độ & Giải quyết phát sinh (Bậc 11)", 
         "Lập báo cáo tiến độ công việc cuối ngày; chủ động đề xuất giải pháp xử lý nhanh các vướng mắc kỹ thuật phát sinh nhằm duy trì sự thông suốt giữa các bộ phận.",
         "Hỏi: 'Nếu nhà cung cấp giao tôn lợp xưởng bị chậm 1 ngày, ảnh hưởng tiến độ lắp dựng, em sẽ đề xuất đổi thứ tự thi công thế nào?'"),
        
        ("12", "Độc lập thi công & Quyết định thực địa (Bậc 12)", 
         "Khả năng độc lập thi công các hạng mục khó; chủ động đưa ra quyết định xử lý kỹ thuật nhỏ tại công trình mà không làm ảnh hưởng đến cấu trúc tổng thể.",
         "Hỏi: 'Khi lắp đặt tôn ngoài hiện trường bị kích thước xô lệch 2cm so với móng, em xử lý cắt/gá thế nào ngay tại chỗ?'"),
        
        ("13", "Làm việc nhóm & Quản lý tiến độ (Bậc 13)", 
         "Khả năng điều phối và làm việc nhóm nhịp nhàng; phân công nhiệm vụ rõ ràng; đảm bảo tiến độ gia công sản xuất đạt chuẩn kỹ thuật bản vẽ và giao hàng đúng hẹn.",
         "Hỏi: 'Làm thế nào để điều phối thợ phụ mài tôn và thợ chính ráp khung phối hợp nhịp nhàng, không bên nào phải đứng chờ?'"),
        
        ("14", "Lãnh đạo đội nhóm & Điều hành sản xuất (Bậc 14)", 
         "Năng lực quản lý và lãnh đạo tổ đội từ 3-5 thợ phụ/chính; bao quát tiến độ và chất lượng toàn tổ; điều phối phân chia công việc tối ưu theo tay nghề của từng công nhân.",
         "Hỏi: 'Nếu sáng hôm nay 1 thợ chính hàn MIG nghỉ ốm đột xuất, trong khi tiến độ ráp khung đang gấp, em điều chuyển công nhân thế nào?'")
    ]

    current_row = 11
    eval_rows = []
    row_bac_map = {}
    
    for row_data in competency_tasks:
        if row_data[0] == "G":
            ws1.row_dimensions[current_row].height = 24
            merge_and_style_header_ws1(ws1, current_row, row_data[1])
        else:
            stt, task, skill, instruction = row_data
            ws1.row_dimensions[current_row].height = 42 # generous spacing for instructions
            
            ws1.cell(current_row, 1, stt).font = font_normal
            ws1.cell(current_row, 1).alignment = align_center
            
            ws1.cell(current_row, 2, task).font = font_bold
            ws1.cell(current_row, 2).alignment = align_left
            
            ws1.cell(current_row, 3, skill).font = font_normal
            ws1.cell(current_row, 3).alignment = align_left
            
            ws1.cell(current_row, 4, instruction).font = font_italic
            ws1.cell(current_row, 4).alignment = align_left
            
            # Candidate Self-Declaration (1/0)
            cell_self = ws1.cell(current_row, 5, 1)
            cell_self.font = font_bold
            cell_self.alignment = align_center
            cell_self.fill = fill_input
            
            # Foreman Practical Test (1/0)
            cell_company = ws1.cell(current_row, 6, 0)
            cell_company.font = font_bold
            cell_company.alignment = align_center
            cell_company.fill = fill_input
            eval_rows.append(current_row)
            
            # Note
            cell_note = ws1.cell(current_row, 7, "[Nhận xét câu hỏi]")
            cell_note.font = font_italic
            cell_note.alignment = align_left
            cell_note.fill = fill_input
            
            for col in range(1, 8):
                ws1.cell(current_row, col).border = border_all_thin
                
            w_map = {
                12: 1, 13: 2, 14: 3, 15: 4, 16: 5,
                18: 6, 19: 7,
                21: 8, 22: 9, 23: 10,
                25: 11, 26: 12, 27: 13, 28: 14
            }
            row_bac_map[current_row] = w_map[current_row]
            
        current_row += 1

    # Place sample test scores (probation pass up to Bậc 7 welding)
    ws1["E12"] = 1  # Bậc 1
    ws1["E13"] = 1  # Bậc 2
    ws1["E14"] = 1  # Bậc 3
    ws1["E15"] = 1  # Bậc 4
    ws1["E16"] = 1  # Bậc 5
    ws1["E18"] = 1  # Bậc 6
    ws1["E19"] = 1  # Bậc 7 (weld)
    ws1["E21"] = 0  # Bậc 8 (blueprints fail)
    ws1["E22"] = 0
    ws1["E23"] = 0
    ws1["E25"] = 0
    ws1["E26"] = 0
    ws1["E27"] = 0
    ws1["E28"] = 0

    ws1.row_dimensions[current_row].height = 15
    current_row += 1

    # -------------------------------------------------------------
    # PRACTICAL TEST ITEMS SECTION (FORM THỰC CHIẾN TAY NGHỀ)
    # -------------------------------------------------------------
    ws1.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
    practical_hdr = ws1.cell(current_row, 1, "PHẦN II: NỘI DUNG THỰC HÀNH THỬ TAY NGHỀ THỰC CHIẾN TẠI XƯỞNG")
    practical_hdr.font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    practical_hdr.fill = PatternFill(start_color="333F48", end_color="333F48", fill_type="solid")
    practical_hdr.alignment = align_center
    ws1.row_dimensions[current_row].height = 26
    current_row += 1
    
    ws1.row_dimensions[current_row].height = 24
    p_headers = [
        ("A", "STT", align_center),
        ("B", "HẠNG MỤC THỰC CHIẾN", align_left),
        ("C", "CHI TIẾT ĐẦU VIỆC THỰC HÀNH", align_left),
        ("D", "TIÊU CHUẨN ĐÁNH GIÁ (ĐẠT)", align_left),
        ("E", "THỰC HIỆN ĐẠT (1 | 0)", align_center),
        ("F", "GHI CHÚ KỸ THUẬT CỦA FOREMAN", align_left),
        ("G", "", align_left) # merged column target
    ]
    for col_let, text, align in p_headers:
        if col_let == "G":
            continue
        cell = ws1[f"{col_let}{current_row}"]
        cell.value = text
        cell.font = font_header
        cell.fill = PatternFill(start_color="5A5A5A", end_color="5A5A5A", fill_type="solid")
        cell.alignment = align
        cell.border = border_all_thin
    ws1.merge_cells(f"F{current_row}:G{current_row}")
    ws1[f"G{current_row}"].border = border_all_thin
    current_row += 1

    practical_tasks = [
        ("T1", "Thử thách 1: Cắt phôi sắt hộp", 
         "Cắt 2 đoạn sắt hộp mạ kẽm 30x30mm dài đúng 500mm bằng máy cắt bàn.", 
         "Dùng thước mét đo lại sai số chiều dài không quá ±1mm; góc cắt phẳng 90 độ, không lệch."),
        
        ("T2", "Thử thách 2: Mài phẳng bavia", 
         "Mài sạch bavia cạnh sắt và làm phẳng mối hàn thô của chi tiết bằng máy mài góc cầm tay.", 
         "Bề mặt mối mài phẳng mịn, sờ tay không gồ ghề, không bị mài lẹm khoét sâu vào phôi thép."),
        
        ("T3", "Thử thách 3: Phun sơn hoàn thiện", 
         "Sử dụng súng phun sơn khí nén sơn phủ màu một bề mặt thép hộp 200x200mm.", 
         "Lớp sơn phủ đều tay, mịn bóng, không bị chảy giọt sơn, không loang lổ mỏng dày."),
        
        ("T4", "Thử thách 4: Hàn đính ráp khung", 
         "Gá lắp và hàn đính góc vuông 90 độ một khung sắt hộp chữ nhật đơn giản.", 
         "Dùng thước ê-ke vuông góc đo lại: góc vuông chuẩn xác, khung phẳng không vặn vẹo."),
        
        ("T5", "Thử thách 5: Hàn que/MIG ngấu lực", 
         "Hàn liên tục một đường hàn dài 100mm trên thép hộp mạ kẽm dày 1.4mm.", 
         "Mối hàn ngấu đều, không bị thủng lỗ, xỉ dễ bong sạch, vảy hàn đều tăm tắp.")
    ]

    p_eval_rows = []
    for stt, p_name, p_detail, p_standard in practical_tasks:
        ws1.row_dimensions[current_row].height = 36
        
        ws1.cell(current_row, 1, stt).font = font_normal
        ws1.cell(current_row, 1).alignment = align_center
        
        ws1.cell(current_row, 2, p_name).font = font_bold
        ws1.cell(current_row, 2).alignment = align_left
        
        ws1.cell(current_row, 3, p_detail).font = font_normal
        ws1.cell(current_row, 3).alignment = align_left
        
        ws1.cell(current_row, 4, p_standard).font = font_italic
        ws1.cell(current_row, 4).alignment = align_left
        
        # Test result (1/0)
        cell_res = ws1.cell(current_row, 5, 1)
        cell_res.font = font_bold
        cell_res.alignment = align_center
        cell_res.fill = fill_input
        p_eval_rows.append(current_row)
        
        # Merge F and G for notes
        ws1.merge_cells(start_row=current_row, start_column=6, end_row=current_row, end_column=7)
        ws1.cell(current_row, 6, "[Ghi chú kỹ thuật thợ]").font = font_italic
        ws1.cell(current_row, 6).fill = fill_input
        ws1.cell(current_row, 6).alignment = align_left
        
        for col in range(1, 8):
            ws1.cell(current_row, col).border = border_all_thin
            
        current_row += 1

    # Place sample test scores for practical trials
    ws1["E36"] = 1  # Cắt phôi đạt
    ws1["E37"] = 1  # Mài phẳng đạt
    ws1["E38"] = 1  # Phun sơn đạt
    ws1["E39"] = 1  # Hàn đính đạt
    ws1["E40"] = 1  # Hàn ngấu đạt (Đạt hết 5 hạng mục cơ bản)

    ws1.row_dimensions[current_row].height = 15
    current_row += 1

    # -------------------------------------------------------------
    # RESULTS DASHBOARD BOX (6-COLUMN FORMAT BELOW THE TABLE)
    # -------------------------------------------------------------
    ws1.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=7)
    r_hdr = ws1.cell(current_row, 2, "KẾT QUẢ PHÂN BẬC LƯƠNG ĐỀ XUẤT SAU THỬ TAY NGHỀ THỰC CHIẾN")
    r_hdr.font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    r_hdr.fill = PatternFill(start_color=STEEL_BLUE, end_color=STEEL_BLUE, fill_type="solid")
    r_hdr.alignment = align_center
    ws1.row_dimensions[current_row].height = 24
    
    for col in range(2, 8):
        ws1.cell(current_row, col).border = Border(top=Side(style="medium", color=STEEL_BLUE))
        
    current_row += 1
    
    # We construct the MAX formula WITHOUT IF functions to prevent Excel array evaluation glitches
    bac_parts = []
    for r_num in eval_rows:
        b_num = row_bac_map[r_num]
        bac_parts.append(f"F{r_num}*{b_num}") # Evaluates F (Thực chiến đạt) column
        
    formula_bac = "=MAX(" + ",".join(bac_parts) + ")"
    
    # 1. Bậc năng lực đạt được
    ws1.row_dimensions[current_row].height = 20
    ws1.cell(current_row, 2, "Bậc năng lực đạt được:").font = font_bold
    ws1.cell(current_row, 2).alignment = align_left
    
    cell_bac = ws1.cell(current_row, 3, formula_bac)
    cell_bac.font = font_bold
    cell_bac.alignment = align_center
    
    ws1.cell(current_row, 4, "Bậc / 14").font = font_italic
    ws1.cell(current_row, 4).alignment = align_left
    
    ws1.cell(current_row, 5, "Kết quả thực chiến:").font = font_bold
    ws1.cell(current_row, 5).alignment = align_right
    
    ws1.merge_cells(start_row=current_row, start_column=6, end_row=current_row, end_column=7)
    cell_p_score = ws1.cell(current_row, 6, f'=IF(SUM(E36:E40)=5, "Đạt thực chiến 5/5", CONCATENATE("Chỉ đạt ", SUM(E36:E40), "/5 hạng mục"))')
    cell_p_score.font = font_bold
    cell_p_score.alignment = align_center
    
    # Style borders
    for col in range(2, 8):
        cell = ws1.cell(current_row, col)
        cell.fill = fill_result_box
        left_s = Side(style="medium", color=STEEL_BLUE) if col == 2 else None
        right_s = Side(style="medium", color=STEEL_BLUE) if col == 7 else None
        cell.border = Border(left=left_s, right=right_s)
        
    bac_cell_ref = f"C{current_row}" # e.g. C43
    current_row += 1
    
    # 2. Chức danh tuyển dụng đề xuất
    ws1.row_dimensions[current_row].height = 20
    ws1.cell(current_row, 2, "Chức danh tuyển dụng đề xuất:").font = font_bold
    ws1.cell(current_row, 2).alignment = align_left
    
    cell_title = ws1.cell(current_row, 3, f"=IF({bac_cell_ref}=0, \"Chưa đạt\", VLOOKUP({bac_cell_ref}, 'Thang Đo & Lương'!A$2:E$15, 2, FALSE))")
    cell_title.font = font_bold
    cell_title.alignment = align_center
    
    for col in range(2, 8):
        cell = ws1.cell(current_row, col)
        cell.fill = fill_result_box
        left_s = Side(style="medium", color=STEEL_BLUE) if col == 2 else None
        right_s = Side(style="medium", color=STEEL_BLUE) if col == 7 else None
        cell.border = Border(left=left_s, right=right_s)
        
    current_row += 1
    
    # 3. Mức lương thực lĩnh (ngày)
    ws1.row_dimensions[current_row].height = 20
    ws1.cell(current_row, 2, "Lương thực lĩnh đề xuất (ngày):").font = font_bold
    ws1.cell(current_row, 2).alignment = align_left
    
    cell_wage_d = ws1.cell(current_row, 3, f"=IF({bac_cell_ref}=0, 0, VLOOKUP({bac_cell_ref}, 'Thang Đo & Lương'!A$2:E$15, 3, FALSE))")
    cell_wage_d.font = font_salary_result
    cell_wage_d.alignment = align_center
    cell_wage_d.number_format = '#,##0" đ"'
    
    ws1.cell(current_row, 4, "(Lương sau thuế/ngày)").font = font_italic
    ws1.cell(current_row, 4).alignment = align_left
    
    for col in range(2, 8):
        cell = ws1.cell(current_row, col)
        cell.fill = fill_result_box
        left_s = Side(style="medium", color=STEEL_BLUE) if col == 2 else None
        right_s = Side(style="medium", color=STEEL_BLUE) if col == 7 else None
        cell.border = Border(left=left_s, right=right_s)
        
    wage_d_ref = f"C{current_row}"
    current_row += 1
    
    # 4. Tổng chi phí chi trả (ngày)
    ws1.row_dimensions[current_row].height = 20
    ws1.cell(current_row, 2, "Tổng chi phí chi trả (ngày):").font = font_bold
    ws1.cell(current_row, 2).alignment = align_left
    
    cell_wage_t = ws1.cell(current_row, 3, f"=IF({bac_cell_ref}=0, 0, VLOOKUP({bac_cell_ref}, 'Thang Đo & Lương'!A$2:E$15, 4, FALSE))")
    cell_wage_t.font = font_bold
    cell_wage_t.alignment = align_center
    cell_wage_t.number_format = '#,##0" đ"'
    
    ws1.cell(current_row, 4, "(Tổng chi phí ngày công)").font = font_italic
    ws1.cell(current_row, 4).alignment = align_left
    
    for col in range(2, 8):
        cell = ws1.cell(current_row, col)
        cell.fill = fill_result_box
        left_s = Side(style="medium", color=STEEL_BLUE) if col == 2 else None
        right_s = Side(style="medium", color=STEEL_BLUE) if col == 7 else None
        cell.border = Border(left=left_s, right=right_s)
        
    current_row += 1
    
    # 5. Dự toán thu nhập tháng (26 ngày công)
    ws1.row_dimensions[current_row].height = 20
    ws1.cell(current_row, 2, "Thu nhập tháng dự kiến (26 công):").font = font_bold
    ws1.cell(current_row, 2).alignment = align_left
    
    cell_wage_m = ws1.cell(current_row, 3, f"={wage_d_ref}*26")
    cell_wage_m.font = font_salary_result
    cell_wage_m.alignment = align_center
    cell_wage_m.number_format = '#,##0" đ"'
    
    ws1.cell(current_row, 4, "(Tương ứng 15-22 triệu/tháng)").font = font_italic
    ws1.cell(current_row, 4).alignment = align_left
    
    for col in range(2, 8):
        cell = ws1.cell(current_row, col)
        cell.fill = fill_result_box
        left_s = Side(style="medium", color=STEEL_BLUE) if col == 2 else None
        right_s = Side(style="medium", color=STEEL_BLUE) if col == 7 else None
        cell.border = Border(left=left_s, right=right_s)
        
    current_row += 1
    
    # 6. Khuyến nghị và Kế hoạch sử dụng
    ws1.row_dimensions[current_row].height = 24
    ws1.cell(current_row, 2, "Khuyến nghị & Phân công:").font = font_bold
    ws1.cell(current_row, 2).alignment = align_left
    
    ws1.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=7)
    rec_cell = ws1.cell(current_row, 3, f"=IF({bac_cell_ref}=0, \"Yêu cầu đào tạo thêm tay nghề hoặc từ chối thử việc\", VLOOKUP({bac_cell_ref}, 'Thang Đo & Lương'!A$2:E$15, 5, FALSE))")
    rec_cell.font = font_bold
    rec_cell.alignment = align_left
    
    for col in range(2, 8):
        cell = ws1.cell(current_row, col)
        cell.fill = fill_result_box
        left_s = Side(style="medium", color=STEEL_BLUE) if col == 2 else None
        right_s = Side(style="medium", color=STEEL_BLUE) if col == 7 else None
        cell.border = Border(left=left_s, right=right_s, bottom=Side(style="medium", color=STEEL_BLUE))

    current_row += 2

    # Slogan block matching company motto
    ws1.row_dimensions[current_row].height = 26
    ws1.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
    slogan_cell = ws1.cell(current_row, 1, 'Khẩu hiệu của chúng tôi: "Chất Lượng Tạo Nên Thương Hiệu. Uy Tín Tạo Nên Sự Thành Công"')
    slogan_cell.font = Font(name="Segoe UI", size=10, bold=True, italic=True, color=STEEL_BLUE)
    slogan_cell.alignment = align_center
    slogan_cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Border for slogan
    for col in range(1, 8):
        ws1.cell(current_row, col).border = Border(
            top=Side(style="thin", color=STEEL_BLUE),
            bottom=Side(style="thin", color=STEEL_BLUE),
            left=Side(style="thin", color=STEEL_BLUE) if col == 1 else None,
            right=Side(style="thin", color=STEEL_BLUE) if col == 7 else None
        )

    current_row += 3
    
    # Signatures
    ws1.row_dimensions[current_row].height = 18
    ws1.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    s1 = ws1.cell(current_row, 1, "ỨNG VIÊN TUYỂN DỤNG")
    s1.font = font_bold
    s1.alignment = align_center
    
    ws1.merge_cells(start_row=current_row, start_column=5, end_row=current_row, end_column=7)
    s2 = ws1.cell(current_row, 5, "QUẢN ĐỐC / TỔ TRƯỞNG DUYỆT")
    s2.font = font_bold
    s2.alignment = align_center
    
    current_row += 1
    ws1.row_dimensions[current_row].height = 15
    ws1.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    sub1 = ws1.cell(current_row, 1, "(Ký và ghi rõ họ tên)")
    sub1.font = font_italic
    sub1.alignment = align_center
    
    ws1.merge_cells(start_row=current_row, start_column=5, end_row=current_row, end_column=7)
    sub2 = ws1.cell(current_row, 5, "(Xác nhận ngạch bậc lương ngày)")
    sub2.font = font_italic
    sub2.alignment = align_center
    
    current_row += 4
    ws1.row_dimensions[current_row].height = 18
    ws1.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    n1 = ws1.cell(current_row, 1, "[Họ tên ứng viên]")
    n1.font = font_normal
    n1.alignment = align_center
    
    ws1.merge_cells(start_row=current_row, start_column=5, end_row=current_row, end_column=7)
    n2 = ws1.cell(current_row, 5, "[Họ tên người đánh giá]")
    n2.font = font_normal
    n2.alignment = align_center

    # Add validations to Column E (Self-Declaration) & Column F (Company Eval) & Column E in trials
    dv_company = DataValidation(type="whole", operator="between", formula1="0", formula2="1", allow_blank=True)
    dv_company.prompt = 'Nhập 1: Đạt, 0: Chưa đạt'
    dv_company.promptTitle = 'Đánh giá'
    ws1.add_data_validation(dv_company)

    for r_num in eval_rows:
        dv_company.add(ws1[f"E{r_num}"])
        dv_company.add(ws1[f"F{r_num}"])
        
    for r_num in p_eval_rows:
        dv_company.add(ws1[f"E{r_num}"])

    # -------------------------------------------------------------
    # SHEET 2: THANG ĐO & LƯƠNG
    # -------------------------------------------------------------
    ws2 = wb.create_sheet(title="Thang Đo & Lương")
    ws2.sheet_view.showGridLines = True
    
    ws2.column_dimensions['A'].width = 10
    ws2.column_dimensions['B'].width = 24
    ws2.column_dimensions['C'].width = 24
    ws2.column_dimensions['D'].width = 24
    ws2.column_dimensions['E'].width = 65
    
    # Headers
    ws2["A1"] = "Bậc"
    ws2["B1"] = "Chức danh đề xuất"
    ws2["C1"] = "Lương thực lĩnh (ngày)"
    ws2["D1"] = "Tổng chi phí công ty (ngày)"
    ws2["E1"] = "Phân công / Vai trò chính trong xưởng"
    
    for col in ["A", "B", "C", "D", "E"]:
        cell = ws2[f"{col}1"]
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_all_thin
        
    scale_data = [
        (1, "Thợ học việc (Cấp 1)", 220000, 300000, "Phụ việc vệ sinh, mài thô, nâng hạ phôi dưới sự chỉ dẫn"),
        (2, "Thợ học việc (Cấp 2)", 220000, 300000, "Thực hiện tốt 5S khu vực hàn gá, đảm bảo an toàn ATLĐ xưởng"),
        (3, "Thợ phụ cơ bản (Cấp 1)", 255000, 335000, "Biết dùng máy cắt cầm tay, khoan tay, thước đo chính xác"),
        (4, "Thợ phụ mài tinh (Cấp 2)", 290000, 370000, "Mài tinh mỹ bề mặt, mài lấy góc sắc nét, mài phá bavia"),
        (5, "Thợ sơn & gá phụ", 350000, 430000, "Đứng buồng sơn phun khí nén; nắm vật tư và hỗ trợ hàn đính gá"),
        (6, "Thợ gá lắp phôi", 380000, 460000, "Nắm rõ các loại vật tư, phôi cắt sắt/inox, chuẩn bị công trình"),
        (7, "Thợ gá hàn liên hợp", 405000, 485000, "Sử dụng thành thạo máy hàn que/MIG/TIG để đính gá kết cấu"),
        (8, "Thợ hàn chính chuyên nghiệp", 430000, 510000, "Đọc bản vẽ sản xuất và tự hàn độc lập sản phẩm chất lượng cao"),
        (9, "Thợ lắp ráp lành nghề", 440000, 520000, "Đọc bản vẽ cơ khí phức tạp, tự bóc tách khai triển phôi gá ráp"),
        (10, "Thợ hoàn thiện độc lập", 470000, 550000, "Tự kiểm soát chất lượng hoàn thiện sản phẩm không cần giám sát"),
        (11, "Thợ dự toán & tiến độ", 480000, 560000, "Ước lượng thời gian/nhân công, lập tiến độ thi công dự án nhỏ"),
        (12, "Tổ phó kỹ thuật", 490000, 570000, "Tự chủ động điều chỉnh chi tiết nhỏ ở hiện trường, thay thế tổ trưởng"),
        (13, "Trưởng nhóm lắp dựng", 500000, 580000, "Điều phối nhóm 2-3 người lắp dựng kết cấu thép/inox ngoài công trường"),
        (14, "Tổ trưởng sản xuất", 550000, 630000, "Phân việc toàn xưởng, chịu trách nhiệm tiến độ & chất lượng tổ đội")
    ]
    
    for r_idx, row_val in enumerate(scale_data, start=2):
        ws2.cell(r_idx, 1, row_val[0]).alignment = align_center
        ws2.cell(r_idx, 1).font = font_bold
        
        ws2.cell(r_idx, 2, row_val[1]).alignment = align_left
        ws2.cell(r_idx, 2).font = font_bold
        
        ws2.cell(r_idx, 3, row_val[2]).number_format = '#,##0" đ"'
        ws2.cell(r_idx, 3).alignment = align_right
        ws2.cell(r_idx, 3).font = font_bold
        
        ws2.cell(r_idx, 4, row_val[3]).number_format = '#,##0" đ"'
        ws2.cell(r_idx, 4).alignment = align_right
        ws2.cell(r_idx, 4).font = font_normal
        
        ws2.cell(r_idx, 5, row_val[4]).alignment = align_left
        ws2.cell(r_idx, 5).font = font_normal
        
        for col in range(1, 6):
            cell = ws2.cell(r_idx, col)
            cell.border = border_all_thin
            if r_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

    # -------------------------------------------------------------
    # SHEET 3: HƯỚNG DẪN SỬ DỤNG
    # -------------------------------------------------------------
    ws3 = wb.create_sheet(title="Hướng Dẫn Sử Dụng")
    ws3.sheet_view.showGridLines = True
    
    ws3.column_dimensions['A'].width = 8
    ws3.column_dimensions['B'].width = 30
    ws3.column_dimensions['C'].width = 75
    
    ws3["B2"] = "HƯỚNG DẪN SỬ DỤNG BỘ CÔNG CỤ TUYỂN DỤNG THỢ CƠ KHÍ"
    ws3["B2"].font = Font(name="Segoe UI", size=14, bold=True, color=STEEL_BLUE)
    
    guide_steps = [
        ("Bước 1", "Phát Form Tự Khai", "HR hoặc Kế toán phát bản in hoặc mở file Excel cho Ứng viên tự nhập '1' (Có thể làm) hoặc '0' (Chưa làm) vào Cột E (ỨNG VIÊN TỰ KHAI) để sàng lọc nhanh lý thuyết."),
        ("Bước 2", "Phỏng vấn nhanh (HR/Kế toán)", "HR hoặc Kế toán hỏi ứng viên các câu hỏi gợi ý ở Cột D (HƯỚNG DẪN KIỂM TRA). Việc này giúp bất kỳ ai cũng có thể phỏng vấn sơ bộ thợ mà không cần có chuyên môn cơ khí."),
        ("Bước 3", "Thử tay nghề thực chiến", "Quản đốc đưa ứng viên xuống xưởng thực hành 5 bài thử tay nghề thực tế (Phần II - Hàng 36 đến 40). Tích '1' (Đạt) hoặc '0' (Không đạt) vào ô kết quả thực chiến tương ứng."),
        ("Bước 4", "Nhập điểm và Nhận kết quả", "Quản đốc hoặc Kế toán nhập kết quả kiểm tra vào Cột F (THỰC CHIẾN ĐẠT). Excel sẽ tự động đối chiếu định mức lương và đề xuất chức danh, mức lương ngày công chuẩn xác ở Bảng Kết Quả phía dưới."),
        ("Lưu ý cho Sếp", "Điều chỉnh mức lương", "Nếu Sếp muốn điều chỉnh định mức lương hay chức danh của các bậc thợ, Sếp chỉ cần mở Sheet 'Thang Đo & Lương' để nhập số tiền mới. Bảng tính chính sẽ tự động cập nhật theo mà Sếp không cần chỉnh sửa bất kỳ công thức nào.")
    ]
    
    ws3["A4"] = "Bước"
    ws3["B4"] = "Tên hoạt động"
    ws3["C4"] = "Nội dung chi tiết quy trình thực hiện"
    for col in ["A", "B", "C"]:
        ws3[f"{col}4"].font = font_header
        ws3[f"{col}4"].fill = fill_header
        ws3[f"{col}4"].alignment = align_center
        ws3[f"{col}4"].border = border_all_thin
        
    for r_idx, (step, act, desc) in enumerate(guide_steps, start=5):
        ws3.cell(r_idx, 1, step).alignment = align_center
        ws3.cell(r_idx, 1).font = font_bold
        ws3.cell(r_idx, 1).border = border_all_thin
        
        ws3.cell(r_idx, 2, act).alignment = align_left
        ws3.cell(r_idx, 2).font = font_bold
        ws3.cell(r_idx, 2).border = border_all_thin
        
        ws3.cell(r_idx, 3, desc).alignment = align_left
        ws3.cell(r_idx, 3).font = font_normal
        ws3.cell(r_idx, 3).border = border_all_thin
        ws3.row_dimensions[r_idx].height = 42

    # Save to the new destination filename
    output_path = "d:/Sao Vàng/Website-SaoVang/CKSV_Bo_Cong_Cu_Tuyen_Dung_Tho_Co_Khi.xlsx"
    wb.save(output_path)
    print(f"New comprehensive recruitment kit created successfully at {output_path}")

if __name__ == "__main__":
    create_recruitment_kit()
