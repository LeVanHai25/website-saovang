import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.drawing.image import Image as OpenpyxlImage

def upgrade_recruitment_template():
    wb = openpyxl.Workbook()
    
    # -------------------------------------------------------------
    # STYLES CONFIGURATION
    # -------------------------------------------------------------
    STEEL_BLUE = "2F5597"
    ICE_BLUE = "D9E1F2"
    LIGHT_GRAY = "F2F2F2"
    BORDER_GRAY = "D9D9D9"
    YELLOW_INPUT = "FFF2CC"
    DARK_GREEN = "385723"
    LIGHT_GREEN = "E2EFDA"
    
    # Fonts
    font_company_name = Font(name="Segoe UI", size=12, bold=True, color=STEEL_BLUE)
    font_company_info = Font(name="Segoe UI", size=9, italic=True, color="595959")
    font_title_main = Font(name="Segoe UI", size=15, bold=True, color=STEEL_BLUE)
    font_title_sub = Font(name="Segoe UI", size=10, italic=True, color="7F7F7F")
    font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    font_group_header = Font(name="Segoe UI", size=11, bold=True, color=STEEL_BLUE)
    font_bold = Font(name="Segoe UI", size=10, bold=True)
    font_normal = Font(name="Segoe UI", size=10)
    font_italic = Font(name="Segoe UI", size=9, italic=True)
    font_result = Font(name="Segoe UI", size=12, bold=True, color=DARK_GREEN)
    
    # Fills
    fill_header = PatternFill(start_color=STEEL_BLUE, end_color=STEEL_BLUE, fill_type="solid")
    fill_group = PatternFill(start_color=ICE_BLUE, end_color=ICE_BLUE, fill_type="solid")
    fill_input = PatternFill(start_color=YELLOW_INPUT, end_color=YELLOW_INPUT, fill_type="solid")
    fill_result_box = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
    fill_sum = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
    
    # Alignments
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    
    # Borders
    thin_side = Side(style="thin", color=BORDER_GRAY)
    border_all_thin = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    border_group = Border(
        left=Side(style="thin", color=STEEL_BLUE),
        right=Side(style="thin", color=STEEL_BLUE),
        top=Side(style="thin", color=STEEL_BLUE),
        bottom=Side(style="thin", color=STEEL_BLUE)
    )

    logo_path = "d:/Sao Vàng/Website-SaoVang/Logo/LogoCTY + SV Aluminium/LogoCTY.png"

    def apply_letterhead(ws, max_col_letter):
        # Merge cells for company info
        ws.row_dimensions[2].height = 20
        ws.row_dimensions[3].height = 15
        ws.row_dimensions[4].height = 15
        
        ws.merge_cells(f"C2:{max_col_letter}2")
        ws["C2"] = "CÔNG TY CỔ PHẦN SẢN XUẤT CƠ KHÍ SAO VÀNG (ASV)"
        ws["C2"].font = font_company_name
        ws["C2"].alignment = Alignment(horizontal="left", vertical="center")
        
        ws.merge_cells(f"C3:{max_col_letter}3")
        ws["C3"] = "Địa chỉ: Tầng 3, TT7-35 Khu đô thị Văn Phú, phường Kiến Hưng, TP Hà Nội, Việt Nam."
        ws["C3"].font = font_company_info
        ws["C3"].alignment = Alignment(horizontal="left", vertical="center")
        
        ws.merge_cells(f"C4:{max_col_letter}4")
        ws["C4"] = "Hotline: 0869 590 279  |  Email: cokhisaovangvn@gmail.com"
        ws["C4"].font = font_company_info
        ws["C4"].alignment = Alignment(horizontal="left", vertical="center")

        try:
            img = OpenpyxlImage(logo_path)
            img.width = 110
            img.height = 45
            ws.add_image(img, "A2")
        except Exception as e:
            print(f"Error adding logo: {e}")

    # -------------------------------------------------------------
    # SHEET 1: THỢ CHÍNH VÀ THỢ PHỤ (UPGRADED MAIN FORM)
    # -------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Tuyển Dụng Thợ Cơ Khí"
    ws1.sheet_view.showGridLines = True
    
    ws1.column_dimensions['A'].width = 6   # STT
    ws1.column_dimensions['B'].width = 20  # NHIỆM VỤ (VAI TRÒ)
    ws1.column_dimensions['C'].width = 18  # THIẾT BỊ / PHÂN LOẠI
    ws1.column_dimensions['D'].width = 54  # KỸ NĂNG LÀM VIỆC CHI TIẾT
    ws1.column_dimensions['E'].width = 42  # KẾT QUẢ BÊN SDLĐ YÊU CẦU ĐẠT ĐƯỢC
    ws1.column_dimensions['F'].width = 16  # KHẢ NĂNG ĐÁP ỨNG (% Tự đánh giá)
    ws1.column_dimensions['G'].width = 16  # CÔNG TY ĐÁNH GIÁ (Đạt: 1 | 0)
    ws1.column_dimensions['H'].width = 25  # GHI CHÚ NHẬN XÉT

    apply_letterhead(ws1, "H")

    # Titles
    ws1.row_dimensions[6].height = 24
    ws1.merge_cells("A6:H6")
    ws1["A6"] = "BẢNG TUYỂN DỤNG VÀ PHÂN CẤP THỢ CƠ KHÍ DÂN DỤNG"
    ws1["A6"].font = font_title_main
    ws1["A6"].alignment = align_center

    ws1.row_dimensions[7].height = 18
    ws1.merge_cells("A7:H7")
    ws1["A7"] = "Hệ thống tiêu chí đánh giá thợ theo cấp bậc: Học việc -> Phụ việc -> Giúp việc -> Thợ chính -> Tổ phó -> Tổ trưởng"
    ws1["A7"].font = font_title_sub
    ws1["A7"].alignment = align_center

    # Candidate profile block
    ws1["B9"] = "Họ và tên ứng viên:"
    ws1["B9"].font = font_bold
    ws1["C9"] = "[Nhập họ tên]"
    ws1["C9"].font = font_normal
    ws1["C9"].border = Border(bottom=Side(style="thin", color="000000"))

    ws1["F9"] = "Ngày đánh giá:"
    ws1["F9"].font = font_bold
    ws1["G9"] = "[Ghi ngày]"
    ws1["G9"].font = font_normal
    ws1["G9"].border = Border(bottom=Side(style="thin", color="000000"))

    ws1["B10"] = "Quê quán/Năm sinh:"
    ws1["B10"].font = font_bold
    ws1["C10"] = "[Quê quán/Năm sinh]"
    ws1["C10"].font = font_normal
    ws1["C10"].border = Border(bottom=Side(style="thin", color="000000"))

    ws1["F10"] = "Số ĐT/CCCD:"
    ws1["F10"].font = font_bold
    ws1["G10"] = "[Số ĐT/CCCD]"
    ws1["G10"].font = font_normal
    ws1["G10"].border = Border(bottom=Side(style="thin", color="000000"))

    ws1.row_dimensions[11].height = 15

    # Main Table Header (Row 12)
    ws1.row_dimensions[12].height = 28
    headers_ws1 = [
        ("A12", "STT", align_center),
        ("B12", "NHIỆM VỤ (VAI TRÒ)", align_left),
        ("C12", "THIẾT BỊ / PHÂN LOẠI", align_center),
        ("D12", "KỸ NĂNG LÀM VIỆC CHI TIẾT (YÊU CẦU THỰC TẾ)", align_left),
        ("E12", "KẾT QUẢ YÊU CẦU CẦN ĐẠT ĐƯỢC (BÊN SDLĐ)", align_left),
        ("F12", "KHẢ NĂNG ĐÁP ỨNG (% Tự đánh giá)", align_center),
        ("G12", "CÔNG TY ĐÁNH GIÁ (Đạt: 1 | 0)", align_center),
        ("H12", "GHI CHÚ / NHẬN XÉT CHI TIẾT", align_left)
    ]
    for cell_ref, text, align in headers_ws1:
        cell = ws1[cell_ref]
        cell.value = text
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align
        cell.border = border_all_thin

    # Structured tasks from THỢ CHÍNH VÀ THỢ PHỤ
    # Format: STT, Vai trò, Thiết bị/Phân loại, Kỹ năng, Kết quả SDLĐ
    tasks_data = [
        ("1", "NV Học việc", "Ý thức & Sức khỏe", "Sức khỏe tốt, nhanh nhẹn, chịu khó học hỏi, có trách nhiệm, phối hợp được với các vị trí trong xưởng.", "Nề nếp, tác phong tốt, đi làm đầy đủ đúng giờ"),
        ("2", "NV Học việc", "Vệ sinh xưởng", "Chủ động thu dọn máy móc dụng cụ làm việc, quét dọn vệ sinh khu vực thi công sau mỗi buổi làm và khi kết thúc công việc.", "Xưởng sạch sẽ, máy móc dụng cụ được sắp xếp đúng nơi quy định"),
        
        ("3", "NV Phụ việc", "Dụng cụ cơ bản", "Nhận biết và có thể sử dụng các loại máy móc và dụng cụ cơ bản trong xưởng sản xuất tùy từng nhu cầu công việc (máy cắt tay, máy cắt bàn, máy khoan, thước đo khoảng cách, dây điện...vv)", "Vận hành an toàn các dụng cụ cơ bản dưới sự giám sát"),
        ("4", "NV Phụ việc", "Máy mài", "Mài phá (yêu cầu cho các vị trí mối hàn trên các bề mặt vật liệu thô, kết cấu hạng nặng, hoặc các mối hàn không đạt tiêu chuẩn về thẩm mỹ, cần tinh chỉnh lại).", "Sạch bavia thô, không còn xỉ hàn nhọn"),
        ("5", "NV Phụ việc", "Máy mài", "Mài lấy mặt phẳng (yêu cầu chung cho hầu hết các sản phẩm cần độ tinh mỹ cao).", "Mối hàn phẳng mượt, sờ không bị gờ tay"),
        ("6", "NV Phụ việc", "Máy mài", "Mài lấy góc (yêu cầu cho các sản phẩm được tổ hợp từ nguyên liệu tấm, hình vv, tổ hợp thành sản phẩm theo thiết kế).", "Góc bo đều, sắc nét theo đúng dưỡng lắp ráp"),
        ("7", "NV Phụ việc", "Máy mài", "Mài hoàn thiện sản phẩm theo yêu cầu về độ tinh mỹ của sản phẩm sắt đen, sắt mạ, inox và vệ sinh bề mặt sản phẩm sau khi hoàn thành phần thô.", "Bề mặt nhẵn bóng, sạch rỉ sét và dầu mỡ"),
        ("8", "NV Phụ việc", "Máy mài", "Mài tinh mỹ (yêu cầu cho sp có độ thẩm mỹ cao, cần mài lấy mặt phẳng tuyệt đối hoặc mài lấy góc sắc nét nếu chi tiết được gá, hàn đạt tiêu chuẩn).", "Góc vuông phẳng đét đạt thẩm mỹ cao"),
        ("9", "NV Phụ việc", "Máy mài", "Mài đánh bóng (yêu cầu cho các sản phẩm sử dụng nguyên liệu inox bóng, hoặc cần làm bóng, mịn cho sp ngoài inox).", "Đạt độ bóng sáng, mịn đều, không xước"),
        ("10", "NV Phụ việc", "Máy cắt", "Có thể sử dụng máy cắt tay, cắt bàn cắt các loại vật liệu không yêu cầu độ chính xác tuyệt đối.", "Cắt đúng kích thước, không lệch góc"),
        ("11", "NV Phụ việc", "Máy khoan", "Có thể sử dụng khoan bắn vít, khoan bê tông cho các phần việc liên quan đến các loại dụng cụ đó.", "Khoan lỗ gọn gàng, định vị chắc chắn"),
        ("12", "NV Phụ việc", "Máy phun sơn", "Sử dụng thành thạo súng phun sơn khí nén, sơn phủ và sơn hoàn thiện sản phẩm cần xử lý sơn.", "Sơn đều màu, mịn, không chảy sệ"),
        
        ("13", "NV Giúp việc", "An toàn & Bao quát", "Có khả năng bao quát quản lý công cụ máy móc khi làm việc & phán đoán tình huống gây mất an toàn lao động có thể xảy ra để kịp thời thông báo và xử lý.", "Phòng ngừa tai nạn, bảo vệ an toàn người và tài sản"),
        ("14", "NV Giúp việc", "Chuẩn bị vật tư", "Nắm bắt và hiểu rõ công năng cho từng loại máy móc, chủ động tính toán chuẩn bị đầy đủ vật tư phụ, máy móc dùng cho công trình khi đã được thông báo đầu việc.", "Đầy đủ dụng cụ đi công trình, không thiếu hụt"),
        ("15", "NV Giúp việc", "Lấy góc vuông", "Sử dụng thành thạo thước mét, công cụ hỗ trợ lấy góc vuông mặt phẳng thăng bằng trong khi sản xuất sản phẩm tại xưởng và công trường.", "Sản phẩm đạt độ thăng bằng, vuông góc ke chuẩn"),
        ("16", "NV Giúp việc", "Chủng loại vật tư", "Nắm rõ cơ bản các loại vật tư, chủng loại vật liệu cơ khí.", "Nhận dạng chính xác độ dày, kích thước thanh thép/inox"),
        ("17", "NV Giúp việc", "Cắt góc mỹ thuật", "Cắt vật liệu bằng máy cắt bàn hoặc cầm tay theo mọi góc độ đúng với kích thước chỉ định hoặc theo bản vẽ kỹ thuật.", "Mép cắt khít góc ghép, chuẩn số đo"),
        ("18", "NV Giúp việc", "Hàn đính gá thô", "Có thể sử dụng máy hàn hàn đính 1 số sản phẩm thô không yêu cầu độ chính xác tuyệt đối.", "Liên kết đính giữ chắc chắn, không lệch phôi"),
        
        ("19", "Thợ chính", "Lập kế hoạch & Bản vẽ", "Biết đọc và triển khai chính xác các thông số trên bản vẽ; ước lượng chuẩn xác 95% lượng nhân công và thời gian để hoàn thành đơn hàng được giao; tự lập triển khai.", "Đảm bảo tiến độ, chất lượng và hao phí vật tư thấp"),
        ("20", "Thợ chính", "Hàn que", "Sử dụng thành thạo máy hàn que (MMA) trên các kết cấu dầm cột thép dày.", "Mối hàn ngấu sâu, chịu lực cao, xỉ bong sạch"),
        ("21", "Thợ chính", "Hàn MIG", "Sử dụng thành thạo máy hàn MIG/MAG bán tự động.", "Mối hàn vảy đều, tốc độ nhanh, ngấu sâu"),
        ("22", "Thợ chính", "Hàn TIG", "Sử dụng thành thạo máy hàn TIG khí bảo vệ.", "Mối hàn trắng sáng, mịn vảy cá, không thủng phôi mỏng"),
        ("23", "Thợ chính", "Hàn Laser", "Sử dụng thành thạo máy hàn laser (cho sắt thép, nhôm, inox).", "Mối hàn cực mịn, không biến dạng nhiệt vật liệu"),
        ("24", "Thợ chính", "Dân dụng tổng hợp", "Biết sử dụng cơ bản các loại máy và phụ kiện đi kèm trong nghề cơ khí dân dụng, xây dựng.", "Hoàn thành tốt các việc lắp đặt xây dựng thực tế"),
        ("25", "Thợ chính", "Hoàn thiện & QC", "Kỹ năng hoàn thiện và đánh giá chất lượng sản phẩm sau khi hoàn thiện.", "Sản phẩm xuất xưởng không lỗi kỹ thuật/thẩm mỹ"),
        
        ("26", "Tổ phó", "Khai triển bản vẽ", "Đọc hiểu bản vẽ sản xuất và có thể triển khai công việc trên bản vẽ khi được giao; tính toán khai triển khoa học, hiệu quả; đề xuất phương án cải tiến.", "Tối ưu hóa năng suất gia công toàn đội"),
        ("27", "Tổ phó", "Độc lập điều chỉnh", "Có thể độc lập trong công việc khi giao; tự quyết định điều chỉnh, thay đổi phương án thi công phù hợp với thực tế không ảnh hưởng chất lượng.", "Xử lý nhanh các phát sinh hiện trường"),
        ("28", "Tổ phó", "Làm việc nhóm", "Khả năng làm việc nhóm tốt, đảm bảo sản xuất các sản phẩm đúng thiết kế, đúng tiến độ.", "Phối hợp nhịp nhàng, tăng tinh thần đoàn kết"),
        
        ("29", "Tổ trưởng", "Lãnh đạo & Điều phối", "Khả năng lãnh đạo đội nhóm, bao quát, điều phối, bố trí công việc phù hợp cho từng thành viên; cam kết thực hiện đúng tiến độ được giao.", "Bố trí đúng người đúng việc, hoàn thành dự án"),
        ("30", "Tổ trưởng", "Báo cáo & Khắc phục", "Báo cáo tình hình tiến độ vào mỗi cuối ngày và các vấn đề vướng mắc gặp phải; đề xuất phương án sửa lỗi nhanh chóng.", "Ban giám đốc nắm bắt thông tin kịp thời, thông suốt"),
        ("31", "Tổ trưởng", "Làm việc với CĐT", "Thay mặt ban giám đốc làm việc trực tiếp với Chủ đầu tư về kỹ thuật, thi công lắp đặt và các thay đổi phát sinh tại hiện trường công trường.", "Tạo uy tín tốt với khách hàng, xử lý việc trôi chảy")
    ]

    current_row = 13
    eval_rows = []
    
    for row in tasks_data:
        stt, role, category, skill, result_req = row
        ws1.row_dimensions[current_row].height = 42
        
        ws1.cell(current_row, 1, int(stt)).alignment = align_center
        ws1.cell(current_row, 1).font = font_normal
        
        ws1.cell(current_row, 2, role).font = font_bold
        ws1.cell(current_row, 2).alignment = align_center
        
        ws1.cell(current_row, 3, category).font = font_bold
        ws1.cell(current_row, 3).alignment = align_center
        
        ws1.cell(current_row, 4, skill).font = font_normal
        ws1.cell(current_row, 4).alignment = align_left
        
        ws1.cell(current_row, 5, result_req).font = font_normal
        ws1.cell(current_row, 5).alignment = align_left
        
        # Candidate self-declaration (% input)
        cell_self = ws1.cell(current_row, 6, 1.0)
        cell_self.font = font_bold
        cell_self.alignment = align_center
        cell_self.fill = fill_input
        cell_self.number_format = '0%'
        
        # Company evaluation (1/0 input)
        cell_comp = ws1.cell(current_row, 7, 0)
        cell_comp.font = font_bold
        cell_comp.alignment = align_center
        cell_comp.fill = fill_input
        eval_rows.append(current_row)
        
        # Note
        cell_note = ws1.cell(current_row, 8, "[Ghi chú]")
        cell_note.font = font_italic
        cell_note.alignment = align_left
        cell_note.fill = fill_input
        
        for col in range(1, 9):
            ws1.cell(current_row, col).border = border_all_thin
            if current_row % 2 == 0:
                if col not in [6, 7, 8]:
                    ws1.cell(current_row, col).fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
                    
        current_row += 1

    # Place sample test scores (probation pass up to Thợ chính)
    for r in range(13, 31): # Up to row 30 (task 18 - NV Giúp việc)
        ws1.cell(r, 7, 1) # Company evaluates as passed
    for r in range(31, 44): # Above row 30 (task 19+ Thợ chính/Tổ trưởng)
        ws1.cell(r, 7, 0) # Company evaluates as not passed yet

    ws1.row_dimensions[current_row].height = 15
    current_row += 1

    # SUMMARY DASHBOARD PANEL AT BOTTOM OF SHEET 1
    ws1.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=8)
    r_hdr = ws1.cell(current_row, 2, "BẢNG ĐỀ XUẤT PHÂN LOẠI CẤP BẬC LƯƠNG SAU PHỎNG VẤN")
    r_hdr.font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    r_hdr.fill = PatternFill(start_color=STEEL_BLUE, end_color=STEEL_BLUE, fill_type="solid")
    r_hdr.alignment = align_center
    ws1.row_dimensions[current_row].height = 24
    
    for col in range(2, 9):
        ws1.cell(current_row, col).border = Border(top=Side(style="medium", color=STEEL_BLUE))
        
    current_row += 1
    
    # 1. Tổng mục đạt
    ws1.row_dimensions[current_row].height = 20
    ws1.cell(current_row, 2, "Tổng số hạng mục phỏng vấn đạt:").font = font_bold
    ws1.cell(current_row, 2).alignment = align_left
    
    cell_total = ws1.cell(current_row, 3, f"=SUM(G13:G43)")
    cell_total.font = font_bold
    cell_total.alignment = align_center
    ws1.cell(current_row, 4, "/ 31 mục tiêu chí").font = font_italic
    ws1.cell(current_row, 4).alignment = align_left
    
    for col in range(2, 9):
        cell = ws1.cell(current_row, col)
        cell.fill = fill_result_box
        left_s = Side(style="medium", color=STEEL_BLUE) if col == 2 else None
        right_s = Side(style="medium", color=STEEL_BLUE) if col == 8 else None
        cell.border = Border(left=left_s, right=right_s)
        
    sum_ref = f"C{current_row}" # C46
    current_row += 1
    
    # 2. Xếp loại cấp bậc tuyển đề xuất
    ws1.row_dimensions[current_row].height = 20
    ws1.cell(current_row, 2, "Xếp hạng cấp bậc đề xuất:").font = font_bold
    ws1.cell(current_row, 2).alignment = align_left
    
    cell_level = ws1.cell(current_row, 3, f'=IF({sum_ref}>=28, "Tổ Trưởng Sản Xuất", IF({sum_ref}>=25, "Tổ Phó Kỹ Thuật", IF({sum_ref}>=18, "Thợ Chính Lành Nghề", IF({sum_ref}>=12, "Thợ Giúp Việc Cứng", IF({sum_ref}>=2, "Thợ Phụ Việc Xưởng", "Thợ Học Việc")))))')
    cell_level.font = font_result
    cell_level.alignment = align_center
    
    ws1.cell(current_row, 4, "(Tự động tính từ số mục đạt)").font = font_italic
    ws1.cell(current_row, 4).alignment = align_left
    
    for col in range(2, 9):
        cell = ws1.cell(current_row, col)
        cell.fill = fill_result_box
        left_s = Side(style="medium", color=STEEL_BLUE) if col == 2 else None
        right_s = Side(style="medium", color=STEEL_BLUE) if col == 8 else None
        cell.border = Border(left=left_s, right=right_s, bottom=Side(style="medium", color=STEEL_BLUE))

    current_row += 2

    # Slogan block matching company motto
    ws1.row_dimensions[current_row].height = 26
    ws1.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
    slogan_cell = ws1.cell(current_row, 1, 'Khẩu hiệu của chúng tôi: "Chất Lượng Tạo Nên Thương Hiệu. Uy Tín Tạo Nên Sự Thành Công"')
    slogan_cell.font = Font(name="Segoe UI", size=10, bold=True, italic=True, color=STEEL_BLUE)
    slogan_cell.alignment = align_center
    slogan_cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    for col in range(1, 9):
        ws1.cell(current_row, col).border = Border(
            top=Side(style="thin", color=STEEL_BLUE),
            bottom=Side(style="thin", color=STEEL_BLUE),
            left=Side(style="thin", color=STEEL_BLUE) if col == 1 else None,
            right=Side(style="thin", color=STEEL_BLUE) if col == 8 else None
        )

    current_row += 3
    
    # Signatures
    ws1.row_dimensions[current_row].height = 18
    ws1.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    s1 = ws1.cell(current_row, 1, "ỨNG VIÊN TUYỂN DỤNG")
    s1.font = font_bold
    s1.alignment = align_center
    
    ws1.merge_cells(start_row=current_row, start_column=5, end_row=current_row, end_column=8)
    s2 = ws1.cell(current_row, 5, "PHÒNG NHÂN SỰ XÁC NHẬN")
    s2.font = font_bold
    s2.alignment = align_center
    
    current_row += 1
    ws1.row_dimensions[current_row].height = 15
    ws1.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    sub1 = ws1.cell(current_row, 1, "(Ký và ghi rõ họ tên)")
    sub1.font = font_italic
    sub1.alignment = align_center
    
    ws1.merge_cells(start_row=current_row, start_column=5, end_row=current_row, end_column=8)
    sub2 = ws1.cell(current_row, 5, "(Ký duyệt phân xếp bậc thử việc)")
    sub2.font = font_italic
    sub2.alignment = align_center

    # Add validations
    dv_self = DataValidation(type="decimal", operator="between", formula1="0", formula2="1", allow_blank=True)
    dv_self.prompt = 'Nhập % khả năng đáp ứng (Ví dụ: 80%)'
    dv_self.promptTitle = 'Ứng viên tự khai'
    ws1.add_data_validation(dv_self)
    
    dv_comp = DataValidation(type="whole", operator="between", formula1="0", formula2="1", allow_blank=True)
    dv_comp.prompt = 'Nhập 1: Đạt, 0: Chưa đạt'
    dv_comp.promptTitle = 'Công ty chấm'
    ws1.add_data_validation(dv_comp)
    
    for r in eval_rows:
        dv_self.add(ws1[f"F{r}"])
        dv_comp.add(ws1[f"G{r}"])

    # -------------------------------------------------------------
    # SHEET 2: Thợ, gốc (CLEANUP & RETENTION OF ORIGINAL STRUCTURE)
    # -------------------------------------------------------------
    ws2 = wb.create_sheet(title="Thợ, gốc")
    ws2.sheet_view.showGridLines = True
    
    ws2.column_dimensions['A'].width = 6   # STT
    ws2.column_dimensions['B'].width = 62  # NHIỆM VỤ
    ws2.column_dimensions['C'].width = 62  # KẾT QUẢ BÊN SDLĐ
    ws2.column_dimensions['D'].width = 16  # KHẢ NĂNG ĐÁP ỨNG
    ws2.column_dimensions['E'].width = 25  # NHẬN XÉT KẾT QUẢ

    apply_letterhead(ws2, "E")

    ws2.row_dimensions[6].height = 24
    ws2.merge_cells("A6:E6")
    ws2["A6"] = "BẢNG TUYỂN DỤNG THỢ CƠ KHÍ DÂN DỤNG (BẢN GỐC)"
    ws2["A6"].font = font_title_main
    ws2["A6"].alignment = align_center

    # Metadata
    ws2.merge_cells("A8:E8")
    ws2["A8"] = "Người ứng tuyển: Họ tên, năm sinh, quê quán, số điện thoại..."
    ws2["A8"].font = font_italic
    ws2["A8"].alignment = align_left

    # Header
    ws2.row_dimensions[10].height = 28
    headers_ws2 = [
        ("A10", "Stt", align_center),
        ("B10", "NHIỆM VỤ (VIỆC CẦN LÀM)", align_left),
        ("C10", "KẾT QUẢ BÊN SDLĐ (CẦN ĐẠT ĐƯỢC)", align_left),
        ("D10", "KHẢ NĂNG ĐÁP ỨNG", align_center),
        ("E10", "NHẬN XÉT KẾT QUẢ SAU THỬ VIỆC", align_left)
    ]
    for cell_ref, text, align in headers_ws2:
        cell = ws2[cell_ref]
        cell.value = text
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align
        cell.border = border_all_thin

    # Keep original 15 items from sheet 1 ("Thợ, gốc")
    original_tasks = [
        ("1", "Sức khỏe tốt, nhanh nhẹn, chịu khó học hỏi, có trách nhiệm, phối hợp được với các vị trí trong xưởng.", "Nề nếp, tác phong tốt, đi làm đầy đủ đúng giờ"),
        ("2", "Biết sử dụng các loại máy móc và dụng cụ cơ bản trong xưởng sản xuất: máy cắt tay, máy cắt bàn, máy khoan, thước đo khoảng cách, dây điện...vv.", "Sử dụng an toàn các thiết bị cơ bản"),
        ("3", "Có thể sử dụng máy mài để mài hoàn thiện sản phẩm theo yêu cầu về độ tinh mỹ của sản phẩm sắt đen, sắt mạ, inox và vệ sinh bề mặt sản phẩm sau khi hoàn thành phần thô (mài bề mặt, mài góc, đánh bóng, đánh rỉ).", "Mối hàn phẳng mượt, góc bo đạt mỹ thuật"),
        ("4", "Sử dụng được các loại máy cơ bản trong công việc máy hàn, máy cắt tay, máy cắt bàn, máy khoan, máy cân bằng laze, thước đo khoảng cách...vv.", "Vận hành thành thạo an toàn"),
        ("5", "Nhận biết được đơn vị đo lường, có thể cắt phôi chính xác theo kích thước được giao.", "Sai số cắt phôi trong phạm vi cho phép"),
        ("6", "Có thể sử dụng máy hàn để hàn đính gá các vị trí cần thiết.", "Liên kết đính ráp giữ chắc phôi"),
        ("7", "Sử dụng thành thạo súng phun sơn khí nén, sơn phủ và sơn hoàn thiện sản phẩm sau quá trình hàn, mài tinh mỹ.", "Màng sơn láng mịn bóng màu"),
        ("8", "Nắm bắt và hiểu rõ hết công năng cho từng loại máy móc, chủ động tính toán chuẩn bị đầy đủ vt phụ, máy móc dùng cho công trình khi đã được thông báo đầu việc.", "Đầy đủ dụng cụ thi công"),
        ("9", "Chủ động thu dọn máy móc dụng cụ làm việc, quét dọn vệ sinh khu vực thi công sau mỗi buổi làm và khi kết thúc công việc ở xưởng cũng như đi công trường.", "Khu vực thi công gọn gàng ngăn nắp"),
        ("10", "Sử dụng thành thạo các loại máy hàn cơ bản: Hàn TIG, Hàn MIG, Hàn que.", "Đường hàn ngấu sâu thẩm mỹ"),
        ("11", "Biết sử dụng cơ bản các loại máy và phụ kiện đi kèm trong nghề cơ khí dân dụng, xây dựng.", "Phối hợp tốt ngoài công trường"),
        ("12", "Nắm rõ cơ bản các loại vật tư, chủng loại vật liệu.", "Nhận diện đúng thép hộp/inox"),
        ("13", "Đọc hiểu bản vẽ sản xuất và có thể triển khai công việc trên bản vẽ khi được giao, biết cách tính toán khai triển khoa học, hiệu quả.", "Thi công đúng thiết kế bản vẽ kỹ thuật"),
        ("14", "Có thể độc lập trong công việc khi giao, có thể tự quyết định điều chỉnh.", "Chủ động xử lý tốt đầu việc đơn lẻ"),
        ("15", "Khả năng làm việc nhóm, đảm bảo sản xuất các sản phẩm đúng thiết kế, đúng tiến độ.", "Phối hợp nhóm nhịp nhàng ăn ý")
    ]

    curr_row = 11
    for row in original_tasks:
        ws2.row_dimensions[curr_row].height = 42
        ws2.cell(curr_row, 1, int(row[0])).alignment = align_center
        ws2.cell(curr_row, 1).border = border_all_thin
        
        ws2.cell(curr_row, 2, row[1]).alignment = align_left
        ws2.cell(curr_row, 2).border = border_all_thin
        ws2.cell(curr_row, 2).font = font_normal
        
        ws2.cell(curr_row, 3, row[2]).alignment = align_left
        ws2.cell(curr_row, 3).border = border_all_thin
        ws2.cell(curr_row, 3).font = font_normal
        
        cell_ans = ws2.cell(curr_row, 4, "[Đánh giá]")
        cell_ans.alignment = align_center
        cell_ans.border = border_all_thin
        cell_ans.fill = fill_input
        cell_ans.font = font_bold
        
        cell_cmt = ws2.cell(curr_row, 5, "[Nhận xét]")
        cell_cmt.alignment = align_left
        cell_cmt.border = border_all_thin
        cell_cmt.fill = fill_input
        cell_cmt.font = font_italic
        
        if curr_row % 2 == 0:
            ws2.cell(curr_row, 1).fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
            ws2.cell(curr_row, 2).fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
            ws2.cell(curr_row, 3).fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
            
        curr_row += 1

    # -------------------------------------------------------------
    # SAVE THE WORKBOOK TO BOTH PATHS
    # -------------------------------------------------------------
    paths = [
        "d:/Sao Vàng/Website-SaoVang/MẪU FOM TUYỂN DỤNG 2024.xlsx",
        "d:/Sao Vàng/Website-SaoVang/MẪU FOM TUYỂN DỤNG 2024_Nang_Cap.xlsx"
    ]
    saved_paths = []
    for path in paths:
        try:
            wb.save(path)
            saved_paths.append("Success")
        except PermissionError:
            print("Permission denied for path. Skipping.")
            
    print("Recruitment Template upgraded successfully.")

if __name__ == "__main__":
    upgrade_recruitment_template()
