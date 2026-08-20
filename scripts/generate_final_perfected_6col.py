import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.drawing.image import Image as OpenpyxlImage

def create_final_perfected_form_with_logo():
    wb = openpyxl.Workbook()
    
    # Define colors
    STEEL_BLUE = "2F5597"
    ICE_BLUE = "D9E1F2"
    LIGHT_GRAY = "F2F2F2"
    BORDER_GRAY = "D9D9D9"
    YELLOW_INPUT = "FFF2CC"
    DARK_GREEN = "385723"
    LIGHT_GREEN = "E2EFDA"
    
    # Fonts
    font_company_name = Font(name="Segoe UI", size=12, bold=True, color=STEEL_BLUE)
    font_company_info = Font(name="Segoe UI", size=9, italic=True, color="595959")
    font_title_main = Font(name="Segoe UI", size=15, bold=True, color=STEEL_BLUE)
    font_title_sub = Font(name="Segoe UI", size=10, italic=True, color="7F7F7F")
    font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    font_group_header = Font(name="Segoe UI", size=11, bold=True, color=STEEL_BLUE)
    font_bold = Font(name="Segoe UI", size=10, bold=True)
    font_normal = Font(name="Segoe UI", size=10)
    font_italic = Font(name="Segoe UI", size=9, italic=True)
    font_salary_result = Font(name="Segoe UI", size=12, bold=True, color=DARK_GREEN)
    
    # Fills
    fill_header = PatternFill(start_color=STEEL_BLUE, end_color=STEEL_BLUE, fill_type="solid")
    fill_group = PatternFill(start_color=ICE_BLUE, end_color=ICE_BLUE, fill_type="solid")
    fill_sum = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
    fill_input = PatternFill(start_color=YELLOW_INPUT, end_color=YELLOW_INPUT, fill_type="solid")
    fill_result_box = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
    
    # Alignments
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    
    # Borders
    thin_side = Side(style="thin", color=BORDER_GRAY)
    border_all_thin = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    border_group = Border(
        left=Side(style="thin", color=STEEL_BLUE),
        right=Side(style="thin", color=STEEL_BLUE),
        top=Side(style="thin", color=STEEL_BLUE),
        bottom=Side(style="thin", color=STEEL_BLUE)
    )
    
    # Column widths (6 Columns layout)
    ws1 = wb.active
    ws1.title = "Đánh Giá Tuyển Dụng"
    ws1.sheet_view.showGridLines = True
    
    ws1.column_dimensions['A'].width = 6   # STT
    ws1.column_dimensions['B'].width = 38  # NHIỆM VỤ ĐÁNH GIÁ (VIỆC CẦN LÀM)
    ws1.column_dimensions['C'].width = 62  # YÊU CẦU KỸ NĂNG VÀ KẾT QUẢ BÊN SDLĐ CẦN ĐẠT ĐƯỢC
    ws1.column_dimensions['D'].width = 16  # ỨNG VIÊN TỰ ĐÁNH GIÁ (% Đáp ứng)
    ws1.column_dimensions['E'].width = 16  # CÔNG TY ĐÁNH GIÁ (Đạt: 1, Chưa: 0)
    ws1.column_dimensions['F'].width = 25  # GHI CHÚ / NHẬN XÉT CHI TIẾT

    # 1. Company Letterhead Row 2-4
    ws1.row_dimensions[2].height = 20
    ws1.row_dimensions[3].height = 15
    ws1.row_dimensions[4].height = 15
    
    # Add Company Name & Info
    ws1.merge_cells("C2:F2")
    ws1["C2"] = "CÔNG TY CỔ PHẦN SẢN XUẤT CƠ KHÍ SAO VÀNG (ASV)"
    ws1["C2"].font = font_company_name
    ws1["C2"].alignment = Alignment(horizontal="left", vertical="center")
    
    ws1.merge_cells("C3:F3")
    ws1["C3"] = "Địa chỉ: Tầng 3, TT7-35 Khu đô thị Văn Phú, phường Kiến Hưng, TP Hà Nội, Việt Nam."
    ws1["C3"].font = font_company_info
    ws1["C3"].alignment = Alignment(horizontal="left", vertical="center")
    
    ws1.merge_cells("C4:F4")
    ws1["C4"] = "Hotline: 0869 590 279  |  Email: cokhisaovangvn@gmail.com"
    ws1["C4"].font = font_company_info
    ws1["C4"].alignment = Alignment(horizontal="left", vertical="center")

    # Add Logo image in A2
    try:
        logo_path = "d:/Sao Vàng/Website-SaoVang/Logo/LogoCTY + SV Aluminium/LogoCTY.png"
        img = OpenpyxlImage(logo_path)
        img.width = 110
        img.height = 45
        ws1.add_image(img, "A2")
    except Exception as e:
        print(f"Error adding logo: {e}")

    # Title Block (shifted to Row 6-7)
    ws1.row_dimensions[6].height = 24
    ws1.merge_cells("A6:F6")
    ws1["A6"] = "FORM TUYỂN DỤNG VÀ ĐÁNH GIÁ NĂNG LỰC THỬ VIỆC THỢ CƠ KHÍ"
    ws1["A6"].font = font_title_main
    ws1["A6"].alignment = align_center

    ws1.row_dimensions[7].height = 18
    ws1.merge_cells("A7:F7")
    ws1["A7"] = "Xác định ngạch bậc & lương ngày dựa trên kỹ năng thực tế của ứng viên"
    ws1["A7"].font = font_title_sub
    ws1["A7"].alignment = align_center

    # Candidate Profile Info (shifted to Row 9-11)
    ws1["B9"] = "Họ và tên ứng viên:"
    ws1["B9"].font = font_bold
    ws1["C9"] = "[Nhập tên ứng viên]"
    ws1["C9"].font = font_normal
    ws1["C9"].border = Border(bottom=Side(style="thin", color="000000"))

    ws1["E9"] = "Ngày phỏng vấn:"
    ws1["E9"].font = font_bold
    ws1["F9"] = "[Nhập ngày]"
    ws1["F9"].font = font_normal
    ws1["F9"].border = Border(bottom=Side(style="thin", color="000000"))

    ws1["B10"] = "Số điện thoại/CCCD:"
    ws1["B10"].font = font_bold
    ws1["C10"] = "[Số điện thoại và CCCD]"
    ws1["C10"].font = font_normal
    ws1["C10"].border = Border(bottom=Side(style="thin", color="000000"))

    ws1["E10"] = "Người chấm điểm:"
    ws1["E10"].font = font_bold
    ws1["F10"] = "[Tên Quản đốc/Tổ trưởng]"
    ws1["F10"].font = font_normal
    ws1["F10"].border = Border(bottom=Side(style="thin", color="000000"))

    ws1["B11"] = "Vị trí ứng tuyển:"
    ws1["B11"].font = font_bold
    ws1["C11"] = "Thợ Cơ khí Dân dụng (Học việc/Phụ/Giúp việc/Thợ chính/Tổ trưởng)"
    ws1["C11"].font = font_italic
    ws1["C11"].border = Border(bottom=Side(style="thin", color="000000"))

    ws1["E11"] = "Kỳ thử việc:"
    ws1["E11"].font = font_bold
    ws1["F11"] = "01 tháng thử việc"
    ws1["F11"].font = font_normal
    ws1["F11"].border = Border(bottom=Side(style="thin", color="000000"))

    ws1.row_dimensions[12].height = 15

    # Table Header (shifted to Row 13)
    headers = [
        ("A13", "STT", align_center),
        ("B13", "NHIỆM VỤ ĐÁNH GIÁ (VIỆC CẦN LÀM)", align_left),
        ("C13", "YÊU CẦU KỸ NĂNG VÀ KẾT QUẢ BÊN SDLĐ CẦN ĐẠT ĐƯỢC", align_left),
        ("D13", "ỨNG VIÊN TỰ ĐÁNH GIÁ (% Đáp ứng)", align_center),
        ("E13", "CÔNG TY ĐÁNH GIÁ (Đạt: 1, Chưa: 0)", align_center),
        ("F13", "GHI CHÚ / NHẬN XÉT CHI TIẾT", align_left)
    ]
    for cell_ref, text, align in headers:
        cell = ws1[cell_ref]
        cell.value = text
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align
        cell.border = border_all_thin
    ws1.row_dimensions[13].height = 28

    # Helper function to write merged headers
    def merge_and_style_header(ws, r_num, title):
        ws.merge_cells(start_row=r_num, start_column=1, end_row=r_num, end_column=6)
        ws.cell(r_num, 1, title).font = font_group_header
        ws.cell(r_num, 1).alignment = align_left
        for col in range(1, 7):
            cell = ws.cell(r_num, col)
            cell.fill = fill_group
            cell.border = border_group

    # Polished tasks list
    distributed_tasks = [
        ("G", "I. HỌC VIỆC VÀ PHỤ VIỆC (Tiêu chuẩn Bậc 1 - Bậc 5)"),
        ("1", "Tác phong, Sức khỏe & Thái độ (Bậc 1)", 
         "Đảm bảo sức khỏe dẻo dai, bền bỉ để làm việc trong môi trường nhà xưởng cơ khí; thái độ trung thực, nhanh nhẹn, chịu khó học hỏi, có ý thức trách nhiệm và phối hợp tốt với đồng nghiệp."),
        
        ("2", "Thực hành Vệ sinh & 5S (Bậc 2)", 
         "Chủ động dọn dẹp máy móc, thiết bị cầm tay gọn gàng sau ca làm việc; quét dọn và vệ sinh sạch sẽ khu vực gia công lắp dựng tại xưởng cũng như khi đi công trình."),
        
        ("3", "Vận hành dụng cụ cầm tay & Mài thô (Bậc 3)", 
         "Sử dụng đúng kỹ thuật, an toàn các dụng cụ cơ bản (máy cắt tay, máy cắt bàn, máy khoan cầm tay, thước đo, thiết bị điện...); có ý thức phòng ngừa rủi ro mất an toàn lao động; mài hoàn thiện bề mặt kết cấu cơ bản (sắt đen, sắt mạ kẽm, inox)."),
        
        ("4", "Kỹ thuật mài tinh mỹ & Đánh bóng (Bậc 4)", 
         "Kỹ thuật gia công mài phẳng tuyệt đối, mài bo góc cạnh sắc nét đạt độ thẩm mỹ cao đối với sản phẩm hoàn thiện; mài xử lý mối hàn thô của kết cấu chịu lực nặng; mài đánh bóng inox đạt độ bóng gương (mirror finish); biết tự mài sắc phục hồi mũi khoan thép."),
        
        ("5", "Sơn phủ hoàn thiện & Hàn đính gá phụ (Bậc 5)", 
         "Vận hành hệ thống phun sơn khí nén, sơn lót chống gỉ và sơn phủ màu bóng mịn đều tay, không chảy sơn; nắm rõ chủng loại vật tư cơ bản; đo cắt phôi chính xác từng mm và làm dưỡng gá loạt phôi; hàn đính gá cơ bản."),
        
        ("G", "II. GIÚP VIỆC (Tiêu chuẩn Bậc 6 - Bậc 7)"),
        ("6", "Chuẩn bị hiện trường & Dụng cụ công trình (Bậc 6)", 
         "Nắm rõ công năng các loại thiết bị cơ khí cầm tay/chuyên dụng; chủ động tính toán, lập danh mục thiết bị phụ trợ và chuẩn bị đầy đủ đồ nghề thi công cần thiết khi đi lắp ráp công trình ngoài hiện trường."),
        
        ("7", "Vận hành máy hàn chuyên nghiệp (Bậc 7)", 
         "Hàn thành thạo các phương pháp hàn hồ quang tay (que MMA), hàn bán tự động MIG trên kết cấu sắt thép chịu lực và hàn TIG trên inox mỏng (mối hàn ngấu sâu, vảy cá đều đẹp, trắng sáng); vận hành thuần thục máy móc xây dựng cơ bản."),
        
        ("G", "III. THỢ CHÍNH - TỰ THỰC HIỆN (Tiêu chuẩn Bậc 8 - Bậc 10)"),
        ("8", "Đọc bản vẽ sản xuất & Bóc tách khai triển (Bậc 8)", 
         "Đọc hiểu bản vẽ thiết kế kỹ thuật cơ khí gia công chế tạo; tự tính toán bóc tách phôi và khai triển kích thước hình học chính xác để gia công sản phẩm thực tế một cách khoa học, tiết kiệm vật tư; đề xuất cải tiến kỹ thuật nâng cao năng suất."),
        
        ("9", "Hoàn thiện & Tự kiểm soát chất lượng QC (Bậc 9)", 
         "Kỹ năng tự kiểm tra chất lượng sản phẩm (Self-QC) sau gia công (độ phẳng, độ vuông góc, tính thẩm mỹ); tinh thần cầu tiến, không giấu dốt, chủ động kiểm tra và tự khắc phục các lỗi kỹ thuật phát sinh."),
        
        ("10", "Dự toán định mức nhân công & Tiến độ (Bậc 10)", 
         "Khả năng đánh giá độ phức tạp của kết cấu bản vẽ; ước lượng chuẩn xác định mức giờ công thợ và lượng nhân lực cần thiết để hoàn thành sản phẩm/công trình theo đúng tiến độ được giao."),
        
        ("G", "IV. TỔ TRƯỞNG VÀ ĐIỀU HÀNH SẢN XUẤT (Tiêu chuẩn Bậc 11 - Bậc 14)"),
        ("11", "Báo cáo tiến độ & Giải quyết phát sinh (Bậc 11)", 
         "Lập báo cáo tiến độ công việc cuối ngày; chủ động đề xuất giải pháp xử lý nhanh các vướng mắc kỹ thuật phát sinh nhằm duy trì sự thông suốt giữa các bộ phận gia công lắp dựng khác."),
        
        ("12", "Độc lập thi công & Quyết định tại hiện trường (Bậc 12)", 
         "Khả năng độc lập thi công các hạng mục khó; chủ động đưa ra quyết định xử lý kỹ thuật nhỏ tại công trình mà không làm ảnh hưởng đến cấu trúc tổng thể và chất lượng sản phẩm."),
        
        ("13", "Làm việc nhóm & Quản lý tiến độ (Bậc 13)", 
         "Khả năng điều phối và làm việc nhóm nhịp nhàng; phân công nhiệm vụ rõ ràng; đảm bảo tiến độ gia công sản xuất đạt chuẩn kỹ thuật bản vẽ và giao hàng đúng hẹn."),
        
        ("14", "Lãnh đạo đội nhóm & Điều hành sản xuất (Bậc 14)", 
         "Năng lực quản lý và lãnh đạo tổ đội từ 3-5 thợ phụ/chính; bao quát tiến độ và chất lượng toàn tổ; điều phối phân chia công việc tối ưu theo tay nghề của từng công nhân khi nhận nhiệm vụ từ Ban Giám đốc.")
    ]

    current_row = 14
    eval_rows = []
    row_bac_map = {}
    
    for row_data in distributed_tasks:
        if row_data[0] == "G":
            ws1.row_dimensions[current_row].height = 24
            merge_and_style_header(ws1, current_row, row_data[1])
        else:
            stt, task, skill = row_data
            ws1.row_dimensions[current_row].height = 36
            
            ws1.cell(current_row, 1, stt).font = font_normal
            ws1.cell(current_row, 1).alignment = align_center
            
            ws1.cell(current_row, 2, task).font = font_bold
            ws1.cell(current_row, 2).alignment = align_left
            
            ws1.cell(current_row, 3, skill).font = font_normal
            ws1.cell(current_row, 3).alignment = align_left
            
            # Candidate Self Eval (%) - Input
            cell_self = ws1.cell(current_row, 4, 1.0)
            cell_self.font = font_bold
            cell_self.alignment = align_center
            cell_self.fill = fill_input
            cell_self.number_format = '0%'
            
            # Company Eval (1: Đạt, 0: Chưa) - Input
            cell_company = ws1.cell(current_row, 5, 0)
            cell_company.font = font_bold
            cell_company.alignment = align_center
            cell_company.fill = fill_input
            eval_rows.append(current_row)
            
            # Note - Input
            cell_note = ws1.cell(current_row, 6, "[Nhận xét nhanh]")
            cell_note.font = font_italic
            cell_note.alignment = align_left
            cell_note.fill = fill_input
            
            for col in range(1, 7):
                ws1.cell(current_row, col).border = border_all_thin
                
            w_map = {
                15: 1, 16: 2, 17: 3, 18: 4, 19: 5,
                21: 6, 22: 7,
                24: 8, 25: 9, 26: 10,
                28: 11, 29: 12, 30: 13, 31: 14
            }
            row_bac_map[current_row] = w_map[current_row]
            
        current_row += 1

    # Place sample test scores (shifted keys)
    ws1["E15"] = 1  # Bậc 1
    ws1["E16"] = 1  # Bậc 2
    ws1["E17"] = 1  # Bậc 3
    ws1["E18"] = 1  # Bậc 4
    ws1["E19"] = 1  # Bậc 5
    ws1["E21"] = 1  # Bậc 6
    ws1["E22"] = 1  # Bậc 7
    ws1["E24"] = 0  # Bậc 8
    ws1["E25"] = 0
    ws1["E26"] = 0
    ws1["E28"] = 0
    ws1["E29"] = 0
    ws1["E30"] = 0
    ws1["E31"] = 0

    ws1.row_dimensions[current_row].height = 15
    current_row += 1

    # RESULTS DASHBOARD BOX
    ws1.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=6)
    r_hdr = ws1.cell(current_row, 2, "KẾT QUẢ PHÂN BẬC LƯƠNG ĐỀ XUẤT SAU THỬ VIỆC")
    r_hdr.font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    r_hdr.fill = PatternFill(start_color=STEEL_BLUE, end_color=STEEL_BLUE, fill_type="solid")
    r_hdr.alignment = align_center
    ws1.row_dimensions[current_row].height = 24
    
    for col in range(2, 7):
        ws1.cell(current_row, col).border = Border(top=Side(style="medium", color=STEEL_BLUE))
        
    current_row += 1
    
    bac_parts = []
    for r_num in eval_rows:
        b_num = row_bac_map[r_num]
        bac_parts.append(f"E{r_num}*{b_num}")
        
    formula_bac = "=MAX(" + ",".join(bac_parts) + ")"
    
    # 1. Bậc năng lực đạt được
    ws1.row_dimensions[current_row].height = 20
    ws1.cell(current_row, 2, "Bậc năng lực đạt được:").font = font_bold
    ws1.cell(current_row, 2).alignment = align_left
    
    cell_bac = ws1.cell(current_row, 3, formula_bac)
    cell_bac.font = font_bold
    cell_bac.alignment = align_center
    
    ws1.cell(current_row, 4, "Bậc / 14").font = font_italic
    ws1.cell(current_row, 4).alignment = align_left
    
    # Style borders
    for col in range(2, 7):
        cell = ws1.cell(current_row, col)
        cell.fill = fill_result_box
        left_s = Side(style="medium", color=STEEL_BLUE) if col == 2 else None
        right_s = Side(style="medium", color=STEEL_BLUE) if col == 6 else None
        cell.border = Border(left=left_s, right=right_s)
        
    bac_cell_ref = f"C{current_row}" # Dynamic cell coordinate
    current_row += 1
    
    # 2. Chức danh tuyển dụng đề xuất
    ws1.row_dimensions[current_row].height = 20
    ws1.cell(current_row, 2, "Chức danh tuyển dụng đề xuất:").font = font_bold
    ws1.cell(current_row, 2).alignment = align_left
    
    cell_title = ws1.cell(current_row, 3, f"=IF({bac_cell_ref}=0, \"Chưa đạt\", VLOOKUP({bac_cell_ref}, 'Thang Đo & Lương'!A$2:E$15, 2, FALSE))")
    cell_title.font = font_bold
    cell_title.alignment = align_center
    
    for col in range(2, 7):
        cell = ws1.cell(current_row, col)
        cell.fill = fill_result_box
        left_s = Side(style="medium", color=STEEL_BLUE) if col == 2 else None
        right_s = Side(style="medium", color=STEEL_BLUE) if col == 6 else None
        cell.border = Border(left=left_s, right=right_s)
        
    current_row += 1
    
    # 3. Mức lương thực lĩnh (ngày)
    ws1.row_dimensions[current_row].height = 20
    ws1.cell(current_row, 2, "Lương thực lĩnh đề xuất (ngày):").font = font_bold
    ws1.cell(current_row, 2).alignment = align_left
    
    cell_wage_d = ws1.cell(current_row, 3, f"=IF({bac_cell_ref}=0, 0, VLOOKUP({bac_cell_ref}, 'Thang Đo & Lương'!A$2:E$15, 3, FALSE))")
    cell_wage_d.font = font_salary_result
    cell_wage_d.alignment = align_center
    cell_wage_d.number_format = '#,##0" đ"'
    
    ws1.cell(current_row, 4, "(Lương sau thuế/ngày)").font = font_italic
    ws1.cell(current_row, 4).alignment = align_left
    
    for col in range(2, 7):
        cell = ws1.cell(current_row, col)
        cell.fill = fill_result_box
        left_s = Side(style="medium", color=STEEL_BLUE) if col == 2 else None
        right_s = Side(style="medium", color=STEEL_BLUE) if col == 6 else None
        cell.border = Border(left=left_s, right=right_s)
        
    wage_d_ref = f"C{current_row}"
    current_row += 1
    
    # 4. Tổng chi phí chi trả (ngày)
    ws1.row_dimensions[current_row].height = 20
    ws1.cell(current_row, 2, "Tổng chi phí chi trả (ngày):").font = font_bold
    ws1.cell(current_row, 2).alignment = align_left
    
    cell_wage_t = ws1.cell(current_row, 3, f"=IF({bac_cell_ref}=0, 0, VLOOKUP({bac_cell_ref}, 'Thang Đo & Lương'!A$2:E$15, 4, FALSE))")
    cell_wage_t.font = font_bold
    cell_wage_t.alignment = align_center
    cell_wage_t.number_format = '#,##0" đ"'
    
    ws1.cell(current_row, 4, "(Tổng chi phí ngày công)").font = font_italic
    ws1.cell(current_row, 4).alignment = align_left
    
    for col in range(2, 7):
        cell = ws1.cell(current_row, col)
        cell.fill = fill_result_box
        left_s = Side(style="medium", color=STEEL_BLUE) if col == 2 else None
        right_s = Side(style="medium", color=STEEL_BLUE) if col == 6 else None
        cell.border = Border(left=left_s, right=right_s)
        
    current_row += 1
    
    # 5. Dự toán thu nhập tháng (26 ngày công)
    ws1.row_dimensions[current_row].height = 20
    ws1.cell(current_row, 2, "Thu nhập tháng dự kiến (26 công):").font = font_bold
    ws1.cell(current_row, 2).alignment = align_left
    
    cell_wage_m = ws1.cell(current_row, 3, f"={wage_d_ref}*26")
    cell_wage_m.font = font_salary_result
    cell_wage_m.alignment = align_center
    cell_wage_m.number_format = '#,##0" đ"'
    
    ws1.cell(current_row, 4, "(Tương ứng 15-22 triệu/tháng)").font = font_italic
    ws1.cell(current_row, 4).alignment = align_left
    
    for col in range(2, 7):
        cell = ws1.cell(current_row, col)
        cell.fill = fill_result_box
        left_s = Side(style="medium", color=STEEL_BLUE) if col == 2 else None
        right_s = Side(style="medium", color=STEEL_BLUE) if col == 6 else None
        cell.border = Border(left=left_s, right=right_s)
        
    current_row += 1
    
    # 6. Khuyến nghị và Kế hoạch sử dụng
    ws1.row_dimensions[current_row].height = 24
    ws1.cell(current_row, 2, "Khuyến nghị & Phân công:").font = font_bold
    ws1.cell(current_row, 2).alignment = align_left
    
    ws1.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=6)
    rec_cell = ws1.cell(current_row, 3, f"=IF({bac_cell_ref}=0, \"Yêu cầu đào tạo thêm tay nghề hoặc từ chối thử việc\", VLOOKUP({bac_cell_ref}, 'Thang Đo & Lương'!A$2:E$15, 5, FALSE))")
    rec_cell.font = font_bold
    rec_cell.alignment = align_left
    
    for col in range(2, 7):
        cell = ws1.cell(current_row, col)
        cell.fill = fill_result_box
        left_s = Side(style="medium", color=STEEL_BLUE) if col == 2 else None
        right_s = Side(style="medium", color=STEEL_BLUE) if col == 6 else None
        cell.border = Border(left=left_s, right=right_s, bottom=Side(style="medium", color=STEEL_BLUE))

    current_row += 2

    # Slogan block matching company motto
    ws1.row_dimensions[current_row].height = 26
    ws1.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    slogan_cell = ws1.cell(current_row, 1, 'Khẩu hiệu của chúng tôi: "Chất Lượng Tạo Nên Thương Hiệu. Uy Tín Tạo Nên Sự Thành Công"')
    slogan_cell.font = Font(name="Segoe UI", size=10, bold=True, italic=True, color=STEEL_BLUE)
    slogan_cell.alignment = align_center
    slogan_cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Border for slogan
    for col in range(1, 7):
        ws1.cell(current_row, col).border = Border(
            top=Side(style="thin", color=STEEL_BLUE),
            bottom=Side(style="thin", color=STEEL_BLUE),
            left=Side(style="thin", color=STEEL_BLUE) if col == 1 else None,
            right=Side(style="thin", color=STEEL_BLUE) if col == 6 else None
        )

    current_row += 3
    
    # Signatures
    ws1.row_dimensions[current_row].height = 18
    ws1.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    s1 = ws1.cell(current_row, 1, "ỨNG VIÊN TUYỂN DỤNG")
    s1.font = font_bold
    s1.alignment = align_center
    
    ws1.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=6)
    s2 = ws1.cell(current_row, 4, "QUẢN ĐỐC / TỔ TRƯỞNG DUYỆT")
    s2.font = font_bold
    s2.alignment = align_center
    
    current_row += 1
    ws1.row_dimensions[current_row].height = 15
    ws1.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    sub1 = ws1.cell(current_row, 1, "(Ký và ghi rõ họ tên)")
    sub1.font = font_italic
    sub1.alignment = align_center
    
    ws1.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=6)
    sub2 = ws1.cell(current_row, 4, "(Xác nhận ngạch bậc lương ngày)")
    sub2.font = font_italic
    sub2.alignment = align_center
    
    current_row += 4
    ws1.row_dimensions[current_row].height = 18
    ws1.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    n1 = ws1.cell(current_row, 1, "[Họ tên ứng viên]")
    n1.font = font_normal
    n1.alignment = align_center
    
    ws1.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=6)
    n2 = ws1.cell(current_row, 4, "[Họ tên người đánh giá]")
    n2.font = font_normal
    n2.alignment = align_center

    # Data validations
    dv_self = DataValidation(type="decimal", operator="between", formula1="0", formula2="1", allow_blank=True)
    dv_self.prompt = 'Nhập % tự đánh giá (Ví dụ: 80%)'
    dv_self.promptTitle = 'Ứng viên tự đánh giá'
    ws1.add_data_validation(dv_self)
    
    dv_company = DataValidation(type="whole", operator="between", formula1="0", formula2="1", allow_blank=True)
    dv_company.prompt = 'Nhập 1: Đạt, 0: Chưa đạt'
    dv_company.promptTitle = 'Quản đốc đánh giá'
    ws1.add_data_validation(dv_company)

    for r_num in eval_rows:
        dv_self.add(ws1[f"D{r_num}"])
        dv_company.add(ws1[f"E{r_num}"])

    # -------------------------------------------------------------
    # SHEET 2: THANG ĐO & LƯƠNG
    # -------------------------------------------------------------
    ws2 = wb.create_sheet(title="Thang Đo & Lương")
    ws2.sheet_view.showGridLines = True
    
    ws2.column_dimensions['A'].width = 10
    ws2.column_dimensions['B'].width = 24
    ws2.column_dimensions['C'].width = 24
    ws2.column_dimensions['D'].width = 24
    ws2.column_dimensions['E'].width = 65
    
    # Headers
    ws2["A1"] = "Bậc"
    ws2["B1"] = "Chức danh đề xuất"
    ws2["C1"] = "Lương thực lĩnh (ngày)"
    ws2["D1"] = "Tổng chi phí công ty (ngày)"
    ws2["E1"] = "Phân công / Vai trò chính trong xưởng"
    
    for col in ["A", "B", "C", "D", "E"]:
        cell = ws2[f"{col}1"]
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_all_thin
        
    scale_data = [
        (1, "Thợ học việc (Cấp 1)", 220000, 300000, "Phụ việc vệ sinh, mài thô, nâng hạ phôi dưới sự chỉ dẫn"),
        (2, "Thợ học việc (Cấp 2)", 220000, 300000, "Thực hiện tốt 5S khu vực hàn gá, đảm bảo an toàn ATLĐ xưởng"),
        (3, "Thợ phụ cơ bản (Cấp 1)", 255000, 335000, "Biết dùng máy cắt cầm tay, khoan tay, thước đo chính xác"),
        (4, "Thợ phụ mài tinh (Cấp 2)", 290000, 370000, "Mài tinh mỹ bề mặt, mài lấy góc sắc nét, mài phá bavia"),
        (5, "Thợ sơn & gá phụ", 350000, 430000, "Đứng buồng sơn phun khí nén; nắm vật tư và hỗ trợ hàn đính gá"),
        (6, "Thợ gá lắp phôi", 380000, 460000, "Nắm rõ các loại vật tư, phôi cắt sắt/inox, chuẩn bị công trình"),
        (7, "Thợ gá hàn liên hợp", 405000, 485000, "Sử dụng thành thạo máy hàn que/MIG/TIG để đính gá kết cấu"),
        (8, "Thợ hàn chính chuyên nghiệp", 430000, 510000, "Đọc bản vẽ sản xuất và tự hàn độc lập sản phẩm chất lượng cao"),
        (9, "Thợ lắp ráp lành nghề", 440000, 520000, "Đọc bản vẽ cơ khí phức tạp, tự bóc tách khai triển phôi gá ráp"),
        (10, "Thợ hoàn thiện độc lập", 470000, 550000, "Tự kiểm soát chất lượng hoàn thiện sản phẩm không cần giám sát"),
        (11, "Thợ dự toán & tiến độ", 480000, 560000, "Ước lượng thời gian/nhân công, lập tiến độ thi công dự án nhỏ"),
        (12, "Tổ phó kỹ thuật", 490000, 570000, "Tự chủ động điều chỉnh chi tiết nhỏ ở hiện trường, thay thế tổ trưởng"),
        (13, "Trưởng nhóm lắp dựng", 500000, 580000, "Điều phối nhóm 2-3 người lắp dựng kết cấu thép/inox ngoài công trường"),
        (14, "Tổ trưởng sản xuất", 550000, 630000, "Phân việc toàn xưởng, chịu trách nhiệm tiến độ & chất lượng tổ đội")
    ]
    
    for r_idx, row_val in enumerate(scale_data, start=2):
        ws2.cell(r_idx, 1, row_val[0]).alignment = align_center
        ws2.cell(r_idx, 1).font = font_bold
        
        ws2.cell(r_idx, 2, row_val[1]).alignment = align_left
        ws2.cell(r_idx, 2).font = font_bold
        
        ws2.cell(r_idx, 3, row_val[2]).number_format = '#,##0" đ"'
        ws2.cell(r_idx, 3).alignment = align_right
        ws2.cell(r_idx, 3).font = font_bold
        
        ws2.cell(r_idx, 4, row_val[3]).number_format = '#,##0" đ"'
        ws2.cell(r_idx, 4).alignment = align_right
        ws2.cell(r_idx, 4).font = font_normal
        
        ws2.cell(r_idx, 5, row_val[4]).alignment = align_left
        ws2.cell(r_idx, 5).font = font_normal
        
        for col in range(1, 6):
            cell = ws2.cell(r_idx, col)
            cell.border = border_all_thin
            if r_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

    paths = [
        "d:/Sao Vàng/Website-SaoVang/CKSV_Form_Tuyen_Dung_Tho_Co_Khi.xlsx",
        "d:/Sao Vàng/Website-SaoVang/CKSV_Form_Tuyen_Dung_Tho_Co_Khi_Phan_Bo.xlsx"
    ]
    saved_paths = []
    for path in paths:
        try:
            wb.save(path)
            saved_paths.append(path)
        except PermissionError:
            print(f"Permission denied for {path}. Skipping.")
            
    print(f"Recruitment Form with logo saved to: {saved_paths}")

if __name__ == "__main__":
    create_final_perfected_form_with_logo()
