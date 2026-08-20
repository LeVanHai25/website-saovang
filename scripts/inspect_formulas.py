import openpyxl

def inspect_formulas():
    wb = openpyxl.load_workbook("d:/Sao Vàng/Website-SaoVang/CKSV_Form_Tuyen_Dung_Tho_Co_Khi.xlsx", data_only=False)
    ws = wb["Đánh Giá Tuyển Dụng"]
    
    wb_data = openpyxl.load_workbook("d:/Sao Vàng/Website-SaoVang/CKSV_Form_Tuyen_Dung_Tho_Co_Khi.xlsx", data_only=True)
    ws_data = wb_data["Đánh Giá Tuyển Dụng"]
    
    lines = []
    lines.append("FORMULAS:")
    for cell in ['C31', 'C32', 'C33', 'C34', 'C35', 'C36']:
        lines.append(f"{cell}: {ws[cell].value}")
        
    lines.append("\nCACHED VALUES:")
    for cell in ['C31', 'C32', 'C33', 'C34', 'C35', 'C36']:
        lines.append(f"{cell}: {ws_data[cell].value}")
        
    lines.append("\nE CELLS:")
    for r in [12, 13, 14, 15, 16, 18, 19, 21, 22, 23, 25, 26, 27, 28]:
        lines.append(f"E{r}: value={ws_data[f'E{r}'].value}, type={type(ws_data[f'E{r}'].value)}")
        
    with open("d:/Sao Vàng/Website-SaoVang/scripts/formula_inspect_results.txt", "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    print("Done")

if __name__ == "__main__":
    inspect_formulas()
