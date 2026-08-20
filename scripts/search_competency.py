import os
import openpyxl

def extract_competency_details():
    base_dir = "d:/Sao Vàng/Website-SaoVang/CKSV"
    sub_dir = os.path.join(base_dir, "bảng duyển dụng")
    
    targets = [
        # (File path, Sheet name)
        (os.path.join(base_dir, "3, CKSV, Nhiệm vụ Tổng đội thi công - 10.9.24.xlsx"), "Sheet1"),
        (os.path.join(sub_dir, "CKSV THANG BẢNG LƯƠNG & FOM TUYỂN DỤNG.xlsx"), "CÔNG NHÂN, TỔ TRƯỞNG"),
        (os.path.join(sub_dir, "CKSV, Tự đánh gia bản thân - 31.7.24.xlsx"), "Năng lực 01.8.24"),
        (os.path.join(base_dir, "00, VAD, Mẫu thẻ điểm --.xlsx"), "KTNB, 10.3.22, 17g")
    ]
    
    output_lines = []
    output_lines.append("COMPETENCY AND EVALUATION CRITERIA FOUND IN EXCEL FILES:")
    
    for filepath, sname in targets:
        if not os.path.exists(filepath):
            output_lines.append(f"\nFile not found: {filepath}")
            continue
            
        output_lines.append("\n" + "="*80)
        output_lines.append(f"FILE: {os.path.basename(filepath)} | SHEET: {sname}")
        output_lines.append("="*80)
        
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            if sname not in wb.sheetnames:
                output_lines.append(f"Sheet {sname} not found. Available: {wb.sheetnames}")
                continue
                
            ws = wb[sname]
            # Read up to 100 rows to extract all criteria
            rows = list(ws.iter_rows(max_row=100, max_col=15, values_only=True))
            for r_idx, r in enumerate(rows):
                if any(x is not None for x in r):
                    clean_row = []
                    for val in r:
                        if val is not None:
                            # Clean up line breaks for display
                            clean_row.append(str(val).replace('\n', ' '))
                        else:
                            clean_row.append("")
                    while clean_row and clean_row[-1] == "":
                        clean_row.pop()
                    if clean_row:
                        output_lines.append(f"  Row {r_idx+1}: {clean_row}")
        except Exception as e:
            output_lines.append(f"Error reading {os.path.basename(filepath)} sheet {sname}: {e}")
            
    output_path = "d:/Sao Vàng/Website-SaoVang/scripts/competency_details.txt"
    with open(output_path, "w", encoding="utf-8") as f_out:
        f_out.write("\n".join(output_lines))
    print(f"Competency details written to {output_path}")

if __name__ == "__main__":
    extract_competency_details()
